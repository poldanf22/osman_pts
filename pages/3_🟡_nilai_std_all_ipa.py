import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill
import tempfile
from PIL import Image

# menghilangkan hamburger
st.markdown("""
<style>
.css-1rs6os.edgvbvh3
{
    visibility:hidden;
}
.css-1lsmgbg.egzxvld0
{
    visibility:hidden;
}
</style>
""", unsafe_allow_html=True)

image = Image.open('logo resmi nf resize.png')
st.image(image)

st.title("Olah Nilai Standar K13")
st.header("SMA, PPLS, RONIN IPA")

col6 = st.container()

with col6:
    KELAS = st.selectbox(
        "KELAS",
        ("--Pilih Kelas--", "10 SMA IPA", "11 SMA IPA", "PPLS IPA", "RONIN IPA"))

col7 = st.container()

with col7:
    SEMESTER = st.selectbox(
        "SEMESTER",
        ("--Pilih Semester--", "SEMESTER 1", "SEMESTER 2"))

col8 = st.container()

with col8:
    PENILAIAN = st.selectbox(
        "PENILAIAN",
        ("--Pilih Penilaian--", "PENILAIAN TENGAH SEMESTER", "PENILAIAN AKHIR SEMESTER", "PENILAIAN AKHIR TAHUN", "TES EVALUASI"))

col9 = st.container()

with col9:
    KURIKULUM = st.selectbox(
        "KURIKULUM",
        ("--Pilih Kurikulum--", "K13"))

TAHUN = st.text_input("Masukkan Tahun Ajaran", placeholder="contoh: 2022-2023")

col1, col2, col3, col4 = st.columns(4)

with col1:
    MTK = st.selectbox(
        "JML. SOAL MAT.",
        (15, 20, 25, 30, 35, 40, 45, 50))

with col2:
    BIO = st.selectbox(
        "JML. SOAL BIO.",
        (15, 20, 25, 30, 35, 40, 45, 50))

with col3:
    FIS = st.selectbox(
        "JML. SOAL FIS.",
        (15, 20, 25, 30, 35, 40, 45, 50))

with col4:
    KIM = st.selectbox(
        "JML. SOAL KIM.",
        (15, 20, 25, 30, 35, 40, 45, 50))


JML_SOAL_MAT = MTK
JML_SOAL_BIO = BIO
JML_SOAL_FIS = FIS
JML_SOAL_KIM = KIM

uploaded_file = st.file_uploader(
    'Letakkan file excel IPA', type='xlsx')

if uploaded_file is not None:
    wb = openpyxl.load_workbook(uploaded_file)
    ws = wb['Sheet1']

    q = len(ws['K'])
    r = len(ws['K'])+2
    s = len(ws['K'])+3
    t = len(ws['K'])+4
    u = len(ws['K'])+5
    v = len(ws['K'])+6
    w = len(ws['K'])+7
    x = len(ws['K'])+8

    ws['G{}'.format(r)] = "=ROUND(AVERAGE(G2:G{}),2)".format(q)
    ws['H{}'.format(r)] = "=ROUND(AVERAGE(H2:H{}),2)".format(q)
    ws['I{}'.format(r)] = "=ROUND(AVERAGE(I2:I{}),2)".format(q)
    ws['J{}'.format(r)] = "=ROUND(AVERAGE(J2:J{}),2)".format(q)
    ws['K{}'.format(r)] = "=ROUND(AVERAGE(K2:K{}),2)".format(q)
    ws['G{}'.format(s)] = "=STDEV(G2:G{})".format(q)
    ws['H{}'.format(s)] = "=STDEV(H2:H{})".format(q)
    ws['I{}'.format(s)] = "=STDEV(I2:I{})".format(q)
    ws['J{}'.format(s)] = "=STDEV(J2:J{})".format(q)
    ws['G{}'.format(t)] = "=MAX(G2:G{})".format(q)
    ws['H{}'.format(t)] = "=MAX(H2:H{})".format(q)
    ws['I{}'.format(t)] = "=MAX(I2:I{})".format(q)
    ws['J{}'.format(t)] = "=MAX(J2:J{})".format(q)
    ws['K{}'.format(t)] = "=MAX(K2:K{})".format(q)
    ws['L{}'.format(r)] = "=MAX(L2:L{})".format(q)
    ws['M{}'.format(r)] = "=MAX(M2:M{})".format(q)
    ws['N{}'.format(r)] = "=MAX(N2:N{})".format(q)
    ws['O{}'.format(r)] = "=MAX(O2:O{})".format(q)
    ws['P{}'.format(r)] = "=MAX(P2:P{})".format(q)
    ws['Q{}'.format(r)] = "=MAX(Q2:Q{})".format(q)
    ws['R{}'.format(r)] = "=MAX(R2:R{})".format(q)
    ws['S{}'.format(r)] = "=MAX(S2:S{})".format(q)
    ws['T{}'.format(r)] = "=ROUND(MAX(T2:T{}),2)".format(q)
    ws['U{}'.format(r)] = "=MAX(U2:U{})".format(q)
    ws['G{}'.format(u)] = "=MIN(G2:G{})".format(q)
    ws['H{}'.format(u)] = "=MIN(H2:H{})".format(q)
    ws['I{}'.format(u)] = "=MIN(I2:I{})".format(q)
    ws['J{}'.format(u)] = "=MIN(J2:J{})".format(q)
    ws['K{}'.format(u)] = "=MIN(K2:K{})".format(q)
    ws['P{}'.format(s)] = "=MIN(P2:P{})".format(q)
    ws['Q{}'.format(s)] = "=MIN(Q2:R{})".format(q)
    ws['R{}'.format(s)] = "=MIN(R2:S{})".format(q)
    ws['S{}'.format(s)] = "=MIN(S2:T{})".format(q)
    ws['T{}'.format(s)] = "=MIN(T2:T{})".format(q)
    ws['P{}'.format(t)] = "=ROUND(AVERAGE(P2:P{}),2)".format(q)
    ws['Q{}'.format(t)] = "=ROUND(AVERAGE(Q2:Q{}),2)".format(q)
    ws['R{}'.format(t)] = "=ROUND(AVERAGE(R2:R{}),2)".format(q)
    ws['S{}'.format(t)] = "=ROUND(AVERAGE(S2:S{}),2)".format(q)
    ws['T{}'.format(t)] = "=ROUND(AVERAGE(T2:T{}),2)".format(q)
    ws['W{}'.format(r)] = "=SUM(W2:W{})".format(q)
    ws['X{}'.format(r)] = "=SUM(X2:X{})".format(q)
    ws['Y{}'.format(r)] = "=SUM(Y2:Y{})".format(q)
    ws['Z{}'.format(r)] = "=SUM(Z2:Z{})".format(q)

    # new
    # iterasi 1 rata-rata - 1

    # MAPEL NORMAL
    ws['AG{}'.format(r)] = "=IF($W${}=0,$G${},$G${}-1)".format(r, r, r)
    ws['AG{}'.format(s)] = "=STDEV(AG2:AG{})".format(q)
    ws['AG{}'.format(t)] = "=MAX(AG2:AG{})".format(q)
    ws['AG{}'.format(u)] = "=MIN(AG2:AG{})".format(q)
    ws['AH{}'.format(r)] = "=IF($X${}=0,$H${},$H${}-1)".format(r, r, r)
    ws['AH{}'.format(s)] = "=STDEV(AH2:AH{})".format(q)
    ws['AH{}'.format(t)] = "=MAX(AH2:AH{})".format(q)
    ws['AH{}'.format(u)] = "=MIN(AH2:AH{})".format(q)
    ws['AI{}'.format(r)] = "=IF($Y${}=0,$I${},$I${}-1)".format(r, r, r)
    ws['AI{}'.format(s)] = "=STDEV(AI2:AI{})".format(q)
    ws['AI{}'.format(t)] = "=MAX(AI2:AI{})".format(q)
    ws['AI{}'.format(u)] = "=MIN(AI2:AI{})".format(q)
    ws['AJ{}'.format(r)] = "=IF($Z${}=0,$J${},$J${}-1)".format(r, r, r)
    ws['AJ{}'.format(s)] = "=STDEV(AJ2:AJ{})".format(q)
    ws['AJ{}'.format(t)] = "=MAX(AJ2:AJ{})".format(q)
    ws['AJ{}'.format(u)] = "=MIN(AJ2:AJ{})".format(q)

    # JUMLAH MAPEL NORMAL
    ws['AK{}'.format(r)] = "=ROUND(AVERAGE(AK2:AK{}),2)".format(q)
    ws['AK{}'.format(t)] = "=MAX(AK2:AK{})".format(q)
    ws['AK{}'.format(u)] = "=MIN(AK2:AK{})".format(q)

    # Z SCORE
    ws['AL{}'.format(r)] = "=MAX(AL2:AL{})".format(q)
    ws['AM{}'.format(r)] = "=MAX(AM2:AM{})".format(q)
    ws['AN{}'.format(r)] = "=MAX(AN2:AN{})".format(q)
    ws['AO{}'.format(r)] = "=MAX(AO2:AO{})".format(q)

    # NILAI STANDAR
    ws['AP{}'.format(r)] = "=MAX(AP2:AP{})".format(q)
    ws['AP{}'.format(s)] = "=MIN(AP2:AP{})".format(q)
    ws['AP{}'.format(t)] = "=ROUND(AVERAGE(AP2:AP{}),2)".format(q)
    ws['AQ{}'.format(r)] = "=MAX(AQ2:AQ{})".format(q)
    ws['AQ{}'.format(s)] = "=MIN(AQ2:AQ{})".format(q)
    ws['AQ{}'.format(t)] = "=ROUND(AVERAGE(AQ2:AQ{}),2)".format(q)
    ws['AR{}'.format(r)] = "=MAX(AR2:AR{})".format(q)
    ws['AR{}'.format(s)] = "=MIN(AR2:AR{})".format(q)
    ws['AR{}'.format(t)] = "=ROUND(AVERAGE(AR2:AR{}),2)".format(q)
    ws['AS{}'.format(r)] = "=MAX(AS2:AS{})".format(q)
    ws['AS{}'.format(s)] = "=MIN(AS2:AS{})".format(q)
    ws['AS{}'.format(t)] = "=ROUND(AVERAGE(AS2:AS{}),2)".format(q)
    ws['AT{}'.format(r)] = "=MAX(AT2:AT{})".format(q)
    ws['AT{}'.format(s)] = "=MIN(AT2:AT{})".format(q)
    ws['AT{}'.format(t)] = "=ROUND(AVERAGE(AT2:AT{}),2)".format(q)

    # INISIASI MAPEL
    ws['AW{}'.format(r)] = "=SUM(AW2:AW{})".format(q)
    ws['AX{}'.format(r)] = "=SUM(AX2:AX{})".format(q)
    ws['AY{}'.format(r)] = "=SUM(AY2:AY{})".format(q)
    ws['AZ{}'.format(r)] = "=SUM(AZ2:AZ{})".format(q)

    # iterasi 2 rata-rata - 1
    # MAPEL NORMAL
    ws['BG{}'.format(r)] = "=IF($AW${}=0,$AG${},$AG${}-1)".format(r, r, r)
    ws['BG{}'.format(s)] = "=STDEV(BG2:BG{})".format(q)
    ws['BG{}'.format(t)] = "=MAX(BG2:BG{})".format(q)
    ws['BG{}'.format(u)] = "=MIN(BG2:BG{})".format(q)
    ws['BH{}'.format(r)] = "=IF($AX${}=0,$AH${},$AH${}-1)".format(r, r, r)
    ws['BH{}'.format(s)] = "=STDEV(BH2:BH{})".format(q)
    ws['BH{}'.format(t)] = "=MAX(BH2:BH{})".format(q)
    ws['BH{}'.format(u)] = "=MIN(BH2:BH{})".format(q)
    ws['BI{}'.format(r)] = "=IF($AY${}=0,$AI${},$AI${}-1)".format(r, r, r)
    ws['BI{}'.format(s)] = "=STDEV(BI2:BI{})".format(q)
    ws['BI{}'.format(t)] = "=MAX(BI2:BI{})".format(q)
    ws['BI{}'.format(u)] = "=MIN(BI2:BI{})".format(q)
    ws['BJ{}'.format(r)] = "=IF($AZ${}=0,$AJ${},$AJ${}-1)".format(r, r, r)
    ws['BJ{}'.format(s)] = "=STDEV(BJ2:BJ{})".format(q)
    ws['BJ{}'.format(t)] = "=MAX(BJ2:BJ{})".format(q)
    ws['BJ{}'.format(u)] = "=MIN(BJ2:BJ{})".format(q)

    # JUMLAH MAPEL NORMAL
    ws['BK{}'.format(r)] = "=ROUND(AVERAGE(BK2:BK{}),2)".format(q)
    ws['BK{}'.format(t)] = "=MAX(BK2:BK{})".format(q)
    ws['BK{}'.format(u)] = "=MIN(BK2:BK{})".format(q)

    # Z SCORE
    ws['BL{}'.format(r)] = "=MAX(BL2:BL{})".format(q)
    ws['BM{}'.format(r)] = "=MAX(BM2:BM{})".format(q)
    ws['BN{}'.format(r)] = "=MAX(BN2:BN{})".format(q)
    ws['BO{}'.format(r)] = "=MAX(BO2:BO{})".format(q)

    # NILAI STANDAR
    ws['BP{}'.format(r)] = "=MAX(BP2:BP{})".format(q)
    ws['BP{}'.format(s)] = "=MIN(BP2:BP{})".format(q)
    ws['BP{}'.format(t)] = "=ROUND(AVERAGE(BP2:BP{}),2)".format(q)
    ws['BQ{}'.format(r)] = "=MAX(BQ2:BQ{})".format(q)
    ws['BQ{}'.format(s)] = "=MIN(BQ2:BQ{})".format(q)
    ws['BQ{}'.format(t)] = "=ROUND(AVERAGE(BQ2:BQ{}),2)".format(q)
    ws['BR{}'.format(r)] = "=MAX(BR2:BR{})".format(q)
    ws['BR{}'.format(s)] = "=MIN(BR2:BR{})".format(q)
    ws['BR{}'.format(t)] = "=ROUND(AVERAGE(BR2:BR{}),2)".format(q)
    ws['BS{}'.format(r)] = "=MAX(BS2:BS{})".format(q)
    ws['BS{}'.format(s)] = "=MIN(BS2:BS{})".format(q)
    ws['BS{}'.format(t)] = "=ROUND(AVERAGE(BS2:BS{}),2)".format(q)
    ws['BT{}'.format(r)] = "=MAX(BT2:BT{})".format(q)
    ws['BT{}'.format(s)] = "=MIN(BT2:BT{})".format(q)
    ws['BT{}'.format(t)] = "=ROUND(AVERAGE(BT2:BT{}),2)".format(q)

    # INISIASI MAPEL
    ws['BW{}'.format(r)] = "=SUM(BW2:BW{})".format(q)
    ws['BX{}'.format(r)] = "=SUM(BX2:BX{})".format(q)
    ws['BY{}'.format(r)] = "=SUM(BY2:BY{})".format(q)
    ws['BZ{}'.format(r)] = "=SUM(BZ2:BZ{})".format(q)

    # iterasi 3 rata-rata - 1
    # MAPEL NORMAL
    ws['CG{}'.format(r)] = "=IF($BW${}=0,$BG${},$BG${}-1)".format(r, r, r)
    ws['CG{}'.format(s)] = "=STDEV(CG2:CG{})".format(q)
    ws['CG{}'.format(t)] = "=MAX(CG2:CG{})".format(q)
    ws['CG{}'.format(u)] = "=MIN(CG2:CG{})".format(q)
    ws['CH{}'.format(r)] = "=IF($BX${}=0,$BH${},$BH${}-1)".format(r, r, r)
    ws['CH{}'.format(s)] = "=STDEV(CH2:CH{})".format(q)
    ws['CH{}'.format(t)] = "=MAX(CH2:CH{})".format(q)
    ws['CH{}'.format(u)] = "=MIN(CH2:CH{})".format(q)
    ws['CI{}'.format(r)] = "=IF($BY${}=0,$BI${},$BI${}-1)".format(r, r, r)
    ws['CI{}'.format(s)] = "=STDEV(CI2:CI{})".format(q)
    ws['CI{}'.format(t)] = "=MAX(CI2:CI{})".format(q)
    ws['CI{}'.format(u)] = "=MIN(CI2:CI{})".format(q)
    ws['CJ{}'.format(r)] = "=IF($BZ${}=0,$BJ${},$BJ${}-1)".format(r, r, r)
    ws['CJ{}'.format(s)] = "=STDEV(CJ2:CJ{})".format(q)
    ws['CJ{}'.format(t)] = "=MAX(CJ2:CJ{})".format(q)
    ws['CJ{}'.format(u)] = "=MIN(CJ2:CJ{})".format(q)

    # JUMLAH MAPEL NORMAL
    ws['CK{}'.format(r)] = "=ROUND(AVERAGE(CK2:CK{}),2)".format(q)
    ws['CK{}'.format(t)] = "=MAX(CK2:CK{})".format(q)
    ws['CK{}'.format(u)] = "=MIN(CK2:CK{})".format(q)

    # Z SCORE
    ws['CL{}'.format(r)] = "=MAX(CL2:CL{})".format(q)
    ws['CM{}'.format(r)] = "=MAX(CM2:CM{})".format(q)
    ws['CN{}'.format(r)] = "=MAX(CN2:CN{})".format(q)
    ws['CO{}'.format(r)] = "=MAX(CO2:CO{})".format(q)

    # NILAI STANDAR
    ws['CP{}'.format(r)] = "=MAX(CP2:CP{})".format(q)
    ws['CP{}'.format(s)] = "=MIN(CP2:CP{})".format(q)
    ws['CP{}'.format(t)] = "=ROUND(AVERAGE(CP2:CP{}),2)".format(q)
    ws['CQ{}'.format(r)] = "=MAX(CQ2:CQ{})".format(q)
    ws['CQ{}'.format(s)] = "=MIN(CQ2:CQ{})".format(q)
    ws['CQ{}'.format(t)] = "=ROUND(AVERAGE(CQ2:CQ{}),2)".format(q)
    ws['CR{}'.format(r)] = "=MAX(CR2:CR{})".format(q)
    ws['CR{}'.format(s)] = "=MIN(CR2:CR{})".format(q)
    ws['CR{}'.format(t)] = "=ROUND(AVERAGE(CR2:CR{}),2)".format(q)
    ws['CS{}'.format(r)] = "=MAX(CS2:CS{})".format(q)
    ws['CS{}'.format(s)] = "=MIN(CS2:CS{})".format(q)
    ws['CS{}'.format(t)] = "=ROUND(AVERAGE(CS2:CS{}),2)".format(q)
    ws['CT{}'.format(r)] = "=MAX(CT2:CT{})".format(q)
    ws['CT{}'.format(s)] = "=MIN(CT2:CT{})".format(q)
    ws['CT{}'.format(t)] = "=ROUND(AVERAGE(CT2:CT{}),2)".format(q)

    # INISIASI MAPEL
    ws['CW{}'.format(r)] = "=SUM(CW2:CW{})".format(q)
    ws['CX{}'.format(r)] = "=SUM(CX2:CX{})".format(q)
    ws['CY{}'.format(r)] = "=SUM(CY2:CY{})".format(q)
    ws['CZ{}'.format(r)] = "=SUM(CZ2:CZ{})".format(q)

    # iterasi 4 rata-rata - 1
    # MAPEL NORMAL
    ws['DG{}'.format(r)] = "=IF($CW${}=0,$CG${},$CG${}-1)".format(r, r, r)
    ws['DG{}'.format(s)] = "=STDEV(DG2:DG{})".format(q)
    ws['DG{}'.format(t)] = "=MAX(DG2:DG{})".format(q)
    ws['DG{}'.format(u)] = "=MIN(DG2:DG{})".format(q)
    ws['DH{}'.format(r)] = "=IF($CX${}=0,$CH${},$CH${}-1)".format(r, r, r)
    ws['DH{}'.format(s)] = "=STDEV(DH2:DH{})".format(q)
    ws['DH{}'.format(t)] = "=MAX(DH2:DH{})".format(q)
    ws['DH{}'.format(u)] = "=MIN(DH2:DH{})".format(q)
    ws['DI{}'.format(r)] = "=IF($CY${}=0,$CI${},$CI${}-1)".format(r, r, r)
    ws['DI{}'.format(s)] = "=STDEV(DI2:DI{})".format(q)
    ws['DI{}'.format(t)] = "=MAX(DI2:DI{})".format(q)
    ws['DI{}'.format(u)] = "=MIN(DI2:DI{})".format(q)
    ws['DJ{}'.format(r)] = "=IF($CZ${}=0,$CJ${},$CJ${}-1)".format(r, r, r)
    ws['DJ{}'.format(s)] = "=STDEV(DJ2:DJ{})".format(q)
    ws['DJ{}'.format(t)] = "=MAX(DJ2:DJ{})".format(q)
    ws['DJ{}'.format(u)] = "=MIN(DJ2:DJ{})".format(q)

    # JUMLAH MAPEL NORMAL
    ws['DK{}'.format(r)] = "=ROUND(AVERAGE(DK2:DK{}),2)".format(q)
    ws['DK{}'.format(t)] = "=MAX(DK2:DK{})".format(q)
    ws['DK{}'.format(u)] = "=MIN(DK2:DK{})".format(q)

    # Z SCORE
    ws['DL{}'.format(r)] = "=MAX(DL2:DL{})".format(q)
    ws['DM{}'.format(r)] = "=MAX(DM2:DM{})".format(q)
    ws['DN{}'.format(r)] = "=MAX(DN2:DN{})".format(q)
    ws['DO{}'.format(r)] = "=MAX(DO2:DO{})".format(q)

    # NILAI STANDAR
    ws['DP{}'.format(r)] = "=MAX(DP2:DP{})".format(q)
    ws['DP{}'.format(s)] = "=MIN(DP2:DP{})".format(q)
    ws['DP{}'.format(t)] = "=ROUND(AVERAGE(DP2:DP{}),2)".format(q)
    ws['DQ{}'.format(r)] = "=MAX(DQ2:DQ{})".format(q)
    ws['DQ{}'.format(s)] = "=MIN(DQ2:DQ{})".format(q)
    ws['DQ{}'.format(t)] = "=ROUND(AVERAGE(DQ2:DQ{}),2)".format(q)
    ws['DR{}'.format(r)] = "=MAX(DR2:DR{})".format(q)
    ws['DR{}'.format(s)] = "=MIN(DR2:DR{})".format(q)
    ws['DR{}'.format(t)] = "=ROUND(AVERAGE(DR2:DR{}),2)".format(q)
    ws['DS{}'.format(r)] = "=MAX(DS2:DS{})".format(q)
    ws['DS{}'.format(s)] = "=MIN(DS2:DS{})".format(q)
    ws['DS{}'.format(t)] = "=ROUND(AVERAGE(DS2:DS{}),2)".format(q)
    ws['DT{}'.format(r)] = "=MAX(DT2:DT{})".format(q)
    ws['DT{}'.format(s)] = "=MIN(DT2:DT{})".format(q)
    ws['DT{}'.format(t)] = "=ROUND(AVERAGE(DT2:DT{}),2)".format(q)

    # INISIASI MAPEL
    ws['DW{}'.format(r)] = "=SUM(DW2:DW{})".format(q)
    ws['DX{}'.format(r)] = "=SUM(DX2:DX{})".format(q)
    ws['DY{}'.format(r)] = "=SUM(DY2:DY{})".format(q)
    ws['DZ{}'.format(r)] = "=SUM(DZ2:DZ{})".format(q)

    # iterasi 5 rata-rata - 1
    # MAPEL NORMAL
    ws['EG{}'.format(r)] = "=IF($DW${}=0,$DG${},$DG${}-1)".format(r, r, r)
    ws['EG{}'.format(s)] = "=STDEV(EG2:EG{})".format(q)
    ws['EG{}'.format(t)] = "=MAX(EG2:EG{})".format(q)
    ws['EG{}'.format(u)] = "=MIN(EG2:EG{})".format(q)
    ws['EH{}'.format(r)] = "=IF($DX${}=0,$DH${},$DH${}-1)".format(r, r, r)
    ws['EH{}'.format(s)] = "=STDEV(EH2:EH{})".format(q)
    ws['EH{}'.format(t)] = "=MAX(EH2:EH{})".format(q)
    ws['EH{}'.format(u)] = "=MIN(EH2:EH{})".format(q)
    ws['EI{}'.format(r)] = "=IF($DY${}=0,$DI${},$DI${}-1)".format(r, r, r)
    ws['EI{}'.format(s)] = "=STDEV(EI2:EI{})".format(q)
    ws['EI{}'.format(t)] = "=MAX(EI2:EI{})".format(q)
    ws['EI{}'.format(u)] = "=MIN(EI2:EI{})".format(q)
    ws['EJ{}'.format(r)] = "=IF($DZ${}=0,$DJ${},$DJ${}-1)".format(r, r, r)
    ws['EJ{}'.format(s)] = "=STDEV(EJ2:EJ{})".format(q)
    ws['EJ{}'.format(t)] = "=MAX(EJ2:EJ{})".format(q)
    ws['EJ{}'.format(u)] = "=MIN(EJ2:EJ{})".format(q)

    # JUMLAH MAPEL NORMAL
    ws['EK{}'.format(r)] = "=ROUND(AVERAGE(EK2:EK{}),2)".format(q)
    ws['EK{}'.format(t)] = "=MAX(EK2:EK{})".format(q)
    ws['EK{}'.format(u)] = "=MIN(EK2:EK{})".format(q)

    # Z SCORE
    ws['EL{}'.format(r)] = "=MAX(EL2:EL{})".format(q)
    ws['EM{}'.format(r)] = "=MAX(EM2:EM{})".format(q)
    ws['EN{}'.format(r)] = "=MAX(EN2:EN{})".format(q)
    ws['EO{}'.format(r)] = "=MAX(EO2:EO{})".format(q)

    # NILAI STANDAR
    ws['EP{}'.format(r)] = "=MAX(EP2:EP{})".format(q)
    ws['EP{}'.format(s)] = "=MIN(EP2:EP{})".format(q)
    ws['EP{}'.format(t)] = "=ROUND(AVERAGE(EP2:EP{}),2)".format(q)
    ws['EQ{}'.format(r)] = "=MAX(EQ2:EQ{})".format(q)
    ws['EQ{}'.format(s)] = "=MIN(EQ2:EQ{})".format(q)
    ws['EQ{}'.format(t)] = "=ROUND(AVERAGE(EQ2:EQ{}),2)".format(q)
    ws['ER{}'.format(r)] = "=MAX(ER2:ER{})".format(q)
    ws['ER{}'.format(s)] = "=MIN(ER2:ER{})".format(q)
    ws['ER{}'.format(t)] = "=ROUND(AVERAGE(ER2:ER{}),2)".format(q)
    ws['ES{}'.format(r)] = "=MAX(ES2:ES{})".format(q)
    ws['ES{}'.format(s)] = "=MIN(ES2:ES{})".format(q)
    ws['ES{}'.format(t)] = "=ROUND(AVERAGE(ES2:ES{}),2)".format(q)
    ws['ET{}'.format(r)] = "=MAX(ET2:ET{})".format(q)
    ws['ET{}'.format(s)] = "=MIN(ET2:ET{})".format(q)
    ws['ET{}'.format(t)] = "=ROUND(AVERAGE(ET2:ET{}),2)".format(q)

    # INISIASI MAPEL
    ws['EW{}'.format(r)] = "=SUM(EW2:EW{})".format(q)
    ws['EX{}'.format(r)] = "=SUM(EX2:EX{})".format(q)
    ws['EY{}'.format(r)] = "=SUM(EY2:EY{})".format(q)
    ws['EZ{}'.format(r)] = "=SUM(EZ2:EZ{})".format(q)

    # Jumlah Soal
    ws['F{}'.format(v)] = 'JUMLAH SOAL'
    ws['G{}'.format(v)] = JML_SOAL_MAT
    ws['H{}'.format(v)] = JML_SOAL_BIO
    ws['I{}'.format(v)] = JML_SOAL_FIS
    ws['J{}'.format(v)] = JML_SOAL_KIM

    # Z Score
    ws['B1'] = 'NAMA SISWA_A'
    ws['C1'] = 'NOMOR NF_A'
    ws['D1'] = 'KELAS_A'
    ws['E1'] = 'NAMA SEKOLAH_A'
    ws['F1'] = 'LOKASI_A'
    ws['G1'] = 'MAT_A'
    ws['H1'] = 'BIO_A'
    ws['I1'] = 'FIS_A'
    ws['J1'] = 'KIM_A'
    ws['K1'] = 'JML_A'
    ws['L1'] = 'Z_MAT_A'
    ws['M1'] = 'Z_BIO_A'
    ws['N1'] = 'Z_FIS_A'
    ws['O1'] = 'Z_KIM_A'
    ws['P1'] = 'S_MAT_A'
    ws['Q1'] = 'S_BIO_A'
    ws['R1'] = 'S_FIS_A'
    ws['S1'] = 'S_KIM_A'
    ws['T1'] = 'S_JML_A'
    ws['U1'] = 'RANK NAS._A'
    ws['V1'] = 'RANK LOK._A'

    ws['L1'].font = Font(bold=False, name='Calibri', size=11)
    ws['M1'].font = Font(bold=False, name='Calibri', size=11)
    ws['N1'].font = Font(bold=False, name='Calibri', size=11)
    ws['O1'].font = Font(bold=False, name='Calibri', size=11)
    ws['P1'].font = Font(bold=False, name='Calibri', size=11)
    ws['Q1'].font = Font(bold=False, name='Calibri', size=11)
    ws['R1'].font = Font(bold=False, name='Calibri', size=11)
    ws['S1'].font = Font(bold=False, name='Calibri', size=11)
    ws['T1'].font = Font(bold=False, name='Calibri', size=11)
    ws['U1'].font = Font(bold=False, name='Calibri', size=11)
    ws['V1'].font = Font(bold=False, name='Calibri', size=11)

    # FILL
    ws['B1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['C1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['D1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['E1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['F1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['G1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['H1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['I1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['J1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['K1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['L1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['M1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['N1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['O1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['P1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['Q1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['R1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['S1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['T1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['U1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['V1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')

    # tambahan
    ws['W1'] = 'MAT_20_A'
    ws['X1'] = 'BIO_20_A'
    ws['Y1'] = 'FIS_20_A'
    ws['Z1'] = 'KIM_20_A'
    ws['W1'].font = Font(bold=False, name='Calibri', size=11)
    ws['X1'].font = Font(bold=False, name='Calibri', size=11)
    ws['Y1'].font = Font(bold=False, name='Calibri', size=11)
    ws['Z1'].font = Font(bold=False, name='Calibri', size=11)
    ws['W1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['X1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['Y1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['Z1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')

    for row in range(2, q+1):
        ws['L{}'.format(
            row)] = '=IFERROR(ROUND(IF(G{}="","",(G{}-G${})/G${}),2),"")'.format(row, row, r, s)
        ws['M{}'.format(
            row)] = '=IFERROR(ROUND(IF(H{}="","",(H{}-H${})/H${}),2),"")'.format(row, row, r, s)
        ws['N{}'.format(
            row)] = '=IFERROR(ROUND(IF(I{}="","",(I{}-I${})/I${}),2),"")'.format(row, row, r, s)
        ws['O{}'.format(
            row)] = '=IFERROR(ROUND(IF(J{}="","",(J{}-J${})/J${}),2),"")'.format(row, row, r, s)
        ws['P{}'.format(
            row)] = '=IFERROR(ROUND(IF(G{}="","",IF(70+30*L{}/$L${}<20,20,70+30*L{}/$L${})),2),"")'.format(row, row, r, row, r)
        ws['Q{}'.format(
            row)] = '=IFERROR(ROUND(IF(H{}="","",IF(70+30*M{}/$M${}<20,20,70+30*M{}/$M${})),2),"")'.format(row, row, r, row, r)
        ws['R{}'.format(
            row)] = '=IFERROR(ROUND(IF(I{}="","",IF(70+30*N{}/$N${}<20,20,70+30*N{}/$N${})),2),"")'.format(row, row, r, row, r)
        ws['S{}'.format(
            row)] = '=IFERROR(ROUND(IF(J{}="","",IF(70+30*O{}/$O${}<20,20,70+30*O{}/$P${})),2),"")'.format(row, row, r, row, r)

        ws['T{}'.format(row)] = '=IF(SUM(P{}:S{})=0,"",SUM(P{}:S{}))'.format(
            row, row, row, row)
        ws['U{}'.format(row)] = '=IF(T{}="","",RANK(T{},$T$2:$T${}))'.format(
            row, row, q)
        ws['V{}'.format(
            row)] = '=IF(U{}="","",COUNTIFS($F$2:$F${},F{},$U$2:$U${},"<"&U{})+1)'.format(row, q, row, q, row)
        # TAMBAHAN
        ws['W{}'.format(row)] = '=IF($G${}=25,IF(AND(G{}>4,P{}=20),1,""),IF($G${}=30,IF(AND(G{}>5,P{}=20),1,""),IF($G${}=35,IF(AND(G{}>6,P{}=20),1,""),IF($G${}=40,IF(AND(G{}>7,P{}=20),1,""),IF($G${}=45,IF(AND(G{}>8,P{}=20),1,""))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['X{}'.format(row)] = '=IF($H${}=25,IF(AND(H{}>4,Q{}=20),1,""),IF($H${}=30,IF(AND(H{}>5,Q{}=20),1,""),IF($H${}=35,IF(AND(H{}>6,Q{}=20),1,""),IF($H${}=40,IF(AND(H{}>7,Q{}=20),1,""),IF($H${}=45,IF(AND(H{}>8,Q{}=20),1,""))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['Y{}'.format(row)] = '=IF($I${}=25,IF(AND(I{}>4,R{}=20),1,""),IF($I${}=30,IF(AND(I{}>5,R{}=20),1,""),IF($I${}=35,IF(AND(I{}>6,R{}=20),1,""),IF($I${}=40,IF(AND(I{}>7,R{}=20),1,""),IF($I${}=45,IF(AND(I{}>8,R{}=20),1,""))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['Z{}'.format(row)] = '=IF($J${}=25,IF(AND(J{}>4,S{}=20),1,""),IF($J${}=30,IF(AND(J{}>5,S{}=20),1,""),IF($J${}=35,IF(AND(J{}>6,S{}=20),1,""),IF($J${}=40,IF(AND(J{}>7,S{}=20),1,""),IF($J${}=45,IF(AND(J{}>8,S{}=20),1,""))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

    # Z Score iterasi 1
    ws['AB1'] = 'NAMA SISWA_B'
    ws['AC1'] = 'NOMOR NF_B'
    ws['AD1'] = 'KELAS_B'
    ws['AE1'] = 'NAMA SEKOLAH_B'
    ws['AF1'] = 'LOKASI_B'
    ws['AG1'] = 'MAT_B'
    ws['AH1'] = 'BIO_B'
    ws['AI1'] = 'FIS_B'
    ws['AJ1'] = 'KIM_B'
    ws['AK1'] = 'JML_B'
    ws['AL1'] = 'Z_MAT_B'
    ws['AM1'] = 'Z_BIO_B'
    ws['AN1'] = 'Z_FIS_B'
    ws['AO1'] = 'Z_KIM_B'
    ws['AP1'] = 'S_MAT_B'
    ws['AQ1'] = 'S_BIO_B'
    ws['AR1'] = 'S_FIS_B'
    ws['AS1'] = 'S_KIM_B'
    ws['AT1'] = 'S_JML_B'
    ws['AU1'] = 'RANK NAS._B'
    ws['AV1'] = 'RANK LOK._B'

    ws['AL1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AM1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AN1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AO1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AP1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AQ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AR1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AS1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AT1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AU1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AV1'].font = Font(bold=False, name='Calibri', size=11)

    # FILL
    ws['AB1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AC1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AD1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AE1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AF1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AG1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AH1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AI1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AJ1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AK1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AL1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AM1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AN1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AO1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AP1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AQ1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AR1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AS1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AT1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AU1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AV1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')

    # tambahan
    ws['AW1'] = 'MAT_20'
    ws['AX1'] = 'BIO_20'
    ws['AY1'] = 'FIS_20'
    ws['AZ1'] = 'KIM_20'
    ws['AW1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AX1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AY1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AZ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AW1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AX1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AY1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AZ1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')

    for row in range(2, q+1):
        # Tambahan
        ws['AB{}'.format(row)] = '=B{}'.format(row)
        ws['AC{}'.format(row)] = '=C{}'.format(row, row)
        ws['AD{}'.format(row)] = '=D{}'.format(row, row)
        ws['AE{}'.format(row)] = '=E{}'.format(row, row)
        ws['AF{}'.format(row)] = '=F{}'.format(row, row)
        ws['AG{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
        ws['AH{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
        ws['AI{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
        ws['AJ{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
        ws['AK{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)

        ws['AL{}'.format(
            row)] = '=IFERROR(ROUND(IF(AG{}="","",(AG{}-AG${})/AG${}),2),"")'.format(row, row, r, s)
        ws['AM{}'.format(
            row)] = '=IFERROR(ROUND(IF(AH{}="","",(AH{}-AH${})/AH${}),2),"")'.format(row, row, r, s)
        ws['AN{}'.format(
            row)] = '=IFERROR(ROUND(IF(AI{}="","",(AI{}-AI${})/AI${}),2),"")'.format(row, row, r, s)
        ws['AO{}'.format(
            row)] = '=IFERROR(ROUND(IF(AJ{}="","",(AJ{}-AJ${})/AJ${}),2),"")'.format(row, row, r, s)

        ws['AP{}'.format(
            row)] = '=IFERROR(ROUND(IF(G{}="","",IF(70+30*AL{}/$AL${}<20,20,70+30*AL{}/$AL${})),2),"")'.format(row, row, r, row, r)
        ws['AQ{}'.format(
            row)] = '=IFERROR(ROUND(IF(H{}="","",IF(70+30*AM{}/$AM{}<20,20,70+30*AM{}/$AM${})),2),"")'.format(row, row, r, row, r)
        ws['AR{}'.format(
            row)] = '=IFERROR(ROUND(IF(I{}="","",IF(70+30*AN{}/$AN${}<20,20,70+30*AN{}/$AN${})),2),"")'.format(row, row, r, row, r)
        ws['AS{}'.format(
            row)] = '=IFERROR(ROUND(IF(J{}="","",IF(70+30*AO{}/$AO${}<20,20,70+30*AO{}/$AO${})),2),"")'.format(row, row, r, row, r)

        ws['AT{}'.format(row)] = '=IF(SUM(AP{}:AS{})=0,"",SUM(AP{}:AS{}))'.format(
            row, row, row, row)
        ws['AU{}'.format(row)] = '=IF(AT{}="","",RANK(AT{},$AT$2:$AT${}))'.format(
            row, row, q)
        ws['AV{}'.format(
            row)] = '=IF(AU{}="","",COUNTIFS($AF$2:$AF${},AF{},$AU$2:$AU${},"<"&AU{})+1)'.format(row, q, row, q, row)
        # TAMBAHAN
        ws['AW{}'.format(row)] = '=IF($G${}=25,IF(AND(AG{}>4,AP{}=20),1,""),IF($G${}=30,IF(AND(AG{}>5,AP{}=20),1,""),IF($G${}=35,IF(AND(AG{}>6,AP{}=20),1,""),IF($G${}=40,IF(AND(AG{}>7,AP{}=20),1,""),IF($G${}=45,IF(AND(AG{}>8,AP{}=20),1,""))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['AX{}'.format(row)] = '=IF($H${}=25,IF(AND(AH{}>4,AQ{}=20),1,""),IF($H${}=30,IF(AND(AH{}>5,AQ{}=20),1,""),IF($H${}=35,IF(AND(AH{}>6,AQ{}=20),1,""),IF($H${}=40,IF(AND(AH{}>7,AQ{}=20),1,""),IF($H${}=45,IF(AND(AH{}>8,AQ{}=20),1,""))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['AY{}'.format(row)] = '=IF($I${}=25,IF(AND(AI{}>4,AR{}=20),1,""),IF($I${}=30,IF(AND(AI{}>5,AR{}=20),1,""),IF($I${}=35,IF(AND(AI{}>6,AR{}=20),1,""),IF($I${}=40,IF(AND(AI{}>7,AR{}=20),1,""),IF($I${}=45,IF(AND(AI{}>8,AR{}=20),1,""))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['AZ{}'.format(row)] = '=IF($J${}=25,IF(AND(AJ{}>4,AS{}=20),1,""),IF($J${}=30,IF(AND(AJ{}>5,AS{}=20),1,""),IF($J${}=35,IF(AND(AJ{}>6,AS{}=20),1,""),IF($J${}=40,IF(AND(AJ{}>7,AS{}=20),1,""),IF($J${}=45,IF(AND(AJ{}>8,AS{}=20),1,""))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

    # Z Score iterasi 2
    ws['BB1'] = 'NAMA SISWA_C'
    ws['BC1'] = 'NOMOR NF_c'
    ws['BD1'] = 'KELAS_C'
    ws['BE1'] = 'NAMA SEKOLAH_C'
    ws['BF1'] = 'LOKASI_C'
    ws['BG1'] = 'MAT_C'
    ws['BH1'] = 'BIO_C'
    ws['BI1'] = 'FIS_C'
    ws['BJ1'] = 'KIM_C'
    ws['BK1'] = 'JML_C'
    ws['BL1'] = 'Z_MAT_C'
    ws['BM1'] = 'Z_BIO_C'
    ws['BN1'] = 'Z_FIS_C'
    ws['BO1'] = 'Z_KIM_C'
    ws['BP1'] = 'S_MAT_C'
    ws['BQ1'] = 'S_BIO_C'
    ws['BR1'] = 'S_FIS_C'
    ws['BS1'] = 'S_KIM_C'
    ws['BT1'] = 'S_JML_C'
    ws['BU1'] = 'RANK NAS._C'
    ws['BV1'] = 'RANK LOK._C'

    ws['BL1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BM1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BN1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BO1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BP1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BQ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BR1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BS1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BT1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BU1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BV1'].font = Font(bold=False, name='Calibri', size=11)

    # FILL
    ws['BB1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['BC1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['BD1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['BE1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['BF1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['BG1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['BH1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['BI1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['BJ1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['BK1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['BL1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['BM1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['BN1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['BO1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['BP1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['BQ1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['BR1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['BS1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['BT1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['BU1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['BV1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')

    # tambahan
    ws['BW1'] = 'MAT_20_C'
    ws['BX1'] = 'BIO_20_C'
    ws['BY1'] = 'FIS_20_C'
    ws['BZ1'] = 'KIM_20_C'
    ws['BW1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BX1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BY1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BZ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BW1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['BX1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['BY1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['BZ1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')

    for row in range(2, q+1):
        # Tambahan
        ws['BB{}'.format(row)] = '=AB{}'.format(row)
        ws['BC{}'.format(row)] = '=AC{}'.format(row, row)
        ws['BD{}'.format(row)] = '=AD{}'.format(row, row)
        ws['BE{}'.format(row)] = '=AE{}'.format(row, row)
        ws['BF{}'.format(row)] = '=AF{}'.format(row, row)
        ws['BG{}'.format(row)] = '=IF(AG{}="","",AG{})'.format(row, row)
        ws['BH{}'.format(row)] = '=IF(AH{}="","",AH{})'.format(row, row)
        ws['BI{}'.format(row)] = '=IF(AI{}="","",AI{})'.format(row, row)
        ws['BJ{}'.format(row)] = '=IF(AJ{}="","",AJ{})'.format(row, row)
        ws['BK{}'.format(row)] = '=IF(AK{}="","",AK{})'.format(row, row)

        ws['BL{}'.format(
            row)] = '=IFERROR(ROUND(IF(BG{}="","",(BG{}-BG${})/BG${}),2),"")'.format(row, row, r, s)
        ws['BM{}'.format(
            row)] = '=IFERROR(ROUND(IF(BH{}="","",(BH{}-BH${})/BH${}),2),"")'.format(row, row, r, s)
        ws['BN{}'.format(
            row)] = '=IFERROR(ROUND(IF(BI{}="","",(BI{}-BI${})/BI${}),2),"")'.format(row, row, r, s)
        ws['BO{}'.format(
            row)] = '=IFERROR(ROUND(IF(BJ{}="","",(BJ{}-BJ${})/BJ${}),2),"")'.format(row, row, r, s)

        ws['BP{}'.format(
            row)] = '=IFERROR(ROUND(IF(BG{}="","",IF(70+30*BL{}/$BL${}<20,20,70+30*BL{}/$BL${})),2),"")'.format(row, row, r, row, r)
        ws['BQ{}'.format(
            row)] = '=IFERROR(ROUND(IF(BH{}="","",IF(70+30*BM{}/$BM{}<20,20,70+30*BM{}/$BM${})),2),"")'.format(row, row, r, row, r)
        ws['BR{}'.format(
            row)] = '=IFERROR(ROUND(IF(BI{}="","",IF(70+30*BN{}/$BN${}<20,20,70+30*BN{}/$BN${})),2),"")'.format(row, row, r, row, r)
        ws['BS{}'.format(
            row)] = '=IFERROR(ROUND(IF(BJ{}="","",IF(70+30*BO{}/$BO${}<20,20,70+30*BO{}/$BO${})),2),"")'.format(row, row, r, row, r)

        ws['BT{}'.format(row)] = '=IF(SUM(BP{}:BS{})=0,"",SUM(BP{}:BS{}))'.format(
            row, row, row, row)
        ws['BU{}'.format(row)] = '=IF(BT{}="","",RANK(BT{},$BT$2:$BT${}))'.format(
            row, row, q)
        ws['BV{}'.format(
            row)] = '=IF(BU{}="","",COUNTIFS($BF$2:$BF${},BF{},$BU$2:$BU${},"<"&BU{})+1)'.format(row, q, row, q, row)
        # TAMBAHAN
        ws['BW{}'.format(row)] = '=IF($G${}=25,IF(AND(BG{}>4,BP{}=20),1,""),IF($G${}=30,IF(AND(BG{}>5,BP{}=20),1,""),IF($G${}=35,IF(AND(BG{}>6,BP{}=20),1,""),IF($G${}=40,IF(AND(BG{}>7,BP{}=20),1,""),IF($G${}=45,IF(AND(BG{}>8,BP{}=20),1,""))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['BX{}'.format(row)] = '=IF($H${}=25,IF(AND(BH{}>4,BQ{}=20),1,""),IF($H${}=30,IF(AND(BH{}>5,BQ{}=20),1,""),IF($H${}=35,IF(AND(BH{}>6,BQ{}=20),1,""),IF($H${}=40,IF(AND(BH{}>7,BQ{}=20),1,""),IF($H${}=45,IF(AND(BH{}>8,BQ{}=20),1,""))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['BY{}'.format(row)] = '=IF($I${}=25,IF(AND(BI{}>4,BR{}=20),1,""),IF($I${}=30,IF(AND(BI{}>5,BR{}=20),1,""),IF($I${}=35,IF(AND(BI{}>6,BR{}=20),1,""),IF($I${}=40,IF(AND(BI{}>7,BR{}=20),1,""),IF($I${}=45,IF(AND(BI{}>8,BR{}=20),1,""))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['BZ{}'.format(row)] = '=IF($J${}=25,IF(AND(BJ{}>4,BS{}=20),1,""),IF($J${}=30,IF(AND(BJ{}>5,BS{}=20),1,""),IF($J${}=35,IF(AND(BJ{}>6,BS{}=20),1,""),IF($J${}=40,IF(AND(BJ{}>7,BS{}=20),1,""),IF($J${}=45,IF(AND(BJ{}>8,BS{}=20),1,""))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

    # Z Score iterasi 3
    ws['CB1'] = 'NAMA SISWA_D'
    ws['CC1'] = 'NOMOR NF_D'
    ws['CD1'] = 'KELAS_D'
    ws['CE1'] = 'NAMA SEKOLAH_D'
    ws['CF1'] = 'LOKASI_D'
    ws['CG1'] = 'MAT_D'
    ws['CH1'] = 'BIO_D'
    ws['CI1'] = 'FIS_D'
    ws['CJ1'] = 'KIM_D'
    ws['CK1'] = 'JML_D'
    ws['CL1'] = 'Z_MAT_D'
    ws['CM1'] = 'Z_BIO_D'
    ws['CN1'] = 'Z_FIS_D'
    ws['CO1'] = 'Z_KIM_D'
    ws['CP1'] = 'S_MAT_D'
    ws['CQ1'] = 'S_BIO_D'
    ws['CR1'] = 'S_FIS_D'
    ws['CS1'] = 'S_KIM_D'
    ws['CT1'] = 'S_JML_D'
    ws['CU1'] = 'RANK NAS._D'
    ws['CV1'] = 'RANK LOK._D'

    ws['CL1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CM1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CN1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CO1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CP1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CQ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CR1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CS1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CT1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CU1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CV1'].font = Font(bold=False, name='Calibri', size=11)

    # FILL
    ws['CB1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['CC1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['CD1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['CE1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['CF1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['CG1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['CH1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['CI1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['CJ1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['CK1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['CL1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['CM1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['CN1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['CO1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['CP1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['CQ1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['CR1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['CS1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['CT1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['CU1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['CV1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')

    # tambahan
    ws['CW1'] = 'MAT_20_D'
    ws['CX1'] = 'BIO_20_D'
    ws['CY1'] = 'FIS_20_D'
    ws['CZ1'] = 'KIM_20_D'
    ws['CW1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CX1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CY1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CZ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CW1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['CX1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['CY1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['CZ1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')

    for row in range(2, q+1):
        ws['CB{}'.format(row)] = '=BB{}'.format(row)
        ws['CC{}'.format(row)] = '=BC{}'.format(row, row)
        ws['CD{}'.format(row)] = '=BD{}'.format(row, row)
        ws['CE{}'.format(row)] = '=BE{}'.format(row, row)
        ws['CF{}'.format(row)] = '=BF{}'.format(row, row)
        ws['CG{}'.format(row)] = '=IF(BG{}="","",BG{})'.format(row, row)
        ws['CH{}'.format(row)] = '=IF(BH{}="","",BH{})'.format(row, row)
        ws['CI{}'.format(row)] = '=IF(BI{}="","",BI{})'.format(row, row)
        ws['CJ{}'.format(row)] = '=IF(BJ{}="","",BJ{})'.format(row, row)
        ws['CK{}'.format(row)] = '=IF(BK{}="","",BK{})'.format(row, row)

        ws['CL{}'.format(
            row)] = '=IFERROR(ROUND(IF(CG{}="","",(CG{}-CG${})/CG${}),2),"")'.format(row, row, r, s)
        ws['CM{}'.format(
            row)] = '=IFERROR(ROUND(IF(CH{}="","",(CH{}-CH${})/CH${}),2),"")'.format(row, row, r, s)
        ws['CN{}'.format(
            row)] = '=IFERROR(ROUND(IF(CI{}="","",(CI{}-CI${})/CI${}),2),"")'.format(row, row, r, s)
        ws['CO{}'.format(
            row)] = '=IFERROR(ROUND(IF(CJ{}="","",(CJ{}-CJ${})/CJ${}),2),"")'.format(row, row, r, s)

        ws['CP{}'.format(
            row)] = '=IFERROR(ROUND(IF(CG{}="","",IF(70+30*CL{}/$CL${}<20,20,70+30*CL{}/$CL${})),2),"")'.format(row, row, r, row, r)
        ws['CQ{}'.format(
            row)] = '=IFERROR(ROUND(IF(CH{}="","",IF(70+30*CM{}/$CM{}<20,20,70+30*CM{}/$CM${})),2),"")'.format(row, row, r, row, r)
        ws['CR{}'.format(
            row)] = '=IFERROR(ROUND(IF(CI{}="","",IF(70+30*CN{}/$CN${}<20,20,70+30*CN{}/$CN${})),2),"")'.format(row, row, r, row, r)
        ws['CS{}'.format(
            row)] = '=IFERROR(ROUND(IF(CJ{}="","",IF(70+30*CO{}/$CO${}<20,20,70+30*CO{}/$CO${})),2),"")'.format(row, row, r, row, r)

        ws['CT{}'.format(row)] = '=IF(SUM(CP{}:CS{})=0,"",SUM(CP{}:CS{}))'.format(
            row, row, row, row)
        ws['CU{}'.format(row)] = '=IF(CT{}="","",RANK(CT{},$CT$2:$CT${}))'.format(
            row, row, q)
        ws['CV{}'.format(
            row)] = '=IF(CU{}="","",COUNTIFS($CF$2:$CF${},CF{},$CU$2:$CU${},"<"&CU{})+1)'.format(row, q, row, q, row)
        # TAMBAHAN
        ws['CW{}'.format(row)] = '=IF($G${}=25,IF(AND(CG{}>4,CP{}=20),1,""),IF($G${}=30,IF(AND(CG{}>5,CP{}=20),1,""),IF($G${}=35,IF(AND(CG{}>6,CP{}=20),1,""),IF($G${}=40,IF(AND(CG{}>7,CP{}=20),1,""),IF($G${}=45,IF(AND(CG{}>8,CP{}=20),1,""))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['CX{}'.format(row)] = '=IF($H${}=25,IF(AND(CH{}>4,CQ{}=20),1,""),IF($H${}=30,IF(AND(CH{}>5,CQ{}=20),1,""),IF($H${}=35,IF(AND(CH{}>6,CQ{}=20),1,""),IF($H${}=40,IF(AND(CH{}>7,CQ{}=20),1,""),IF($H${}=45,IF(AND(CH{}>8,CQ{}=20),1,""))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['CY{}'.format(row)] = '=IF($I${}=25,IF(AND(CI{}>4,CR{}=20),1,""),IF($I${}=30,IF(AND(CI{}>5,CR{}=20),1,""),IF($I${}=35,IF(AND(CI{}>6,CR{}=20),1,""),IF($I${}=40,IF(AND(CI{}>7,CR{}=20),1,""),IF($I${}=45,IF(AND(CI{}>8,CR{}=20),1,""))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['CZ{}'.format(row)] = '=IF($J${}=25,IF(AND(CJ{}>4,CS{}=20),1,""),IF($J${}=30,IF(AND(CJ{}>5,CS{}=20),1,""),IF($J${}=35,IF(AND(CJ{}>6,CS{}=20),1,""),IF($J${}=40,IF(AND(CJ{}>7,CS{}=20),1,""),IF($J${}=45,IF(AND(CJ{}>8,CS{}=20),1,""))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

    # Z Score iterasi 4
    ws['DB1'] = 'NAMA SISWA_E'
    ws['DC1'] = 'NOMOR NF_E'
    ws['DD1'] = 'KELAS_E'
    ws['DE1'] = 'NAMA SEKOLAH_E'
    ws['DF1'] = 'LOKASI_E'
    ws['DG1'] = 'MAT_E'
    ws['DH1'] = 'BIO_E'
    ws['DI1'] = 'FIS_E'
    ws['DJ1'] = 'KIM_E'
    ws['DK1'] = 'JML_E'
    ws['DL1'] = 'Z_MAT_E'
    ws['DM1'] = 'Z_BIO_E'
    ws['DN1'] = 'Z_FIS_E'
    ws['DO1'] = 'Z_KIM_E'
    ws['DP1'] = 'S_MAT_E'
    ws['DQ1'] = 'S_BIO_E'
    ws['DR1'] = 'S_FIS_E'
    ws['DS1'] = 'S_KIM_E'
    ws['DT1'] = 'S_JML_E'
    ws['DU1'] = 'RANK NAS._E'
    ws['DV1'] = 'RANK LOK._E'

    ws['DL1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DM1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DN1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DO1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DP1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DQ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DR1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DS1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DT1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DU1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DV1'].font = Font(bold=False, name='Calibri', size=11)

    # FILL
    ws['DB1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['DC1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['DD1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['DE1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['DF1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['DG1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['DH1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['DI1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['DJ1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['DK1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['DL1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['DM1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['DN1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['DO1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['DP1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['DQ1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['DR1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['DS1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['DT1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['DU1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['DV1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')

    # tambahan
    ws['DW1'] = 'MAT_20'
    ws['DX1'] = 'BIO_20'
    ws['DY1'] = 'FIS_20'
    ws['DZ1'] = 'KIM_20'
    ws['DW1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DX1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DY1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DZ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DW1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['DX1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['DY1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['DZ1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')

    for row in range(2, q+1):
        # Tambahan
        ws['DB{}'.format(row)] = '=CB{}'.format(row)
        ws['DC{}'.format(row)] = '=CC{}'.format(row, row)
        ws['DD{}'.format(row)] = '=CD{}'.format(row, row)
        ws['DE{}'.format(row)] = '=CE{}'.format(row, row)
        ws['DF{}'.format(row)] = '=CF{}'.format(row, row)
        ws['DG{}'.format(row)] = '=IF(CG{}="","",CG{})'.format(row, row)
        ws['DH{}'.format(row)] = '=IF(CH{}="","",CH{})'.format(row, row)
        ws['DI{}'.format(row)] = '=IF(CI{}="","",CI{})'.format(row, row)
        ws['DJ{}'.format(row)] = '=IF(CJ{}="","",CJ{})'.format(row, row)
        ws['DK{}'.format(row)] = '=IF(CK{}="","",CK{})'.format(row, row)

        ws['DL{}'.format(
            row)] = '=IFERROR(ROUND(IF(DG{}="","",(DG{}-DG${})/DG${}),2),"")'.format(row, row, r, s)
        ws['DM{}'.format(
            row)] = '=IFERROR(ROUND(IF(DH{}="","",(DH{}-DH${})/DH${}),2),"")'.format(row, row, r, s)
        ws['DN{}'.format(
            row)] = '=IFERROR(ROUND(IF(DI{}="","",(DI{}-DI${})/DI${}),2),"")'.format(row, row, r, s)
        ws['DO{}'.format(
            row)] = '=IFERROR(ROUND(IF(DJ{}="","",(DJ{}-DJ${})/DJ${}),2),"")'.format(row, row, r, s)

        ws['DP{}'.format(
            row)] = '=IFERROR(ROUND(IF(DG{}="","",IF(70+30*DL{}/$DL${}<20,20,70+30*DL{}/$DL${})),2),"")'.format(row, row, r, row, r)
        ws['DQ{}'.format(
            row)] = '=IFERROR(ROUND(IF(DH{}="","",IF(70+30*DM{}/$DM{}<20,20,70+30*DM{}/$DM${})),2),"")'.format(row, row, r, row, r)
        ws['DR{}'.format(
            row)] = '=IFERROR(ROUND(IF(DI{}="","",IF(70+30*DN{}/$DN${}<20,20,70+30*DN{}/$DN${})),2),"")'.format(row, row, r, row, r)
        ws['DS{}'.format(
            row)] = '=IFERROR(ROUND(IF(DJ{}="","",IF(70+30*DO{}/$DO${}<20,20,70+30*DO{}/$DO${})),2),"")'.format(row, row, r, row, r)

        ws['DT{}'.format(row)] = '=IF(SUM(DP{}:DS{})=0,"",SUM(DP{}:DS{}))'.format(
            row, row, row, row)
        ws['DU{}'.format(row)] = '=IF(DT{}="","",RANK(DT{},$DT$2:$DT${}))'.format(
            row, row, q)
        ws['DV{}'.format(
            row)] = '=IF(DU{}="","",COUNTIFS($DF$2:$DF${},DF{},$DU$2:$DU${},"<"&DU{})+1)'.format(row, q, row, q, row)
        # TAMBAHAN
        ws['DW{}'.format(row)] = '=IF($G${}=25,IF(AND(DG{}>4,DP{}=20),1,""),IF($G${}=30,IF(AND(DG{}>5,DP{}=20),1,""),IF($G${}=35,IF(AND(DG{}>6,DP{}=20),1,""),IF($G${}=40,IF(AND(DG{}>7,DP{}=20),1,""),IF($G${}=45,IF(AND(DG{}>8,DP{}=20),1,""))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['DX{}'.format(row)] = '=IF($H${}=25,IF(AND(DH{}>4,DQ{}=20),1,""),IF($H${}=30,IF(AND(DH{}>5,DQ{}=20),1,""),IF($H${}=35,IF(AND(DH{}>6,DQ{}=20),1,""),IF($H${}=40,IF(AND(DH{}>7,DQ{}=20),1,""),IF($H${}=45,IF(AND(DH{}>8,DQ{}=20),1,""))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['DY{}'.format(row)] = '=IF($I${}=25,IF(AND(DI{}>4,DR{}=20),1,""),IF($I${}=30,IF(AND(DI{}>5,DR{}=20),1,""),IF($I${}=35,IF(AND(DI{}>6,DR{}=20),1,""),IF($I${}=40,IF(AND(DI{}>7,DR{}=20),1,""),IF($I${}=45,IF(AND(DI{}>8,DR{}=20),1,""))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['DZ{}'.format(row)] = '=IF($J${}=25,IF(AND(DJ{}>4,DS{}=20),1,""),IF($J${}=30,IF(AND(DJ{}>5,DS{}=20),1,""),IF($J${}=35,IF(AND(DJ{}>6,DS{}=20),1,""),IF($J${}=40,IF(AND(DJ{}>7,DS{}=20),1,""),IF($J${}=45,IF(AND(DJ{}>8,DS{}=20),1,""))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

    # Z Score iterasi 5
    ws['EB1'] = 'NAMA SISWA'
    ws['EC1'] = 'NOMOR NF'
    ws['ED1'] = 'KELAS'
    ws['EE1'] = 'NAMA SEKOLAH'
    ws['EF1'] = 'LOKASI'
    ws['EG1'] = 'MAT'
    ws['EH1'] = 'BIO'
    ws['EI1'] = 'FIS'
    ws['EJ1'] = 'KIM'
    ws['EK1'] = 'JML'
    ws['EL1'] = 'Z_MAT'
    ws['EM1'] = 'Z_BIO'
    ws['EN1'] = 'Z_FIS'
    ws['EO1'] = 'Z_KIM'
    ws['EP1'] = 'S_MAT'
    ws['EQ1'] = 'S_BIO'
    ws['ER1'] = 'S_FIS'
    ws['ES1'] = 'S_KIM'
    ws['ET1'] = 'S_JML'
    ws['EU1'] = 'RANK NAS.'
    ws['EV1'] = 'RANK LOK.'

    ws['EL1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EM1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EN1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EO1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EP1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EQ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['ER1'].font = Font(bold=False, name='Calibri', size=11)
    ws['ES1'].font = Font(bold=False, name='Calibri', size=11)
    ws['ET1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EU1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EV1'].font = Font(bold=False, name='Calibri', size=11)

    # FILL
    ws['EB1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['EC1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['ED1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['EE1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['EF1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['EG1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['EH1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['EI1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['EJ1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['EK1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['EL1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['EM1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['EN1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['EO1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['EP1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['EQ1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['ER1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['ES1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['ET1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['EU1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['EV1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')

    # tambahan
    ws['EW1'] = 'MAT_20'
    ws['EX1'] = 'BIO_20'
    ws['EY1'] = 'FIS_20'
    ws['EZ1'] = 'KIM_20'
    ws['EW1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EX1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EY1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EZ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EW1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['EX1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['EY1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['EZ1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')

    for row in range(2, q+1):
        # Tambahan
        ws['EB{}'.format(row)] = '=DB{}'.format(row)
        ws['EC{}'.format(row)] = '=DC{}'.format(row, row)
        ws['ED{}'.format(row)] = '=DD{}'.format(row, row)
        ws['EE{}'.format(row)] = '=DE{}'.format(row, row)
        ws['EF{}'.format(row)] = '=DF{}'.format(row, row)
        ws['EG{}'.format(row)] = '=IF(DG{}="","",DG{})'.format(row, row)
        ws['EH{}'.format(row)] = '=IF(DH{}="","",DH{})'.format(row, row)
        ws['EI{}'.format(row)] = '=IF(DI{}="","",DI{})'.format(row, row)
        ws['EJ{}'.format(row)] = '=IF(DJ{}="","",DJ{})'.format(row, row)
        ws['EK{}'.format(row)] = '=IF(DK{}="","",DK{})'.format(row, row)

        ws['EL{}'.format(
            row)] = '=IFERROR(ROUND(IF(EG{}="","",(EG{}-EG${})/EG${}),2),"")'.format(row, row, r, s)
        ws['EM{}'.format(
            row)] = '=IFERROR(ROUND(IF(EH{}="","",(EH{}-EH${})/EH${}),2),"")'.format(row, row, r, s)
        ws['EN{}'.format(
            row)] = '=IFERROR(ROUND(IF(EI{}="","",(EI{}-EI${})/EI${}),2),"")'.format(row, row, r, s)
        ws['EO{}'.format(
            row)] = '=IFERROR(ROUND(IF(EJ{}="","",(EJ{}-EJ${})/EJ${}),2),"")'.format(row, row, r, s)

        ws['EP{}'.format(
            row)] = '=IFERROR(ROUND(IF(EG{}="","",IF(70+30*EL{}/$EL${}<20,20,70+30*EL{}/$EL${})),2),"")'.format(row, row, r, row, r)
        ws['EQ{}'.format(
            row)] = '=IFERROR(ROUND(IF(EH{}="","",IF(70+30*EM{}/$EM{}<20,20,70+30*EM{}/$EM${})),2),"")'.format(row, row, r, row, r)
        ws['ER{}'.format(
            row)] = '=IFERROR(ROUND(IF(EI{}="","",IF(70+30*EN{}/$EN${}<20,20,70+30*EN{}/$EN${})),2),"")'.format(row, row, r, row, r)
        ws['ES{}'.format(
            row)] = '=IFERROR(ROUND(IF(EJ{}="","",IF(70+30*EO{}/$EO${}<20,20,70+30*EO{}/$EO${})),2),"")'.format(row, row, r, row, r)

        ws['ET{}'.format(row)] = '=IF(SUM(EP{}:ES{})=0,"",SUM(EP{}:ES{}))'.format(
            row, row, row, row)
        ws['EU{}'.format(row)] = '=IF(ET{}="","",RANK(ET{},$ET$2:$ET${}))'.format(
            row, row, q)
        ws['EV{}'.format(
            row)] = '=IF(EU{}="","",COUNTIFS($EF$2:$EF${},EF{},$EU$2:$EU${},"<"&EU{})+1)'.format(row, q, row, q, row)
        # TAMBAHAN
        ws['EW{}'.format(row)] = '=IF($G${}=25,IF(AND(EG{}>4,EP{}=20),1,""),IF($G${}=30,IF(AND(EG{}>5,EP{}=20),1,""),IF($G${}=35,IF(AND(EG{}>6,EP{}=20),1,""),IF($G${}=40,IF(AND(EG{}>7,EP{}=20),1,""),IF($G${}=45,IF(AND(EG{}>8,EP{}=20),1,""))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['EX{}'.format(row)] = '=IF($H${}=25,IF(AND(EH{}>4,EQ{}=20),1,""),IF($H${}=30,IF(AND(EH{}>5,EQ{}=20),1,""),IF($H${}=35,IF(AND(EH{}>6,EQ{}=20),1,""),IF($H${}=40,IF(AND(EH{}>7,EQ{}=20),1,""),IF($H${}=45,IF(AND(EH{}>8,EQ{}=20),1,""))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['EY{}'.format(row)] = '=IF($I${}=25,IF(AND(EI{}>4,ER{}=20),1,""),IF($I${}=30,IF(AND(EI{}>5,ER{}=20),1,""),IF($I${}=35,IF(AND(EI{}>6,ER{}=20),1,""),IF($I${}=40,IF(AND(EI{}>7,ER{}=20),1,""),IF($I${}=45,IF(AND(EI{}>8,ER{}=20),1,""))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['EZ{}'.format(row)] = '=IF($J${}=25,IF(AND(EJ{}>4,ES{}=20),1,""),IF($J${}=30,IF(AND(EJ{}>5,ES{}=20),1,""),IF($J${}=35,IF(AND(EJ{}>6,ES{}=20),1,""),IF($J${}=40,IF(AND(EJ{}>7,ES{}=20),1,""),IF($J${}=45,IF(AND(EJ{}>8,ES{}=20),1,""))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

    # Mengubah 'KELAS' sesuai dengan nilai yang dipilih dari selectbox 'KELAS'
    kelas = KELAS.lower().replace(" ", "")
    semester = SEMESTER.lower()
    tahun = TAHUN.replace("-", "")
    penilaian = PENILAIAN.lower()
    kurikulum = KURIKULUM.lower()

    path_file = f"{kelas}_{penilaian}_{semester}_{kurikulum}_{tahun}_nilai_std.xlsx"

    # Simpan file ke direktori temporer
    temp_dir = tempfile.gettempdir()
    file_path = temp_dir + '/' + path_file
    wb.save(file_path)

    st.success(
        "File siap diunduh!")

    # Tombol unduh file
    with open(file_path, "rb") as f:
        bytes_data = f.read()
    st.download_button(label="Unduh File", data=bytes_data,
                       file_name=path_file)

    st.warning(
        "Buka file unduhan, klik 'Enable Editing' dan 'Save'")
