import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill
import tempfile
from PIL import Image

# menghilangkan hamburger
st.markdown("""
<style>
.css-1rs6os.edgvbvh3
{
    visibility:hidden;
}
.css-1lsmgbg.egzxvld0
{
    visibility:hidden;
}
</style>
""", unsafe_allow_html=True)

image = Image.open('logo resmi nf resize.png')
st.image(image)

st.title("Olah Nilai Standar")
st.header("10-11 SMA IPS")

col6 = st.container()

with col6:
    KELAS = st.selectbox(
        "KELAS",
        ("--Pilih Kelas--", "10 SMA IPS", "11 SMA IPS"))

col7 = st.container()

with col7:
    SEMESTER = st.selectbox(
        "SEMESTER",
        ("--Pilih Semester--", "SEMESTER 1", "SEMESTER 2"))

col8 = st.container()

with col8:
    PENILAIAN = st.selectbox(
        "PENILAIAN",
        ("--Pilih Penilaian--", "PENILAIAN TENGAH SEMESTER", "PENILAIAN AKHIR SEMESTER", "PENILAIAN AKHIR TAHUN"))

col9 = st.container()

with col9:
    KURIKULUM = st.selectbox(
        "KURIKULUM",
        ("--Pilih Kurikulum--", "K13", "KM"))

TAHUN = st.text_input("Masukkan Tahun Ajaran", placeholder="contoh: 2022-2023")

col1, col2, col3, col4, col5, col6, col7 = st.columns(7)

with col1:
    MTK = st.selectbox(
        "JML. SOAL MAT.",
        (20, 25, 30, 35, 40, 45))

with col2:
    IND = st.selectbox(
        "JML. SOAL IND.",
        (20, 25, 30, 35, 40, 45))

with col3:
    ENG = st.selectbox(
        "JML. SOAL ENG.",
        (20, 25, 30, 35, 40, 45))

with col4:
    SEJ = st.selectbox(
        "JML. SOAL SEJ.",
        (20, 25, 30, 35, 40, 45))

with col5:
    GEO = st.selectbox(
        "JML. SOAL GEO.",
        (20, 25, 30, 35, 40, 45))

with col6:
    EKO = st.selectbox(
        "JML. SOAL EKO.",
        (20, 25, 30, 35, 40, 45))

with col7:
    SOS = st.selectbox(
        "JML. SOAL SOS.",
        (20, 25, 30, 35, 40, 45))


JML_SOAL_MAT = MTK
JML_SOAL_IND = IND
JML_SOAL_ENG = ENG
JML_SOAL_SEJ = SEJ
JML_SOAL_GEO = GEO
JML_SOAL_EKO = EKO
JML_SOAL_SOS = SOS

uploaded_file = st.file_uploader('Letakkan file excel IPS', type='xlsx')

if uploaded_file is not None:
    wb = openpyxl.load_workbook(uploaded_file)
    ws = wb['Sheet1']

    q = len(ws['K'])
    r = len(ws['K'])+2
    s = len(ws['K'])+3
    t = len(ws['K'])+4
    u = len(ws['K'])+5
    v = len(ws['K'])+6
    w = len(ws['K'])+7
    x = len(ws['K'])+8

    ws['G{}'.format(r)] = "=ROUND(AVERAGE(G2:G{}),2)".format(q)
    ws['H{}'.format(r)] = "=ROUND(AVERAGE(H2:H{}),2)".format(q)
    ws['I{}'.format(r)] = "=ROUND(AVERAGE(I2:I{}),2)".format(q)
    ws['J{}'.format(r)] = "=ROUND(AVERAGE(J2:J{}),2)".format(q)
    ws['K{}'.format(r)] = "=ROUND(AVERAGE(K2:K{}),2)".format(q)
    ws['L{}'.format(r)] = "=ROUND(AVERAGE(L2:L{}),2)".format(q)
    ws['M{}'.format(r)] = "=ROUND(AVERAGE(M2:M{}),2)".format(q)
    ws['N{}'.format(r)] = "=ROUND(AVERAGE(N2:N{}),2)".format(q)

    ws['G{}'.format(s)] = "=STDEV(G2:G{})".format(q)
    ws['H{}'.format(s)] = "=STDEV(H2:H{})".format(q)
    ws['I{}'.format(s)] = "=STDEV(I2:I{})".format(q)
    ws['J{}'.format(s)] = "=STDEV(J2:J{})".format(q)
    ws['K{}'.format(s)] = "=STDEV(K2:K{})".format(q)
    ws['L{}'.format(s)] = "=STDEV(L2:L{})".format(q)
    ws['M{}'.format(s)] = "=STDEV(M2:M{})".format(q)

    ws['G{}'.format(t)] = "=MAX(G2:G{})".format(q)
    ws['H{}'.format(t)] = "=MAX(H2:H{})".format(q)
    ws['I{}'.format(t)] = "=MAX(I2:I{})".format(q)
    ws['J{}'.format(t)] = "=MAX(J2:J{})".format(q)
    ws['K{}'.format(t)] = "=MAX(K2:K{})".format(q)
    ws['L{}'.format(t)] = "=MAX(L2:L{})".format(q)
    ws['M{}'.format(t)] = "=MAX(M2:M{})".format(q)

    ws['N{}'.format(r)] = "=MAX(N2:N{})".format(q)
    ws['O{}'.format(r)] = "=MAX(O2:O{})".format(q)
    ws['P{}'.format(r)] = "=MAX(P2:P{})".format(q)
    ws['Q{}'.format(r)] = "=MAX(Q2:Q{})".format(q)
    ws['R{}'.format(r)] = "=MAX(R2:R{})".format(q)
    ws['S{}'.format(r)] = "=MAX(S2:S{})".format(q)
    ws['T{}'.format(r)] = "=MAX(T2:T{})".format(q)
    ws['U{}'.format(r)] = "=MAX(U2:U{})".format(q)
    ws['V{}'.format(r)] = "=MAX(V2:V{})".format(q)
    ws['W{}'.format(r)] = "=MAX(W2:W{})".format(q)
    ws['X{}'.format(r)] = "=MAX(X2:X{})".format(q)
    ws['Y{}'.format(r)] = "=MAX(Y2:Y{})".format(q)
    ws['Z{}'.format(r)] = "=MAX(Z2:Z{})".format(q)
    ws['AA{}'.format(r)] = "=MAX(AA2:AA{})".format(q)

    ws['AB{}'.format(r)] = "=ROUND(MAX(AB2:AB{}),2)".format(q)
    ws['AC{}'.format(r)] = "=MAX(AC2:AC{})".format(q)

    ws['G{}'.format(u)] = "=MIN(G2:G{})".format(q)
    ws['H{}'.format(u)] = "=MIN(H2:H{})".format(q)
    ws['I{}'.format(u)] = "=MIN(I2:I{})".format(q)
    ws['J{}'.format(u)] = "=MIN(J2:J{})".format(q)
    ws['K{}'.format(u)] = "=MIN(K2:K{})".format(q)
    ws['L{}'.format(u)] = "=MIN(L2:L{})".format(q)
    ws['M{}'.format(u)] = "=MIN(M2:M{})".format(q)
    ws['N{}'.format(u)] = "=MIN(N2:N{})".format(q)

    ws['V{}'.format(s)] = "=MIN(V2:V{})".format(q)
    ws['W{}'.format(s)] = "=MIN(W2:W{})".format(q)
    ws['X{}'.format(s)] = "=MIN(X2:X{})".format(q)
    ws['Y{}'.format(s)] = "=MIN(Y2:Y{})".format(q)
    ws['Z{}'.format(s)] = "=MIN(Z2:Z{})".format(q)
    ws['AA{}'.format(s)] = "=MIN(AA2:AA{})".format(q)
    ws['AB{}'.format(s)] = "=MIN(AB2:AB{})".format(q)
    ws['AC{}'.format(s)] = "=MIN(AC2:AC{})".format(q)

    ws['V{}'.format(t)] = "=ROUND(AVERAGE(V2:V{}),2)".format(q)
    ws['W{}'.format(t)] = "=ROUND(AVERAGE(W2:W{}),2)".format(q)
    ws['X{}'.format(t)] = "=ROUND(AVERAGE(X2:X{}),2)".format(q)
    ws['Y{}'.format(t)] = "=ROUND(AVERAGE(Y2:Y{}),2)".format(q)
    ws['Z{}'.format(t)] = "=ROUND(AVERAGE(Z2:Z{}),2)".format(q)
    ws['AA{}'.format(t)] = "=ROUND(AVERAGE(AA2:AA{}),2)".format(q)
    ws['AB{}'.format(t)] = "=ROUND(AVERAGE(AB2:AB{}),2)".format(q)
    ws['AC{}'.format(t)] = "=ROUND(AVERAGE(AC2:AC{}),2)".format(q)

    ws['AF{}'.format(r)] = "=SUM(AF2:AF{})".format(q)
    ws['AG{}'.format(r)] = "=SUM(AG2:AG{})".format(q)
    ws['AH{}'.format(r)] = "=SUM(AH2:AH{})".format(q)
    ws['AI{}'.format(r)] = "=SUM(AI2:AI{})".format(q)
    ws['AJ{}'.format(r)] = "=SUM(AJ2:AJ{})".format(q)
    ws['AK{}'.format(r)] = "=SUM(AK2:AK{})".format(q)
    ws['AL{}'.format(r)] = "=SUM(AL2:AL{})".format(q)

    # Jumlah Soal
    ws['F{}'.format(v)] = 'JUMLAH SOAL'
    ws['G{}'.format(v)] = JML_SOAL_MAT
    ws['H{}'.format(v)] = JML_SOAL_IND
    ws['I{}'.format(v)] = JML_SOAL_ENG
    ws['J{}'.format(v)] = JML_SOAL_SEJ
    ws['K{}'.format(v)] = JML_SOAL_GEO
    ws['L{}'.format(v)] = JML_SOAL_EKO
    ws['M{}'.format(v)] = JML_SOAL_SOS

    # new
    # iterasi 1 rata-rata - 1
    # rata" MTK ke MTK tambahan dan mapel MTK awal
    ws['AS{}'.format(r)] = "=IF($AF${}=0,$G${},$G${}-1)".format(r, r, r)
    ws['AS{}'.format(s)] = "=STDEV(AS2:AS{})".format(q)
    ws['AS{}'.format(t)] = "=MAX(AS2:AS{})".format(q)
    ws['AS{}'.format(u)] = "=MIN(AS2:AS{})".format(q)
    # rata" IND ke IND tambahan dan mapel IND awal
    ws['AT{}'.format(r)] = "=IF($AG${}=0,$H${},$H${}-1)".format(r, r, r)
    ws['AT{}'.format(s)] = "=STDEV(AT2:AT{})".format(q)
    ws['AT{}'.format(t)] = "=MAX(AT2:AT{})".format(q)
    ws['AT{}'.format(u)] = "=MIN(AT2:AT{})".format(q)
    # rata" ENG ke ENG tambahan dan mapel ENG awal
    ws['AU{}'.format(r)] = "=IF($AH${}=0,$I${},$I${}-1)".format(r, r, r)
    ws['AU{}'.format(s)] = "=STDEV(AU2:AU{})".format(q)
    ws['AU{}'.format(t)] = "=MAX(AU2:AU{})".format(q)
    ws['AU{}'.format(u)] = "=MIN(AU2:AU{})".format(q)
    # rata" SEJ ke SEJ tambahan dan mapel SEJ awal
    ws['AV{}'.format(r)] = "=IF($AI${}=0,$J${},$J${}-1)".format(r, r, r)
    ws['AV{}'.format(s)] = "=STDEV(AV2:AV{})".format(q)
    ws['AV{}'.format(t)] = "=MAX(AV2:AV{})".format(q)
    ws['AV{}'.format(u)] = "=MIN(AV2:AV{})".format(q)
    # rata" GEO ke GEO tambahan dan mapel GEO awal
    ws['AW{}'.format(r)] = "=IF($AJ${}=0,$K${},$K${}-1)".format(r, r, r)
    ws['AW{}'.format(s)] = "=STDEV(AW2:AW{})".format(q)
    ws['AW{}'.format(t)] = "=MAX(AW2:AW{})".format(q)
    ws['AW{}'.format(u)] = "=MIN(AW2:AW{})".format(q)
    # rata" EKO ke EKO tambahan dan mapel EKO awal
    ws['AX{}'.format(r)] = "=IF($AK${}=0,$L${},$L${}-1)".format(r, r, r)
    ws['AX{}'.format(s)] = "=STDEV(AX2:AX{})".format(q)
    ws['AX{}'.format(t)] = "=MAX(AX2:AX{})".format(q)
    ws['AX{}'.format(u)] = "=MIN(AX2:AX{})".format(q)
    # rata" SOS ke SOS tambahan dan mapel SOS awal
    ws['AY{}'.format(r)] = "=IF($AL${}=0,$M${},$M${}-1)".format(r, r, r)
    ws['AY{}'.format(s)] = "=STDEV(AY2:AY{})".format(q)
    ws['AY{}'.format(t)] = "=MAX(AY2:AY{})".format(q)
    ws['AY{}'.format(u)] = "=MIN(AY2:AY{})".format(q)
    # jml MAPEL
    ws['AZ{}'.format(r)] = "=ROUND(AVERAGE(AZ2:AZ{}),2)".format(q)
    ws['AZ{}'.format(t)] = "=MAX(AZ2:AZ{})".format(q)
    ws['AZ{}'.format(u)] = "=MIN(AZ2:AZ{})".format(q)
    # MAX Z SCORE
    ws['BA{}'.format(r)] = "=MAX(BA2:BA{})".format(q)
    ws['BB{}'.format(r)] = "=MAX(BB2:BB{})".format(q)
    ws['BC{}'.format(r)] = "=MAX(BC2:BC{})".format(q)
    ws['BD{}'.format(r)] = "=MAX(BD2:BD{})".format(q)
    ws['BE{}'.format(r)] = "=MAX(BE2:BE{})".format(q)
    ws['BF{}'.format(r)] = "=MAX(BF2:BF{})".format(q)
    ws['BG{}'.format(r)] = "=MAX(BG2:BG{})".format(q)
    # NILAI STANDAR MTK
    ws['BH{}'.format(r)] = "=MAX(BH2:BH{})".format(q)
    ws['BH{}'.format(s)] = "=MIN(BH2:BH{})".format(q)
    ws['BH{}'.format(t)] = "=ROUND(AVERAGE(BH2:BH{}),2)".format(q)
    # NILAI STANDAR IND
    ws['BI{}'.format(r)] = "=MAX(BI2:BI{})".format(q)
    ws['BI{}'.format(s)] = "=MIN(BI2:BI{})".format(q)
    ws['BI{}'.format(t)] = "=ROUND(AVERAGE(BI2:BI{}),2)".format(q)
    # NILAI STANDAR ENG
    ws['BJ{}'.format(r)] = "=MAX(BJ2:BJ{})".format(q)
    ws['BJ{}'.format(s)] = "=MIN(BJ2:BJ{})".format(q)
    ws['BJ{}'.format(t)] = "=ROUND(AVERAGE(BJ2:BJ{}),2)".format(q)
    # NILAI STANDAR SEJ
    ws['BK{}'.format(r)] = "=MAX(BK2:BK{})".format(q)
    ws['BK{}'.format(s)] = "=MIN(BK2:BK{})".format(q)
    ws['BK{}'.format(t)] = "=ROUND(AVERAGE(BK2:BK{}),2)".format(q)
    # NILAI STANDAR GEO
    ws['BL{}'.format(r)] = "=MAX(BL2:BL{})".format(q)
    ws['BL{}'.format(s)] = "=MIN(BL2:BL{})".format(q)
    ws['BL{}'.format(t)] = "=ROUND(AVERAGE(BL2:BL{}),2)".format(q)
    # NILAI STANDAR EKO
    ws['BM{}'.format(r)] = "=MAX(BM2:BM{})".format(q)
    ws['BM{}'.format(s)] = "=MIN(BM2:BM{})".format(q)
    ws['BM{}'.format(t)] = "=ROUND(AVERAGE(BM2:BM{}),2)".format(q)
    # NILAI STANDAR SOS
    ws['BN{}'.format(r)] = "=MAX(BN2:BN{})".format(q)
    ws['BN{}'.format(s)] = "=MIN(BN2:BN{})".format(q)
    ws['BN{}'.format(t)] = "=ROUND(AVERAGE(BN2:BN{}),2)".format(q)
    # NILAI STANDAR S.JML
    ws['BO{}'.format(r)] = "=MAX(BO2:BO{})".format(q)
    ws['BO{}'.format(s)] = "=MIN(BO2:BO{})".format(q)
    ws['BO{}'.format(t)] = "=ROUND(AVERAGE(BO2:BO{}),2)".format(q)

    # TAMBAHAN
    ws['BR{}'.format(r)] = "=SUM(BR2:BR{})".format(q)
    ws['BS{}'.format(r)] = "=SUM(BS2:BS{})".format(q)
    ws['BT{}'.format(r)] = "=SUM(BT2:BT{})".format(q)
    ws['BU{}'.format(r)] = "=SUM(BU2:BU{})".format(q)
    ws['BV{}'.format(r)] = "=SUM(BV2:BV{})".format(q)
    ws['BW{}'.format(r)] = "=SUM(BW2:BW{})".format(q)
    ws['BX{}'.format(r)] = "=SUM(BX2:BX{})".format(q)

    # iterasi 2 rata-rata - 2
    # rata" MTK ke MTK tambahan dan mapel MTK awal
    ws['CE{}'.format(r)] = "=IF($BR${}=0,$AS${},$AS${}-1)".format(r, r, r)
    ws['CE{}'.format(s)] = "=STDEV(CE2:CE{})".format(q)
    ws['CE{}'.format(t)] = "=MAX(CE2:CE{})".format(q)
    ws['CE{}'.format(u)] = "=MIN(CE2:CE{})".format(q)
    # rata" IND ke IND tambahan dan mapel IND awal
    ws['CF{}'.format(r)] = "=IF($BS${}=0,$AT${},$AT${}-1)".format(r, r, r)
    ws['CF{}'.format(s)] = "=STDEV(CF2:CF{})".format(q)
    ws['CF{}'.format(t)] = "=MAX(CF2:CF{})".format(q)
    ws['CF{}'.format(u)] = "=MIN(CF2:CF{})".format(q)
    # rata" ENG ke ENG tambahan dan mapel ENG awal
    ws['CG{}'.format(r)] = "=IF($BT${}=0,$AU${},$AU${}-1)".format(r, r, r)
    ws['CG{}'.format(s)] = "=STDEV(CG2:CG{})".format(q)
    ws['CG{}'.format(t)] = "=MAX(CG2:CG{})".format(q)
    ws['CG{}'.format(u)] = "=MIN(CG2:CG{})".format(q)
    # rata" SEJ ke SEJ tambahan dan mapel SEJ awal
    ws['CH{}'.format(r)] = "=IF($BU${}=0,$AV${},$AV${}-1)".format(r, r, r)
    ws['CH{}'.format(s)] = "=STDEV(CH2:CH{})".format(q)
    ws['CH{}'.format(t)] = "=MAX(CH2:CH{})".format(q)
    ws['CH{}'.format(u)] = "=MIN(CH2:CH{})".format(q)
    # rata" GEO ke GEO tambahan dan mapel GEO awal
    ws['CI{}'.format(r)] = "=IF($BV${}=0,$AW${},$AW${}-1)".format(r, r, r)
    ws['CI{}'.format(s)] = "=STDEV(CI2:CI{})".format(q)
    ws['CI{}'.format(t)] = "=MAX(CI2:CI{})".format(q)
    ws['CI{}'.format(u)] = "=MIN(CI2:CI{})".format(q)
    # rata" EKO ke EKO tambahan dan mapel EKO awal
    ws['CJ{}'.format(r)] = "=IF($BW${}=0,$AX${},$AX${}-1)".format(r, r, r)
    ws['CJ{}'.format(s)] = "=STDEV(CJ2:CJ{})".format(q)
    ws['CJ{}'.format(t)] = "=MAX(CJ2:CJ{})".format(q)
    ws['CJ{}'.format(u)] = "=MIN(CJ2:CJ{})".format(q)
    # rata" SOS ke SOS tambahan dan mapel SOS awal
    ws['CK{}'.format(r)] = "=IF($BX${}=0,$AY${},$AY${}-1)".format(r, r, r)
    ws['CK{}'.format(s)] = "=STDEV(CK2:CK{})".format(q)
    ws['CK{}'.format(t)] = "=MAX(CK2:CK{})".format(q)
    ws['CK{}'.format(u)] = "=MIN(CK2:CK{})".format(q)
    # jml MAPEL
    ws['CL{}'.format(r)] = "=ROUND(AVERAGE(CL2:CL{}),2)".format(q)
    ws['CL{}'.format(t)] = "=MAX(CL2:CL{})".format(q)
    ws['CL{}'.format(u)] = "=MIN(CL2:CL{})".format(q)
    # MAX Z SCORE
    ws['CM{}'.format(r)] = "=MAX(CM2:CM{})".format(q)
    ws['CN{}'.format(r)] = "=MAX(CN2:CN{})".format(q)
    ws['CO{}'.format(r)] = "=MAX(CO2:CO{})".format(q)
    ws['CP{}'.format(r)] = "=MAX(CP2:CP{})".format(q)
    ws['CQ{}'.format(r)] = "=MAX(CQ2:CQ{})".format(q)
    ws['CR{}'.format(r)] = "=MAX(CR2:CR{})".format(q)
    ws['CS{}'.format(r)] = "=MAX(CS2:CS{})".format(q)
    # NILAI STANDAR MTK
    ws['CT{}'.format(r)] = "=MAX(CT2:CT{})".format(q)
    ws['CT{}'.format(s)] = "=MIN(CT2:CT{})".format(q)
    ws['CT{}'.format(t)] = "=ROUND(AVERAGE(CT2:CT{}),2)".format(q)
    # NILAI STANDAR IND
    ws['CU{}'.format(r)] = "=MAX(CU2:CU{})".format(q)
    ws['CU{}'.format(s)] = "=MIN(CU2:CU{})".format(q)
    ws['CU{}'.format(t)] = "=ROUND(AVERAGE(CU2:CU{}),2)".format(q)
    # NILAI STANDAR ENG
    ws['CV{}'.format(r)] = "=MAX(CV2:CV{})".format(q)
    ws['CV{}'.format(s)] = "=MIN(CV2:CV{})".format(q)
    ws['CV{}'.format(t)] = "=ROUND(AVERAGE(CV2:CV{}),2)".format(q)
    # NILAI STANDAR SEJ
    ws['CW{}'.format(r)] = "=MAX(CW2:CW{})".format(q)
    ws['CW{}'.format(s)] = "=MIN(CW2:CW{})".format(q)
    ws['CW{}'.format(t)] = "=ROUND(AVERAGE(CW2:CW{}),2)".format(q)
    # NILAI STANDAR GEO
    ws['CX{}'.format(r)] = "=MAX(CX2:CX{})".format(q)
    ws['CX{}'.format(s)] = "=MIN(CX2:CX{})".format(q)
    ws['CX{}'.format(t)] = "=ROUND(AVERAGE(CX2:CX{}),2)".format(q)
    # NILAI STANDAR EKO
    ws['CY{}'.format(r)] = "=MAX(CY2:CY{})".format(q)
    ws['CY{}'.format(s)] = "=MIN(CY2:CY{})".format(q)
    ws['CY{}'.format(t)] = "=ROUND(AVERAGE(CY2:CY{}),2)".format(q)
    # NILAI STANDAR SOS
    ws['CZ{}'.format(r)] = "=MAX(CZ2:CZ{})".format(q)
    ws['CZ{}'.format(s)] = "=MIN(CZ2:CZ{})".format(q)
    ws['CZ{}'.format(t)] = "=ROUND(AVERAGE(CZ2:CZ{}),2)".format(q)
    # NILAI STANDAR S.JML
    ws['DA{}'.format(r)] = "=MAX(DA2:DA{})".format(q)
    ws['DA{}'.format(s)] = "=MIN(DA2:DA{})".format(q)
    ws['DA{}'.format(t)] = "=ROUND(AVERAGE(DA2:DA{}),2)".format(q)

    # TAMBAHAN
    ws['DD{}'.format(r)] = "=SUM(DD2:DD{})".format(q)
    ws['DE{}'.format(r)] = "=SUM(DE2:DE{})".format(q)
    ws['DF{}'.format(r)] = "=SUM(DF2:DF{})".format(q)
    ws['DG{}'.format(r)] = "=SUM(DG2:DG{})".format(q)
    ws['DH{}'.format(r)] = "=SUM(DH2:DH{})".format(q)
    ws['DI{}'.format(r)] = "=SUM(DI2:DI{})".format(q)
    ws['DJ{}'.format(r)] = "=SUM(DJ2:DJ{})".format(q)

    # iterasi 3 rata-rata - 3
    # rata" MTK ke MTK tambahan dan mapel MTK awal
    ws['DQ{}'.format(r)] = "=IF($DD${}=0,$CE${},$CE${}-1)".format(r, r, r)
    ws['DQ{}'.format(s)] = "=STDEV(DQ2:DQ{})".format(q)
    ws['DQ{}'.format(t)] = "=MAX(DQ2:DQ{})".format(q)
    ws['DQ{}'.format(u)] = "=MIN(DQ2:DQ{})".format(q)
    # rata" IND ke IND tambahan dan mapel IND awal
    ws['DR{}'.format(r)] = "=IF($DE${}=0,$CF${},$CF${}-1)".format(r, r, r)
    ws['DR{}'.format(s)] = "=STDEV(DR2:DR{})".format(q)
    ws['DR{}'.format(t)] = "=MAX(DR2:DR{})".format(q)
    ws['DR{}'.format(u)] = "=MIN(DR2:DR{})".format(q)
    # rata" ENG ke ENG tambahan dan mapel ENG awal
    ws['DS{}'.format(r)] = "=IF($DF${}=0,$CG${},$CG{}-1)".format(r, r, r)
    ws['DS{}'.format(s)] = "=STDEV(DS2:DS{})".format(q)
    ws['DS{}'.format(t)] = "=MAX(DS2:DS{})".format(q)
    ws['DS{}'.format(u)] = "=MIN(DS2:DS{})".format(q)
    # rata" SEJ ke SEJ tambahan dan mapel SEJ awal
    ws['DT{}'.format(r)] = "=IF($DG${}=0,$CH${},$CH${}-1)".format(r, r, r)
    ws['DT{}'.format(s)] = "=STDEV(DT2:DT{})".format(q)
    ws['DT{}'.format(t)] = "=MAX(DT2:DT{})".format(q)
    ws['DT{}'.format(u)] = "=MIN(DT2:DT{})".format(q)
    # rata" GEO ke GEO tambahan dan mapel GEO awal
    ws['DU{}'.format(r)] = "=IF($DH${}=0,$CI${},$CI${}-1)".format(r, r, r)
    ws['DU{}'.format(s)] = "=STDEV(DU2:DU{})".format(q)
    ws['DU{}'.format(t)] = "=MAX(DU2:DU{})".format(q)
    ws['DU{}'.format(u)] = "=MIN(DU2:DU{})".format(q)
    # rata" EKO ke EKO tambahan dan mapel EKO awal
    ws['DV{}'.format(r)] = "=IF($DI${}=0,$CJ${},$CJ${}-1)".format(r, r, r)
    ws['DV{}'.format(s)] = "=STDEV(DV2:DV{})".format(q)
    ws['DV{}'.format(t)] = "=MAX(DV2:DV{})".format(q)
    ws['DV{}'.format(u)] = "=MIN(DV2:DV{})".format(q)
    # rata" SOS ke SOS tambahan dan mapel SOS awal
    ws['DW{}'.format(r)] = "=IF($DJ${}=0,$CK${},$CK${}-1)".format(r, r, r)
    ws['DW{}'.format(s)] = "=STDEV(DW2:DW{})".format(q)
    ws['DW{}'.format(t)] = "=MAX(DW2:DW{})".format(q)
    ws['DW{}'.format(u)] = "=MIN(DW2:DW{})".format(q)
    # jml MAPEL
    ws['DX{}'.format(r)] = "=ROUND(AVERAGE(DX2:DX{}),2)".format(q)
    ws['DX{}'.format(t)] = "=MAX(DX2:DX{})".format(q)
    ws['DX{}'.format(u)] = "=MIN(DX2:DX{})".format(q)
    # MAX Z SCORE
    ws['DY{}'.format(r)] = "=MAX(DY2:DY{})".format(q)
    ws['DZ{}'.format(r)] = "=MAX(DZ2:DZ{})".format(q)
    ws['EA{}'.format(r)] = "=MAX(EA2:EA{})".format(q)
    ws['EB{}'.format(r)] = "=MAX(EB2:EB{})".format(q)
    ws['EC{}'.format(r)] = "=MAX(EC2:EC{})".format(q)
    ws['ED{}'.format(r)] = "=MAX(ED2:ED{})".format(q)
    ws['EE{}'.format(r)] = "=MAX(EE2:EE{})".format(q)
    # NILAI STANDAR MTK
    ws['EF{}'.format(r)] = "=MAX(EF2:EF{})".format(q)
    ws['EF{}'.format(s)] = "=MIN(EF2:EF{})".format(q)
    ws['EF{}'.format(t)] = "=ROUND(AVERAGE(EF2:EF{}),2)".format(q)
    # NILAI STANDAR IND
    ws['EG{}'.format(r)] = "=MAX(EG2:EG{})".format(q)
    ws['EG{}'.format(s)] = "=MIN(EG2:EG{})".format(q)
    ws['EG{}'.format(t)] = "=ROUND(AVERAGE(EG2:EG{}),2)".format(q)
    # NILAI STANDAR ENG
    ws['EH{}'.format(r)] = "=MAX(EH2:EH{})".format(q)
    ws['EH{}'.format(s)] = "=MIN(EH2:EH{})".format(q)
    ws['EH{}'.format(t)] = "=ROUND(AVERAGE(EH2:EH{}),2)".format(q)
    # NILAI STANDAR SEJ
    ws['EI{}'.format(r)] = "=MAX(EI2:EI{})".format(q)
    ws['EI{}'.format(s)] = "=MIN(EI2:EI{})".format(q)
    ws['EI{}'.format(t)] = "=ROUND(AVERAGE(EI2:EI{}),2)".format(q)
    # NILAI STANDAR GEO
    ws['EJ{}'.format(r)] = "=MAX(EJ2:EJ{})".format(q)
    ws['EJ{}'.format(s)] = "=MIN(EJ2:EJ{})".format(q)
    ws['EJ{}'.format(t)] = "=ROUND(AVERAGE(EJ2:EJ{}),2)".format(q)
    # NILAI STANDAR EKO
    ws['EK{}'.format(r)] = "=MAX(EK2:EK{})".format(q)
    ws['EK{}'.format(s)] = "=MIN(EK2:EK{})".format(q)
    ws['EK{}'.format(t)] = "=ROUND(AVERAGE(EK2:EK{}),2)".format(q)
    # NILAI STANDAR SOS
    ws['EL{}'.format(r)] = "=MAX(EL2:EL{})".format(q)
    ws['EL{}'.format(s)] = "=MIN(EL2:EL{})".format(q)
    ws['EL{}'.format(t)] = "=ROUND(AVERAGE(EL2:EL{}),2)".format(q)
    # NILAI STANDAR S.JML
    ws['EM{}'.format(r)] = "=MAX(EM2:EM{})".format(q)
    ws['EM{}'.format(s)] = "=MIN(EM2:EM{})".format(q)
    ws['EM{}'.format(t)] = "=ROUND(AVERAGE(EM2:EM{}),2)".format(q)

    # TAMBAHAN
    ws['EP{}'.format(r)] = "=SUM(EP2:EP{})".format(q)
    ws['EQ{}'.format(r)] = "=SUM(EQ2:EQ{})".format(q)
    ws['ER{}'.format(r)] = "=SUM(ER2:ER{})".format(q)
    ws['ES{}'.format(r)] = "=SUM(ES2:ES{})".format(q)
    ws['ET{}'.format(r)] = "=SUM(ET2:ET{})".format(q)
    ws['EU{}'.format(r)] = "=SUM(EU2:EU{})".format(q)
    ws['EV{}'.format(r)] = "=SUM(EV2:EV{})".format(q)

    # iterasi 4 rata-rata - 4
    # rata" MTK ke MTK tambahan dan mapel MTK awal
    ws['FC{}'.format(r)] = "=IF($EP${}=0,$DQ${},$DQ${}-1)".format(r, r, r)
    ws['FC{}'.format(s)] = "=STDEV(FC2:FC{})".format(q)
    ws['FC{}'.format(t)] = "=MAX(FC2:FC{})".format(q)
    ws['FC{}'.format(u)] = "=MIN(FC2:FC{})".format(q)
    # rata" IND ke IND tambahan dan mapel IND awal
    ws['FD{}'.format(r)] = "=IF($EQ${}=0,$DR${},$DR${}-1)".format(r, r, r)
    ws['FD{}'.format(s)] = "=STDEV(FD2:FD{})".format(q)
    ws['FD{}'.format(t)] = "=MAX(FD2:FD{})".format(q)
    ws['FD{}'.format(u)] = "=MIN(FD2:FD{})".format(q)
    # rata" ENG ke ENG tambahan dan mapel ENG awal
    ws['FE{}'.format(r)] = "=IF($ER${}=0,$DS${},$DS{}-1)".format(r, r, r)
    ws['FE{}'.format(s)] = "=STDEV(FE2:FE{})".format(q)
    ws['FE{}'.format(t)] = "=MAX(FE2:FE{})".format(q)
    ws['FE{}'.format(u)] = "=MIN(FE2:FE{})".format(q)
    # rata" SEJ ke SEJ tambahan dan mapel SEJ awal
    ws['FF{}'.format(r)] = "=IF($ES${}=0,$DT${},$DT${}-1)".format(r, r, r)
    ws['FF{}'.format(s)] = "=STDEV(FF2:FF{})".format(q)
    ws['FF{}'.format(t)] = "=MAX(FF2:FF{})".format(q)
    ws['FF{}'.format(u)] = "=MIN(FF2:FF{})".format(q)
    # rata" GEO ke GEO tambahan dan mapel GEO awal
    ws['FG{}'.format(r)] = "=IF($ET${}=0,$DU${},$DU${}-1)".format(r, r, r)
    ws['FG{}'.format(s)] = "=STDEV(FG2:FG{})".format(q)
    ws['FG{}'.format(t)] = "=MAX(FG2:FG{})".format(q)
    ws['FG{}'.format(u)] = "=MIN(FG2:FG{})".format(q)
    # rata" EKO ke EKO tambahan dan mapel EKO awal
    ws['FH{}'.format(r)] = "=IF($EU${}=0,$DV${},$DV${}-1)".format(r, r, r)
    ws['FH{}'.format(s)] = "=STDEV(FH2:FH{})".format(q)
    ws['FH{}'.format(t)] = "=MAX(FH2:FH{})".format(q)
    ws['FH{}'.format(u)] = "=MIN(FH2:FH{})".format(q)
    # rata" SOS ke SOS tambahan dan mapel SOS awal
    ws['FI{}'.format(r)] = "=IF($EV${}=0,$DW${},$DW${}-1)".format(r, r, r)
    ws['FI{}'.format(s)] = "=STDEV(FI2:FI{})".format(q)
    ws['FI{}'.format(t)] = "=MAX(FI2:FI{})".format(q)
    ws['FI{}'.format(u)] = "=MIN(FI2:FI{})".format(q)
    # jml MAPEL
    ws['FJ{}'.format(r)] = "=ROUND(AVERAGE(FJ2:FJ{}),2)".format(q)
    ws['FJ{}'.format(t)] = "=MAX(FJ2:FJ{})".format(q)
    ws['FJ{}'.format(u)] = "=MIN(FJ2:FJ{})".format(q)
    # MAX Z SCORE
    ws['FK{}'.format(r)] = "=MAX(FK2:FK{})".format(q)
    ws['FL{}'.format(r)] = "=MAX(FL2:FL{})".format(q)
    ws['FM{}'.format(r)] = "=MAX(FM2:FM{})".format(q)
    ws['FN{}'.format(r)] = "=MAX(FN2:FN{})".format(q)
    ws['FO{}'.format(r)] = "=MAX(FO2:FO{})".format(q)
    ws['FP{}'.format(r)] = "=MAX(FP2:FP{})".format(q)
    ws['FQ{}'.format(r)] = "=MAX(FQ2:FQ{})".format(q)
    # NILAI STANDAR MTK
    ws['FR{}'.format(r)] = "=MAX(FR2:FR{})".format(q)
    ws['FR{}'.format(s)] = "=MIN(FR2:FR{})".format(q)
    ws['FR{}'.format(t)] = "=ROUND(AVERAGE(FR2:FR{}),2)".format(q)
    # NILAI STANDAR IND
    ws['FS{}'.format(r)] = "=MAX(FS2:FS{})".format(q)
    ws['FS{}'.format(s)] = "=MIN(FS2:FS{})".format(q)
    ws['FS{}'.format(t)] = "=ROUND(AVERAGE(FS2:FS{}),2)".format(q)
    # NILAI STANDAR ENG
    ws['FT{}'.format(r)] = "=MAX(FT2:FT{})".format(q)
    ws['FT{}'.format(s)] = "=MIN(FT2:FT{})".format(q)
    ws['FT{}'.format(t)] = "=ROUND(AVERAGE(FT2:FT{}),2)".format(q)
    # NILAI STANDAR SEJ
    ws['FU{}'.format(r)] = "=MAX(FU2:FU{})".format(q)
    ws['FU{}'.format(s)] = "=MIN(FU2:FU{})".format(q)
    ws['FU{}'.format(t)] = "=ROUND(AVERAGE(FU2:FU{}),2)".format(q)
    # NILAI STANDAR GEO
    ws['FV{}'.format(r)] = "=MAX(FV2:FV{})".format(q)
    ws['FV{}'.format(s)] = "=MIN(FV2:FV{})".format(q)
    ws['FV{}'.format(t)] = "=ROUND(AVERAGE(FV2:FV{}),2)".format(q)
    # NILAI STANDAR EKO
    ws['FW{}'.format(r)] = "=MAX(FW2:FW{})".format(q)
    ws['FW{}'.format(s)] = "=MIN(FW2:FW{})".format(q)
    ws['FW{}'.format(t)] = "=ROUND(AVERAGE(FW2:FW{}),2)".format(q)
    # NILAI STANDAR SOS
    ws['FX{}'.format(r)] = "=MAX(FX2:FX{})".format(q)
    ws['FX{}'.format(s)] = "=MIN(FX2:FX{})".format(q)
    ws['FX{}'.format(t)] = "=ROUND(AVERAGE(FX2:FX{}),2)".format(q)
    # NILAI STANDAR S.JML
    ws['FY{}'.format(r)] = "=MAX(FY2:FY{})".format(q)
    ws['FY{}'.format(s)] = "=MIN(FY2:FY{})".format(q)
    ws['FY{}'.format(t)] = "=ROUND(AVERAGE(FY2:FY{}),2)".format(q)

    # TAMBAHAN
    ws['GB{}'.format(r)] = "=SUM(GB2:GB{})".format(q)
    ws['GC{}'.format(r)] = "=SUM(GC2:GC{})".format(q)
    ws['GD{}'.format(r)] = "=SUM(GD2:GD{})".format(q)
    ws['GE{}'.format(r)] = "=SUM(GE2:GE{})".format(q)
    ws['GF{}'.format(r)] = "=SUM(GF2:GF{})".format(q)
    ws['GG{}'.format(r)] = "=SUM(GG2:GG{})".format(q)
    ws['GH{}'.format(r)] = "=SUM(GH2:GH{})".format(q)

    # iterasi 5 rata-rata - 5
    # rata" MTK ke MTK tambahan dan mapel MTK awal
    ws['GO{}'.format(r)] = "=IF($GB${}=0,$FC${},$FC${}-1)".format(r, r, r)
    ws['GO{}'.format(s)] = "=STDEV(GO2:GO{})".format(q)
    ws['GO{}'.format(t)] = "=MAX(GO2:GO{})".format(q)
    ws['GO{}'.format(u)] = "=MIN(GO2:GO{})".format(q)
    # rata" IND ke IND tambahan dan mapel IND awal
    ws['GP{}'.format(r)] = "=IF($GC${}=0,$FD${},$FD${}-1)".format(r, r, r)
    ws['GP{}'.format(s)] = "=STDEV(GP2:GP{})".format(q)
    ws['GP{}'.format(t)] = "=MAX(GP2:GP{})".format(q)
    ws['GP{}'.format(u)] = "=MIN(GP2:GP{})".format(q)
    # rata" ENG ke ENG tambahan dan mapel ENG awal
    ws['GQ{}'.format(r)] = "=IF($GD${}=0,$FE${},$FE{}-1)".format(r, r, r)
    ws['GQ{}'.format(s)] = "=STDEV(GQ2:GQ{})".format(q)
    ws['GQ{}'.format(t)] = "=MAX(GQ2:GQ{})".format(q)
    ws['GQ{}'.format(u)] = "=MIN(GQ2:GQ{})".format(q)
    # rata" SEJ ke SEJ tambahan dan mapel SEJ awal
    ws['GR{}'.format(r)] = "=IF($GE${}=0,$FF${},$FF${}-1)".format(r, r, r)
    ws['GR{}'.format(s)] = "=STDEV(GR2:GR{})".format(q)
    ws['GR{}'.format(t)] = "=MAX(GR2:GR{})".format(q)
    ws['GR{}'.format(u)] = "=MIN(GR2:GR{})".format(q)
    # rata" GEO ke GEO tambahan dan mapel GEO awal
    ws['GS{}'.format(r)] = "=IF($GF${}=0,$FG${},$FG${}-1)".format(r, r, r)
    ws['GS{}'.format(s)] = "=STDEV(GS2:GS{})".format(q)
    ws['GS{}'.format(t)] = "=MAX(GS2:GS{})".format(q)
    ws['GS{}'.format(u)] = "=MIN(GS2:GS{})".format(q)
    # rata" EKO ke EKO tambahan dan mapel EKO awal
    ws['GT{}'.format(r)] = "=IF($GG${}=0,$FH${},$FH${}-1)".format(r, r, r)
    ws['GT{}'.format(s)] = "=STDEV(GT2:GT{})".format(q)
    ws['GT{}'.format(t)] = "=MAX(GT2:GT{})".format(q)
    ws['GT{}'.format(u)] = "=MIN(GT2:GT{})".format(q)
    # rata" SOS ke SOS tambahan dan mapel SOS awal
    ws['GU{}'.format(r)] = "=IF($GH${}=0,$FI${},$FI${}-1)".format(r, r, r)
    ws['GU{}'.format(s)] = "=STDEV(GU2:GU{})".format(q)
    ws['GU{}'.format(t)] = "=MAX(GU2:GU{})".format(q)
    ws['GU{}'.format(u)] = "=MIN(GU2:GU{})".format(q)
    # jml MAPEL
    ws['GV{}'.format(r)] = "=ROUND(AVERAGE(GV2:GV{}),2)".format(q)
    ws['GV{}'.format(t)] = "=MAX(GV2:GV{})".format(q)
    ws['GV{}'.format(u)] = "=MIN(GV2:GV{})".format(q)
    # MAX Z SCORE
    ws['GW{}'.format(r)] = "=MAX(GW2:GW{})".format(q)
    ws['GX{}'.format(r)] = "=MAX(GX2:GX{})".format(q)
    ws['GY{}'.format(r)] = "=MAX(GY2:GY{})".format(q)
    ws['GZ{}'.format(r)] = "=MAX(GZ2:GZ{})".format(q)
    ws['HA{}'.format(r)] = "=MAX(HA2:HA{})".format(q)
    ws['HB{}'.format(r)] = "=MAX(HB2:HB{})".format(q)
    ws['HC{}'.format(r)] = "=MAX(HC2:HC{})".format(q)
    # NILAI STANDAR MTK
    ws['HD{}'.format(r)] = "=MAX(HD2:HD{})".format(q)
    ws['HD{}'.format(s)] = "=MIN(HD2:HD{})".format(q)
    ws['HD{}'.format(t)] = "=ROUND(AVERAGE(HD2:HD{}),2)".format(q)
    # NILAI STANDAR IND
    ws['HE{}'.format(r)] = "=MAX(HE2:HE{})".format(q)
    ws['HE{}'.format(s)] = "=MIN(HE2:HE{})".format(q)
    ws['HE{}'.format(t)] = "=ROUND(AVERAGE(HE2:HE{}),2)".format(q)
    # NILAI STANDAR ENG
    ws['HF{}'.format(r)] = "=MAX(HF2:HF{})".format(q)
    ws['HF{}'.format(s)] = "=MIN(HF2:HF{})".format(q)
    ws['HF{}'.format(t)] = "=ROUND(AVERAGE(HF2:HF{}),2)".format(q)
    # NILAI STANDAR SEJ
    ws['HG{}'.format(r)] = "=MAX(HG2:HG{})".format(q)
    ws['HG{}'.format(s)] = "=MIN(HG2:HG{})".format(q)
    ws['HG{}'.format(t)] = "=ROUND(AVERAGE(HG2:HG{}),2)".format(q)
    # NILAI STANDAR GEO
    ws['HH{}'.format(r)] = "=MAX(HH2:HH{})".format(q)
    ws['HH{}'.format(s)] = "=MIN(HH2:HH{})".format(q)
    ws['HH{}'.format(t)] = "=ROUND(AVERAGE(HH2:HH{}),2)".format(q)
    # NILAI STANDAR EKO
    ws['HI{}'.format(r)] = "=MAX(HI2:HI{})".format(q)
    ws['HI{}'.format(s)] = "=MIN(HI2:HI{})".format(q)
    ws['HI{}'.format(t)] = "=ROUND(AVERAGE(HI2:HI{}),2)".format(q)
    # NILAI STANDAR SOS
    ws['HJ{}'.format(r)] = "=MAX(HJ2:HJ{})".format(q)
    ws['HJ{}'.format(s)] = "=MIN(HJ2:HJ{})".format(q)
    ws['HJ{}'.format(t)] = "=ROUND(AVERAGE(HJ2:HJ{}),2)".format(q)
    # NILAI STANDAR S.JML
    ws['HK{}'.format(r)] = "=MAX(HK2:HK{})".format(q)
    ws['HK{}'.format(s)] = "=MIN(HK2:HK{})".format(q)
    ws['HK{}'.format(t)] = "=ROUND(AVERAGE(HK2:HK{}),2)".format(q)

    # TAMBAHAN
    ws['HN{}'.format(r)] = "=SUM(HN2:HN{})".format(q)
    ws['HO{}'.format(r)] = "=SUM(HO2:HO{})".format(q)
    ws['HP{}'.format(r)] = "=SUM(HP2:HP{})".format(q)
    ws['HQ{}'.format(r)] = "=SUM(HQ2:HQ{})".format(q)
    ws['HR{}'.format(r)] = "=SUM(HR2:HR{})".format(q)
    ws['HS{}'.format(r)] = "=SUM(HS2:HS{})".format(q)
    ws['HT{}'.format(r)] = "=SUM(HT2:HT{})".format(q)

    # Z Score [1]
    ws['B1'] = 'NAMA SISWA_A'
    ws['C1'] = 'NOMOR NF_A'
    ws['D1'] = 'KELAS_A'
    ws['E1'] = 'NAMA SEKOLAH_A'
    ws['F1'] = 'LOKASI_A'

    ws['G1'] = 'MAT_A'
    ws['H1'] = 'IND_A'
    ws['I1'] = 'ENG_A'
    ws['J1'] = 'SEJ_A'
    ws['K1'] = 'GEO_A'
    ws['L1'] = 'EKO_A'
    ws['M1'] = 'SOS_A'
    ws['N1'] = 'JML_A'

    ws['O1'] = 'Z_MAT_A'
    ws['P1'] = 'Z_IND_A'
    ws['Q1'] = 'Z_ENG_A'
    ws['R1'] = 'Z_SEJ_A'
    ws['S1'] = 'Z_GEO_A'
    ws['T1'] = 'Z_EKO_A'
    ws['U1'] = 'Z_SOS_A'

    ws['V1'] = 'S_MAT_A'
    ws['W1'] = 'S_IND_A'
    ws['X1'] = 'S_ENG_A'
    ws['Y1'] = 'S_SEJ_A'
    ws['Z1'] = 'S_GEO_A'
    ws['AA1'] = 'S_EKO_A'
    ws['AB1'] = 'S_SOS_A'
    ws['AC1'] = 'S_JML_A'

    ws['AD1'] = 'RANK NAS._A'
    ws['AE1'] = 'RANK LOK._A'

    ws['O1'].font = Font(bold=False, name='Calibri', size=11)
    ws['P1'].font = Font(bold=False, name='Calibri', size=11)
    ws['Q1'].font = Font(bold=False, name='Calibri', size=11)
    ws['R1'].font = Font(bold=False, name='Calibri', size=11)
    ws['S1'].font = Font(bold=False, name='Calibri', size=11)
    ws['T1'].font = Font(bold=False, name='Calibri', size=11)
    ws['U1'].font = Font(bold=False, name='Calibri', size=11)
    ws['V1'].font = Font(bold=False, name='Calibri', size=11)
    ws['W1'].font = Font(bold=False, name='Calibri', size=11)
    ws['X1'].font = Font(bold=False, name='Calibri', size=11)
    ws['Y1'].font = Font(bold=False, name='Calibri', size=11)
    ws['Z1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AA1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AB1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AC1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AD1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AE1'].font = Font(bold=False, name='Calibri', size=11)

    # FILL
    ws['B1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['C1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['D1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['E1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['F1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['G1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['H1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['I1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['J1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['K1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['L1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['M1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['N1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['O1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['P1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['Q1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['R1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['S1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['T1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['U1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['V1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['W1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['X1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['Y1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['Z1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AA1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AB1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AC1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AD1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AE1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')

    # tambahan
    ws['AF1'] = 'MAT_20_A'
    ws['AG1'] = 'IND_20_A'
    ws['AH1'] = 'ENG_20_A'
    ws['AI1'] = 'SEJ_20_A'
    ws['AJ1'] = 'GEO_20_A'
    ws['AK1'] = 'EKO_20_A'
    ws['AL1'] = 'SOS_20_A'

    ws['AF1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AG1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AH1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AI1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AJ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AK1'].font = Font(bold=False, name='Calibri', size=11)
    ws['AL1'].font = Font(bold=False, name='Calibri', size=11)

    ws['AF1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AG1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AH1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AI1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AJ1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AK1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')
    ws['AL1'].fill = PatternFill(
        fill_type='solid', start_color='00FF6600', end_color='00FF6600')

    for row in range(2, q+1):
        ws['O{}'.format(
            row)] = '=IFERROR(ROUND(IF(G{}="","",(G{}-G${})/G${}),2),"")'.format(row, row, r, s)
        ws['P{}'.format(
            row)] = '=IFERROR(ROUND(IF(H{}="","",(H{}-H${})/H${}),2),"")'.format(row, row, r, s)
        ws['Q{}'.format(
            row)] = '=IFERROR(ROUND(IF(I{}="","",(I{}-I${})/I${}),2),"")'.format(row, row, r, s)
        ws['R{}'.format(
            row)] = '=IFERROR(ROUND(IF(J{}="","",(J{}-J${})/J${}),2),"")'.format(row, row, r, s)
        ws['S{}'.format(
            row)] = '=IFERROR(ROUND(IF(K{}="","",(K{}-K${})/K${}),2),"")'.format(row, row, r, s)
        ws['T{}'.format(
            row)] = '=IFERROR(ROUND(IF(L{}="","",(L{}-L${})/L${}),2),"")'.format(row, row, r, s)
        ws['U{}'.format(
            row)] = '=IFERROR(ROUND(IF(M{}="","",(M{}-M${})/M${}),2),"")'.format(row, row, r, s)

        ws['V{}'.format(
            row)] = '=IFERROR(ROUND(IF(G{}="","",IF(70+30*O{}/$O${}<20,20,70+30*O{}/$O${})),2),"")'.format(row, row, r, row, r)
        ws['W{}'.format(
            row)] = '=IFERROR(ROUND(IF(H{}="","",IF(70+30*P{}/$P${}<20,20,70+30*P{}/$P${})),2),"")'.format(row, row, r, row, r)
        ws['X{}'.format(
            row)] = '=IFERROR(ROUND(IF(I{}="","",IF(70+30*Q{}/$Q${}<20,20,70+30*Q{}/$Q${})),2),"")'.format(row, row, r, row, r)
        ws['Y{}'.format(
            row)] = '=IFERROR(ROUND(IF(J{}="","",IF(70+30*R{}/$R${}<20,20,70+30*R{}/$R${})),2),"")'.format(row, row, r, row, r)
        ws['Z{}'.format(
            row)] = '=IFERROR(ROUND(IF(K{}="","",IF(70+30*S{}/$S${}<20,20,70+30*S{}/$S${})),2),"")'.format(row, row, r, row, r)
        ws['AA{}'.format(
            row)] = '=IFERROR(ROUND(IF(L{}="","",IF(70+30*T{}/$T${}<20,20,70+30*T{}/$T${})),2),"")'.format(row, row, r, row, r)
        ws['AB{}'.format(
            row)] = '=IFERROR(ROUND(IF(M{}="","",IF(70+30*U{}/$U${}<20,20,70+30*U{}/$U${})),2),"")'.format(row, row, r, row, r)

        ws['AC{}'.format(row)] = '=IF(SUM(V{}:AB{})=0,"",SUM(V{}:AB{}))'.format(
            row, row, row, row)
        ws['AD{}'.format(row)] = '=IF(AC{}="","",RANK(AC{},$AC$2:$AC${}))'.format(
            row, row, q)
        ws['AE{}'.format(
            row)] = '=IF(AD{}="","",COUNTIFS($F$2:$F${},F{},$AD$2:$AD${},"<"&AD{})+1)'.format(row, q, row, q, row)
    # TAMBAHAN
        ws['AF{}'.format(row)] = '=IF($G${}=20,IF(AND(G{}>3,V{}=20),1,""),IF($G${}=25,IF(AND(G{}>4,V{}=20),1,""),IF($G${}=30,IF(AND(G{}>5,V{}=20),1,""),IF($G${}=35,IF(AND(G{}>6,V{}=20),1,""),IF($G${}=40,IF(AND(G{}>7,V{}=20),1,""),IF($G${}=45,IF(AND(G{}>8,V{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['AG{}'.format(row)] = '=IF($H${}=20,IF(AND(H{}>3,W{}=20),1,""),IF($H${}=25,IF(AND(H{}>4,W{}=20),1,""),IF($H${}=30,IF(AND(H{}>5,W{}=20),1,""),IF($H${}=35,IF(AND(H{}>6,W{}=20),1,""),IF($H${}=40,IF(AND(H{}>7,W{}=20),1,""),IF($H${}=45,IF(AND(H{}>8,W{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['AH{}'.format(row)] = '=IF($I${}=20,IF(AND(I{}>3,X{}=20),1,""),IF($I${}=25,IF(AND(I{}>4,X{}=20),1,""),IF($I${}=30,IF(AND(I{}>5,X{}=20),1,""),IF($I${}=35,IF(AND(I{}>6,X{}=20),1,""),IF($I${}=40,IF(AND(I{}>7,X{}=20),1,""),IF($I${}=45,IF(AND(I{}>8,X{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['AI{}'.format(row)] = '=IF($J${}=20,IF(AND(J{}>3,Y{}=20),1,""),IF($J${}=25,IF(AND(J{}>4,Y{}=20),1,""),IF($J${}=30,IF(AND(J{}>5,Y{}=20),1,""),IF($J${}=35,IF(AND(J{}>6,Y{}=20),1,""),IF($J${}=40,IF(AND(J{}>7,Y{}=20),1,""),IF($J${}=45,IF(AND(J{}>8,Y{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['AJ{}'.format(row)] = '=IF($K${}=20,IF(AND(K{}>3,Z{}=20),1,""),IF($K${}=25,IF(AND(K{}>4,Z{}=20),1,""),IF($K${}=30,IF(AND(K{}>5,Z{}=20),1,""),IF($K${}=35,IF(AND(K{}>6,Z{}=20),1,""),IF($K${}=40,IF(AND(K{}>7,Z{}=20),1,""),IF($K${}=45,IF(AND(K{}>8,Z{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['AK{}'.format(row)] = '=IF($L${}=20,IF(AND(L{}>3,AA{}=20),1,""),IF($L${}=25,IF(AND(L{}>4,AA{}=20),1,""),IF($L${}=30,IF(AND(L{}>5,AA{}=20),1,""),IF($L${}=35,IF(AND(L{}>6,AA{}=20),1,""),IF($L${}=40,IF(AND(L{}>7,AA{}=20),1,""),IF($L${}=45,IF(AND(L{}>8,AA{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['AL{}'.format(row)] = '=IF($M${}=20,IF(AND(M{}>3,AB{}=20),1,""),IF($M${}=25,IF(AND(M{}>4,AB{}=20),1,""),IF($M${}=30,IF(AND(M{}>5,AB{}=20),1,""),IF($M${}=35,IF(AND(M{}>6,AB{}=20),1,""),IF($M${}=40,IF(AND(M{}>7,AB{}=20),1,""),IF($M${}=45,IF(AND(M{}>8,AB{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

    # Z Score [2]
    ws['AN1'] = 'NAMA SISWA_B'
    ws['AO1'] = 'NOMOR NF_B'
    ws['AP1'] = 'KELAS_B'
    ws['AQ1'] = 'NAMA SEKOLAH_B'
    ws['AR1'] = 'LOKASI_B'

    ws['AS1'] = 'MAT_B'
    ws['AT1'] = 'IND_B'
    ws['AU1'] = 'ENG_B'
    ws['AV1'] = 'SEJ_B'
    ws['AW1'] = 'GEO_B'
    ws['AX1'] = 'EKO_B'
    ws['AY1'] = 'SOS_B'
    ws['AZ1'] = 'JML_B'

    ws['BA1'] = 'Z_MAT_B'
    ws['BB1'] = 'Z_IND_B'
    ws['BC1'] = 'Z_ENG_B'
    ws['BD1'] = 'Z_SEJ_B'
    ws['BE1'] = 'Z_GEO_B'
    ws['BF1'] = 'Z_EKO_B'
    ws['BG1'] = 'Z_SOS_B'

    ws['BH1'] = 'S_MAT_B'
    ws['BI1'] = 'S_IND_B'
    ws['BJ1'] = 'S_ENG_B'
    ws['BK1'] = 'S_SEJ_B'
    ws['BL1'] = 'S_GEO_B'
    ws['BM1'] = 'S_EKO_B'
    ws['BN1'] = 'S_SOS_B'
    ws['BO1'] = 'S_JML_B'

    ws['BP1'] = 'RANK NAS._B'
    ws['BQ1'] = 'RANK LOK._B'

    ws['BA1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BB1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BC1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BD1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BE1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BF1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BG1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BH1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BI1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BJ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BK1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BL1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BM1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BN1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BO1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BP1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BQ1'].font = Font(bold=False, name='Calibri', size=11)

    # FILL
    ws['AN1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AO1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AP1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AQ1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AR1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AS1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AT1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AU1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AV1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AW1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AX1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AY1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['AZ1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BA1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BB1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BC1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BD1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BE1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BF1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BG1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BH1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BI1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BJ1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BK1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BL1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BM1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BN1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BO1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BP1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BQ1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')

    # tambahan
    ws['BR1'] = 'MAT_20_B'
    ws['BS1'] = 'IND_20_B'
    ws['BT1'] = 'ENG_20_B'
    ws['BU1'] = 'SEJ_20_B'
    ws['BV1'] = 'GEO_20_B'
    ws['BW1'] = 'EKO_20_B'
    ws['BX1'] = 'SOS_20_B'

    ws['BR1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BS1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BT1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BU1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BV1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BW1'].font = Font(bold=False, name='Calibri', size=11)
    ws['BX1'].font = Font(bold=False, name='Calibri', size=11)

    ws['BR1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BS1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BT1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BU1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BV1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BW1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')
    ws['BX1'].fill = PatternFill(
        fill_type='solid', start_color='31E1F7', end_color='31E1F7')

    for row in range(2, q+1):
        ws['AN{}'.format(row)] = '=B{}'.format(row)
        ws['AO{}'.format(row)] = '=C{}'.format(row, row)
        ws['AP{}'.format(row)] = '=D{}'.format(row, row)
        ws['AQ{}'.format(row)] = '=E{}'.format(row, row)
        ws['AR{}'.format(row)] = '=F{}'.format(row, row)
        ws['AS{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
        ws['AT{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
        ws['AU{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
        ws['AV{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
        ws['AW{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
        ws['AX{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
        ws['AY{}'.format(row)] = '=IF(M{}="","",M{})'.format(row, row)
        ws['AZ{}'.format(row)] = '=IF(N{}="","",N{})'.format(row, row)
    # Z Ke mapel
        ws['BA{}'.format(
            row)] = '=IFERROR(ROUND(IF(AS{}="","",(AS{}-AS${})/AS${}),2),"")'.format(row, row, r, s)
        ws['BB{}'.format(
            row)] = '=IFERROR(ROUND(IF(AT{}="","",(AT{}-AT${})/AT${}),2),"")'.format(row, row, r, s)
        ws['BC{}'.format(
            row)] = '=IFERROR(ROUND(IF(AU{}="","",(AU{}-AU${})/AU${}),2),"")'.format(row, row, r, s)
        ws['BD{}'.format(
            row)] = '=IFERROR(ROUND(IF(AV{}="","",(AV{}-AV${})/AV${}),2),"")'.format(row, row, r, s)
        ws['BE{}'.format(
            row)] = '=IFERROR(ROUND(IF(AW{}="","",(AW{}-AW${})/AW${}),2),"")'.format(row, row, r, s)
        ws['BF{}'.format(
            row)] = '=IFERROR(ROUND(IF(AX{}="","",(AX{}-AX${})/AX${}),2),"")'.format(row, row, r, s)
        ws['BG{}'.format(
            row)] = '=IFERROR(ROUND(IF(AY{}="","",(AY{}-AY${})/AY${}),2),"")'.format(row, row, r, s)
    # NILAI STANDAR ke mapel dan Z score
        ws['BH{}'.format(
            row)] = '=IFERROR(ROUND(IF(AS{}="","",IF(70+30*BA{}/$BA${}<20,20,70+30*BA{}/$BA${})),2),"")'.format(row, row, r, row, r)
        ws['BI{}'.format(
            row)] = '=IFERROR(ROUND(IF(AT{}="","",IF(70+30*BB{}/$BB${}<20,20,70+30*BB{}/$BB${})),2),"")'.format(row, row, r, row, r)
        ws['BJ{}'.format(
            row)] = '=IFERROR(ROUND(IF(AU{}="","",IF(70+30*BC{}/$BC${}<20,20,70+30*BC{}/$BC${})),2),"")'.format(row, row, r, row, r)
        ws['BK{}'.format(
            row)] = '=IFERROR(ROUND(IF(AV{}="","",IF(70+30*BD{}/$BD${}<20,20,70+30*BD{}/$BD${})),2),"")'.format(row, row, r, row, r)
        ws['BL{}'.format(
            row)] = '=IFERROR(ROUND(IF(AW{}="","",IF(70+30*BE{}/$BE${}<20,20,70+30*BE{}/$BE${})),2),"")'.format(row, row, r, row, r)
        ws['BM{}'.format(
            row)] = '=IFERROR(ROUND(IF(AX{}="","",IF(70+30*BF{}/$BF${}<20,20,70+30*BF{}/$BF${})),2),"")'.format(row, row, r, row, r)
        ws['BN{}'.format(
            row)] = '=IFERROR(ROUND(IF(AY{}="","",IF(70+30*BG{}/$BG${}<20,20,70+30*BG{}/$BG${})),2),"")'.format(row, row, r, row, r)
    # JUMLAH SELURUH NILAI STANDAR
        ws['BO{}'.format(row)] = '=IF(SUM(BH{}:BN{})=0,"",SUM(BH{}:BN{}))'.format(
            row, row, row, row)
        ws['BP{}'.format(row)] = '=IF(BO{}="","",RANK(BO{},$BO$2:$BO${}))'.format(
            row, row, q)
        ws['BQ{}'.format(
            row)] = '=IF(BP{}="","",COUNTIFS($AR$2:$AR${},AR{},$BP$2:$BP${},"<"&BP{})+1)'.format(row, q, row, q, row)
    # TAMBAHAN, MAPEL DAN NILAI STANDAR
        ws['BR{}'.format(row)] = '=IF($G${}=20,IF(AND(AS{}>3,BH{}=20),1,""),IF($G${}=25,IF(AND(AS{}>4,BH{}=20),1,""),IF($G${}=30,IF(AND(AS{}>5,BH{}=20),1,""),IF($G${}=35,IF(AND(AS{}>6,BH{}=20),1,""),IF($G${}=40,IF(AND(AS{}>7,BH{}=20),1,""),IF($G${}=45,IF(AND(AS{}>8,BH{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['BS{}'.format(row)] = '=IF($H${}=20,IF(AND(AT{}>3,BI{}=20),1,""),IF($H${}=25,IF(AND(AT{}>4,BI{}=20),1,""),IF($H${}=30,IF(AND(AT{}>5,BI{}=20),1,""),IF($H${}=35,IF(AND(AT{}>6,BI{}=20),1,""),IF($H${}=40,IF(AND(AT{}>7,BI{}=20),1,""),IF($H${}=45,IF(AND(AT{}>8,BI{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['BT{}'.format(row)] = '=IF($I${}=20,IF(AND(AU{}>3,BJ{}=20),1,""),IF($I${}=25,IF(AND(AU{}>4,BJ{}=20),1,""),IF($I${}=30,IF(AND(AU{}>5,BJ{}=20),1,""),IF($I${}=35,IF(AND(AU{}>6,BJ{}=20),1,""),IF($I${}=40,IF(AND(AU{}>7,BJ{}=20),1,""),IF($I${}=45,IF(AND(AU{}>8,BJ{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['BU{}'.format(row)] = '=IF($J${}=20,IF(AND(AV{}>3,BK{}=20),1,""),IF($J${}=25,IF(AND(AV{}>4,BK{}=20),1,""),IF($J${}=30,IF(AND(AV{}>5,BK{}=20),1,""),IF($J${}=35,IF(AND(AV{}>6,BK{}=20),1,""),IF($J${}=40,IF(AND(AV{}>7,BK{}=20),1,""),IF($J${}=45,IF(AND(AV{}>8,BK{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['BV{}'.format(row)] = '=IF($K${}=20,IF(AND(AW{}>3,BL{}=20),1,""),IF($K${}=25,IF(AND(AW{}>4,BL{}=20),1,""),IF($K${}=30,IF(AND(AW{}>5,BL{}=20),1,""),IF($K${}=35,IF(AND(AW{}>6,BL{}=20),1,""),IF($K${}=40,IF(AND(AW{}>7,BL{}=20),1,""),IF($K${}=45,IF(AND(AW{}>8,BL{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['BW{}'.format(row)] = '=IF($L${}=20,IF(AND(AX{}>3,BM{}=20),1,""),IF($L${}=25,IF(AND(AX{}>4,BM{}=20),1,""),IF($L${}=30,IF(AND(AX{}>5,BM{}=20),1,""),IF($L${}=35,IF(AND(AX{}>6,BM{}=20),1,""),IF($L${}=40,IF(AND(AX{}>7,BM{}=20),1,""),IF($L${}=45,IF(AND(AX{}>8,BM{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['BX{}'.format(row)] = '=IF($M${}=20,IF(AND(AY{}>3,BN{}=20),1,""),IF($M${}=25,IF(AND(AY{}>4,BN{}=20),1,""),IF($M${}=30,IF(AND(AY{}>5,BN{}=20),1,""),IF($M${}=35,IF(AND(AY{}>6,BN{}=20),1,""),IF($M${}=40,IF(AND(AY{}>7,BN{}=20),1,""),IF($M${}=45,IF(AND(AY{}>8,BN{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

    # Z Score [3]
    ws['BZ1'] = 'NAMA SISWA_C'
    ws['CA1'] = 'NOMOR NF_C'
    ws['CB1'] = 'KELAS_C'
    ws['CC1'] = 'NAMA SEKOLAH_C'
    ws['CD1'] = 'LOKASI_C'

    ws['CE1'] = 'MAT_C'
    ws['CF1'] = 'IND_C'
    ws['CG1'] = 'ENG_C'
    ws['CH1'] = 'SEJ_C'
    ws['CI1'] = 'GEO_C'
    ws['CJ1'] = 'EKO_C'
    ws['CK1'] = 'SOS_C'
    ws['CL1'] = 'JML_C'

    ws['CM1'] = 'Z_MAT_C'
    ws['CN1'] = 'Z_IND_C'
    ws['CO1'] = 'Z_ENG_C'
    ws['CP1'] = 'Z_SEJ_C'
    ws['CQ1'] = 'Z_GEO_C'
    ws['CR1'] = 'Z_EKO_C'
    ws['CS1'] = 'Z_SOS_C'

    ws['CT1'] = 'S_MAT_C'
    ws['CU1'] = 'S_IND_C'
    ws['CV1'] = 'S_ENG_C'
    ws['CW1'] = 'S_SEJ_C'
    ws['CX1'] = 'S_GEO_C'
    ws['CY1'] = 'S_EKO_C'
    ws['CZ1'] = 'S_SOS_C'
    ws['DA1'] = 'S_JML_C'

    ws['DB1'] = 'RANK NAS._C'
    ws['DC1'] = 'RANK LOK._C'

    ws['CM1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CN1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CO1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CP1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CQ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CR1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CS1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CT1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CU1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CV1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CW1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CX1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CY1'].font = Font(bold=False, name='Calibri', size=11)
    ws['CZ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DA1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DB1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DC1'].font = Font(bold=False, name='Calibri', size=11)

    # FILL
    ws['BZ1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['CA1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['CB1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['CC1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['CD1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['CE1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['CF1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['CG1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['CH1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['CI1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['CJ1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['CK1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['CL1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['CM1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['CN1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['CO1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['CP1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['CQ1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['CR1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['CS1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['CT1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['CU1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['CV1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['CW1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['CX1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['CY1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['CZ1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['DA1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['DB1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['DC1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')

    # tambahan
    ws['DD1'] = 'MAT_20_C'
    ws['DE1'] = 'IND_20_C'
    ws['DF1'] = 'ENG_20_C'
    ws['DG1'] = 'SEJ_20_C'
    ws['DH1'] = 'GEO_20_C'
    ws['DI1'] = 'EKO_20_C'
    ws['DJ1'] = 'SOS_20_C'

    ws['DD1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DE1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DF1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DG1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DH1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DI1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DJ1'].font = Font(bold=False, name='Calibri', size=11)

    ws['DD1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['DE1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['DF1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['DG1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['DH1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['DI1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')
    ws['DJ1'].fill = PatternFill(
        fill_type='solid', start_color='A1C298', end_color='A1C298')

    for row in range(2, q+1):
        ws['BZ{}'.format(row)] = '=AN{}'.format(row)
        ws['CA{}'.format(row)] = '=AO{}'.format(row, row)
        ws['CB{}'.format(row)] = '=AP{}'.format(row, row)
        ws['CC{}'.format(row)] = '=AQ{}'.format(row, row)
        ws['CD{}'.format(row)] = '=AR{}'.format(row, row)
        ws['CE{}'.format(row)] = '=IF(AS{}="","",AS{})'.format(row, row)
        ws['CF{}'.format(row)] = '=IF(AT{}="","",AT{})'.format(row, row)
        ws['CG{}'.format(row)] = '=IF(AU{}="","",AU{})'.format(row, row)
        ws['CH{}'.format(row)] = '=IF(AV{}="","",AV{})'.format(row, row)
        ws['CI{}'.format(row)] = '=IF(AW{}="","",AW{})'.format(row, row)
        ws['CJ{}'.format(row)] = '=IF(AX{}="","",AX{})'.format(row, row)
        ws['CK{}'.format(row)] = '=IF(AY{}="","",AY{})'.format(row, row)
        ws['CL{}'.format(row)] = '=IF(AZ{}="","",AZ{})'.format(row, row)
    # Z Ke mapel
        ws['CM{}'.format(
            row)] = '=IFERROR(ROUND(IF(CE{}="","",(CE{}-CE${})/CE${}),2),"")'.format(row, row, r, s)
        ws['CN{}'.format(
            row)] = '=IFERROR(ROUND(IF(CF{}="","",(CF{}-CF${})/CF${}),2),"")'.format(row, row, r, s)
        ws['CO{}'.format(
            row)] = '=IFERROR(ROUND(IF(CG{}="","",(CG{}-CG${})/CG${}),2),"")'.format(row, row, r, s)
        ws['CP{}'.format(
            row)] = '=IFERROR(ROUND(IF(CH{}="","",(CH{}-CH${})/CH${}),2),"")'.format(row, row, r, s)
        ws['CQ{}'.format(
            row)] = '=IFERROR(ROUND(IF(CI{}="","",(CI{}-CI${})/CI${}),2),"")'.format(row, row, r, s)
        ws['CR{}'.format(
            row)] = '=IFERROR(ROUND(IF(CJ{}="","",(CJ{}-CJ${})/CJ${}),2),"")'.format(row, row, r, s)
        ws['CS{}'.format(
            row)] = '=IFERROR(ROUND(IF(CK{}="","",(CK{}-CK${})/CK${}),2),"")'.format(row, row, r, s)
    # NILAI STANDAR ke mapel dan Z score
        ws['CT{}'.format(
            row)] = '=IFERROR(ROUND(IF(CE{}="","",IF(70+30*CM{}/$CM${}<20,20,70+30*CM{}/$CM${})),2),"")'.format(row, row, r, row, r)
        ws['CU{}'.format(
            row)] = '=IFERROR(ROUND(IF(CF{}="","",IF(70+30*CN{}/$CN${}<20,20,70+30*CN{}/$CN${})),2),"")'.format(row, row, r, row, r)
        ws['CV{}'.format(
            row)] = '=IFERROR(ROUND(IF(CG{}="","",IF(70+30*CO{}/$CO${}<20,20,70+30*CO{}/$CO${})),2),"")'.format(row, row, r, row, r)
        ws['CW{}'.format(
            row)] = '=IFERROR(ROUND(IF(CH{}="","",IF(70+30*CP{}/$CP${}<20,20,70+30*CP{}/$CP${})),2),"")'.format(row, row, r, row, r)
        ws['CX{}'.format(
            row)] = '=IFERROR(ROUND(IF(CI{}="","",IF(70+30*CQ{}/$CQ${}<20,20,70+30*CQ{}/$CQ${})),2),"")'.format(row, row, r, row, r)
        ws['CY{}'.format(
            row)] = '=IFERROR(ROUND(IF(CJ{}="","",IF(70+30*CR{}/$CR${}<20,20,70+30*CR{}/$CR${})),2),"")'.format(row, row, r, row, r)
        ws['CZ{}'.format(
            row)] = '=IFERROR(ROUND(IF(CK{}="","",IF(70+30*CS{}/$CS${}<20,20,70+30*CS{}/$CS${})),2),"")'.format(row, row, r, row, r)
    # JUMLAH SELURUH NILAI STANDAR
        ws['DA{}'.format(row)] = '=IF(SUM(CT{}:CZ{})=0,"",SUM(CT{}:CZ{}))'.format(
            row, row, row, row)
        ws['DB{}'.format(row)] = '=IF(DA{}="","",RANK(DA{},$DA$2:$DA${}))'.format(
            row, row, q)
        ws['DC{}'.format(
            row)] = '=IF(DB{}="","",COUNTIFS($CD$2:$CD${},CD{},$DB$2:$DB${},"<"&DB{})+1)'.format(row, q, row, q, row)
    # TAMBAHAN, MAPEL DAN NILAI STANDAR
        ws['DD{}'.format(row)] = '=IF($G${}=20,IF(AND(CE{}>3,CT{}=20),1,""),IF($G${}=25,IF(AND(CE{}>4,CT{}=20),1,""),IF($G${}=30,IF(AND(CE{}>5,CT{}=20),1,""),IF($G${}=35,IF(AND(CE{}>6,CT{}=20),1,""),IF($G${}=40,IF(AND(CE{}>7,CT{}=20),1,""),IF($G${}=45,IF(AND(CE{}>8,CT{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['DE{}'.format(row)] = '=IF($H${}=20,IF(AND(CF{}>3,CU{}=20),1,""),IF($H${}=25,IF(AND(CF{}>4,CU{}=20),1,""),IF($H${}=30,IF(AND(CF{}>5,CU{}=20),1,""),IF($H${}=35,IF(AND(CF{}>6,CU{}=20),1,""),IF($H${}=40,IF(AND(CF{}>7,CU{}=20),1,""),IF($H${}=45,IF(AND(CF{}>8,CU{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['DF{}'.format(row)] = '=IF($I${}=20,IF(AND(CG{}>3,CV{}=20),1,""),IF($I${}=25,IF(AND(CG{}>4,CV{}=20),1,""),IF($I${}=30,IF(AND(CG{}>5,CV{}=20),1,""),IF($I${}=35,IF(AND(CG{}>6,CV{}=20),1,""),IF($I${}=40,IF(AND(CG{}>7,CV{}=20),1,""),IF($I${}=45,IF(AND(CG{}>8,CV{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['DG{}'.format(row)] = '=IF($J${}=20,IF(AND(CH{}>3,CW{}=20),1,""),IF($J${}=25,IF(AND(CH{}>4,CW{}=20),1,""),IF($J${}=30,IF(AND(CH{}>5,CW{}=20),1,""),IF($J${}=35,IF(AND(CH{}>6,CW{}=20),1,""),IF($J${}=40,IF(AND(CH{}>7,CW{}=20),1,""),IF($J${}=45,IF(AND(CH{}>8,CW{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['DH{}'.format(row)] = '=IF($K${}=20,IF(AND(CI{}>3,CX{}=20),1,""),IF($K${}=25,IF(AND(CI{}>4,CX{}=20),1,""),IF($K${}=30,IF(AND(CI{}>5,CX{}=20),1,""),IF($K${}=35,IF(AND(CI{}>6,CX{}=20),1,""),IF($K${}=40,IF(AND(CI{}>7,CX{}=20),1,""),IF($K${}=45,IF(AND(CI{}>8,CX{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['DI{}'.format(row)] = '=IF($L${}=20,IF(AND(CJ{}>3,CY{}=20),1,""),IF($L${}=25,IF(AND(CJ{}>4,CY{}=20),1,""),IF($L${}=30,IF(AND(CJ{}>5,CY{}=20),1,""),IF($L${}=35,IF(AND(CJ{}>6,CY{}=20),1,""),IF($L${}=40,IF(AND(CJ{}>7,CY{}=20),1,""),IF($L${}=45,IF(AND(CJ{}>8,CY{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['DJ{}'.format(row)] = '=IF($M${}=20,IF(AND(CK{}>3,CZ{}=20),1,""),IF($M${}=25,IF(AND(CK{}>4,CZ{}=20),1,""),IF($M${}=30,IF(AND(CK{}>5,CZ{}=20),1,""),IF($M${}=35,IF(AND(CK{}>6,CZ{}=20),1,""),IF($M${}=40,IF(AND(CK{}>7,CZ{}=20),1,""),IF($M${}=45,IF(AND(CK{}>8,CZ{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

    # Z Score [4]
    ws['DL1'] = 'NAMA SISWA_D'
    ws['DM1'] = 'NOMOR NF_D'
    ws['DN1'] = 'KELAS_D'
    ws['DO1'] = 'NAMA SEKOLAH_D'
    ws['DP1'] = 'LOKASI_D'

    ws['DQ1'] = 'MAT_D'
    ws['DR1'] = 'IND_D'
    ws['DS1'] = 'ENG_D'
    ws['DT1'] = 'SEJ_D'
    ws['DU1'] = 'GEO_D'
    ws['DV1'] = 'EKO_D'
    ws['DW1'] = 'SOS_D'
    ws['DX1'] = 'JML_D'

    ws['DY1'] = 'Z_MAT_D'
    ws['DZ1'] = 'Z_IND_D'
    ws['EA1'] = 'Z_ENG_D'
    ws['EB1'] = 'Z_SEJ_D'
    ws['EC1'] = 'Z_GEO_D'
    ws['ED1'] = 'Z_EKO_D'
    ws['EE1'] = 'Z_SOS_D'

    ws['EF1'] = 'S_MAT_D'
    ws['EG1'] = 'S_IND_D'
    ws['EH1'] = 'S_ENG_D'
    ws['EI1'] = 'S_SEJ_D'
    ws['EJ1'] = 'S_GEO_D'
    ws['EK1'] = 'S_EKO_D'
    ws['EL1'] = 'S_SOS_D'
    ws['EM1'] = 'S_JML_D'

    ws['EN1'] = 'RANK NAS._D'
    ws['EO1'] = 'RANK LOK._D'

    ws['DY1'].font = Font(bold=False, name='Calibri', size=11)
    ws['DZ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EA1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EB1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EC1'].font = Font(bold=False, name='Calibri', size=11)
    ws['ED1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EE1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EF1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EG1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EH1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EI1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EJ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EK1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EL1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EM1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EN1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EO1'].font = Font(bold=False, name='Calibri', size=11)

    # FILL
    ws['DL1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['DM1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['DN1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['DO1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['DP1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['DQ1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['DR1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['DS1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['DT1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['DU1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['DV1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['DW1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['DX1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['DY1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['DZ1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['EA1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['EB1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['EC1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['ED1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['EE1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['EF1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['EG1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['EH1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['EI1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['EJ1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['EK1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['EL1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['EM1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['EN1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['EO1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')

    # tambahan
    ws['EP1'] = 'MAT_20_D'
    ws['EQ1'] = 'IND_20_D'
    ws['ER1'] = 'ENG_20_D'
    ws['ES1'] = 'SEJ_20_D'
    ws['ET1'] = 'GEO_20_D'
    ws['EU1'] = 'EKO_20_D'
    ws['EV1'] = 'SOS_20_D'

    ws['EP1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EQ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['ER1'].font = Font(bold=False, name='Calibri', size=11)
    ws['ES1'].font = Font(bold=False, name='Calibri', size=11)
    ws['ET1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EU1'].font = Font(bold=False, name='Calibri', size=11)
    ws['EV1'].font = Font(bold=False, name='Calibri', size=11)

    ws['EP1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['EQ1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['ER1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['ES1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['ET1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['EU1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
    ws['EV1'].fill = PatternFill(
        fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')

    for row in range(2, q+1):
        ws['DL{}'.format(row)] = '=BZ{}'.format(row)
        ws['DM{}'.format(row)] = '=CA{}'.format(row, row)
        ws['DN{}'.format(row)] = '=CB{}'.format(row, row)
        ws['DO{}'.format(row)] = '=CC{}'.format(row, row)
        ws['DP{}'.format(row)] = '=CD{}'.format(row, row)
        ws['DQ{}'.format(row)] = '=IF(CE{}="","",CE{})'.format(row, row)
        ws['DR{}'.format(row)] = '=IF(CF{}="","",CF{})'.format(row, row)
        ws['DS{}'.format(row)] = '=IF(CG{}="","",CG{})'.format(row, row)
        ws['DT{}'.format(row)] = '=IF(CH{}="","",CH{})'.format(row, row)
        ws['DU{}'.format(row)] = '=IF(CI{}="","",CI{})'.format(row, row)
        ws['DV{}'.format(row)] = '=IF(CJ{}="","",CJ{})'.format(row, row)
        ws['DW{}'.format(row)] = '=IF(CK{}="","",CK{})'.format(row, row)
        ws['DX{}'.format(row)] = '=IF(CL{}="","",CL{})'.format(row, row)
    # Z Ke mapel
        ws['DY{}'.format(
            row)] = '=IFERROR(ROUND(IF(DQ{}="","",(DQ{}-DQ${})/DQ${}),2),"")'.format(row, row, r, s)
        ws['DZ{}'.format(
            row)] = '=IFERROR(ROUND(IF(DR{}="","",(DR{}-DR${})/DR${}),2),"")'.format(row, row, r, s)
        ws['EA{}'.format(
            row)] = '=IFERROR(ROUND(IF(DS{}="","",(DS{}-DS${})/DS${}),2),"")'.format(row, row, r, s)
        ws['EB{}'.format(
            row)] = '=IFERROR(ROUND(IF(DT{}="","",(DT{}-DT${})/DT${}),2),"")'.format(row, row, r, s)
        ws['EC{}'.format(
            row)] = '=IFERROR(ROUND(IF(DU{}="","",(DU{}-DU${})/DU${}),2),"")'.format(row, row, r, s)
        ws['ED{}'.format(
            row)] = '=IFERROR(ROUND(IF(DV{}="","",(DV{}-DV${})/DV${}),2),"")'.format(row, row, r, s)
        ws['EE{}'.format(
            row)] = '=IFERROR(ROUND(IF(DW{}="","",(DW{}-DW${})/DW${}),2),"")'.format(row, row, r, s)
    # NILAI STANDAR ke mapel dan Z score
        ws['EF{}'.format(
            row)] = '=IFERROR(ROUND(IF(DQ{}="","",IF(70+30*DY{}/$DY${}<20,20,70+30*DY{}/$DY${})),2),"")'.format(row, row, r, row, r)
        ws['EG{}'.format(
            row)] = '=IFERROR(ROUND(IF(DR{}="","",IF(70+30*DZ{}/$DZ${}<20,20,70+30*DZ{}/$DZ${})),2),"")'.format(row, row, r, row, r)
        ws['EH{}'.format(
            row)] = '=IFERROR(ROUND(IF(DS{}="","",IF(70+30*EA{}/$EA${}<20,20,70+30*EA{}/$EA${})),2),"")'.format(row, row, r, row, r)
        ws['EI{}'.format(
            row)] = '=IFERROR(ROUND(IF(DT{}="","",IF(70+30*EB{}/$EB${}<20,20,70+30*EB{}/$EB${})),2),"")'.format(row, row, r, row, r)
        ws['EJ{}'.format(
            row)] = '=IFERROR(ROUND(IF(DU{}="","",IF(70+30*EC{}/$EC${}<20,20,70+30*EC{}/$EC${})),2),"")'.format(row, row, r, row, r)
        ws['EK{}'.format(
            row)] = '=IFERROR(ROUND(IF(DV{}="","",IF(70+30*ED{}/$ED${}<20,20,70+30*ED{}/$ED${})),2),"")'.format(row, row, r, row, r)
        ws['EL{}'.format(
            row)] = '=IFERROR(ROUND(IF(DW{}="","",IF(70+30*EE{}/$EE${}<20,20,70+30*EE{}/$EE${})),2),"")'.format(row, row, r, row, r)
    # JUMLAH SELURUH NILAI STANDAR
        ws['EM{}'.format(row)] = '=IF(SUM(EF{}:EL{})=0,"",SUM(EF{}:EL{}))'.format(
            row, row, row, row)
        ws['EN{}'.format(row)] = '=IF(EM{}="","",RANK(EM{},$EM$2:$EM${}))'.format(
            row, row, q)
        ws['EO{}'.format(
            row)] = '=IF(EN{}="","",COUNTIFS($DP$2:$DP${},DP{},$EN$2:$EN${},"<"&EN{})+1)'.format(row, q, row, q, row)
    # TAMBAHAN, MAPEL DAN NILAI STANDAR
        ws['EP{}'.format(row)] = '=IF($G${}=20,IF(AND(DQ{}>3,EF{}=20),1,""),IF($G${}=25,IF(AND(DQ{}>4,EF{}=20),1,""),IF($G${}=30,IF(AND(DQ{}>5,EF{}=20),1,""),IF($G${}=35,IF(AND(DQ{}>6,EF{}=20),1,""),IF($G${}=40,IF(AND(DQ{}>7,EF{}=20),1,""),IF($G${}=45,IF(AND(DQ{}>8,EF{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['EQ{}'.format(row)] = '=IF($H${}=20,IF(AND(DR{}>3,EG{}=20),1,""),IF($H${}=25,IF(AND(DR{}>4,EG{}=20),1,""),IF($H${}=30,IF(AND(DR{}>5,EG{}=20),1,""),IF($H${}=35,IF(AND(DR{}>6,EG{}=20),1,""),IF($H${}=40,IF(AND(DR{}>7,EG{}=20),1,""),IF($H${}=45,IF(AND(DR{}>8,EG{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['ER{}'.format(row)] = '=IF($I${}=20,IF(AND(DS{}>3,EH{}=20),1,""),IF($I${}=25,IF(AND(DS{}>4,EH{}=20),1,""),IF($I${}=30,IF(AND(DS{}>5,EH{}=20),1,""),IF($I${}=35,IF(AND(DS{}>6,EH{}=20),1,""),IF($I${}=40,IF(AND(DS{}>7,EH{}=20),1,""),IF($I${}=45,IF(AND(DS{}>8,EH{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['ES{}'.format(row)] = '=IF($J${}=20,IF(AND(DT{}>3,EI{}=20),1,""),IF($J${}=25,IF(AND(DT{}>4,EI{}=20),1,""),IF($J${}=30,IF(AND(DT{}>5,EI{}=20),1,""),IF($J${}=35,IF(AND(DT{}>6,EI{}=20),1,""),IF($J${}=40,IF(AND(DT{}>7,EI{}=20),1,""),IF($J${}=45,IF(AND(DT{}>8,EI{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['ET{}'.format(row)] = '=IF($K${}=20,IF(AND(DU{}>3,EJ{}=20),1,""),IF($K${}=25,IF(AND(DU{}>4,EJ{}=20),1,""),IF($K${}=30,IF(AND(DU{}>5,EJ{}=20),1,""),IF($K${}=35,IF(AND(DU{}>6,EJ{}=20),1,""),IF($K${}=40,IF(AND(DU{}>7,EJ{}=20),1,""),IF($K${}=45,IF(AND(DU{}>8,EJ{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['EU{}'.format(row)] = '=IF($L${}=20,IF(AND(DV{}>3,EK{}=20),1,""),IF($L${}=25,IF(AND(DV{}>4,EK{}=20),1,""),IF($L${}=30,IF(AND(DV{}>5,EK{}=20),1,""),IF($L${}=35,IF(AND(DV{}>6,EK{}=20),1,""),IF($L${}=40,IF(AND(DV{}>7,EK{}=20),1,""),IF($L${}=45,IF(AND(DV{}>8,EK{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['EV{}'.format(row)] = '=IF($M${}=20,IF(AND(DW{}>3,EL{}=20),1,""),IF($M${}=25,IF(AND(DW{}>4,EL{}=20),1,""),IF($M${}=30,IF(AND(DW{}>5,EL{}=20),1,""),IF($M${}=35,IF(AND(DW{}>6,EL{}=20),1,""),IF($M${}=40,IF(AND(DW{}>7,EL{}=20),1,""),IF($M${}=45,IF(AND(DW{}>8,EL{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

    # Z Score [5]
    ws['EX1'] = 'NAMA SISWA_E'
    ws['EY1'] = 'NOMOR NF_E'
    ws['EZ1'] = 'KELAS_E'
    ws['FA1'] = 'NAMA SEKOLAH_E'
    ws['FB1'] = 'LOKASI_E'

    ws['FC1'] = 'MAT_E'
    ws['FD1'] = 'IND_E'
    ws['FE1'] = 'ENG_E'
    ws['FF1'] = 'SEJ_E'
    ws['FG1'] = 'GEO_E'
    ws['FH1'] = 'EKO_E'
    ws['FI1'] = 'SOS_E'
    ws['FJ1'] = 'JML_E'

    ws['FK1'] = 'Z_MAT_E'
    ws['FL1'] = 'Z_IND_E'
    ws['FM1'] = 'Z_ENG_E'
    ws['FN1'] = 'Z_SEJ_E'
    ws['FO1'] = 'Z_GEO_E'
    ws['FP1'] = 'Z_EKO_E'
    ws['FQ1'] = 'Z_SOS_E'

    ws['FR1'] = 'S_MAT_E'
    ws['FS1'] = 'S_IND_E'
    ws['FT1'] = 'S_ENG_E'
    ws['FU1'] = 'S_SEJ_E'
    ws['FV1'] = 'S_GEO_E'
    ws['FW1'] = 'S_EKO_E'
    ws['FX1'] = 'S_SOS_E'
    ws['FY1'] = 'S_JML_E'

    ws['FZ1'] = 'RANK NAS._E'
    ws['GA1'] = 'RANK LOK._E'

    ws['FK1'].font = Font(bold=False, name='Calibri', size=11)
    ws['FL1'].font = Font(bold=False, name='Calibri', size=11)
    ws['FM1'].font = Font(bold=False, name='Calibri', size=11)
    ws['FN1'].font = Font(bold=False, name='Calibri', size=11)
    ws['FO1'].font = Font(bold=False, name='Calibri', size=11)
    ws['FP1'].font = Font(bold=False, name='Calibri', size=11)
    ws['FQ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['FR1'].font = Font(bold=False, name='Calibri', size=11)
    ws['FS1'].font = Font(bold=False, name='Calibri', size=11)
    ws['FT1'].font = Font(bold=False, name='Calibri', size=11)
    ws['FU1'].font = Font(bold=False, name='Calibri', size=11)
    ws['FV1'].font = Font(bold=False, name='Calibri', size=11)
    ws['FW1'].font = Font(bold=False, name='Calibri', size=11)
    ws['FX1'].font = Font(bold=False, name='Calibri', size=11)
    ws['FY1'].font = Font(bold=False, name='Calibri', size=11)
    ws['FZ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GA1'].font = Font(bold=False, name='Calibri', size=11)

    # FILL
    ws['EX1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['EY1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['EZ1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['FA1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['FB1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['FC1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['FD1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['FE1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['FF1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['FG1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['FH1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['FI1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['FJ1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['FK1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['FL1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['FM1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['FN1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['FO1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['FP1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['FQ1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['FR1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['FS1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['FT1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['FU1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['FV1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['FW1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['FX1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['FY1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['FZ1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['GA1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')

    # tambahan
    ws['GB1'] = 'MAT_20_E'
    ws['GC1'] = 'IND_20_E'
    ws['GD1'] = 'ENG_20_E'
    ws['GE1'] = 'SEJ_20_E'
    ws['GF1'] = 'GEO_20_E'
    ws['GG1'] = 'EKO_20_E'
    ws['GH1'] = 'SOS_20_E'

    ws['GB1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GC1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GD1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GE1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GF1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GG1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GH1'].font = Font(bold=False, name='Calibri', size=11)

    ws['GB1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['GC1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['GD1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['GE1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['GF1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['GG1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
    ws['GH1'].fill = PatternFill(
        fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')

    for row in range(2, q+1):
        ws['EX{}'.format(row)] = '=DL{}'.format(row)
        ws['EY{}'.format(row)] = '=DM{}'.format(row, row)
        ws['EZ{}'.format(row)] = '=DN{}'.format(row, row)
        ws['FA{}'.format(row)] = '=DO{}'.format(row, row)
        ws['FB{}'.format(row)] = '=DP{}'.format(row, row)
        ws['FC{}'.format(row)] = '=IF(DQ{}="","",DQ{})'.format(row, row)
        ws['FD{}'.format(row)] = '=IF(DR{}="","",DR{})'.format(row, row)
        ws['FE{}'.format(row)] = '=IF(DS{}="","",DS{})'.format(row, row)
        ws['FF{}'.format(row)] = '=IF(DT{}="","",DT{})'.format(row, row)
        ws['FG{}'.format(row)] = '=IF(DU{}="","",DU{})'.format(row, row)
        ws['FH{}'.format(row)] = '=IF(DV{}="","",DV{})'.format(row, row)
        ws['FI{}'.format(row)] = '=IF(DW{}="","",DW{})'.format(row, row)
        ws['FJ{}'.format(row)] = '=IF(DX{}="","",DX{})'.format(row, row)
    # Z Ke mapel
        ws['FK{}'.format(
            row)] = '=IFERROR(ROUND(IF(FC{}="","",(FC{}-FC${})/FC${}),2),"")'.format(row, row, r, s)
        ws['FL{}'.format(
            row)] = '=IFERROR(ROUND(IF(FD{}="","",(FD{}-FD${})/FD${}),2),"")'.format(row, row, r, s)
        ws['FM{}'.format(
            row)] = '=IFERROR(ROUND(IF(FE{}="","",(FE{}-FE${})/FE${}),2),"")'.format(row, row, r, s)
        ws['FN{}'.format(
            row)] = '=IFERROR(ROUND(IF(FF{}="","",(FF{}-FF${})/FF${}),2),"")'.format(row, row, r, s)
        ws['FO{}'.format(
            row)] = '=IFERROR(ROUND(IF(FG{}="","",(FG{}-FG${})/FG${}),2),"")'.format(row, row, r, s)
        ws['FP{}'.format(
            row)] = '=IFERROR(ROUND(IF(FH{}="","",(FH{}-FH${})/FH${}),2),"")'.format(row, row, r, s)
        ws['FQ{}'.format(
            row)] = '=IFERROR(ROUND(IF(FI{}="","",(FI{}-FI${})/FI${}),2),"")'.format(row, row, r, s)
    # NILAI STANDAR ke mapel dan Z score
        ws['FR{}'.format(
            row)] = '=IFERROR(ROUND(IF(FC{}="","",IF(70+30*FK{}/$FK${}<20,20,70+30*FK{}/$FK${})),2),"")'.format(row, row, r, row, r)
        ws['FS{}'.format(
            row)] = '=IFERROR(ROUND(IF(FD{}="","",IF(70+30*FL{}/$FL${}<20,20,70+30*FL{}/$FL${})),2),"")'.format(row, row, r, row, r)
        ws['FT{}'.format(
            row)] = '=IFERROR(ROUND(IF(FE{}="","",IF(70+30*FM{}/$FM${}<20,20,70+30*FM{}/$FM${})),2),"")'.format(row, row, r, row, r)
        ws['FU{}'.format(
            row)] = '=IFERROR(ROUND(IF(FF{}="","",IF(70+30*FN{}/$FN${}<20,20,70+30*FN{}/$FN${})),2),"")'.format(row, row, r, row, r)
        ws['FV{}'.format(
            row)] = '=IFERROR(ROUND(IF(FG{}="","",IF(70+30*FO{}/$FO${}<20,20,70+30*FO{}/$FO${})),2),"")'.format(row, row, r, row, r)
        ws['FW{}'.format(
            row)] = '=IFERROR(ROUND(IF(FH{}="","",IF(70+30*FP{}/$FP${}<20,20,70+30*FP{}/$FP${})),2),"")'.format(row, row, r, row, r)
        ws['FX{}'.format(
            row)] = '=IFERROR(ROUND(IF(FI{}="","",IF(70+30*FQ{}/$FQ${}<20,20,70+30*FQ{}/$FQ${})),2),"")'.format(row, row, r, row, r)
    # JUMLAH SELURUH NILAI STANDAR
        ws['FY{}'.format(row)] = '=IF(SUM(FR{}:FX{})=0,"",SUM(FR{}:FX{}))'.format(
            row, row, row, row)
        ws['FZ{}'.format(row)] = '=IF(FY{}="","",RANK(FY{},$FY$2:$FY${}))'.format(
            row, row, q)
        ws['GA{}'.format(
            row)] = '=IF(FZ{}="","",COUNTIFS($FB$2:$FB${},FB{},$FZ$2:$FZ${},"<"&FZ{})+1)'.format(row, q, row, q, row)
    # TAMBAHAN, MAPEL DAN NILAI STANDAR
        ws['GB{}'.format(row)] = '=IF($G${}=20,IF(AND(FC{}>3,FR{}=20),1,""),IF($G${}=25,IF(AND(FC{}>4,FR{}=20),1,""),IF($G${}=30,IF(AND(FC{}>5,FR{}=20),1,""),IF($G${}=35,IF(AND(FC{}>6,FR{}=20),1,""),IF($G${}=40,IF(AND(FC{}>7,FR{}=20),1,""),IF($G${}=45,IF(AND(FC{}>8,FR{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['GC{}'.format(row)] = '=IF($H${}=20,IF(AND(FD{}>3,FS{}=20),1,""),IF($H${}=25,IF(AND(FD{}>4,FS{}=20),1,""),IF($H${}=30,IF(AND(FD{}>5,FS{}=20),1,""),IF($H${}=35,IF(AND(FD{}>6,FS{}=20),1,""),IF($H${}=40,IF(AND(FD{}>7,FS{}=20),1,""),IF($H${}=45,IF(AND(FD{}>8,FS{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['GD{}'.format(row)] = '=IF($I${}=20,IF(AND(FE{}>3,FT{}=20),1,""),IF($I${}=25,IF(AND(FE{}>4,FT{}=20),1,""),IF($I${}=30,IF(AND(FE{}>5,FT{}=20),1,""),IF($I${}=35,IF(AND(FE{}>6,FT{}=20),1,""),IF($I${}=40,IF(AND(FE{}>7,FT{}=20),1,""),IF($I${}=45,IF(AND(FE{}>8,FT{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['GE{}'.format(row)] = '=IF($J${}=20,IF(AND(FF{}>3,FU{}=20),1,""),IF($J${}=25,IF(AND(FF{}>4,FU{}=20),1,""),IF($J${}=30,IF(AND(FF{}>5,FU{}=20),1,""),IF($J${}=35,IF(AND(FF{}>6,FU{}=20),1,""),IF($J${}=40,IF(AND(FF{}>7,FU{}=20),1,""),IF($J${}=45,IF(AND(FF{}>8,FU{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['GF{}'.format(row)] = '=IF($K${}=20,IF(AND(FG{}>3,FV{}=20),1,""),IF($K${}=25,IF(AND(FG{}>4,FV{}=20),1,""),IF($K${}=30,IF(AND(FG{}>5,FV{}=20),1,""),IF($K${}=35,IF(AND(FG{}>6,FV{}=20),1,""),IF($K${}=40,IF(AND(FG{}>7,FV{}=20),1,""),IF($K${}=45,IF(AND(FG{}>8,FV{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['GG{}'.format(row)] = '=IF($L${}=20,IF(AND(FH{}>3,FW{}=20),1,""),IF($L${}=25,IF(AND(FH{}>4,FW{}=20),1,""),IF($L${}=30,IF(AND(FH{}>5,FW{}=20),1,""),IF($L${}=35,IF(AND(FH{}>6,FW{}=20),1,""),IF($L${}=40,IF(AND(FH{}>7,FW{}=20),1,""),IF($L${}=45,IF(AND(FH{}>8,FW{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['GH{}'.format(row)] = '=IF($M${}=20,IF(AND(FI{}>3,FX{}=20),1,""),IF($M${}=25,IF(AND(FI{}>4,FX{}=20),1,""),IF($M${}=30,IF(AND(FI{}>5,FX{}=20),1,""),IF($M${}=35,IF(AND(FI{}>6,FX{}=20),1,""),IF($M${}=40,IF(AND(FI{}>7,FX{}=20),1,""),IF($M${}=45,IF(AND(FI{}>8,FX{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

    # Z Score
    ws['GJ1'] = 'NAMA SISWA'
    ws['GK1'] = 'NOMOR NF'
    ws['GL1'] = 'KELAS'
    ws['GM1'] = 'NAMA SEKOLAH'
    ws['GN1'] = 'LOKASI'

    ws['GO1'] = 'MAT'
    ws['GP1'] = 'IND'
    ws['GQ1'] = 'ENG'
    ws['GR1'] = 'SEJ'
    ws['GS1'] = 'GEO'
    ws['GT1'] = 'EKO'
    ws['GU1'] = 'SOS'
    ws['GV1'] = 'JML'

    ws['GW1'] = 'Z_MAT'
    ws['GX1'] = 'Z_IND'
    ws['GY1'] = 'Z_ENG'
    ws['GZ1'] = 'Z_SEJ'
    ws['HA1'] = 'Z_GEO'
    ws['HB1'] = 'Z_EKO'
    ws['HC1'] = 'Z_SOS'

    ws['HD1'] = 'S_MAT'
    ws['HE1'] = 'S_IND'
    ws['HF1'] = 'S_ENG'
    ws['HG1'] = 'S_SEJ'
    ws['HH1'] = 'S_GEO'
    ws['HI1'] = 'S_EKO'
    ws['HJ1'] = 'S_SOS'
    ws['HK1'] = 'S_JML'

    ws['HL1'] = 'RANK NAS.'
    ws['HM1'] = 'RANK LOK.'

    ws['GW1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GX1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GY1'].font = Font(bold=False, name='Calibri', size=11)
    ws['GZ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['HA1'].font = Font(bold=False, name='Calibri', size=11)
    ws['HB1'].font = Font(bold=False, name='Calibri', size=11)
    ws['HC1'].font = Font(bold=False, name='Calibri', size=11)
    ws['HD1'].font = Font(bold=False, name='Calibri', size=11)
    ws['HE1'].font = Font(bold=False, name='Calibri', size=11)
    ws['HF1'].font = Font(bold=False, name='Calibri', size=11)
    ws['HG1'].font = Font(bold=False, name='Calibri', size=11)
    ws['HH1'].font = Font(bold=False, name='Calibri', size=11)
    ws['HI1'].font = Font(bold=False, name='Calibri', size=11)
    ws['HJ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['HK1'].font = Font(bold=False, name='Calibri', size=11)
    ws['HL1'].font = Font(bold=False, name='Calibri', size=11)
    ws['HM1'].font = Font(bold=False, name='Calibri', size=11)

    # FILL
    ws['GJ1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['GK1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['GL1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['GM1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['GN1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['GO1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['GP1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['GQ1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['GR1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['GS1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['GT1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['GU1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['GV1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['GW1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['GX1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['GY1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['GZ1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['HA1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['HB1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['HC1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['HD1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['HE1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['HF1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['HG1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['HH1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['HI1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['HJ1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['HK1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['HL1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['HM1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')

    # tambahan
    ws['HN1'] = 'MAT_20'
    ws['HO1'] = 'IND_20'
    ws['HP1'] = 'ENG_20'
    ws['HQ1'] = 'SEJ_20'
    ws['HR1'] = 'GEO_20'
    ws['HS1'] = 'EKO_20'
    ws['HT1'] = 'SOS_20'

    ws['HN1'].font = Font(bold=False, name='Calibri', size=11)
    ws['HO1'].font = Font(bold=False, name='Calibri', size=11)
    ws['HP1'].font = Font(bold=False, name='Calibri', size=11)
    ws['HQ1'].font = Font(bold=False, name='Calibri', size=11)
    ws['HR1'].font = Font(bold=False, name='Calibri', size=11)
    ws['HS1'].font = Font(bold=False, name='Calibri', size=11)
    ws['HT1'].font = Font(bold=False, name='Calibri', size=11)

    ws['HN1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['HO1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['HP1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['HQ1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['HR1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['HS1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
    ws['HT1'].fill = PatternFill(
        fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')

    for row in range(2, q+1):
        ws['GJ{}'.format(row)] = '=EX{}'.format(row)
        ws['GK{}'.format(row)] = '=EY{}'.format(row, row)
        ws['GL{}'.format(row)] = '=EZ{}'.format(row, row)
        ws['GM{}'.format(row)] = '=FA{}'.format(row, row)
        ws['GN{}'.format(row)] = '=FB{}'.format(row, row)
        ws['GO{}'.format(row)] = '=IF(FC{}="","",FC{})'.format(row, row)
        ws['GP{}'.format(row)] = '=IF(FD{}="","",FD{})'.format(row, row)
        ws['GQ{}'.format(row)] = '=IF(FE{}="","",FE{})'.format(row, row)
        ws['GR{}'.format(row)] = '=IF(FF{}="","",FF{})'.format(row, row)
        ws['GS{}'.format(row)] = '=IF(FG{}="","",FG{})'.format(row, row)
        ws['GT{}'.format(row)] = '=IF(FH{}="","",FH{})'.format(row, row)
        ws['GU{}'.format(row)] = '=IF(FI{}="","",FI{})'.format(row, row)
        ws['GV{}'.format(row)] = '=IF(FJ{}="","",FJ{})'.format(row, row)
    # Z Ke mapel
        ws['GW{}'.format(
            row)] = '=IFERROR(ROUND(IF(GO{}="","",(GO{}-GO${})/GO${}),2),"")'.format(row, row, r, s)
        ws['GX{}'.format(
            row)] = '=IFERROR(ROUND(IF(GP{}="","",(GP{}-GP${})/GP${}),2),"")'.format(row, row, r, s)
        ws['GY{}'.format(
            row)] = '=IFERROR(ROUND(IF(GQ{}="","",(GQ{}-GQ${})/GQ${}),2),"")'.format(row, row, r, s)
        ws['GZ{}'.format(
            row)] = '=IFERROR(ROUND(IF(GR{}="","",(GR{}-GR${})/GR${}),2),"")'.format(row, row, r, s)
        ws['HA{}'.format(
            row)] = '=IFERROR(ROUND(IF(GS{}="","",(GS{}-GS${})/GS${}),2),"")'.format(row, row, r, s)
        ws['HB{}'.format(
            row)] = '=IFERROR(ROUND(IF(GT{}="","",(GT{}-GT${})/GT${}),2),"")'.format(row, row, r, s)
        ws['HC{}'.format(
            row)] = '=IFERROR(ROUND(IF(GU{}="","",(GU{}-GU${})/GU${}),2),"")'.format(row, row, r, s)
    # NILAI STANDAR ke mapel dan Z score
        ws['HD{}'.format(
            row)] = '=IFERROR(ROUND(IF(GO{}="","",IF(70+30*GW{}/$GW${}<20,20,70+30*GW{}/$GW${})),2),"")'.format(row, row, r, row, r)
        ws['HE{}'.format(
            row)] = '=IFERROR(ROUND(IF(GP{}="","",IF(70+30*GX{}/$GX${}<20,20,70+30*GX{}/$GX${})),2),"")'.format(row, row, r, row, r)
        ws['HF{}'.format(
            row)] = '=IFERROR(ROUND(IF(GQ{}="","",IF(70+30*GY{}/$GY${}<20,20,70+30*GY{}/$GY${})),2),"")'.format(row, row, r, row, r)
        ws['HG{}'.format(
            row)] = '=IFERROR(ROUND(IF(GR{}="","",IF(70+30*GZ{}/$GZ${}<20,20,70+30*GZ{}/$GZ${})),2),"")'.format(row, row, r, row, r)
        ws['HH{}'.format(
            row)] = '=IFERROR(ROUND(IF(GS{}="","",IF(70+30*HA{}/$HA${}<20,20,70+30*HA{}/$HA${})),2),"")'.format(row, row, r, row, r)
        ws['HI{}'.format(
            row)] = '=IFERROR(ROUND(IF(GT{}="","",IF(70+30*HB{}/$HB${}<20,20,70+30*HB{}/$HB${})),2),"")'.format(row, row, r, row, r)
        ws['HJ{}'.format(
            row)] = '=IFERROR(ROUND(IF(GU{}="","",IF(70+30*HC{}/$HC${}<20,20,70+30*HC{}/$HC${})),2),"")'.format(row, row, r, row, r)
    # JUMLAH SELURUH NILAI STANDAR
        ws['HK{}'.format(row)] = '=IF(SUM(HD{}:HJ{})=0,"",SUM(HD{}:HJ{}))'.format(
            row, row, row, row)
        ws['HL{}'.format(row)] = '=IF(HK{}="","",RANK(HK{},$HK$2:$HK${}))'.format(
            row, row, q)
        ws['HM{}'.format(
            row)] = '=IF(HL{}="","",COUNTIFS($GN$2:$GN${},GN{},$HL$2:$HL${},"<"&HL{})+1)'.format(row, q, row, q, row)
    # TAMBAHAN, MAPEL DAN NILAI STANDAR
        ws['HN{}'.format(row)] = '=IF($G${}=20,IF(AND(GO{}>3,HD{}=20),1,""),IF($G${}=25,IF(AND(GO{}>4,HD{}=20),1,""),IF($G${}=30,IF(AND(GO{}>5,HD{}=20),1,""),IF($G${}=35,IF(AND(GO{}>6,HD{}=20),1,""),IF($G${}=40,IF(AND(GO{}>7,HD{}=20),1,""),IF($G${}=45,IF(AND(GO{}>8,HD{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['HO{}'.format(row)] = '=IF($H${}=20,IF(AND(GP{}>3,HE{}=20),1,""),IF($H${}=25,IF(AND(GP{}>4,HE{}=20),1,""),IF($H${}=30,IF(AND(GP{}>5,HE{}=20),1,""),IF($H${}=35,IF(AND(GP{}>6,HE{}=20),1,""),IF($H${}=40,IF(AND(GP{}>7,HE{}=20),1,""),IF($H${}=45,IF(AND(GP{}>8,HE{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['HP{}'.format(row)] = '=IF($I${}=20,IF(AND(GQ{}>3,HF{}=20),1,""),IF($I${}=25,IF(AND(GQ{}>4,HF{}=20),1,""),IF($I${}=30,IF(AND(GQ{}>5,HF{}=20),1,""),IF($I${}=35,IF(AND(GQ{}>6,HF{}=20),1,""),IF($I${}=40,IF(AND(GQ{}>7,HF{}=20),1,""),IF($I${}=45,IF(AND(GQ{}>8,HF{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['HQ{}'.format(row)] = '=IF($J${}=20,IF(AND(GR{}>3,HG{}=20),1,""),IF($J${}=25,IF(AND(GR{}>4,HG{}=20),1,""),IF($J${}=30,IF(AND(GR{}>5,HG{}=20),1,""),IF($J${}=35,IF(AND(GR{}>6,HG{}=20),1,""),IF($J${}=40,IF(AND(GR{}>7,HG{}=20),1,""),IF($J${}=45,IF(AND(GR{}>8,HG{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['HR{}'.format(row)] = '=IF($K${}=20,IF(AND(GS{}>3,HH{}=20),1,""),IF($K${}=25,IF(AND(GS{}>4,HH{}=20),1,""),IF($K${}=30,IF(AND(GS{}>5,HH{}=20),1,""),IF($K${}=35,IF(AND(GS{}>6,HH{}=20),1,""),IF($K${}=40,IF(AND(GS{}>7,HH{}=20),1,""),IF($K${}=45,IF(AND(GS{}>8,HH{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['HS{}'.format(row)] = '=IF($L${}=20,IF(AND(GT{}>3,HI{}=20),1,""),IF($L${}=25,IF(AND(GT{}>4,HI{}=20),1,""),IF($L${}=30,IF(AND(GT{}>5,HI{}=20),1,""),IF($L${}=35,IF(AND(GT{}>6,HI{}=20),1,""),IF($L${}=40,IF(AND(GT{}>7,HI{}=20),1,""),IF($L${}=45,IF(AND(GT{}>8,HI{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
        ws['HT{}'.format(row)] = '=IF($M${}=20,IF(AND(GU{}>3,HJ{}=20),1,""),IF($M${}=25,IF(AND(GU{}>4,HJ{}=20),1,""),IF($M${}=30,IF(AND(GU{}>5,HJ{}=20),1,""),IF($M${}=35,IF(AND(GU{}>6,HJ{}=20),1,""),IF($M${}=40,IF(AND(GU{}>7,HJ{}=20),1,""),IF($M${}=45,IF(AND(GU{}>8,HJ{}=20),1,"")))))))'.format(
            v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

    # Mengubah 'KELAS' sesuai dengan nilai yang dipilih dari selectbox 'KELAS'
    kelas = KELAS.lower().replace(" ", "")
    semester = SEMESTER.lower()
    tahun = TAHUN.replace("-", "")
    penilaian = PENILAIAN.lower()
    kurikulum = KURIKULUM.lower()

    path_file = f"{kelas}_{penilaian}_{semester}_{kurikulum}_{tahun}_nilai_std.xlsx"

    # Simpan file ke direktori temporer
    temp_dir = tempfile.gettempdir()
    file_path = temp_dir + '/' + path_file
    wb.save(file_path)

    st.success(
        "File siap diunduh!")

    # Tombol unduh file
	with open(file_path, "rb") as f:
        bytes_data = f.read()
    st.download_button(label="Unduh File", data=bytes_data,
                       file_name=path_file)

    st.warning(
        "Buka file unduhan, klik 'Enable Editing' dan 'Save'")
