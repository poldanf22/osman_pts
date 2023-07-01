import pickle
import streamlit as st
import streamlit_authenticator as stauth
from pathlib import Path
from PIL import Image
import pandas as pd
from streamlit_option_menu import option_menu
import openpyxl
from openpyxl.styles import Font, PatternFill
import tempfile

# User Authentication
names = ["TI Polda NF 1", "TI Polda NF 2"]
usernames = ["admin1", "admin2"]

# load hashed kd_akses
file_path = Path(__file__).parent / "hashed_pw.pkl"
with file_path.open("rb") as file:
    hashed_kd_akses = pickle.load(file)

authenticator = stauth.Authenticate(
    names, usernames, hashed_kd_akses, "lookup", "abcdef")
name, authentication_status, username = authenticator.login("Login", "main")

if authentication_status == False:
    st.error("Username/kode akses salah")

if authentication_status == None:
    st.warning("Silahkan masukan username dan kode akses")

url = "https://osman2-8bdgvgq3z54.streamlit.app/"

if authentication_status:
    authenticator.logout("Logout", "sidebar")
    with st.sidebar:
        with st.sidebar:
            st.markdown(
                f'''<a href={url}><button style="background-color:GreenYellow;">Untuk Lok.</button></a>''', unsafe_allow_html=True)
        selected_file = option_menu(
            menu_title="Pilih file:",
            options=["Pivot", "Nilai Std. SD, SMP, 10KM", "Nilai Std. All IPA",
                     "Nilai Std. 10, 11 IPS", "Nilai Std. PPLS, RONIN IPS", "Nilai Std. 11KM"],
        )
    if selected_file == "Pivot":
        # kurikulum - kelas - mapel
        # 4sd k13
        k13_4sd_mat = 'LHG94EEQ'
        k13_4sd_ind = 'LHG9KCRA'
        k13_4sd_eng = 'LHGA44Y9'
        k13_4sd_ipa = 'LHGALT9N'
        k13_4sd_ips = 'LHH0F32F'
        k13_4sd = [k13_4sd_mat, k13_4sd_ind,
                   k13_4sd_eng, k13_4sd_ipa, k13_4sd_ips]
        column_order_k13_4sd = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAT_4SD', 'IND_4SD',
                                'ENG_4SD', 'IPA_4SD', 'IPS_4SD']

        # 4sd km
        km_4sd_mat = 'LHH0U12P'
        km_4sd_ind = 'LHH19TQN'
        km_4sd_eng = 'LHH47YLV'
        km_4sd_ipas = 'LHH4U3Q0'
        km_4sd = [km_4sd_mat, km_4sd_ind, km_4sd_eng, km_4sd_ipas]
        column_order_km_4sd = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAT_4KM', 'IND_4KM',
                               'ENG_4KM', 'IPAS_4KM']

        # 5sd k13
        k13_5sd_mat = 'LHH5V62M'
        k13_5sd_ind = 'LHH6WL2C'
        k13_5sd_eng = 'LHH7NAB5'
        k13_5sd_ipa = 'LHHCO0Q4'
        k13_5sd_ips = 'LHHDAY7I'
        k13_5sd = [k13_5sd_mat, k13_5sd_ind,
                   k13_5sd_eng, k13_5sd_ipa, k13_5sd_ips]
        column_order_k13_5sd = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAT_5SD', 'IND_5SD',
                                'ENG_5SD', 'IPA_5SD', 'IPS_5SD']

        # 7smp k13
        k13_7smp_mat = 'LHHDRBXZ'
        k13_7smp_ind = 'LHHDUWKS'
        k13_7smp_eng = 'LHHDX6U7'
        k13_7smp_ipa = 'LHHDZC8Y'
        k13_7smp_ips = 'LHHE476J'
        k13_7smp = [k13_7smp_mat, k13_7smp_ind,
                    k13_7smp_eng, k13_7smp_ipa, k13_7smp_ips]
        column_order_k13_7smp = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAT_7SMP', 'IND_7SMP',
                                 'ENG_7SMP', 'IPA_7SMP', 'IPS_7SMP']

        # 8smp k13
        k13_8smp_mat = 'LHH6H3F6'
        k13_8smp_ind = 'LHH6TEO5'
        k13_8smp_eng = 'LHHN9AZH'
        k13_8smp_ipa = 'LHHNDOAI'
        k13_8smp_ips = 'LHHNFJ3E'
        k13_8smp = [k13_8smp_mat, k13_8smp_ind,
                    k13_8smp_eng, k13_8smp_ipa, k13_8smp_ips]
        column_order_k13_8smp = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAT_8SMP', 'IND_8SMP',
                                 'ENG_8SMP', 'IPA_8SMP', 'IPS_8SMP']

        # 7smp km
        km_7smp_mat = 'LHHE7GC8'
        km_7smp_ind = 'LHHEAQWK'
        km_7smp_eng = 'LHHEEEB5'
        km_7smp_ipa = 'LHHF9Q62'
        km_7smp_ips = 'LHHFBCWT'
        km_7smp = [km_7smp_mat, km_7smp_ind,
                   km_7smp_eng, km_7smp_ipa, km_7smp_ips]
        column_order_km_7smp = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAT_7KM', 'IND_7KM',
                                'ENG_7KM', 'IPA_7KM', 'IPS_7KM']

        # 10sma ipa k13
        k13_10ipa_mat = 'LHHO4J0W'
        k13_10ipa_bio = 'LHHO78FV'
        k13_10ipa_fis = 'LHHOB3L0'
        k13_10ipa_kim = 'LHHODJIH'
        k13_10ipa = [k13_10ipa_mat, k13_10ipa_bio,
                     k13_10ipa_fis, k13_10ipa_kim]
        column_order_k13_10ipa = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAT_10IPA', 'FIS_10IPA',
                                  'KIM_10IPA', 'BIO_10IPA']

        # 10sma ips k13
        k13_10ips_mat = 'LHHOHQW3'
        k13_10ips_sos = 'LHHOL2GH'
        k13_10ips_eng = 'LHHOOPEJ'
        k13_10ips_eko = 'LHHOR6Q5'
        k13_10ips_ind = 'LHHOUB5D'
        k13_10ips_sej = 'LHHOXG3D'
        k13_10ips_geo = 'LHHP0FDK'
        k13_10ips = [k13_10ips_mat, k13_10ips_sos, k13_10ips_eng,
                     k13_10ips_eko, k13_10ips_ind, k13_10ips_sej, k13_10ips_geo]
        column_order_k13_10ips = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAT_10IPS', 'IND_10IPS',
                                  'ENG_10IPS', 'SEJ_10IPS', 'GEO_10IPS', 'EKO_10IPS', 'SOS_10IPS']

        # 10sma km
        km_10sma_mat = 'LHHP5ZFU'
        km_10sma_ind = 'LHHP8HQU'
        km_10sma_eng = 'LHHPD245'
        km_10sma_ipa = 'LHHPGDX7'
        km_10sma_ips = 'LHHPI776'
        km_10sma = [km_10sma_mat, km_10sma_ind,
                    km_10sma_eng, km_10sma_ipa, km_10sma_ips]
        column_order_km_10sma = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAT_10KM', 'IND_10KM',
                                 'ENG_10KM', 'IPA_10KM', 'IPS_10KM']

        # 11sma ipa k13
        k13_11ipa_mat = 'LHHPMD5N'
        k13_11ipa_bio = 'LHHPPGY1'
        k13_11ipa_fis = 'LHHPSAH8'
        k13_11ipa_kim = 'LHHPVWEU'
        k13_11ipa = [k13_11ipa_mat, k13_11ipa_bio,
                     k13_11ipa_fis, k13_11ipa_kim]
        column_order_k13_11ipa = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAT_11IPA', 'FIS_11IPA',
                                  'KIM_11IPA', 'BIO_11IPA']

        # 11sma ips k13
        k13_11ips_mat = 'LHHPY4DQ'
        k13_11ips_sos = 'LHHQ0GEG'
        k13_11ips_eng = 'LHHQ2SOF'
        k13_11ips_eko = 'LHHQ5KA7'
        k13_11ips_ind = 'LHHQ9G2X'
        k13_11ips_sej = 'LHHQBZF1'
        k13_11ips_geo = 'LHHQEL75'
        k13_11ips = [k13_11ips_mat, k13_11ips_sos, k13_11ips_eng,
                     k13_11ips_eko, k13_11ips_ind, k13_11ips_sej, k13_11ips_geo]
        column_order_k13_11ips = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAT_11IPS', 'IND_11IPS',
                                  'ENG_11IPS', 'SEJ_11IPS', 'GEO_11IPS', 'EKO_11IPS', 'SOS_11IPS']

        # 11sma km
        km_11sma_maw = 'LHHQHX38'
        km_11sma_map = 'LHPP32X4'
        km_11sma_ind = 'LHHR041C'
        km_11sma_eng = 'LHHQTDZN'
        km_11sma_sej = 'LHHR2936'
        km_11sma_geo = 'LHHR7D5N'
        km_11sma_eko = 'LHHQXWPS'
        km_11sma_sos = 'LHHQNOEV'
        km_11sma_fis = 'LHHQVJ2O'
        km_11sma_kim = 'LHHR50DU'
        km_11sma_bio = 'LHHQQ7WE'
        km_11sma = [km_11sma_maw, km_11sma_map, km_11sma_ind, km_11sma_eng,
                    km_11sma_sej, km_11sma_geo, km_11sma_eko, km_11sma_sos,
                    km_11sma_fis, km_11sma_kim, km_11sma_bio]
        column_order_km_11sma = ['IDTAHUN', 'NAMA', 'NONF', 'KELAS', 'NAMA_SKLH', 'KD_LOK', 'MAW_11KM', 'MAP_11KM', 'IND_11KM',
                                 'ENG_11KM', 'SEJ_11KM', 'GEO_11KM', 'EKO_11KM', 'SOS_11KM', 'FIS_11KM', 'KIM_11KM', 'BIO_11KM']

        image = Image.open('logo resmi nf resize.png')
        st.image(image)

        st.title("PIVOT - PAT")

        col1 = st.container()
        with col1:
            KURIKULUM = st.selectbox(
                "KURIKULUM",
                ("--Pilih Kurikulum--", "K13", "KM"))

        col2 = st.container()
        with col2:
            KELAS = st.selectbox(
                "KELAS",
                ("--Pilih Kelas--", "4 SD", "5 SD", "7 SMP", "8 SMP", "10 IPA", "10 IPS", "10 SMA", "11 IPA", "11 IPS", "11 SMA"))

        TAHUN = st.text_input("Masukkan Tahun Ajaran",
                              placeholder="contoh: 2022-2023")

        uploaded_bobot = st.file_uploader(
            'Letakkan file excel bobot TO', type='xlsx')
        uploaded_jwb = st.file_uploader(
            'Letakkan file excel jwb TO', type='xlsx')

        bobot = None
        jwb = None

        if uploaded_bobot is not None:
            bobot = pd.read_excel(uploaded_bobot)

        if uploaded_jwb is not None:
            jwb = pd.read_excel(uploaded_jwb)

        if bobot is not None and jwb is not None:
            bobot = bobot.drop(['id', 'jns_pkt', 'jns_tes', 'kel_studi', 'nama_tes', 'no_soal', 'bobot', 'kd_studi', 'bab', 'eigen', 'kode_soal', 'st_eigen',
                                'modified_time', 'kode_naskah', 'group_tes', 'kunci', 'sequence', 'label', 'item_id'], axis=1)  # Menghilangkan kolom sebelum dilakukan merge

            result = pd.merge(bobot, jwb[['kode', 'nama', 'nonf', 'kd_lok',
                                          'nama_sklh', 'kelas', 'jml_benar']], on='kode', how='left')
            # Menghapus nilai NaN dari kolom 'nonf'
            result = result.dropna(subset=['nonf'])

            # k13
            if KELAS == "4 SD" and KURIKULUM == "K13":
                kode_kls_kur = k13_4sd
                column_order = column_order_k13_4sd
            elif KELAS == "5 SD" and KURIKULUM == "K13":
                kode_kls_kur = k13_5sd
                column_order = column_order_k13_5sd
            elif KELAS == "7 SMP" and KURIKULUM == "K13":
                kode_kls_kur = k13_7smp
                column_order = column_order_k13_7smp
            elif KELAS == "8 SMP" and KURIKULUM == "K13":
                kode_kls_kur = k13_8smp
                column_order = column_order_k13_8smp
            elif KELAS == "10 IPA" and KURIKULUM == "K13":
                kode_kls_kur = k13_10ipa
                column_order = column_order_k13_10ipa
            elif KELAS == "11 IPA" and KURIKULUM == "K13":
                kode_kls_kur = k13_11ipa
                column_order = column_order_k13_11ipa
            elif KELAS == "10 IPS" and KURIKULUM == "K13":
                kode_kls_kur = k13_10ips
                column_order = column_order_k13_10ips
            elif KELAS == "11 IPS" and KURIKULUM == "K13":
                kode_kls_kur = k13_11ips
                column_order = column_order_k13_11ips
            # km
            elif KELAS == "4 SD" and KURIKULUM == "KM":
                kode_kls_kur = km_4sd
                column_order = column_order_km_4sd
            elif KELAS == "7 SMP" and KURIKULUM == "KM":
                kode_kls_kur = km_7smp
                column_order = column_order_km_7smp
            elif KELAS == "10 SMA" and KURIKULUM == "KM":
                kode_kls_kur = km_10sma
                column_order = column_order_km_10sma
            elif KELAS == "11 SMA" and KURIKULUM == "KM":
                kode_kls_kur = km_11sma
                column_order = column_order_km_11sma

            result_filtered = result[result['kode'].isin(kode_kls_kur)]
            result_filtered.drop_duplicates(
                subset=['nama', 'kode'], keep='first', inplace=True)

            # Menggunakan pivot_table untuk menjadikan konten kolom 'studi' sebagai header dan menghilangkan duplikat
            result_pivot = pd.pivot_table(result_filtered, index=[
                'nama', 'nonf', 'kd_lok', 'nama_sklh', 'kelas', 'idtahun'], columns='kode', values='jml_benar', aggfunc='first')
            result_pivot.reset_index(inplace=True)  # Mengatur ulang indeks

            # Ubah nama kolom
            result_pivot = result_pivot.rename(
                columns={'nama': 'NAMA', 'nonf': 'NONF', 'kd_lok': 'KD_LOK', 'nama_sklh': 'NAMA_SKLH', 'kelas': 'KELAS', 'idtahun': 'IDTAHUN',
                         'LHHQHX38': 'MAW_11KM', 'LHHQNOEV': 'SOS_11KM', 'LHHQQ7WE': 'BIO_11KM', 'LHHQTDZN': 'ENG_11KM', 'LHHQVJ2O': 'FIS_11KM', 'LHHQXWPS': 'EKO_11KM', 'LHHR041C': 'IND_11KM', 'LHHR2936': 'SEJ_11KM', 'LHHR50DU': 'KIM_11KM', 'LHHR7D5N': 'GEO_11KM', 'LHPP32X4': 'MAP_11KM',
                         'LHHPY4DQ': 'MAT_11IPS', 'LHHQ0GEG': 'SOS_11IPS', 'LHHQ2SOF': 'ENG_11IPS', 'LHHQ5KA7': 'EKO_11IPS', 'LHHQ9G2X': 'IND_11IPS', 'LHHQBZF1': 'SEJ_11IPS', 'LHHQEL75': 'GEO_11IPS',
                         'LHHPMD5N': 'MAT_11IPA', 'LHHPPGY1': 'BIO_11IPA', 'LHHPSAH8': 'FIS_11IPA', 'LHHPVWEU': 'KIM_11IPA',
                         'LHHP5ZFU': 'MAT_10KM', 'LHHP8HQU': 'IND_10KM', 'LHHPD245': 'ENG_10KM', 'LHHPGDX7': 'IPA_10KM', 'LHHPI776': 'IPS_10KM',
                         'LHHOHQW3': 'MAT_10IPS', 'LHHOL2GH': 'SOS_10IPS', 'LHHOOPEJ': 'ENG_10IPS', 'LHHOR6Q5': 'EKO_10IPS', 'LHHOUB5D': 'IND_10IPS', 'LHHOXG3D': 'SEJ_10IPS', 'LHHP0FDK': 'GEO_10IPS',
                         'LHHO4J0W': 'MAT_10IPA', 'LHHO78FV': 'BIO_10IPA', 'LHHOB3L0': 'FIS_10IPA', 'LHHODJIH': 'KIM_10IPA',
                         'LHH6H3F6': 'MAT_8SMP', 'LHH6TEO5': 'IND_8SMP', 'LHHN9AZH': 'ENG_8SMP', 'LHHNDOAI': 'IPA_8SMP', 'LHHNFJ3E': 'IPS_8SMP',
                         'LHHE7GC8': 'MAT_7KM', 'LHHEAQWK': 'IND_7KM', 'LHHEEEB5': 'ENG_7KM', 'LHHF9Q62': 'IPA_7KM', 'LHHFBCWT': 'IPS_7KM',
                         'LHHDRBXZ': 'MAT_7SMP', 'LHHDUWKS': 'IND_7SMP', 'LHHDX6U7': 'ENG_7SMP', 'LHHDZC8Y': 'IPA_7SMP', 'LHHE476J': 'IPS_7SMP',
                         'LHH5V62M': 'MAT_5SD', 'LHH6WL2C': 'IND_5SD', 'LHH7NAB5': 'ENG_5SD', 'LHHCO0Q4': 'IPA_5SD', 'LHHDAY7I': 'IPS_5SD',
                         'LHH0U12P': 'MAT_4KM', 'LHH19TQN': 'IND_4KM', 'LHH47YLV': 'ENG_4KM', 'LHH4U3Q0': 'IPAS_4KM',
                         'LHG94EEQ': 'MAT_4SD', 'LHG9KCRA': 'IND_4SD', 'LHGA44Y9': 'ENG_4SD', 'LHGALT9N': 'IPA_4SD', 'LHH0F32F': 'IPS_4SD'})

            result_pivot = result_pivot.reindex(columns=column_order)

            kelas = KELAS.lower().replace(" ", "")
            kurikulum = KURIKULUM.lower()
            tahun = TAHUN.replace("-", "")

            path_file = f"{kelas}_pat_sm2_{kurikulum}_{tahun}_pivot.xlsx"

            # Simpan file ke direktori temporer
            temp_dir = tempfile.gettempdir()
            file_path = temp_dir + '/' + path_file
            # wb.save(file_path)

            # Menyimpan DataFrame ke file Excel
            result_pivot.to_excel(file_path, index=False)
            st.success("File siap diunduh!")

            # Tombol unduh file
            with open(file_path, "rb") as f:
                bytes_data = f.read()
            st.download_button(label="Unduh File", data=bytes_data,
                               file_name=path_file)
    if selected_file == "Nilai Std. SD, SMP, 10KM":
        # menghilangkan hamburger
        st.markdown("""
        <style>
        .css-1rs6os.edgvbvh3
        {
            visibility:hidden;
        }
        .css-1lsmgbg.egzxvld0
        {
            visibility:hidden;
        }
        </style>
        """, unsafe_allow_html=True)

        image = Image.open('logo resmi nf resize.png')
        st.image(image)

        st.title("Olah Nilai Standar K13")

        # url = 'https://docs.google.com/document/d/1xjkgcq86pMfLieqwBGTmV0kB6mWO_R1L7042razShlk/edit?usp=sharing'

        # st.warning("Harap dibaca terlebih dahulu panduannya")
        # if st.button("Panduan"):
        #     webbrowser.open_new_tab(url)

        st.header("SD-SMP-10KM")

        col6 = st.container()

        with col6:
            KELAS = st.selectbox(
                "KELAS",
                ("--Pilih Kelas--", "4 SD", "5 SD", "6 SD", "7 SMP", "8 SMP", "9 SMP", "10 SMA"))

        col7 = st.container()

        with col7:
            SEMESTER = st.selectbox(
                "SEMESTER",
                ("--Pilih Semester--", "SEMESTER 1", "SEMESTER 2"))

        col8 = st.container()

        with col8:
            PENILAIAN = st.selectbox(
                "PENILAIAN",
                ("--Pilih Penilaian--", "PENILAIAN TENGAH SEMESTER", "PENILAIAN AKHIR SEMESTER", "PENILAIAN AKHIR TAHUN"))

        col9 = st.container()

        with col9:
            KURIKULUM = st.selectbox(
                "KURIKULUM",
                ("--Pilih Kurikulum--", "K13", "KM"))

        TAHUN = st.text_input("Masukkan Tahun Ajaran",
                              placeholder="contoh: 2022-2023")

        col1, col2, col3, col4, col5 = st.columns(5)

        with col1:
            MTK = st.selectbox(
                "JML. SOAL MAT.",
                ("--Pilih--", 25, 30, 35, 40, 45))

        with col2:
            IND = st.selectbox(
                "JML. SOAL IND.",
                ("--Pilih--", 25, 30, 35, 40, 45))

        with col3:
            ENG = st.selectbox(
                "JML. SOAL ENG.",
                ("--Pilih--", 25, 30, 35, 40, 45))

        with col4:
            IPA = st.selectbox(
                "JML. SOAL IPA.",
                ("--Pilih--", 25, 30, 35, 40, 45))

        with col5:
            IPS = st.selectbox(
                "JML. SOAL IPS.",
                ("--Pilih--", 25, 30, 35, 40, 45))

        JML_SOAL_MAT = MTK
        JML_SOAL_IND = IND
        JML_SOAL_ENG = ENG
        JML_SOAL_IPA = IPA
        JML_SOAL_IPS = IPS

        uploaded_file = st.file_uploader(
            'Letakkan file excel', type='xlsx')

        if uploaded_file is not None:

            wb = openpyxl.load_workbook(uploaded_file)
            ws = wb['Sheet1']

            q = len(ws['K'])
            r = len(ws['K'])+2
            s = len(ws['K'])+3
            t = len(ws['K'])+4
            u = len(ws['K'])+5
            v = len(ws['K'])+6
            w = len(ws['K'])+7
            x = len(ws['K'])+8

            ws['G{}'.format(r)] = "=ROUND(AVERAGE(G2:G{}),2)".format(q)  # mat
            ws['H{}'.format(r)] = "=ROUND(AVERAGE(H2:H{}),2)".format(q)  # ind
            ws['I{}'.format(r)] = "=ROUND(AVERAGE(I2:I{}),2)".format(q)  # eng
            ws['J{}'.format(r)] = "=ROUND(AVERAGE(J2:J{}),2)".format(q)  # ipa
            ws['K{}'.format(r)] = "=ROUND(AVERAGE(K2:K{}),2)".format(q)  # ips
            ws['L{}'.format(r)] = "=ROUND(AVERAGE(L2:L{}),2)".format(q)  # jml
            ws['G{}'.format(s)] = "=STDEV(G2:G{})".format(q)
            ws['H{}'.format(s)] = "=STDEV(H2:H{})".format(q)
            ws['I{}'.format(s)] = "=STDEV(I2:I{})".format(q)
            ws['J{}'.format(s)] = "=STDEV(J2:J{})".format(q)
            ws['K{}'.format(s)] = "=STDEV(K2:K{})".format(q)
            ws['G{}'.format(t)] = "=MAX(G2:G{})".format(q)
            ws['H{}'.format(t)] = "=MAX(H2:H{})".format(q)
            ws['I{}'.format(t)] = "=MAX(I2:I{})".format(q)
            ws['J{}'.format(t)] = "=MAX(J2:J{})".format(q)
            ws['K{}'.format(t)] = "=MAX(K2:K{})".format(q)
            ws['L{}'.format(t)] = "=MAX(L2:L{})".format(q)
            ws['M{}'.format(r)] = "=MAX(M2:M{})".format(q)
            ws['N{}'.format(r)] = "=MAX(N2:N{})".format(q)
            ws['O{}'.format(r)] = "=MAX(O2:O{})".format(q)
            ws['P{}'.format(r)] = "=MAX(P2:P{})".format(q)
            ws['Q{}'.format(r)] = "=MAX(Q2:Q{})".format(q)
            ws['R{}'.format(r)] = "=MAX(R2:R{})".format(q)
            ws['S{}'.format(r)] = "=MAX(S2:S{})".format(q)
            ws['T{}'.format(r)] = "=MAX(T2:T{})".format(q)
            ws['U{}'.format(r)] = "=MAX(U2:U{})".format(q)
            ws['V{}'.format(r)] = "=MAX(V2:V{})".format(q)
            ws['W{}'.format(r)] = "=ROUND(MAX(W2:W{}),2)".format(q)
            ws['G{}'.format(u)] = "=MIN(G2:G{})".format(q)
            ws['H{}'.format(u)] = "=MIN(H2:H{})".format(q)
            ws['I{}'.format(u)] = "=MIN(I2:I{})".format(q)
            ws['J{}'.format(u)] = "=MIN(J2:J{})".format(q)
            ws['K{}'.format(u)] = "=MIN(K2:K{})".format(q)
            ws['L{}'.format(u)] = "=MIN(L2:L{})".format(q)
            ws['R{}'.format(s)] = "=MIN(R2:R{})".format(q)
            ws['S{}'.format(s)] = "=MIN(S2:S{})".format(q)
            ws['T{}'.format(s)] = "=MIN(T2:T{})".format(q)
            ws['U{}'.format(s)] = "=MIN(U2:U{})".format(q)
            ws['V{}'.format(s)] = "=MIN(V2:V{})".format(q)
            ws['W{}'.format(s)] = "=MIN(W2:W{})".format(q)
            ws['R{}'.format(t)] = "=ROUND(AVERAGE(R2:R{}),2)".format(q)
            ws['S{}'.format(t)] = "=ROUND(AVERAGE(S2:S{}),2)".format(q)
            ws['T{}'.format(t)] = "=ROUND(AVERAGE(T2:T{}),2)".format(q)
            ws['U{}'.format(t)] = "=ROUND(AVERAGE(U2:U{}),2)".format(q)
            ws['V{}'.format(t)] = "=ROUND(AVERAGE(V2:V{}),2)".format(q)
            ws['W{}'.format(t)] = "=ROUND(AVERAGE(W2:W{}),2)".format(q)
            ws['X{}'.format(r)] = "=MAX(X2:X{})".format(q)
            ws['Z{}'.format(r)] = "=SUM(Z2:Z{})".format(q)
            ws['AA{}'.format(r)] = "=SUM(AA2:AA{})".format(q)
            ws['AB{}'.format(r)] = "=SUM(AB2:AB{})".format(q)
            ws['AC{}'.format(r)] = "=SUM(AC2:AC{})".format(q)
            ws['AD{}'.format(r)] = "=SUM(AD2:AD{})".format(q)
            # new
            # iterasi 1 rata-rata - 1
            ws['F{}'.format(v)] = 'JUMLAH SOAL'
            ws['G{}'.format(v)] = JML_SOAL_MAT
            ws['H{}'.format(v)] = JML_SOAL_IND
            ws['I{}'.format(v)] = JML_SOAL_ENG
            ws['J{}'.format(v)] = JML_SOAL_IPA
            ws['K{}'.format(v)] = JML_SOAL_IPS
            ws['AK{}'.format(r)] = "=IF($Z${}=0,$G${},$G${}-1)".format(r, r, r)
            ws['AK{}'.format(s)] = "=STDEV(AK2:AK{})".format(q)
            ws['AK{}'.format(t)] = "=MAX(AK2:AK{})".format(q)
            ws['AK{}'.format(u)] = "=MIN(AK2:AK{})".format(q)
            ws['AL{}'.format(
                r)] = "=IF($AA${}=0,$H${},$H${}-1)".format(r, r, r)
            ws['AL{}'.format(s)] = "=STDEV(AL2:AL{})".format(q)
            ws['AL{}'.format(t)] = "=MAX(AL2:AL{})".format(q)
            ws['AL{}'.format(u)] = "=MIN(AL2:AL{})".format(q)
            ws['AM{}'.format(
                r)] = "=IF($AB${}=0,$I${},$I${}-1)".format(r, r, r)
            ws['AM{}'.format(s)] = "=STDEV(AM2:AM{})".format(q)
            ws['AM{}'.format(t)] = "=MAX(AM2:AM{})".format(q)
            ws['AM{}'.format(u)] = "=MIN(AM2:AM{})".format(q)
            ws['AN{}'.format(
                r)] = "=IF($AC${}=0,$J${},$J${}-1)".format(r, r, r)
            ws['AN{}'.format(s)] = "=STDEV(AN2:AN{})".format(q)
            ws['AN{}'.format(t)] = "=MAX(AN2:AN{})".format(q)
            ws['AN{}'.format(u)] = "=MIN(AN2:AN{})".format(q)
            ws['AO{}'.format(
                r)] = "=IF($AD${}=0,$K${},$K${}-1)".format(r, r, r)
            ws['AO{}'.format(s)] = "=STDEV(AO2:AO{})".format(q)
            ws['AO{}'.format(t)] = "=MAX(AO2:AO{})".format(q)
            ws['AO{}'.format(u)] = "=MIN(AO2:AO{})".format(q)
            ws['AP{}'.format(r)] = "=ROUND(AVERAGE(AP2:AP{}),2)".format(q)
            ws['AP{}'.format(t)] = "=MAX(AP2:AP{})".format(q)
            ws['AP{}'.format(u)] = "=MIN(AP2:AP{})".format(q)
            ws['AQ{}'.format(r)] = "=MAX(AQ2:AQ{})".format(q)
            ws['AR{}'.format(r)] = "=MAX(AR2:AR{})".format(q)
            ws['AS{}'.format(r)] = "=MAX(AS2:AS{})".format(q)
            ws['AT{}'.format(r)] = "=MAX(AT2:AT{})".format(q)
            ws['AU{}'.format(r)] = "=MAX(AU2:AU{})".format(q)
            ws['AV{}'.format(r)] = "=MAX(AV2:AV{})".format(q)
            ws['AV{}'.format(s)] = "=MIN(AV2:AV{})".format(q)
            ws['AV{}'.format(t)] = "=ROUND(AVERAGE(AV2:AV{}),2)".format(q)
            ws['AW{}'.format(r)] = "=MAX(AW2:AW{})".format(q)
            ws['AW{}'.format(s)] = "=MIN(AW2:AW{})".format(q)
            ws['AW{}'.format(t)] = "=ROUND(AVERAGE(AW2:AW{}),2)".format(q)
            ws['AX{}'.format(r)] = "=MAX(AX2:AX{})".format(q)
            ws['AX{}'.format(s)] = "=MIN(AX2:AX{})".format(q)
            ws['AX{}'.format(t)] = "=ROUND(AVERAGE(AX2:AX{}),2)".format(q)
            ws['AY{}'.format(r)] = "=MAX(AY2:AY{})".format(q)
            ws['AY{}'.format(s)] = "=MIN(AY2:AY{})".format(q)
            ws['AY{}'.format(t)] = "=ROUND(AVERAGE(AY2:AY{}),2)".format(q)
            ws['AZ{}'.format(r)] = "=MAX(AZ2:AZ{})".format(q)
            ws['AZ{}'.format(s)] = "=MIN(AZ2:AZ{})".format(q)
            ws['AZ{}'.format(t)] = "=ROUND(AVERAGE(AZ2:AZ{}),2)".format(q)
            ws['BA{}'.format(r)] = "=MAX(BA2:BA{})".format(q)
            ws['BA{}'.format(s)] = "=MIN(BA2:BA{})".format(q)
            ws['BA{}'.format(t)] = "=ROUND(AVERAGE(BA2:BA{}),2)".format(q)
            ws['BD{}'.format(r)] = "=SUM(BD2:BD{})".format(q)
            ws['BE{}'.format(r)] = "=SUM(BE2:BE{})".format(q)
            ws['BF{}'.format(r)] = "=SUM(BF2:BF{})".format(q)
            ws['BG{}'.format(r)] = "=SUM(BG2:BG{})".format(q)
            ws['BH{}'.format(r)] = "=SUM(BH2:BH{})".format(q)

            # iterasi 2 rata-rata - 1
            ws['BO{}'.format(
                r)] = "=IF($BD${}=0,$AK${},$AK${}-1)".format(r, r, r)
            ws['BO{}'.format(s)] = "=STDEV(BO2:BO{})".format(q)
            ws['BO{}'.format(t)] = "=MAX(BO2:BO{})".format(q)
            ws['BO{}'.format(u)] = "=MIN(BO2:BO{})".format(q)
            ws['BP{}'.format(
                r)] = "=IF($BE${}=0,$AL${},$AL${}-1)".format(r, r, r)
            ws['BP{}'.format(s)] = "=STDEV(BP2:BP{})".format(q)
            ws['BP{}'.format(t)] = "=MAX(BP2:BP{})".format(q)
            ws['BP{}'.format(u)] = "=MIN(BP2:BP{})".format(q)
            ws['BQ{}'.format(
                r)] = "=IF($BF${}=0,$AM${},$AM${}-1)".format(r, r, r)
            ws['BQ{}'.format(s)] = "=STDEV(BQ2:BQ{})".format(q)
            ws['BQ{}'.format(t)] = "=MAX(BQ2:BQ{})".format(q)
            ws['BQ{}'.format(u)] = "=MIN(BQ2:BQ{})".format(q)
            ws['BR{}'.format(
                r)] = "=IF($BG${}=0,$AN${},$AN${}-1)".format(r, r, r)
            ws['BR{}'.format(s)] = "=STDEV(BR2:BR{})".format(q)
            ws['BR{}'.format(t)] = "=MAX(BR2:BR{})".format(q)
            ws['BR{}'.format(u)] = "=MIN(BR2:BR{})".format(q)
            ws['BS{}'.format(
                r)] = "=IF($BH${}=0,$AO${},$AO${}-1)".format(r, r, r)
            ws['BS{}'.format(s)] = "=STDEV(BS2:BS{})".format(q)
            ws['BS{}'.format(t)] = "=MAX(BS2:BS{})".format(q)
            ws['BS{}'.format(u)] = "=MIN(BS2:BS{})".format(q)
            ws['BT{}'.format(r)] = "=ROUND(AVERAGE(BT2:BT{}),2)".format(q)
            ws['BT{}'.format(t)] = "=MAX(BT2:BT{})".format(q)
            ws['BT{}'.format(u)] = "=MIN(BT2:BT{})".format(q)
            ws['BU{}'.format(r)] = "=MAX(BU2:BU{})".format(q)
            ws['BV{}'.format(r)] = "=MAX(BV2:BV{})".format(q)
            ws['BW{}'.format(r)] = "=MAX(BW2:BW{})".format(q)
            ws['BX{}'.format(r)] = "=MAX(BX2:BX{})".format(q)
            ws['BY{}'.format(r)] = "=MAX(BY2:BY{})".format(q)
            ws['BZ{}'.format(r)] = "=MAX(BZ2:BZ{})".format(q)
            ws['BZ{}'.format(s)] = "=MIN(BZ2:BZ{})".format(q)
            ws['BZ{}'.format(t)] = "=ROUND(AVERAGE(BZ2:BZ{}),2)".format(q)
            ws['CA{}'.format(r)] = "=MAX(CA2:CA{})".format(q)
            ws['CA{}'.format(s)] = "=MIN(CA2:CA{})".format(q)
            ws['CA{}'.format(t)] = "=ROUND(AVERAGE(CA2:CA{}),2)".format(q)
            ws['CB{}'.format(r)] = "=MAX(CB2:CB{})".format(q)
            ws['CB{}'.format(s)] = "=MIN(CB2:CB{})".format(q)
            ws['CB{}'.format(t)] = "=ROUND(AVERAGE(CB2:CB{}),2)".format(q)
            ws['CC{}'.format(r)] = "=MAX(CC2:CC{})".format(q)
            ws['CC{}'.format(s)] = "=MIN(CC2:CC{})".format(q)
            ws['CC{}'.format(t)] = "=ROUND(AVERAGE(CC2:CC{}),2)".format(q)
            ws['CD{}'.format(r)] = "=MAX(CD2:CD{})".format(q)
            ws['CD{}'.format(s)] = "=MIN(CD2:CD{})".format(q)
            ws['CD{}'.format(t)] = "=ROUND(AVERAGE(CD2:CD{}),2)".format(q)
            ws['CE{}'.format(r)] = "=MAX(CE2:CE{})".format(q)
            ws['CE{}'.format(s)] = "=MIN(CE2:CE{})".format(q)
            ws['CE{}'.format(t)] = "=ROUND(AVERAGE(CE2:CE{}),2)".format(q)
            ws['CH{}'.format(r)] = "=SUM(CH2:CH{})".format(q)
            ws['CI{}'.format(r)] = "=SUM(CI2:CI{})".format(q)
            ws['CJ{}'.format(r)] = "=SUM(CJ2:CJ{})".format(q)
            ws['CK{}'.format(r)] = "=SUM(CK2:CK{})".format(q)
            ws['CL{}'.format(r)] = "=SUM(CL2:CL{})".format(q)

            # iterasi 3 rata-rata - 1
            ws['CS{}'.format(
                r)] = "=IF($CH${}=0,$BO${},$BO${}-1)".format(r, r, r)
            ws['CS{}'.format(s)] = "=STDEV(CS2:CS{})".format(q)
            ws['CS{}'.format(t)] = "=MAX(CS2:CS{})".format(q)
            ws['CS{}'.format(u)] = "=MIN(CS2:CS{})".format(q)
            ws['CT{}'.format(
                r)] = "=IF($CI${}=0,$BP${},$BP${}-1)".format(r, r, r)
            ws['CT{}'.format(s)] = "=STDEV(CT2:CT{})".format(q)
            ws['CT{}'.format(t)] = "=MAX(CT2:CT{})".format(q)
            ws['CT{}'.format(u)] = "=MIN(CT2:CT{})".format(q)
            ws['CU{}'.format(
                r)] = "=IF($CJ${}=0,$BQ${},$BQ${}-1)".format(r, r, r)
            ws['CU{}'.format(s)] = "=STDEV(CU2:CU{})".format(q)
            ws['CU{}'.format(t)] = "=MAX(CU2:CU{})".format(q)
            ws['CU{}'.format(u)] = "=MIN(CU2:CU{})".format(q)
            ws['CV{}'.format(
                r)] = "=IF($CK${}=0,$BR${},$BR${}-1)".format(r, r, r)
            ws['CV{}'.format(s)] = "=STDEV(CV2:CV{})".format(q)
            ws['CV{}'.format(t)] = "=MAX(CV2:CV{})".format(q)
            ws['CV{}'.format(u)] = "=MIN(CV2:CV{})".format(q)
            ws['CW{}'.format(
                r)] = "=IF($CL${}=0,$BS${},$BS${}-1)".format(r, r, r)
            ws['CW{}'.format(s)] = "=STDEV(CW2:CW{})".format(q)
            ws['CW{}'.format(t)] = "=MAX(CW2:CW{})".format(q)
            ws['CW{}'.format(u)] = "=MIN(CW2:CW{})".format(q)
            ws['CX{}'.format(r)] = "=ROUND(AVERAGE(CX2:CX{}),2)".format(q)
            ws['CX{}'.format(t)] = "=MAX(CX2:CX{})".format(q)
            ws['CX{}'.format(u)] = "=MIN(CX2:CX{})".format(q)
            ws['CY{}'.format(r)] = "=MAX(CY2:CY{})".format(q)
            ws['CZ{}'.format(r)] = "=MAX(CZ2:CZ{})".format(q)
            ws['DA{}'.format(r)] = "=MAX(DA2:DA{})".format(q)
            ws['DB{}'.format(r)] = "=MAX(DB2:DB{})".format(q)
            ws['DC{}'.format(r)] = "=MAX(DC2:DC{})".format(q)
            ws['DD{}'.format(r)] = "=MAX(DD2:DD{})".format(q)
            ws['DD{}'.format(s)] = "=MIN(DD2:DD{})".format(q)
            ws['DD{}'.format(t)] = "=ROUND(AVERAGE(DD2:DD{}),2)".format(q)
            ws['DE{}'.format(r)] = "=MAX(DE2:DE{})".format(q)
            ws['DE{}'.format(s)] = "=MIN(DE2:DE{})".format(q)
            ws['DE{}'.format(t)] = "=ROUND(AVERAGE(DE2:DE{}),2)".format(q)
            ws['DF{}'.format(r)] = "=MAX(DF2:DF{})".format(q)
            ws['DF{}'.format(s)] = "=MIN(DF2:DF{})".format(q)
            ws['DF{}'.format(t)] = "=ROUND(AVERAGE(DF2:DF{}),2)".format(q)
            ws['DG{}'.format(r)] = "=MAX(DG2:DG{})".format(q)
            ws['DG{}'.format(s)] = "=MIN(DG2:DG{})".format(q)
            ws['DG{}'.format(t)] = "=ROUND(AVERAGE(DG2:DG{}),2)".format(q)
            ws['DH{}'.format(r)] = "=MAX(DH2:DH{})".format(q)
            ws['DH{}'.format(s)] = "=MIN(DH2:DH{})".format(q)
            ws['DH{}'.format(t)] = "=ROUND(AVERAGE(DH2:DH{}),2)".format(q)
            ws['DI{}'.format(r)] = "=MAX(DI2:DI{})".format(q)
            ws['DI{}'.format(s)] = "=MIN(DI2:DI{})".format(q)
            ws['DI{}'.format(t)] = "=ROUND(AVERAGE(DI2:DI{}),2)".format(q)
            ws['DL{}'.format(r)] = "=SUM(DL2:DL{})".format(q)
            ws['DM{}'.format(r)] = "=SUM(DM2:DM{})".format(q)
            ws['DN{}'.format(r)] = "=SUM(DN2:DN{})".format(q)
            ws['DO{}'.format(r)] = "=SUM(DO2:DO{})".format(q)
            ws['DP{}'.format(r)] = "=SUM(DP2:DP{})".format(q)

            # iterasi 4 rata-rata - 1
            ws['DW{}'.format(
                r)] = "=IF($DL${}=0,$CS${},$CS${}-1)".format(r, r, r)
            ws['DW{}'.format(s)] = "=STDEV(DW2:DW{})".format(q)
            ws['DW{}'.format(t)] = "=MAX(DW2:DW{})".format(q)
            ws['DW{}'.format(u)] = "=MIN(DW2:DW{})".format(q)
            ws['DX{}'.format(
                r)] = "=IF($DM${}=0,$CT${},$CT${}-1)".format(r, r, r)
            ws['DX{}'.format(s)] = "=STDEV(DX2:DX{})".format(q)
            ws['DX{}'.format(t)] = "=MAX(DX2:DX{})".format(q)
            ws['DX{}'.format(u)] = "=MIN(DX2:DX{})".format(q)
            ws['DY{}'.format(
                r)] = "=IF($DN${}=0,$CU${},$CU${}-1)".format(r, r, r)
            ws['DY{}'.format(s)] = "=STDEV(DY2:DY{})".format(q)
            ws['DY{}'.format(t)] = "=MAX(DY2:DY{})".format(q)
            ws['DY{}'.format(u)] = "=MIN(DY2:DY{})".format(q)
            ws['DZ{}'.format(
                r)] = "=IF($DO${}=0,$CV${},$CV${}-1)".format(r, r, r)
            ws['DZ{}'.format(s)] = "=STDEV(DZ2:DZ{})".format(q)
            ws['DZ{}'.format(t)] = "=MAX(DZ2:DZ{})".format(q)
            ws['DZ{}'.format(u)] = "=MIN(DZ2:DZ{})".format(q)
            ws['EA{}'.format(
                r)] = "=IF($DP${}=0,$CW${},$CW${}-1)".format(r, r, r)
            ws['EA{}'.format(s)] = "=STDEV(EA2:EA{})".format(q)
            ws['EA{}'.format(t)] = "=MAX(EA2:EA{})".format(q)
            ws['EA{}'.format(u)] = "=MIN(EA2:EA{})".format(q)
            ws['EB{}'.format(r)] = "=ROUND(AVERAGE(EB2:EB{}),2)".format(q)
            ws['EB{}'.format(t)] = "=MAX(EB2:EB{})".format(q)
            ws['EB{}'.format(u)] = "=MIN(EB2:EB{})".format(q)
            ws['EC{}'.format(r)] = "=MAX(EC2:EC{})".format(q)
            ws['ED{}'.format(r)] = "=MAX(ED2:ED{})".format(q)
            ws['EE{}'.format(r)] = "=MAX(EE2:EE{})".format(q)
            ws['EF{}'.format(r)] = "=MAX(EF2:EF{})".format(q)
            ws['EG{}'.format(r)] = "=MAX(EG2:EG{})".format(q)
            ws['EH{}'.format(r)] = "=MAX(EH2:EH{})".format(q)
            ws['EH{}'.format(s)] = "=MIN(EH2:EH{})".format(q)
            ws['EH{}'.format(t)] = "=ROUND(AVERAGE(EH2:EH{}),2)".format(q)
            ws['EI{}'.format(r)] = "=MAX(EI2:EI{})".format(q)
            ws['EI{}'.format(s)] = "=MIN(EI2:EI{})".format(q)
            ws['EI{}'.format(t)] = "=ROUND(AVERAGE(EI2:EI{}),2)".format(q)
            ws['EJ{}'.format(r)] = "=MAX(EJ2:EJ{})".format(q)
            ws['EJ{}'.format(s)] = "=MIN(EJ2:EJ{})".format(q)
            ws['EJ{}'.format(t)] = "=ROUND(AVERAGE(EJ2:EJ{}),2)".format(q)
            ws['EK{}'.format(r)] = "=MAX(EK2:EK{})".format(q)
            ws['EK{}'.format(s)] = "=MIN(EK2:EK{})".format(q)
            ws['EK{}'.format(t)] = "=ROUND(AVERAGE(EK2:EK{}),2)".format(q)
            ws['EL{}'.format(r)] = "=MAX(EL2:EL{})".format(q)
            ws['EL{}'.format(s)] = "=MIN(EL2:EL{})".format(q)
            ws['EL{}'.format(t)] = "=ROUND(AVERAGE(EL2:EL{}),2)".format(q)
            ws['EM{}'.format(r)] = "=MAX(EM2:EM{})".format(q)
            ws['EM{}'.format(s)] = "=MIN(EM2:EM{})".format(q)
            ws['EM{}'.format(t)] = "=ROUND(AVERAGE(EM2:EM{}),2)".format(q)
            ws['EP{}'.format(r)] = "=SUM(EP2:EP{})".format(q)
            ws['EQ{}'.format(r)] = "=SUM(EQ2:EQ{})".format(q)
            ws['ER{}'.format(r)] = "=SUM(ER2:ER{})".format(q)
            ws['ES{}'.format(r)] = "=SUM(ES2:ES{})".format(q)
            ws['ET{}'.format(r)] = "=SUM(ET2:ET{})".format(q)

            # iterasi 5 rata-rata - 1
            ws['FA{}'.format(
                r)] = "=IF($EP${}=0,$DW${},$DW${}-1)".format(r, r, r)
            ws['FA{}'.format(s)] = "=STDEV(FA2:FA{})".format(q)
            ws['FA{}'.format(t)] = "=MAX(FA2:FA{})".format(q)
            ws['FA{}'.format(u)] = "=MIN(FA2:FA{})".format(q)
            ws['FB{}'.format(
                r)] = "=IF($EQ${}=0,$DX${},$DX${}-1)".format(r, r, r)
            ws['FB{}'.format(s)] = "=STDEV(FB2:FB{})".format(q)
            ws['FB{}'.format(t)] = "=MAX(FB2:FB{})".format(q)
            ws['FB{}'.format(u)] = "=MIN(FB2:FB{})".format(q)
            ws['FC{}'.format(
                r)] = "=IF($ER${}=0,$DY${},$DY${}-1)".format(r, r, r)
            ws['FC{}'.format(s)] = "=STDEV(FC2:FC{})".format(q)
            ws['FC{}'.format(t)] = "=MAX(FC2:FC{})".format(q)
            ws['FC{}'.format(u)] = "=MIN(FC2:FC{})".format(q)
            ws['FD{}'.format(
                r)] = "=IF($ES${}=0,$DZ${},$DZ${}-1)".format(r, r, r)
            ws['FD{}'.format(s)] = "=STDEV(FD2:FD{})".format(q)
            ws['FD{}'.format(t)] = "=MAX(FD2:FD{})".format(q)
            ws['FD{}'.format(u)] = "=MIN(FD2:FD{})".format(q)
            ws['FE{}'.format(
                r)] = "=IF($ET${}=0,$EA${},$EA${}-1)".format(r, r, r)
            ws['FE{}'.format(s)] = "=STDEV(FE2:FE{})".format(q)
            ws['FE{}'.format(t)] = "=MAX(FE2:FE{})".format(q)
            ws['FE{}'.format(u)] = "=MIN(FE2:FE{})".format(q)
            ws['FF{}'.format(r)] = "=ROUND(AVERAGE(FF2:FF{}),2)".format(q)
            ws['FF{}'.format(t)] = "=MAX(FF2:FF{})".format(q)
            ws['FF{}'.format(u)] = "=MIN(FF2:FF{})".format(q)
            ws['FG{}'.format(r)] = "=MAX(FG2:FG{})".format(q)
            ws['FH{}'.format(r)] = "=MAX(FH2:FH{})".format(q)
            ws['FI{}'.format(r)] = "=MAX(FI2:FI{})".format(q)
            ws['FJ{}'.format(r)] = "=MAX(FJ2:FJ{})".format(q)
            ws['FK{}'.format(r)] = "=MAX(FK2:FK{})".format(q)
            ws['FL{}'.format(r)] = "=MAX(FL2:FL{})".format(q)
            ws['FL{}'.format(s)] = "=MIN(FL2:FL{})".format(q)
            ws['FL{}'.format(t)] = "=ROUND(AVERAGE(FL2:FL{}),2)".format(q)
            ws['FM{}'.format(r)] = "=MAX(FM2:FM{})".format(q)
            ws['FM{}'.format(s)] = "=MIN(FM2:FM{})".format(q)
            ws['FM{}'.format(t)] = "=ROUND(AVERAGE(FM2:FM{}),2)".format(q)
            ws['FN{}'.format(r)] = "=MAX(FN2:FN{})".format(q)
            ws['FN{}'.format(s)] = "=MIN(FN2:FN{})".format(q)
            ws['FN{}'.format(t)] = "=ROUND(AVERAGE(FN2:FN{}),2)".format(q)
            ws['FO{}'.format(r)] = "=MAX(FO2:FO{})".format(q)
            ws['FO{}'.format(s)] = "=MIN(FO2:FO{})".format(q)
            ws['FO{}'.format(t)] = "=ROUND(AVERAGE(FO2:FO{}),2)".format(q)
            ws['FP{}'.format(r)] = "=MAX(FP2:FP{})".format(q)
            ws['FP{}'.format(s)] = "=MIN(FP2:FP{})".format(q)
            ws['FP{}'.format(t)] = "=ROUND(AVERAGE(FP2:FP{}),2)".format(q)
            ws['FQ{}'.format(r)] = "=MAX(FQ2:FQ{})".format(q)
            ws['FQ{}'.format(s)] = "=MIN(FQ2:FQ{})".format(q)
            ws['FQ{}'.format(t)] = "=ROUND(AVERAGE(FQ2:FQ{}),2)".format(q)
            ws['FT{}'.format(r)] = "=SUM(FT2:FT{})".format(q)
            ws['FU{}'.format(r)] = "=SUM(FU2:FU{})".format(q)
            ws['FV{}'.format(r)] = "=SUM(FV2:FV{})".format(q)
            ws['FW{}'.format(r)] = "=SUM(FW2:FW{})".format(q)
            ws['FX{}'.format(r)] = "=SUM(FX2:FX{})".format(q)

            # Z Score
            ws['B1'] = 'NAMA_SISWA_1'
            ws['C1'] = 'NOMOR_NF_1'
            ws['D1'] = 'KELAS_1'
            ws['E1'] = 'NAMA_SEKOLAH_1'
            ws['F1'] = 'LOKASI_1'
            ws['G1'] = 'MAT_1'
            ws['H1'] = 'IND_1'
            ws['I1'] = 'ENG_1'
            ws['J1'] = 'IPA_1'
            ws['K1'] = 'IPS_1'
            ws['L1'] = 'JML_1'
            ws['M1'] = 'Z_MAT_1'
            ws['N1'] = 'Z_IND_1'
            ws['O1'] = 'Z_ENG_1'
            ws['P1'] = 'Z_IPA_1'
            ws['Q1'] = 'Z_IPS_1'
            ws['R1'] = 'S_MAT_1'
            ws['S1'] = 'S_IND_1'
            ws['T1'] = 'S_ENG_1'
            ws['U1'] = 'S_IPA_1'
            ws['V1'] = 'S_IPS_1'
            ws['W1'] = 'S_JML_1'
            ws['X1'] = 'RANK_NAS._1'
            ws['Y1'] = 'RANK_LOK._1'
            ws['M1'].font = Font(bold=False, name='Calibri', size=11)
            ws['N1'].font = Font(bold=False, name='Calibri', size=11)
            ws['O1'].font = Font(bold=False, name='Calibri', size=11)
            ws['P1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Q1'].font = Font(bold=False, name='Calibri', size=11)
            ws['R1'].font = Font(bold=False, name='Calibri', size=11)
            ws['S1'].font = Font(bold=False, name='Calibri', size=11)
            ws['T1'].font = Font(bold=False, name='Calibri', size=11)
            ws['U1'].font = Font(bold=False, name='Calibri', size=11)
            ws['V1'].font = Font(bold=False, name='Calibri', size=11)
            ws['W1'].font = Font(bold=False, name='Calibri', size=11)
            ws['X1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Y1'].font = Font(bold=False, name='Calibri', size=11)
        # FILL
            ws['B1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['C1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['D1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['E1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['F1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['G1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['H1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['I1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['J1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['K1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['L1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['M1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['N1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['O1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['P1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Q1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['R1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['S1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['T1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['U1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['V1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['W1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['X1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Y1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            # tambahan
            ws['Z1'] = 'MAT_20_1'
            ws['AA1'] = 'IND_20_1'
            ws['AB1'] = 'ENG_20_1'
            ws['AC1'] = 'IPA_20_1'
            ws['AD1'] = 'IPS_20_1'
            ws['Z1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Z1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AA1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AB1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AC1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AD1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')

            for row in range(2, q+1):
                ws['L{}'.format(
                    row)] = '=SUM(G{}:K{})'.format(row, row, row)
                ws['M{}'.format(
                    row)] = '=IFERROR(ROUND(IF(G{}="","",(G{}-G${})/G${}),2),"")'.format(row, row, r, s)
                ws['N{}'.format(
                    row)] = '=IFERROR(ROUND(IF(H{}="","",(H{}-H${})/H${}),2),"")'.format(row, row, r, s)
                ws['O{}'.format(
                    row)] = '=IFERROR(ROUND(IF(I{}="","",(I{}-I${})/I${}),2),"")'.format(row, row, r, s)
                ws['P{}'.format(
                    row)] = '=IFERROR(ROUND(IF(J{}="","",(J{}-J${})/J${}),2),"")'.format(row, row, r, s)
                ws['Q{}'.format(
                    row)] = '=IFERROR(ROUND(IF(K{}="","",(K{}-K${})/K${}),2),"")'.format(row, row, r, s)
                ws['R{}'.format(
                    row)] = '=IFERROR(ROUND(IF(G{}="","",IF(70+30*M{}/$M${}<20,20,70+30*M{}/$M${})),2),"")'.format(row, row, r, row, r)
                ws['S{}'.format(
                    row)] = '=IFERROR(ROUND(IF(H{}="","",IF(70+30*N{}/$N${}<20,20,70+30*N{}/$N${})),2),"")'.format(row, row, r, row, r)
                ws['T{}'.format(
                    row)] = '=IFERROR(ROUND(IF(I{}="","",IF(70+30*O{}/$O${}<20,20,70+30*O{}/$O${})),2),"")'.format(row, row, r, row, r)
                ws['U{}'.format(
                    row)] = '=IFERROR(ROUND(IF(J{}="","",IF(70+30*P{}/$P${}<20,20,70+30*P{}/$P${})),2),"")'.format(row, row, r, row, r)
                ws['V{}'.format(
                    row)] = '=IFERROR(ROUND(IF(K{}="","",IF(70+30*Q{}/$Q${}<20,20,70+30*Q{}/$Q${})),2),"")'.format(row, row, r, row, r)

                ws['W{}'.format(row)] = '=IF(SUM(R{}:V{})=0,"",SUM(R{}:V{}))'.format(
                    row, row, row, row)
                ws['X{}'.format(row)] = '=IF(W{}="","",RANK(W{},$W$2:$W${}))'.format(
                    row, row, q)
                ws['Y{}'.format(
                    row)] = '=IF(X{}="","",COUNTIFS($F$2:$F${},F{},$X$2:$X${},"<"&X{})+1)'.format(row, q, row, q, row)
                # TAMBAHAN
                ws['Z{}'.format(row)] = '=IF($G${}=25,IF(AND(G{}>4,R{}=20),1,""),IF($G${}=30,IF(AND(G{}>5,R{}=20),1,""),IF($G${}=35,IF(AND(G{}>6,R{}=20),1,""),IF($G${}=40,IF(AND(G{}>7,R{}=20),1,""),IF($G${}=45,IF(AND(G{}>8,R{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AA{}'.format(row)] = '=IF($H${}=25,IF(AND(H{}>4,S{}=20),1,""),IF($H${}=30,IF(AND(H{}>5,S{}=20),1,""),IF($H${}=35,IF(AND(H{}>6,S{}=20),1,""),IF($H${}=40,IF(AND(H{}>7,S{}=20),1,""),IF($H${}=45,IF(AND(H{}>8,S{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AB{}'.format(row)] = '=IF($I${}=25,IF(AND(I{}>4,T{}=20),1,""),IF($I${}=30,IF(AND(I{}>5,T{}=20),1,""),IF($I${}=35,IF(AND(I{}>6,T{}=20),1,""),IF($I${}=40,IF(AND(I{}>7,T{}=20),1,""),IF($I${}=45,IF(AND(I{}>8,T{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AC{}'.format(row)] = '=IF($J${}=25,IF(AND(J{}>4,U{}=20),1,""),IF($J${}=30,IF(AND(J{}>5,U{}=20),1,""),IF($J${}=35,IF(AND(J{}>6,U{}=20),1,""),IF($J${}=40,IF(AND(J{}>7,U{}=20),1,""),IF($J${}=45,IF(AND(J{}>8,U{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AD{}'.format(row)] = '=IF($K${}=25,IF(AND(K{}>4,V{}=20),1,""),IF($K${}=30,IF(AND(K{}>5,V{}=20),1,""),IF($K${}=35,IF(AND(K{}>6,V{}=20),1,""),IF($K${}=40,IF(AND(K{}>7,V{}=20),1,""),IF($K${}=45,IF(AND(K{}>8,V{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

        # new Z Score
            ws['AF1'] = 'NAMA_SISWA_2'
            ws['AG1'] = 'NOMOR_NF_2'
            ws['AH1'] = 'KELAS_2'
            ws['AI1'] = 'NAMA_SEKOLAH_2'
            ws['AJ1'] = 'LOKASI_2'
            ws['AK1'] = 'MAT_2'
            ws['AL1'] = 'IND_2'
            ws['AM1'] = 'ENG_2'
            ws['AN1'] = 'IPA_2'
            ws['AO1'] = 'IPS_2'
            ws['AP1'] = 'JML_2'
            ws['AQ1'] = 'Z_MAT_2'
            ws['AR1'] = 'Z_IND_2'
            ws['AS1'] = 'Z_ENG_2'
            ws['AT1'] = 'Z_IPA_2'
            ws['AU1'] = 'Z_IPS_2'
            ws['AV1'] = 'S_MAT_2'
            ws['AW1'] = 'S_IND_2'
            ws['AX1'] = 'S_ENG_2'
            ws['AY1'] = 'S_IPA_2'
            ws['AZ1'] = 'S_IPS_2'
            ws['BA1'] = 'S_JML_2'
            ws['BB1'] = 'RANK_NAS._2'
            ws['BC1'] = 'RANK_LOK._2'
            ws['AQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BC1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['AF1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AG1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AH1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AI1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AJ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AK1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AL1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AM1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AN1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AO1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AP1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AQ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AR1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AS1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AT1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AU1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AV1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AW1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AX1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AY1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AZ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BA1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BB1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BC1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')

            # tambahan
            ws['BD1'] = 'MAT_20_2'
            ws['BE1'] = 'IND_20_2'
            ws['BF1'] = 'ENG_20_2'
            ws['BG1'] = 'IPA_20_2'
            ws['BH1'] = 'IPS_20_2'
            ws['BD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BD1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BE1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BF1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BG1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BH1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')

            for row in range(2, q+1):
                ws['AF{}'.format(row)] = '=B{}'.format(row)
                ws['AG{}'.format(row)] = '=C{}'.format(row, row)
                ws['AH{}'.format(row)] = '=D{}'.format(row, row)
                ws['AI{}'.format(row)] = '=E{}'.format(row, row)
                ws['AJ{}'.format(row)] = '=F{}'.format(row, row)
                ws['AK{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['AL{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['AM{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['AN{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['AO{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['AP{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['AQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AK{}="","",(AK{}-AK${})/AK${}),2),"")'.format(row, row, r, s)
                ws['AR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AL{}="","",(AL{}-AL${})/AL${}),2),"")'.format(row, row, r, s)
                ws['AS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AM{}="","",(AM{}-AM${})/AM${}),2),"")'.format(row, row, r, s)
                ws['AT{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AN{}="","",(AN{}-AN${})/AN${}),2),"")'.format(row, row, r, s)
                ws['AU{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AO{}="","",(AO{}-AO${})/AO${}),2),"")'.format(row, row, r, s)
                ws['AV{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AK{}="","",IF(70+30*AQ{}/$AQ${}<20,20,70+30*AQ{}/$AQ${})),2),"")'.format(row, row, r, row, r)
                ws['AW{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AL{}="","",IF(70+30*AR{}/$AR${}<20,20,70+30*AR{}/$AR${})),2),"")'.format(row, row, r, row, r)
                ws['AX{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AM{}="","",IF(70+30*AS{}/$AS${}<20,20,70+30*AS{}/$AS${})),2),"")'.format(row, row, r, row, r)
                ws['AY{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AN{}="","",IF(70+30*AT{}/$AT${}<20,20,70+30*AT{}/$AT${})),2),"")'.format(row, row, r, row, r)
                ws['AZ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AO{}="","",IF(70+30*AU{}/$AU${}<20,20,70+30*AU{}/$AU${})),2),"")'.format(row, row, r, row, r)

                ws['BA{}'.format(row)] = '=IF(SUM(AV{}:AZ{})=0,"",SUM(AV{}:AZ{}))'.format(
                    row, row, row, row)
                ws['BB{}'.format(row)] = '=IF(BA{}="","",RANK(BA{},$BA$2:$BA${}))'.format(
                    row, row, q)
                ws['BC{}'.format(
                    row)] = '=IF(BB{}="","",COUNTIFS($AJ$2:$AJ${},F{},$BB$2:$BB${},"<"&BB{})+1)'.format(row, q, row, q, row)
            #     TAMBAHAN
                ws['BD{}'.format(row)] = '=IF($G${}=25,IF(AND(AK{}>4,AV{}=20),1,""),IF($G${}=30,IF(AND(AK{}>5,AV{}=20),1,""),IF($G${}=35,IF(AND(AK{}>6,AV{}=20),1,""),IF($G${}=40,IF(AND(AK{}>7,AV{}=20),1,""),IF($G${}=45,IF(AND(AK{}>8,AV{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BE{}'.format(row)] = '=IF($H${}=25,IF(AND(AL{}>4,AW{}=20),1,""),IF($H${}=30,IF(AND(AL{}>5,AW{}=20),1,""),IF($H${}=35,IF(AND(AL{}>6,AW{}=20),1,""),IF($H${}=40,IF(AND(AL{}>7,AW{}=20),1,""),IF($H${}=45,IF(AND(AL{}>8,AW{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BF{}'.format(row)] = '=IF($I${}=25,IF(AND(AM{}>4,AX{}=20),1,""),IF($I${}=30,IF(AND(AM{}>5,AX{}=20),1,""),IF($I${}=35,IF(AND(AM{}>6,AX{}=20),1,""),IF($I${}=40,IF(AND(AM{}>7,AX{}=20),1,""),IF($I${}=45,IF(AND(AM{}>8,AX{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BG{}'.format(row)] = '=IF($J${}=25,IF(AND(AN{}>4,AY{}=20),1,""),IF($J${}=30,IF(AND(AN{}>5,AY{}=20),1,""),IF($J${}=35,IF(AND(AN{}>6,AY{}=20),1,""),IF($J${}=40,IF(AND(AN{}>7,AY{}=20),1,""),IF($J${}=45,IF(AND(AN{}>8,AY{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BH{}'.format(row)] = '=IF($K${}=25,IF(AND(AO{}>4,AZ{}=20),1,""),IF($K${}=30,IF(AND(AO{}>5,AZ{}=20),1,""),IF($K${}=35,IF(AND(AO{}>6,AZ{}=20),1,""),IF($K${}=40,IF(AND(AO{}>7,AZ{}=20),1,""),IF($K${}=45,IF(AND(AO{}>8,AZ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

                # new Z Score [2]
            ws['BJ1'] = 'NAMA_SISWA_3'
            ws['BK1'] = 'NOMOR_NF_3'
            ws['BL1'] = 'KELAS_3'
            ws['BM1'] = 'NAMA_SEKOLAH_3'
            ws['BN1'] = 'LOKASI_3'
            ws['BO1'] = 'MAT_3'
            ws['BP1'] = 'IND_3'
            ws['BQ1'] = 'ENG_3'
            ws['BR1'] = 'IPA_3'
            ws['BS1'] = 'IPS_3'
            ws['BT1'] = 'JML_3'
            ws['BU1'] = 'Z_MAT_3'
            ws['BV1'] = 'Z_IND_3'
            ws['BW1'] = 'Z_ENG_3'
            ws['BX1'] = 'Z_IPA_3'
            ws['BY1'] = 'Z_IPS_3'
            ws['BZ1'] = 'S_MAT_3'
            ws['CA1'] = 'S_IND_3'
            ws['CB1'] = 'S_ENG_3'
            ws['CC1'] = 'S_IPA_3'
            ws['CD1'] = 'S_IPS_3'
            ws['CE1'] = 'S_JML_3'
            ws['CF1'] = 'RANK_NAS._3'
            ws['CG1'] = 'RANK_LOK._3'
            ws['BU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CG1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['BJ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BK1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BL1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BM1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BN1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BO1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BP1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BQ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BR1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BS1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BT1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BU1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BV1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BW1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BX1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BY1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BZ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CA1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CB1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CC1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CD1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CE1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CF1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CG1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')

            # tambahan
            ws['CH1'] = 'MAT_20_3'
            ws['CI1'] = 'IND_20_3'
            ws['CJ1'] = 'ENG_20_3'
            ws['CK1'] = 'IPA_20_3'
            ws['CL1'] = 'IPS_20_3'
            ws['CH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CH1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CI1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CJ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CK1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CL1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')

            for row in range(2, q+1):
                ws['BJ{}'.format(row)] = '=B{}'.format(row)
                ws['BK{}'.format(row)] = '=C{}'.format(row, row)
                ws['BL{}'.format(row)] = '=D{}'.format(row, row)
                ws['BM{}'.format(row)] = '=E{}'.format(row, row)
                ws['BN{}'.format(row)] = '=F{}'.format(row, row)
                ws['BO{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['BP{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['BQ{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['BR{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['BS{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['BT{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['BU{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BO{}="","",(BO{}-BO${})/BO${}),2),"")'.format(row, row, r, s)
                ws['BV{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BP{}="","",(BP{}-BP${})/BP${}),2),"")'.format(row, row, r, s)
                ws['BW{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BQ{}="","",(BQ{}-BQ${})/BQ${}),2),"")'.format(row, row, r, s)
                ws['BX{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BR{}="","",(BR{}-BR${})/BR${}),2),"")'.format(row, row, r, s)
                ws['BY{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BS{}="","",(BS{}-BS${})/BS${}),2),"")'.format(row, row, r, s)
                ws['BZ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BO{}="","",IF(70+30*BU{}/$BU${}<20,20,70+30*BU{}/$BU${})),2),"")'.format(row, row, r, row, r)
                ws['CA{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BP{}="","",IF(70+30*BV{}/$BV${}<20,20,70+30*BV{}/$BV${})),2),"")'.format(row, row, r, row, r)
                ws['CB{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BQ{}="","",IF(70+30*BW{}/$BW${}<20,20,70+30*BW{}/$BW${})),2),"")'.format(row, row, r, row, r)
                ws['CC{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BR{}="","",IF(70+30*BX{}/$BX${}<20,20,70+30*BX{}/$BX${})),2),"")'.format(row, row, r, row, r)
                ws['CD{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BS{}="","",IF(70+30*BY{}/$BY${}<20,20,70+30*BY{}/$BY${})),2),"")'.format(row, row, r, row, r)

                ws['CE{}'.format(row)] = '=IF(SUM(BZ{}:CD{})=0,"",SUM(BZ{}:CD{}))'.format(
                    row, row, row, row)
                ws['CF{}'.format(row)] = '=IF(CE{}="","",RANK(CE{},$CE$2:$CE${}))'.format(
                    row, row, q)
                ws['CG{}'.format(
                    row)] = '=IF(CF{}="","",COUNTIFS($BN$2:$BN${},F{},$CF$2:$CF${},"<"&CF{})+1)'.format(row, q, row, q, row)
                #     TAMBAHAN
                ws['CH{}'.format(row)] = '=IF($G${}=25,IF(AND(BO{}>4,BZ{}=20),1,""),IF($G${}=30,IF(AND(BO{}>5,BZ{}=20),1,""),IF($G${}=35,IF(AND(BO{}>6,BZ{}=20),1,""),IF($G${}=40,IF(AND(BO{}>7,BZ{}=20),1,""),IF($G${}=45,IF(AND(BO{}>8,BZ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CI{}'.format(row)] = '=IF($H${}=25,IF(AND(BP{}>4,CA{}=20),1,""),IF($H${}=30,IF(AND(BP{}>5,CA{}=20),1,""),IF($H${}=35,IF(AND(BP{}>6,CA{}=20),1,""),IF($H${}=40,IF(AND(BP{}>7,CA{}=20),1,""),IF($H${}=45,IF(AND(BP{}>8,CA{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CJ{}'.format(row)] = '=IF($I${}=25,IF(AND(BQ{}>4,CB{}=20),1,""),IF($I${}=30,IF(AND(BQ{}>5,CB{}=20),1,""),IF($I${}=35,IF(AND(BQ{}>6,CB{}=20),1,""),IF($I${}=40,IF(AND(BQ{}>7,CB{}=20),1,""),IF($I${}=45,IF(AND(BQ{}>8,CB{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CK{}'.format(row)] = '=IF($J${}=25,IF(AND(BR{}>4,CC{}=20),1,""),IF($J${}=30,IF(AND(BR{}>5,CC{}=20),1,""),IF($J${}=35,IF(AND(BR{}>6,CC{}=20),1,""),IF($J${}=40,IF(AND(BR{}>7,CC{}=20),1,""),IF($J${}=45,IF(AND(BR{}>8,CC{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CL{}'.format(row)] = '=IF($K${}=25,IF(AND(BS{}>4,CD{}=20),1,""),IF($K${}=30,IF(AND(BS{}>5,CD{}=20),1,""),IF($K${}=35,IF(AND(BS{}>6,CD{}=20),1,""),IF($K${}=40,IF(AND(BS{}>7,CD{}=20),1,""),IF($K${}=45,IF(AND(BS{}>8,CD{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

                # new Z Score [3]
            ws['CN1'] = 'NAMA_SISWA_4'
            ws['CO1'] = 'NOMOR_NF_4'
            ws['CP1'] = 'KELAS_4'
            ws['CQ1'] = 'NAMA_SEKOLAH_4'
            ws['CR1'] = 'LOKASI_4'
            ws['CS1'] = 'MAT_4'
            ws['CT1'] = 'IND_4'
            ws['CU1'] = 'ENG_4'
            ws['CV1'] = 'IPA_4'
            ws['CW1'] = 'IPS_4'
            ws['CX1'] = 'JML_4'
            ws['CY1'] = 'Z_MAT_4'
            ws['CZ1'] = 'Z_IND_4'
            ws['DA1'] = 'Z_ENG_4'
            ws['DB1'] = 'Z_IPA_4'
            ws['DC1'] = 'Z_IPS_4'
            ws['DD1'] = 'S_MAT_4'
            ws['DE1'] = 'S_IND_4'
            ws['DF1'] = 'S_ENG_4'
            ws['DG1'] = 'S_IPA_4'
            ws['DH1'] = 'S_IPS_4'
            ws['DI1'] = 'S_JML_4'
            ws['DJ1'] = 'RANK_NAS._4'
            ws['DK1'] = 'RANK_LOK._4'
            ws['CY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DK1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['CN1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CO1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CP1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CQ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CR1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CS1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CT1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CU1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CV1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CW1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CX1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CY1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CZ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DA1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DB1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DC1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DD1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DE1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DF1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DG1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DH1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DI1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DJ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DK1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')

            # tambahan
            ws['DL1'] = 'MAT_20_4'
            ws['DM1'] = 'IND_20_4'
            ws['DN1'] = 'ENG_20_4'
            ws['DO1'] = 'IPA_20_4'
            ws['DP1'] = 'IPS_20_4'
            ws['DL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DL1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DM1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DN1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DO1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DP1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')

            for row in range(2, q+1):
                ws['CN{}'.format(row)] = '=B{}'.format(row)
                ws['CO{}'.format(row)] = '=C{}'.format(row, row)
                ws['CP{}'.format(row)] = '=D{}'.format(row, row)
                ws['CQ{}'.format(row)] = '=E{}'.format(row, row)
                ws['CR{}'.format(row)] = '=F{}'.format(row, row)
                ws['CS{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['CT{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['CU{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['CV{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['CW{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['CX{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['CY{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CS{}="","",(CS{}-CS${})/CS${}),2),"")'.format(row, row, r, s)
                ws['CZ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CT{}="","",(CT{}-CT${})/CT${}),2),"")'.format(row, row, r, s)
                ws['DA{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CU{}="","",(CU{}-CU${})/CU${}),2),"")'.format(row, row, r, s)
                ws['DB{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CV{}="","",(CV{}-CV${})/CV${}),2),"")'.format(row, row, r, s)
                ws['DC{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CW{}="","",(CW{}-CW${})/CW${}),2),"")'.format(row, row, r, s)
                ws['DD{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CY{}="","",IF(70+30*CY{}/$CY${}<20,20,70+30*CY{}/$CY${})),2),"")'.format(row, row, r, row, r)
                ws['DE{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CZ{}="","",IF(70+30*CZ{}/$CZ${}<20,20,70+30*CZ{}/$CZ${})),2),"")'.format(row, row, r, row, r)
                ws['DF{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DA{}="","",IF(70+30*DA{}/$DA${}<20,20,70+30*DA{}/$DA${})),2),"")'.format(row, row, r, row, r)
                ws['DG{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DB{}="","",IF(70+30*DB{}/$DB${}<20,20,70+30*DB{}/$DB${})),2),"")'.format(row, row, r, row, r)
                ws['DH{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DC{}="","",IF(70+30*DC{}/$DC${}<20,20,70+30*DC{}/$DC${})),2),"")'.format(row, row, r, row, r)

                ws['DI{}'.format(row)] = '=IF(SUM(DD{}:DH{})=0,"",SUM(DD{}:DH{}))'.format(
                    row, row, row, row)
                ws['DJ{}'.format(row)] = '=IF(DI{}="","",RANK(DI{},$DI$2:$DI${}))'.format(
                    row, row, q)
                ws['DK{}'.format(
                    row)] = '=IF(DJ{}="","",COUNTIFS($CR$2:$CR${},F{},$DJ$2:$DJ${},"<"&DJ{})+1)'.format(row, q, row, q, row)
                #     TAMBAHAN
                ws['DL{}'.format(row)] = '=IF($G${}=25,IF(AND(CS{}>4,DD{}=20),1,""),IF($G${}=30,IF(AND(CS{}>5,DD{}=20),1,""),IF($G${}=35,IF(AND(CS{}>6,DD{}=20),1,""),IF($G${}=40,IF(AND(CS{}>7,DD{}=20),1,""),IF($G${}=45,IF(AND(CS{}>8,DD{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DM{}'.format(row)] = '=IF($H${}=25,IF(AND(CT{}>4,DE{}=20),1,""),IF($H${}=30,IF(AND(CT{}>5,DE{}=20),1,""),IF($H${}=35,IF(AND(CT{}>6,DE{}=20),1,""),IF($H${}=40,IF(AND(CT{}>7,DE{}=20),1,""),IF($H${}=45,IF(AND(CT{}>8,DE{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DN{}'.format(row)] = '=IF($I${}=25,IF(AND(CU{}>4,DF{}=20),1,""),IF($I${}=30,IF(AND(CU{}>5,DF{}=20),1,""),IF($I${}=35,IF(AND(CU{}>6,DF{}=20),1,""),IF($I${}=40,IF(AND(CU{}>7,DF{}=20),1,""),IF($I${}=45,IF(AND(CU{}>8,DF{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DO{}'.format(row)] = '=IF($J${}=25,IF(AND(CV{}>4,DG{}=20),1,""),IF($J${}=30,IF(AND(CV{}>5,DG{}=20),1,""),IF($J${}=35,IF(AND(CV{}>6,DG{}=20),1,""),IF($J${}=40,IF(AND(CV{}>7,DG{}=20),1,""),IF($J${}=45,IF(AND(CV{}>8,DG{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DP{}'.format(row)] = '=IF($K${}=25,IF(AND(CW{}>4,DH{}=20),1,""),IF($K${}=30,IF(AND(CW{}>5,DH{}=20),1,""),IF($K${}=35,IF(AND(CW{}>6,DH{}=20),1,""),IF($K${}=40,IF(AND(CW{}>7,DH{}=20),1,""),IF($K${}=45,IF(AND(CW{}>8,DH{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # new Z Score [4]
            ws['DR1'] = 'NAMA_SISWA_5'
            ws['DS1'] = 'NOMOR_NF_5'
            ws['DT1'] = 'KELAS_5'
            ws['DU1'] = 'NAMA_SEKOLAH_5'
            ws['DV1'] = 'LOKASI_5'
            ws['DW1'] = 'MAT_5'
            ws['DX1'] = 'IND_5'
            ws['DY1'] = 'ENG_5'
            ws['DZ1'] = 'IPA_5'
            ws['EA1'] = 'IPS_5'
            ws['EB1'] = 'JML_5'
            ws['EC1'] = 'Z_MAT_5'
            ws['ED1'] = 'Z_IND_5'
            ws['EE1'] = 'Z_ENG_5'
            ws['EF1'] = 'Z_IPA_5'
            ws['EG1'] = 'Z_IPS_5'
            ws['EH1'] = 'S_MAT_5'
            ws['EI1'] = 'S_IND_5'
            ws['EJ1'] = 'S_ENG_5'
            ws['EK1'] = 'S_IPA_5'
            ws['EL1'] = 'S_IPS_5'
            ws['EM1'] = 'S_JML_5'
            ws['EN1'] = 'RANK_NAS._5'
            ws['EO1'] = 'RANK_LOK._5'
            ws['EC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ED1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EO1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['DR1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DS1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DT1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DU1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DV1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DW1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DX1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DY1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DZ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EA1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EB1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EC1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['ED1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EE1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EF1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EG1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EH1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EI1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EJ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EK1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EL1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EM1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EN1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EO1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')

            # tambahan
            ws['EP1'] = 'MAT_20_5'
            ws['EQ1'] = 'IND_20_5'
            ws['ER1'] = 'ENG_20_5'
            ws['ES1'] = 'IPA_20_5'
            ws['ET1'] = 'IPS_20_5'
            ws['EP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ER1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ES1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ET1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EP1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EQ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['ER1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['ES1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['ET1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')

            for row in range(2, q+1):
                ws['DR{}'.format(row)] = '=B{}'.format(row)
                ws['DS{}'.format(row)] = '=C{}'.format(row, row)
                ws['DT{}'.format(row)] = '=D{}'.format(row, row)
                ws['DU{}'.format(row)] = '=E{}'.format(row, row)
                ws['DV{}'.format(row)] = '=F{}'.format(row, row)
                ws['DW{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['DX{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['DY{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['DZ{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['EA{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['EB{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['EC{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DW{}="","",(DW{}-DW${})/DW${}),2),"")'.format(row, row, r, s)
                ws['ED{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DX{}="","",(DX{}-DX${})/DX${}),2),"")'.format(row, row, r, s)
                ws['EE{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DY{}="","",(DY{}-DY${})/DY${}),2),"")'.format(row, row, r, s)
                ws['EF{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DZ{}="","",(DZ{}-DZ${})/DZ${}),2),"")'.format(row, row, r, s)
                ws['EG{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EA{}="","",(EA{}-EA${})/EA${}),2),"")'.format(row, row, r, s)
                ws['EH{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EC{}="","",IF(70+30*EC{}/$EC${}<20,20,70+30*EC{}/$EC${})),2),"")'.format(row, row, r, row, r)
                ws['EI{}'.format(
                    row)] = '=IFERROR(ROUND(IF(ED{}="","",IF(70+30*ED{}/$ED${}<20,20,70+30*ED{}/$ED${})),2),"")'.format(row, row, r, row, r)
                ws['EJ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EE{}="","",IF(70+30*EE{}/$EE${}<20,20,70+30*EE{}/$EE${})),2),"")'.format(row, row, r, row, r)
                ws['EK{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EF{}="","",IF(70+30*EF{}/$EF${}<20,20,70+30*EF{}/$EF${})),2),"")'.format(row, row, r, row, r)
                ws['EL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EG{}="","",IF(70+30*EG{}/$EG${}<20,20,70+30*EG{}/$EG${})),2),"")'.format(row, row, r, row, r)

                ws['EM{}'.format(row)] = '=IF(SUM(EH{}:EL{})=0,"",SUM(EH{}:EL{}))'.format(
                    row, row, row, row)
                ws['EN{}'.format(row)] = '=IF(EM{}="","",RANK(EM{},$EM$2:$EM${}))'.format(
                    row, row, q)
                ws['EO{}'.format(
                    row)] = '=IF(EN{}="","",COUNTIFS($DV$2:$DV${},F{},$EN$2:$EN${},"<"&EN{})+1)'.format(row, q, row, q, row)
                #     TAMBAHAN
                ws['EP{}'.format(row)] = '=IF($G${}=25,IF(AND(DW{}>4,EH{}=20),1,""),IF($G${}=30,IF(AND(DW{}>5,EH{}=20),1,""),IF($G${}=35,IF(AND(DW{}>6,EH{}=20),1,""),IF($G${}=40,IF(AND(DW{}>7,EH{}=20),1,""),IF($G${}=45,IF(AND(DW{}>8,EH{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['EQ{}'.format(row)] = '=IF($H${}=25,IF(AND(DX{}>4,EI{}=20),1,""),IF($H${}=30,IF(AND(DX{}>5,EI{}=20),1,""),IF($H${}=35,IF(AND(DX{}>6,EI{}=20),1,""),IF($H${}=40,IF(AND(DX{}>7,EI{}=20),1,""),IF($H${}=45,IF(AND(DX{}>8,EI{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['ER{}'.format(row)] = '=IF($I${}=25,IF(AND(DY{}>4,EJ{}=20),1,""),IF($I${}=30,IF(AND(DY{}>5,EJ{}=20),1,""),IF($I${}=35,IF(AND(DY{}>6,EJ{}=20),1,""),IF($I${}=40,IF(AND(DY{}>7,EJ{}=20),1,""),IF($I${}=45,IF(AND(DY{}>8,EJ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['ES{}'.format(row)] = '=IF($J${}=25,IF(AND(DZ{}>4,EK{}=20),1,""),IF($J${}=30,IF(AND(DZ{}>5,EK{}=20),1,""),IF($J${}=35,IF(AND(DZ{}>6,EK{}=20),1,""),IF($J${}=40,IF(AND(DZ{}>7,EK{}=20),1,""),IF($J${}=45,IF(AND(DZ{}>8,EK{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['ET{}'.format(row)] = '=IF($K${}=25,IF(AND(EA{}>4,EL{}=20),1,""),IF($K${}=30,IF(AND(EA{}>5,EL{}=20),1,""),IF($K${}=35,IF(AND(EA{}>6,EL{}=20),1,""),IF($K${}=40,IF(AND(EA{}>7,EL{}=20),1,""),IF($K${}=45,IF(AND(EA{}>8,EL{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

                # new Z Score [5]
            ws['EV1'] = 'NAMA SISWA'
            ws['EW1'] = 'NOMOR NF'
            ws['EX1'] = 'KELAS'
            ws['EY1'] = 'NAMA SEKOLAH'
            ws['EZ1'] = 'LOKASI'
            ws['FA1'] = 'MAT'
            ws['FB1'] = 'IND'
            ws['FC1'] = 'ENG'
            ws['FD1'] = 'IPA'
            ws['FE1'] = 'IPS'
            ws['FF1'] = 'JML'
            ws['FG1'] = 'Z_MAT'
            ws['FH1'] = 'Z_IND'
            ws['FI1'] = 'Z_ENG'
            ws['FJ1'] = 'Z_IPA'
            ws['FK1'] = 'Z_IPS'
            ws['FL1'] = 'S_MAT'
            ws['FM1'] = 'S_IND'
            ws['FN1'] = 'S_ENG'
            ws['FO1'] = 'S_IPA'
            ws['FP1'] = 'S_IPS'
            ws['FQ1'] = 'S_JML'
            ws['FR1'] = 'RANK NAS.'
            ws['FS1'] = 'RANK LOK.'
            ws['FG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FS1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['EV1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EW1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EX1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EY1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EZ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FA1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FB1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FC1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FD1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FE1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FF1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FG1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FH1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FI1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FJ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FK1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FL1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FM1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FN1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FO1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FP1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FQ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FR1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FS1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')

            # tambahan
            ws['FT1'] = 'MAT_20'
            ws['FU1'] = 'IND_20'
            ws['FV1'] = 'ENG_20'
            ws['FW1'] = 'IPA_20'
            ws['FX1'] = 'IPS_20'
            ws['FT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FT1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FU1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FV1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FW1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FX1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')

            for row in range(2, q+1):
                ws['EV{}'.format(row)] = '=B{}'.format(row)
                ws['EW{}'.format(row)] = '=C{}'.format(row, row)
                ws['EX{}'.format(row)] = '=D{}'.format(row, row)
                ws['EY{}'.format(row)] = '=E{}'.format(row, row)
                ws['EZ{}'.format(row)] = '=F{}'.format(row, row)
                ws['FA{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['FB{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['FC{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['FD{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['FE{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['FF{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['FG{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FA{}="","",(FA{}-FA${})/FA${}),2),"")'.format(row, row, r, s)
                ws['FH{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FB{}="","",(FB{}-FB${})/FB${}),2),"")'.format(row, row, r, s)
                ws['FI{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FC{}="","",(FC{}-FC${})/FC${}),2),"")'.format(row, row, r, s)
                ws['FJ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FD{}="","",(FD{}-FD${})/FD${}),2),"")'.format(row, row, r, s)
                ws['FK{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FE{}="","",(FE{}-FE${})/FE${}),2),"")'.format(row, row, r, s)
                ws['FL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FG{}="","",IF(70+30*FG{}/$FG${}<20,20,70+30*FG{}/$FG${})),2),"")'.format(row, row, r, row, r)
                ws['FM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FH{}="","",IF(70+30*FH{}/$FH${}<20,20,70+30*FH{}/$FH${})),2),"")'.format(row, row, r, row, r)
                ws['FN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FI{}="","",IF(70+30*FI{}/$FI${}<20,20,70+30*FI{}/$FI${})),2),"")'.format(row, row, r, row, r)
                ws['FO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FJ{}="","",IF(70+30*FJ{}/$FJ${}<20,20,70+30*FJ{}/$FJ${})),2),"")'.format(row, row, r, row, r)
                ws['FP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FK{}="","",IF(70+30*FK{}/$FK${}<20,20,70+30*FK{}/$FK${})),2),"")'.format(row, row, r, row, r)

                ws['FQ{}'.format(row)] = '=IF(SUM(FL{}:FP{})=0,"",SUM(FL{}:FP{}))'.format(
                    row, row, row, row)
                ws['FR{}'.format(row)] = '=IF(FQ{}="","",RANK(FQ{},$FQ$2:$FQ${}))'.format(
                    row, row, q)
                ws['FS{}'.format(
                    row)] = '=IF(FR{}="","",COUNTIFS($EZ$2:$EZ${},F{},$FR$2:$FR${},"<"&FR{})+1)'.format(row, q, row, q, row)
                #     TAMBAHAN
                ws['FT{}'.format(row)] = '=IF($G${}=25,IF(AND(FA{}>4,FL{}=20),1,""),IF($G${}=30,IF(AND(FA{}>5,FL{}=20),1,""),IF($G${}=35,IF(AND(FA{}>6,FL{}=20),1,""),IF($G${}=40,IF(AND(FA{}>7,FL{}=20),1,""),IF($G${}=45,IF(AND(FA{}>8,FL{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['FU{}'.format(row)] = '=IF($H${}=25,IF(AND(FB{}>4,FM{}=20),1,""),IF($H${}=30,IF(AND(FB{}>5,FM{}=20),1,""),IF($H${}=35,IF(AND(FB{}>6,FM{}=20),1,""),IF($H${}=40,IF(AND(FB{}>7,FM{}=20),1,""),IF($H${}=45,IF(AND(FB{}>8,FM{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['FV{}'.format(row)] = '=IF($I${}=25,IF(AND(FC{}>4,FN{}=20),1,""),IF($I${}=30,IF(AND(FC{}>5,FN{}=20),1,""),IF($I${}=35,IF(AND(FC{}>6,FN{}=20),1,""),IF($I${}=40,IF(AND(FC{}>7,FN{}=20),1,""),IF($I${}=45,IF(AND(FC{}>8,FN{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['FW{}'.format(row)] = '=IF($J${}=25,IF(AND(FD{}>4,FO{}=20),1,""),IF($J${}=30,IF(AND(FD{}>5,FO{}=20),1,""),IF($J${}=35,IF(AND(FD{}>6,FO{}=20),1,""),IF($J${}=40,IF(AND(FD{}>7,FO{}=20),1,""),IF($J${}=45,IF(AND(FD{}>8,FO{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['FX{}'.format(row)] = '=IF($K${}=25,IF(AND(FE{}>4,FP{}=20),1,""),IF($K${}=30,IF(AND(FE{}>5,FP{}=20),1,""),IF($K${}=35,IF(AND(FE{}>6,FP{}=20),1,""),IF($K${}=40,IF(AND(FE{}>7,FP{}=20),1,""),IF($K${}=45,IF(AND(FE{}>8,FP{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Mengubah 'KELAS' sesuai dengan nilai yang dipilih dari selectbox 'KELAS'
            kelas = KELAS.lower().replace(" ", "")
            semester = SEMESTER.lower()
            tahun = TAHUN.replace("-", "")
            penilaian = PENILAIAN.lower()
            kurikulum = KURIKULUM.lower()

            path_file = f"{kelas}_{penilaian}_{semester}_{kurikulum}_{tahun}_nilai_std.xlsx"

            # Simpan file ke direktori temporer
            temp_dir = tempfile.gettempdir()
            file_path = temp_dir + '/' + path_file
            wb.save(file_path)

            st.success(
                "File siap diunduh!")

            # Tombol unduh file
            with open(file_path, "rb") as f:
                bytes_data = f.read()
            st.download_button(label="Unduh File", data=bytes_data,
                               file_name=path_file)

            st.warning(
                "Buka file unduhan, klik 'Enable Editing' dan 'Save'")
    if selected_file == "Nilai Std. All IPA":
        # menghilangkan hamburger
        st.markdown("""
        <style>
        .css-1rs6os.edgvbvh3
        {
            visibility:hidden;
        }
        .css-1lsmgbg.egzxvld0
        {
            visibility:hidden;
        }
        </style>
        """, unsafe_allow_html=True)

        image = Image.open('logo resmi nf resize.png')
        st.image(image)

        st.title("Olah Nilai Standar K13")
        st.header("SMA, PPLS, RONIN IPA")

        col6 = st.container()

        with col6:
            KELAS = st.selectbox(
                "KELAS",
                ("--Pilih Kelas--", "10 SMA IPA", "11 SMA IPA", "PPLS IPA", "RONIN IPA"))

        col7 = st.container()

        with col7:
            SEMESTER = st.selectbox(
                "SEMESTER",
                ("--Pilih Semester--", "SEMESTER 1", "SEMESTER 2"))

        col8 = st.container()

        with col8:
            PENILAIAN = st.selectbox(
                "PENILAIAN",
                ("--Pilih Penilaian--", "PENILAIAN TENGAH SEMESTER", "PENILAIAN AKHIR SEMESTER", "PENILAIAN AKHIR TAHUN", "TES EVALUASI"))

        col9 = st.container()

        with col9:
            KURIKULUM = st.selectbox(
                "KURIKULUM",
                ("--Pilih Kurikulum--", "K13"))

        TAHUN = st.text_input("Masukkan Tahun Ajaran",
                              placeholder="contoh: 2022-2023")

        col1, col2, col3, col4 = st.columns(4)

        with col1:
            MTK = st.selectbox(
                "JML. SOAL MAT.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col2:
            FIS = st.selectbox(
                "JML. SOAL FIS.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col3:
            KIM = st.selectbox(
                "JML. SOAL KIM.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col4:
            BIO = st.selectbox(
                "JML. SOAL BIO.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        JML_SOAL_MAT = MTK
        JML_SOAL_BIO = BIO
        JML_SOAL_FIS = FIS
        JML_SOAL_KIM = KIM

        uploaded_file = st.file_uploader(
            'Letakkan file excel IPA', type='xlsx')

        if uploaded_file is not None:
            wb = openpyxl.load_workbook(uploaded_file)
            ws = wb['Sheet1']

            q = len(ws['K'])
            r = len(ws['K'])+2
            s = len(ws['K'])+3
            t = len(ws['K'])+4
            u = len(ws['K'])+5
            v = len(ws['K'])+6
            w = len(ws['K'])+7
            x = len(ws['K'])+8

            ws['G{}'.format(r)] = "=ROUND(AVERAGE(G2:G{}),2)".format(q)
            ws['H{}'.format(r)] = "=ROUND(AVERAGE(H2:H{}),2)".format(q)
            ws['I{}'.format(r)] = "=ROUND(AVERAGE(I2:I{}),2)".format(q)
            ws['J{}'.format(r)] = "=ROUND(AVERAGE(J2:J{}),2)".format(q)
            ws['K{}'.format(r)] = "=ROUND(AVERAGE(K2:K{}),2)".format(q)
            ws['G{}'.format(s)] = "=STDEV(G2:G{})".format(q)
            ws['H{}'.format(s)] = "=STDEV(H2:H{})".format(q)
            ws['I{}'.format(s)] = "=STDEV(I2:I{})".format(q)
            ws['J{}'.format(s)] = "=STDEV(J2:J{})".format(q)
            ws['G{}'.format(t)] = "=MAX(G2:G{})".format(q)
            ws['H{}'.format(t)] = "=MAX(H2:H{})".format(q)
            ws['I{}'.format(t)] = "=MAX(I2:I{})".format(q)
            ws['J{}'.format(t)] = "=MAX(J2:J{})".format(q)
            ws['K{}'.format(t)] = "=MAX(K2:K{})".format(q)
            ws['L{}'.format(r)] = "=MAX(L2:L{})".format(q)
            ws['M{}'.format(r)] = "=MAX(M2:M{})".format(q)
            ws['N{}'.format(r)] = "=MAX(N2:N{})".format(q)
            ws['O{}'.format(r)] = "=MAX(O2:O{})".format(q)
            ws['P{}'.format(r)] = "=MAX(P2:P{})".format(q)
            ws['Q{}'.format(r)] = "=MAX(Q2:Q{})".format(q)
            ws['R{}'.format(r)] = "=MAX(R2:R{})".format(q)
            ws['S{}'.format(r)] = "=MAX(S2:S{})".format(q)
            ws['T{}'.format(r)] = "=ROUND(MAX(T2:T{}),2)".format(q)
            ws['U{}'.format(r)] = "=MAX(U2:U{})".format(q)
            ws['G{}'.format(u)] = "=MIN(G2:G{})".format(q)
            ws['H{}'.format(u)] = "=MIN(H2:H{})".format(q)
            ws['I{}'.format(u)] = "=MIN(I2:I{})".format(q)
            ws['J{}'.format(u)] = "=MIN(J2:J{})".format(q)
            ws['K{}'.format(u)] = "=MIN(K2:K{})".format(q)
            ws['P{}'.format(s)] = "=MIN(P2:P{})".format(q)
            ws['Q{}'.format(s)] = "=MIN(Q2:R{})".format(q)
            ws['R{}'.format(s)] = "=MIN(R2:S{})".format(q)
            ws['S{}'.format(s)] = "=MIN(S2:T{})".format(q)
            ws['T{}'.format(s)] = "=MIN(T2:T{})".format(q)
            ws['P{}'.format(t)] = "=ROUND(AVERAGE(P2:P{}),2)".format(q)
            ws['Q{}'.format(t)] = "=ROUND(AVERAGE(Q2:Q{}),2)".format(q)
            ws['R{}'.format(t)] = "=ROUND(AVERAGE(R2:R{}),2)".format(q)
            ws['S{}'.format(t)] = "=ROUND(AVERAGE(S2:S{}),2)".format(q)
            ws['T{}'.format(t)] = "=ROUND(AVERAGE(T2:T{}),2)".format(q)
            ws['W{}'.format(r)] = "=SUM(W2:W{})".format(q)
            ws['X{}'.format(r)] = "=SUM(X2:X{})".format(q)
            ws['Y{}'.format(r)] = "=SUM(Y2:Y{})".format(q)
            ws['Z{}'.format(r)] = "=SUM(Z2:Z{})".format(q)

            # new
            # iterasi 1 rata-rata - 1

            # MAPEL NORMAL
            ws['AG{}'.format(r)] = "=IF($W${}=0,$G${},$G${}-1)".format(r, r, r)
            ws['AG{}'.format(s)] = "=STDEV(AG2:AG{})".format(q)
            ws['AG{}'.format(t)] = "=MAX(AG2:AG{})".format(q)
            ws['AG{}'.format(u)] = "=MIN(AG2:AG{})".format(q)
            ws['AH{}'.format(r)] = "=IF($X${}=0,$H${},$H${}-1)".format(r, r, r)
            ws['AH{}'.format(s)] = "=STDEV(AH2:AH{})".format(q)
            ws['AH{}'.format(t)] = "=MAX(AH2:AH{})".format(q)
            ws['AH{}'.format(u)] = "=MIN(AH2:AH{})".format(q)
            ws['AI{}'.format(r)] = "=IF($Y${}=0,$I${},$I${}-1)".format(r, r, r)
            ws['AI{}'.format(s)] = "=STDEV(AI2:AI{})".format(q)
            ws['AI{}'.format(t)] = "=MAX(AI2:AI{})".format(q)
            ws['AI{}'.format(u)] = "=MIN(AI2:AI{})".format(q)
            ws['AJ{}'.format(r)] = "=IF($Z${}=0,$J${},$J${}-1)".format(r, r, r)
            ws['AJ{}'.format(s)] = "=STDEV(AJ2:AJ{})".format(q)
            ws['AJ{}'.format(t)] = "=MAX(AJ2:AJ{})".format(q)
            ws['AJ{}'.format(u)] = "=MIN(AJ2:AJ{})".format(q)

            # JUMLAH MAPEL NORMAL
            ws['AK{}'.format(r)] = "=ROUND(AVERAGE(AK2:AK{}),2)".format(q)
            ws['AK{}'.format(t)] = "=MAX(AK2:AK{})".format(q)
            ws['AK{}'.format(u)] = "=MIN(AK2:AK{})".format(q)

            # Z SCORE
            ws['AL{}'.format(r)] = "=MAX(AL2:AL{})".format(q)
            ws['AM{}'.format(r)] = "=MAX(AM2:AM{})".format(q)
            ws['AN{}'.format(r)] = "=MAX(AN2:AN{})".format(q)
            ws['AO{}'.format(r)] = "=MAX(AO2:AO{})".format(q)

            # NILAI STANDAR
            ws['AP{}'.format(r)] = "=MAX(AP2:AP{})".format(q)
            ws['AP{}'.format(s)] = "=MIN(AP2:AP{})".format(q)
            ws['AP{}'.format(t)] = "=ROUND(AVERAGE(AP2:AP{}),2)".format(q)
            ws['AQ{}'.format(r)] = "=MAX(AQ2:AQ{})".format(q)
            ws['AQ{}'.format(s)] = "=MIN(AQ2:AQ{})".format(q)
            ws['AQ{}'.format(t)] = "=ROUND(AVERAGE(AQ2:AQ{}),2)".format(q)
            ws['AR{}'.format(r)] = "=MAX(AR2:AR{})".format(q)
            ws['AR{}'.format(s)] = "=MIN(AR2:AR{})".format(q)
            ws['AR{}'.format(t)] = "=ROUND(AVERAGE(AR2:AR{}),2)".format(q)
            ws['AS{}'.format(r)] = "=MAX(AS2:AS{})".format(q)
            ws['AS{}'.format(s)] = "=MIN(AS2:AS{})".format(q)
            ws['AS{}'.format(t)] = "=ROUND(AVERAGE(AS2:AS{}),2)".format(q)
            ws['AT{}'.format(r)] = "=MAX(AT2:AT{})".format(q)
            ws['AT{}'.format(s)] = "=MIN(AT2:AT{})".format(q)
            ws['AT{}'.format(t)] = "=ROUND(AVERAGE(AT2:AT{}),2)".format(q)

            # INISIASI MAPEL
            ws['AW{}'.format(r)] = "=SUM(AW2:AW{})".format(q)
            ws['AX{}'.format(r)] = "=SUM(AX2:AX{})".format(q)
            ws['AY{}'.format(r)] = "=SUM(AY2:AY{})".format(q)
            ws['AZ{}'.format(r)] = "=SUM(AZ2:AZ{})".format(q)

            # iterasi 2 rata-rata - 1
            # MAPEL NORMAL
            ws['BG{}'.format(
                r)] = "=IF($AW${}=0,$AG${},$AG${}-1)".format(r, r, r)
            ws['BG{}'.format(s)] = "=STDEV(BG2:BG{})".format(q)
            ws['BG{}'.format(t)] = "=MAX(BG2:BG{})".format(q)
            ws['BG{}'.format(u)] = "=MIN(BG2:BG{})".format(q)
            ws['BH{}'.format(
                r)] = "=IF($AX${}=0,$AH${},$AH${}-1)".format(r, r, r)
            ws['BH{}'.format(s)] = "=STDEV(BH2:BH{})".format(q)
            ws['BH{}'.format(t)] = "=MAX(BH2:BH{})".format(q)
            ws['BH{}'.format(u)] = "=MIN(BH2:BH{})".format(q)
            ws['BI{}'.format(
                r)] = "=IF($AY${}=0,$AI${},$AI${}-1)".format(r, r, r)
            ws['BI{}'.format(s)] = "=STDEV(BI2:BI{})".format(q)
            ws['BI{}'.format(t)] = "=MAX(BI2:BI{})".format(q)
            ws['BI{}'.format(u)] = "=MIN(BI2:BI{})".format(q)
            ws['BJ{}'.format(
                r)] = "=IF($AZ${}=0,$AJ${},$AJ${}-1)".format(r, r, r)
            ws['BJ{}'.format(s)] = "=STDEV(BJ2:BJ{})".format(q)
            ws['BJ{}'.format(t)] = "=MAX(BJ2:BJ{})".format(q)
            ws['BJ{}'.format(u)] = "=MIN(BJ2:BJ{})".format(q)

            # JUMLAH MAPEL NORMAL
            ws['BK{}'.format(r)] = "=ROUND(AVERAGE(BK2:BK{}),2)".format(q)
            ws['BK{}'.format(t)] = "=MAX(BK2:BK{})".format(q)
            ws['BK{}'.format(u)] = "=MIN(BK2:BK{})".format(q)

            # Z SCORE
            ws['BL{}'.format(r)] = "=MAX(BL2:BL{})".format(q)
            ws['BM{}'.format(r)] = "=MAX(BM2:BM{})".format(q)
            ws['BN{}'.format(r)] = "=MAX(BN2:BN{})".format(q)
            ws['BO{}'.format(r)] = "=MAX(BO2:BO{})".format(q)

            # NILAI STANDAR
            ws['BP{}'.format(r)] = "=MAX(BP2:BP{})".format(q)
            ws['BP{}'.format(s)] = "=MIN(BP2:BP{})".format(q)
            ws['BP{}'.format(t)] = "=ROUND(AVERAGE(BP2:BP{}),2)".format(q)
            ws['BQ{}'.format(r)] = "=MAX(BQ2:BQ{})".format(q)
            ws['BQ{}'.format(s)] = "=MIN(BQ2:BQ{})".format(q)
            ws['BQ{}'.format(t)] = "=ROUND(AVERAGE(BQ2:BQ{}),2)".format(q)
            ws['BR{}'.format(r)] = "=MAX(BR2:BR{})".format(q)
            ws['BR{}'.format(s)] = "=MIN(BR2:BR{})".format(q)
            ws['BR{}'.format(t)] = "=ROUND(AVERAGE(BR2:BR{}),2)".format(q)
            ws['BS{}'.format(r)] = "=MAX(BS2:BS{})".format(q)
            ws['BS{}'.format(s)] = "=MIN(BS2:BS{})".format(q)
            ws['BS{}'.format(t)] = "=ROUND(AVERAGE(BS2:BS{}),2)".format(q)
            ws['BT{}'.format(r)] = "=MAX(BT2:BT{})".format(q)
            ws['BT{}'.format(s)] = "=MIN(BT2:BT{})".format(q)
            ws['BT{}'.format(t)] = "=ROUND(AVERAGE(BT2:BT{}),2)".format(q)

            # INISIASI MAPEL
            ws['BW{}'.format(r)] = "=SUM(BW2:BW{})".format(q)
            ws['BX{}'.format(r)] = "=SUM(BX2:BX{})".format(q)
            ws['BY{}'.format(r)] = "=SUM(BY2:BY{})".format(q)
            ws['BZ{}'.format(r)] = "=SUM(BZ2:BZ{})".format(q)

            # iterasi 3 rata-rata - 1
            # MAPEL NORMAL
            ws['CG{}'.format(
                r)] = "=IF($BW${}=0,$BG${},$BG${}-1)".format(r, r, r)
            ws['CG{}'.format(s)] = "=STDEV(CG2:CG{})".format(q)
            ws['CG{}'.format(t)] = "=MAX(CG2:CG{})".format(q)
            ws['CG{}'.format(u)] = "=MIN(CG2:CG{})".format(q)
            ws['CH{}'.format(
                r)] = "=IF($BX${}=0,$BH${},$BH${}-1)".format(r, r, r)
            ws['CH{}'.format(s)] = "=STDEV(CH2:CH{})".format(q)
            ws['CH{}'.format(t)] = "=MAX(CH2:CH{})".format(q)
            ws['CH{}'.format(u)] = "=MIN(CH2:CH{})".format(q)
            ws['CI{}'.format(
                r)] = "=IF($BY${}=0,$BI${},$BI${}-1)".format(r, r, r)
            ws['CI{}'.format(s)] = "=STDEV(CI2:CI{})".format(q)
            ws['CI{}'.format(t)] = "=MAX(CI2:CI{})".format(q)
            ws['CI{}'.format(u)] = "=MIN(CI2:CI{})".format(q)
            ws['CJ{}'.format(
                r)] = "=IF($BZ${}=0,$BJ${},$BJ${}-1)".format(r, r, r)
            ws['CJ{}'.format(s)] = "=STDEV(CJ2:CJ{})".format(q)
            ws['CJ{}'.format(t)] = "=MAX(CJ2:CJ{})".format(q)
            ws['CJ{}'.format(u)] = "=MIN(CJ2:CJ{})".format(q)

            # JUMLAH MAPEL NORMAL
            ws['CK{}'.format(r)] = "=ROUND(AVERAGE(CK2:CK{}),2)".format(q)
            ws['CK{}'.format(t)] = "=MAX(CK2:CK{})".format(q)
            ws['CK{}'.format(u)] = "=MIN(CK2:CK{})".format(q)

            # Z SCORE
            ws['CL{}'.format(r)] = "=MAX(CL2:CL{})".format(q)
            ws['CM{}'.format(r)] = "=MAX(CM2:CM{})".format(q)
            ws['CN{}'.format(r)] = "=MAX(CN2:CN{})".format(q)
            ws['CO{}'.format(r)] = "=MAX(CO2:CO{})".format(q)

            # NILAI STANDAR
            ws['CP{}'.format(r)] = "=MAX(CP2:CP{})".format(q)
            ws['CP{}'.format(s)] = "=MIN(CP2:CP{})".format(q)
            ws['CP{}'.format(t)] = "=ROUND(AVERAGE(CP2:CP{}),2)".format(q)
            ws['CQ{}'.format(r)] = "=MAX(CQ2:CQ{})".format(q)
            ws['CQ{}'.format(s)] = "=MIN(CQ2:CQ{})".format(q)
            ws['CQ{}'.format(t)] = "=ROUND(AVERAGE(CQ2:CQ{}),2)".format(q)
            ws['CR{}'.format(r)] = "=MAX(CR2:CR{})".format(q)
            ws['CR{}'.format(s)] = "=MIN(CR2:CR{})".format(q)
            ws['CR{}'.format(t)] = "=ROUND(AVERAGE(CR2:CR{}),2)".format(q)
            ws['CS{}'.format(r)] = "=MAX(CS2:CS{})".format(q)
            ws['CS{}'.format(s)] = "=MIN(CS2:CS{})".format(q)
            ws['CS{}'.format(t)] = "=ROUND(AVERAGE(CS2:CS{}),2)".format(q)
            ws['CT{}'.format(r)] = "=MAX(CT2:CT{})".format(q)
            ws['CT{}'.format(s)] = "=MIN(CT2:CT{})".format(q)
            ws['CT{}'.format(t)] = "=ROUND(AVERAGE(CT2:CT{}),2)".format(q)

            # INISIASI MAPEL
            ws['CW{}'.format(r)] = "=SUM(CW2:CW{})".format(q)
            ws['CX{}'.format(r)] = "=SUM(CX2:CX{})".format(q)
            ws['CY{}'.format(r)] = "=SUM(CY2:CY{})".format(q)
            ws['CZ{}'.format(r)] = "=SUM(CZ2:CZ{})".format(q)

            # iterasi 4 rata-rata - 1
            # MAPEL NORMAL
            ws['DG{}'.format(
                r)] = "=IF($CW${}=0,$CG${},$CG${}-1)".format(r, r, r)
            ws['DG{}'.format(s)] = "=STDEV(DG2:DG{})".format(q)
            ws['DG{}'.format(t)] = "=MAX(DG2:DG{})".format(q)
            ws['DG{}'.format(u)] = "=MIN(DG2:DG{})".format(q)
            ws['DH{}'.format(
                r)] = "=IF($CX${}=0,$CH${},$CH${}-1)".format(r, r, r)
            ws['DH{}'.format(s)] = "=STDEV(DH2:DH{})".format(q)
            ws['DH{}'.format(t)] = "=MAX(DH2:DH{})".format(q)
            ws['DH{}'.format(u)] = "=MIN(DH2:DH{})".format(q)
            ws['DI{}'.format(
                r)] = "=IF($CY${}=0,$CI${},$CI${}-1)".format(r, r, r)
            ws['DI{}'.format(s)] = "=STDEV(DI2:DI{})".format(q)
            ws['DI{}'.format(t)] = "=MAX(DI2:DI{})".format(q)
            ws['DI{}'.format(u)] = "=MIN(DI2:DI{})".format(q)
            ws['DJ{}'.format(
                r)] = "=IF($CZ${}=0,$CJ${},$CJ${}-1)".format(r, r, r)
            ws['DJ{}'.format(s)] = "=STDEV(DJ2:DJ{})".format(q)
            ws['DJ{}'.format(t)] = "=MAX(DJ2:DJ{})".format(q)
            ws['DJ{}'.format(u)] = "=MIN(DJ2:DJ{})".format(q)

            # JUMLAH MAPEL NORMAL
            ws['DK{}'.format(r)] = "=ROUND(AVERAGE(DK2:DK{}),2)".format(q)
            ws['DK{}'.format(t)] = "=MAX(DK2:DK{})".format(q)
            ws['DK{}'.format(u)] = "=MIN(DK2:DK{})".format(q)

            # Z SCORE
            ws['DL{}'.format(r)] = "=MAX(DL2:DL{})".format(q)
            ws['DM{}'.format(r)] = "=MAX(DM2:DM{})".format(q)
            ws['DN{}'.format(r)] = "=MAX(DN2:DN{})".format(q)
            ws['DO{}'.format(r)] = "=MAX(DO2:DO{})".format(q)

            # NILAI STANDAR
            ws['DP{}'.format(r)] = "=MAX(DP2:DP{})".format(q)
            ws['DP{}'.format(s)] = "=MIN(DP2:DP{})".format(q)
            ws['DP{}'.format(t)] = "=ROUND(AVERAGE(DP2:DP{}),2)".format(q)
            ws['DQ{}'.format(r)] = "=MAX(DQ2:DQ{})".format(q)
            ws['DQ{}'.format(s)] = "=MIN(DQ2:DQ{})".format(q)
            ws['DQ{}'.format(t)] = "=ROUND(AVERAGE(DQ2:DQ{}),2)".format(q)
            ws['DR{}'.format(r)] = "=MAX(DR2:DR{})".format(q)
            ws['DR{}'.format(s)] = "=MIN(DR2:DR{})".format(q)
            ws['DR{}'.format(t)] = "=ROUND(AVERAGE(DR2:DR{}),2)".format(q)
            ws['DS{}'.format(r)] = "=MAX(DS2:DS{})".format(q)
            ws['DS{}'.format(s)] = "=MIN(DS2:DS{})".format(q)
            ws['DS{}'.format(t)] = "=ROUND(AVERAGE(DS2:DS{}),2)".format(q)
            ws['DT{}'.format(r)] = "=MAX(DT2:DT{})".format(q)
            ws['DT{}'.format(s)] = "=MIN(DT2:DT{})".format(q)
            ws['DT{}'.format(t)] = "=ROUND(AVERAGE(DT2:DT{}),2)".format(q)

            # INISIASI MAPEL
            ws['DW{}'.format(r)] = "=SUM(DW2:DW{})".format(q)
            ws['DX{}'.format(r)] = "=SUM(DX2:DX{})".format(q)
            ws['DY{}'.format(r)] = "=SUM(DY2:DY{})".format(q)
            ws['DZ{}'.format(r)] = "=SUM(DZ2:DZ{})".format(q)

            # iterasi 5 rata-rata - 1
            # MAPEL NORMAL
            ws['EG{}'.format(
                r)] = "=IF($DW${}=0,$DG${},$DG${}-1)".format(r, r, r)
            ws['EG{}'.format(s)] = "=STDEV(EG2:EG{})".format(q)
            ws['EG{}'.format(t)] = "=MAX(EG2:EG{})".format(q)
            ws['EG{}'.format(u)] = "=MIN(EG2:EG{})".format(q)
            ws['EH{}'.format(
                r)] = "=IF($DX${}=0,$DH${},$DH${}-1)".format(r, r, r)
            ws['EH{}'.format(s)] = "=STDEV(EH2:EH{})".format(q)
            ws['EH{}'.format(t)] = "=MAX(EH2:EH{})".format(q)
            ws['EH{}'.format(u)] = "=MIN(EH2:EH{})".format(q)
            ws['EI{}'.format(
                r)] = "=IF($DY${}=0,$DI${},$DI${}-1)".format(r, r, r)
            ws['EI{}'.format(s)] = "=STDEV(EI2:EI{})".format(q)
            ws['EI{}'.format(t)] = "=MAX(EI2:EI{})".format(q)
            ws['EI{}'.format(u)] = "=MIN(EI2:EI{})".format(q)
            ws['EJ{}'.format(
                r)] = "=IF($DZ${}=0,$DJ${},$DJ${}-1)".format(r, r, r)
            ws['EJ{}'.format(s)] = "=STDEV(EJ2:EJ{})".format(q)
            ws['EJ{}'.format(t)] = "=MAX(EJ2:EJ{})".format(q)
            ws['EJ{}'.format(u)] = "=MIN(EJ2:EJ{})".format(q)

            # JUMLAH MAPEL NORMAL
            ws['EK{}'.format(r)] = "=ROUND(AVERAGE(EK2:EK{}),2)".format(q)
            ws['EK{}'.format(t)] = "=MAX(EK2:EK{})".format(q)
            ws['EK{}'.format(u)] = "=MIN(EK2:EK{})".format(q)

            # Z SCORE
            ws['EL{}'.format(r)] = "=MAX(EL2:EL{})".format(q)
            ws['EM{}'.format(r)] = "=MAX(EM2:EM{})".format(q)
            ws['EN{}'.format(r)] = "=MAX(EN2:EN{})".format(q)
            ws['EO{}'.format(r)] = "=MAX(EO2:EO{})".format(q)

            # NILAI STANDAR
            ws['EP{}'.format(r)] = "=MAX(EP2:EP{})".format(q)
            ws['EP{}'.format(s)] = "=MIN(EP2:EP{})".format(q)
            ws['EP{}'.format(t)] = "=ROUND(AVERAGE(EP2:EP{}),2)".format(q)
            ws['EQ{}'.format(r)] = "=MAX(EQ2:EQ{})".format(q)
            ws['EQ{}'.format(s)] = "=MIN(EQ2:EQ{})".format(q)
            ws['EQ{}'.format(t)] = "=ROUND(AVERAGE(EQ2:EQ{}),2)".format(q)
            ws['ER{}'.format(r)] = "=MAX(ER2:ER{})".format(q)
            ws['ER{}'.format(s)] = "=MIN(ER2:ER{})".format(q)
            ws['ER{}'.format(t)] = "=ROUND(AVERAGE(ER2:ER{}),2)".format(q)
            ws['ES{}'.format(r)] = "=MAX(ES2:ES{})".format(q)
            ws['ES{}'.format(s)] = "=MIN(ES2:ES{})".format(q)
            ws['ES{}'.format(t)] = "=ROUND(AVERAGE(ES2:ES{}),2)".format(q)
            ws['ET{}'.format(r)] = "=MAX(ET2:ET{})".format(q)
            ws['ET{}'.format(s)] = "=MIN(ET2:ET{})".format(q)
            ws['ET{}'.format(t)] = "=ROUND(AVERAGE(ET2:ET{}),2)".format(q)

            # INISIASI MAPEL
            ws['EW{}'.format(r)] = "=SUM(EW2:EW{})".format(q)
            ws['EX{}'.format(r)] = "=SUM(EX2:EX{})".format(q)
            ws['EY{}'.format(r)] = "=SUM(EY2:EY{})".format(q)
            ws['EZ{}'.format(r)] = "=SUM(EZ2:EZ{})".format(q)

            # Jumlah Soal
            ws['F{}'.format(v)] = 'JUMLAH SOAL'
            ws['G{}'.format(v)] = JML_SOAL_MAT
            ws['H{}'.format(v)] = JML_SOAL_FIS
            ws['I{}'.format(v)] = JML_SOAL_KIM
            ws['J{}'.format(v)] = JML_SOAL_BIO

            # Z Score
            ws['B1'] = 'NAMA SISWA_A'
            ws['C1'] = 'NOMOR NF_A'
            ws['D1'] = 'KELAS_A'
            ws['E1'] = 'NAMA SEKOLAH_A'
            ws['F1'] = 'LOKASI_A'
            ws['G1'] = 'MAT_A'
            ws['H1'] = 'FIS_A'
            ws['I1'] = 'KIM_A'
            ws['J1'] = 'BIO_A'
            ws['K1'] = 'JML_A'
            ws['L1'] = 'Z_MAT_A'
            ws['M1'] = 'Z_FIS_A'
            ws['N1'] = 'Z_KIM_A'
            ws['O1'] = 'Z_BIO_A'
            ws['P1'] = 'S_MAT_A'
            ws['Q1'] = 'S_FIS_A'
            ws['R1'] = 'S_KIM_A'
            ws['S1'] = 'S_BIO_A'
            ws['T1'] = 'S_JML_A'
            ws['U1'] = 'RANK NAS._A'
            ws['V1'] = 'RANK LOK._A'

            ws['L1'].font = Font(bold=False, name='Calibri', size=11)
            ws['M1'].font = Font(bold=False, name='Calibri', size=11)
            ws['N1'].font = Font(bold=False, name='Calibri', size=11)
            ws['O1'].font = Font(bold=False, name='Calibri', size=11)
            ws['P1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Q1'].font = Font(bold=False, name='Calibri', size=11)
            ws['R1'].font = Font(bold=False, name='Calibri', size=11)
            ws['S1'].font = Font(bold=False, name='Calibri', size=11)
            ws['T1'].font = Font(bold=False, name='Calibri', size=11)
            ws['U1'].font = Font(bold=False, name='Calibri', size=11)
            ws['V1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['B1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['C1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['D1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['E1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['F1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['G1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['H1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['I1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['J1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['K1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['L1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['M1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['N1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['O1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['P1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Q1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['R1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['S1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['T1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['U1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['V1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')

            # tambahan
            ws['W1'] = 'MAT_20_A'
            ws['X1'] = 'FIS_20_A'
            ws['Y1'] = 'KIM_20_A'
            ws['Z1'] = 'BIO_20_A'
            ws['W1'].font = Font(bold=False, name='Calibri', size=11)
            ws['X1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Y1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Z1'].font = Font(bold=False, name='Calibri', size=11)
            ws['W1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['X1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Y1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Z1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')

            for row in range(2, q+1):
                ws['K{}'.format(
                    row)] = '=SUM(G{}:J{})'.format(row, row, row)
                ws['L{}'.format(
                    row)] = '=IFERROR(ROUND(IF(G{}="","",(G{}-G${})/G${}),2),"")'.format(row, row, r, s)
                ws['M{}'.format(
                    row)] = '=IFERROR(ROUND(IF(H{}="","",(H{}-H${})/H${}),2),"")'.format(row, row, r, s)
                ws['N{}'.format(
                    row)] = '=IFERROR(ROUND(IF(I{}="","",(I{}-I${})/I${}),2),"")'.format(row, row, r, s)
                ws['O{}'.format(
                    row)] = '=IFERROR(ROUND(IF(J{}="","",(J{}-J${})/J${}),2),"")'.format(row, row, r, s)
                ws['P{}'.format(
                    row)] = '=IFERROR(ROUND(IF(G{}="","",IF(70+30*L{}/$L${}<20,20,70+30*L{}/$L${})),2),"")'.format(row, row, r, row, r)
                ws['Q{}'.format(
                    row)] = '=IFERROR(ROUND(IF(H{}="","",IF(70+30*M{}/$M${}<20,20,70+30*M{}/$M${})),2),"")'.format(row, row, r, row, r)
                ws['R{}'.format(
                    row)] = '=IFERROR(ROUND(IF(I{}="","",IF(70+30*N{}/$N${}<20,20,70+30*N{}/$N${})),2),"")'.format(row, row, r, row, r)
                ws['S{}'.format(
                    row)] = '=IFERROR(ROUND(IF(J{}="","",IF(70+30*O{}/$O${}<20,20,70+30*O{}/$P${})),2),"")'.format(row, row, r, row, r)

                ws['T{}'.format(row)] = '=IF(SUM(P{}:S{})=0,"",SUM(P{}:S{}))'.format(
                    row, row, row, row)
                ws['U{}'.format(row)] = '=IF(T{}="","",RANK(T{},$T$2:$T${}))'.format(
                    row, row, q)
                ws['V{}'.format(
                    row)] = '=IF(U{}="","",COUNTIFS($F$2:$F${},F{},$U$2:$U${},"<"&U{})+1)'.format(row, q, row, q, row)
                # TAMBAHAN
                ws['W{}'.format(row)] = '=IF($G${}=25,IF(AND(G{}>4,P{}=20),1,""),IF($G${}=30,IF(AND(G{}>5,P{}=20),1,""),IF($G${}=35,IF(AND(G{}>6,P{}=20),1,""),IF($G${}=40,IF(AND(G{}>7,P{}=20),1,""),IF($G${}=45,IF(AND(G{}>8,P{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['X{}'.format(row)] = '=IF($H${}=25,IF(AND(H{}>4,Q{}=20),1,""),IF($H${}=30,IF(AND(H{}>5,Q{}=20),1,""),IF($H${}=35,IF(AND(H{}>6,Q{}=20),1,""),IF($H${}=40,IF(AND(H{}>7,Q{}=20),1,""),IF($H${}=45,IF(AND(H{}>8,Q{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['Y{}'.format(row)] = '=IF($I${}=25,IF(AND(I{}>4,R{}=20),1,""),IF($I${}=30,IF(AND(I{}>5,R{}=20),1,""),IF($I${}=35,IF(AND(I{}>6,R{}=20),1,""),IF($I${}=40,IF(AND(I{}>7,R{}=20),1,""),IF($I${}=45,IF(AND(I{}>8,R{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['Z{}'.format(row)] = '=IF($J${}=25,IF(AND(J{}>4,S{}=20),1,""),IF($J${}=30,IF(AND(J{}>5,S{}=20),1,""),IF($J${}=35,IF(AND(J{}>6,S{}=20),1,""),IF($J${}=40,IF(AND(J{}>7,S{}=20),1,""),IF($J${}=45,IF(AND(J{}>8,S{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score iterasi 1
            ws['AB1'] = 'NAMA SISWA_B'
            ws['AC1'] = 'NOMOR NF_B'
            ws['AD1'] = 'KELAS_B'
            ws['AE1'] = 'NAMA SEKOLAH_B'
            ws['AF1'] = 'LOKASI_B'
            ws['AG1'] = 'MAT_B'
            ws['AH1'] = 'FIS_B'
            ws['AI1'] = 'KIM_B'
            ws['AJ1'] = 'BIO_B'
            ws['AK1'] = 'JML_B'
            ws['AL1'] = 'Z_MAT_B'
            ws['AM1'] = 'Z_FIS_B'
            ws['AN1'] = 'Z_KIM_B'
            ws['AO1'] = 'Z_BIO_B'
            ws['AP1'] = 'S_MAT_B'
            ws['AQ1'] = 'S_FIS_B'
            ws['AR1'] = 'S_KIM_B'
            ws['AS1'] = 'S_BIO_B'
            ws['AT1'] = 'S_JML_B'
            ws['AU1'] = 'RANK NAS._B'
            ws['AV1'] = 'RANK LOK._B'

            ws['AL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AV1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['AB1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AC1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AD1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AE1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AF1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AG1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AH1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AI1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AJ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AK1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AL1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AM1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AN1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AO1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AP1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AQ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AR1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AS1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AT1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AU1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AV1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')

            # tambahan
            ws['AW1'] = 'MAT_20'
            ws['AX1'] = 'FIS_20'
            ws['AY1'] = 'KIM_20'
            ws['AZ1'] = 'BIO_20'
            ws['AW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AW1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AX1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AY1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AZ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')

            for row in range(2, q+1):
                # Tambahan
                ws['AB{}'.format(row)] = '=B{}'.format(row)
                ws['AC{}'.format(row)] = '=C{}'.format(row, row)
                ws['AD{}'.format(row)] = '=D{}'.format(row, row)
                ws['AE{}'.format(row)] = '=E{}'.format(row, row)
                ws['AF{}'.format(row)] = '=F{}'.format(row, row)
                ws['AG{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['AH{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['AI{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['AJ{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['AK{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)

                ws['AL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AG{}="","",(AG{}-AG${})/AG${}),2),"")'.format(row, row, r, s)
                ws['AM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AH{}="","",(AH{}-AH${})/AH${}),2),"")'.format(row, row, r, s)
                ws['AN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AI{}="","",(AI{}-AI${})/AI${}),2),"")'.format(row, row, r, s)
                ws['AO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AJ{}="","",(AJ{}-AJ${})/AJ${}),2),"")'.format(row, row, r, s)

                ws['AP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(G{}="","",IF(70+30*AL{}/$AL${}<20,20,70+30*AL{}/$AL${})),2),"")'.format(row, row, r, row, r)
                ws['AQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(H{}="","",IF(70+30*AM{}/$AM{}<20,20,70+30*AM{}/$AM${})),2),"")'.format(row, row, r, row, r)
                ws['AR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(I{}="","",IF(70+30*AN{}/$AN${}<20,20,70+30*AN{}/$AN${})),2),"")'.format(row, row, r, row, r)
                ws['AS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(J{}="","",IF(70+30*AO{}/$AO${}<20,20,70+30*AO{}/$AO${})),2),"")'.format(row, row, r, row, r)

                ws['AT{}'.format(row)] = '=IF(SUM(AP{}:AS{})=0,"",SUM(AP{}:AS{}))'.format(
                    row, row, row, row)
                ws['AU{}'.format(row)] = '=IF(AT{}="","",RANK(AT{},$AT$2:$AT${}))'.format(
                    row, row, q)
                ws['AV{}'.format(
                    row)] = '=IF(AU{}="","",COUNTIFS($AF$2:$AF${},AF{},$AU$2:$AU${},"<"&AU{})+1)'.format(row, q, row, q, row)
                # TAMBAHAN
                ws['AW{}'.format(row)] = '=IF($G${}=25,IF(AND(AG{}>4,AP{}=20),1,""),IF($G${}=30,IF(AND(AG{}>5,AP{}=20),1,""),IF($G${}=35,IF(AND(AG{}>6,AP{}=20),1,""),IF($G${}=40,IF(AND(AG{}>7,AP{}=20),1,""),IF($G${}=45,IF(AND(AG{}>8,AP{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AX{}'.format(row)] = '=IF($H${}=25,IF(AND(AH{}>4,AQ{}=20),1,""),IF($H${}=30,IF(AND(AH{}>5,AQ{}=20),1,""),IF($H${}=35,IF(AND(AH{}>6,AQ{}=20),1,""),IF($H${}=40,IF(AND(AH{}>7,AQ{}=20),1,""),IF($H${}=45,IF(AND(AH{}>8,AQ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AY{}'.format(row)] = '=IF($I${}=25,IF(AND(AI{}>4,AR{}=20),1,""),IF($I${}=30,IF(AND(AI{}>5,AR{}=20),1,""),IF($I${}=35,IF(AND(AI{}>6,AR{}=20),1,""),IF($I${}=40,IF(AND(AI{}>7,AR{}=20),1,""),IF($I${}=45,IF(AND(AI{}>8,AR{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AZ{}'.format(row)] = '=IF($J${}=25,IF(AND(AJ{}>4,AS{}=20),1,""),IF($J${}=30,IF(AND(AJ{}>5,AS{}=20),1,""),IF($J${}=35,IF(AND(AJ{}>6,AS{}=20),1,""),IF($J${}=40,IF(AND(AJ{}>7,AS{}=20),1,""),IF($J${}=45,IF(AND(AJ{}>8,AS{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score iterasi 2
            ws['BB1'] = 'NAMA SISWA_C'
            ws['BC1'] = 'NOMOR NF_c'
            ws['BD1'] = 'KELAS_C'
            ws['BE1'] = 'NAMA SEKOLAH_C'
            ws['BF1'] = 'LOKASI_C'
            ws['BG1'] = 'MAT_C'
            ws['BH1'] = 'FIS_C'
            ws['BI1'] = 'KIM_C'
            ws['BJ1'] = 'BIO_C'
            ws['BK1'] = 'JML_C'
            ws['BL1'] = 'Z_MAT_C'
            ws['BM1'] = 'Z_FIS_C'
            ws['BN1'] = 'Z_KIM_C'
            ws['BO1'] = 'Z_BIO_C'
            ws['BP1'] = 'S_MAT_C'
            ws['BQ1'] = 'S_FIS_C'
            ws['BR1'] = 'S_KIM_C'
            ws['BS1'] = 'S_BIO_C'
            ws['BT1'] = 'S_JML_C'
            ws['BU1'] = 'RANK NAS._C'
            ws['BV1'] = 'RANK LOK._C'

            ws['BL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BV1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['BB1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BC1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BD1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BE1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BF1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BG1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BH1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BI1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BJ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BK1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BL1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BM1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BN1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BO1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BP1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BQ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BR1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BS1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BT1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BU1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BV1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')

            # tambahan
            ws['BW1'] = 'MAT_20_C'
            ws['BX1'] = 'FIS_20_C'
            ws['BY1'] = 'KIM_20_C'
            ws['BZ1'] = 'BIO_20_C'
            ws['BW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BW1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BX1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BY1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BZ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')

            for row in range(2, q+1):
                # Tambahan
                ws['BB{}'.format(row)] = '=AB{}'.format(row)
                ws['BC{}'.format(row)] = '=AC{}'.format(row, row)
                ws['BD{}'.format(row)] = '=AD{}'.format(row, row)
                ws['BE{}'.format(row)] = '=AE{}'.format(row, row)
                ws['BF{}'.format(row)] = '=AF{}'.format(row, row)
                ws['BG{}'.format(row)] = '=IF(AG{}="","",AG{})'.format(
                    row, row)
                ws['BH{}'.format(row)] = '=IF(AH{}="","",AH{})'.format(
                    row, row)
                ws['BI{}'.format(row)] = '=IF(AI{}="","",AI{})'.format(
                    row, row)
                ws['BJ{}'.format(row)] = '=IF(AJ{}="","",AJ{})'.format(
                    row, row)
                ws['BK{}'.format(row)] = '=IF(AK{}="","",AK{})'.format(
                    row, row)

                ws['BL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BG{}="","",(BG{}-BG${})/BG${}),2),"")'.format(row, row, r, s)
                ws['BM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BH{}="","",(BH{}-BH${})/BH${}),2),"")'.format(row, row, r, s)
                ws['BN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BI{}="","",(BI{}-BI${})/BI${}),2),"")'.format(row, row, r, s)
                ws['BO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BJ{}="","",(BJ{}-BJ${})/BJ${}),2),"")'.format(row, row, r, s)

                ws['BP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BG{}="","",IF(70+30*BL{}/$BL${}<20,20,70+30*BL{}/$BL${})),2),"")'.format(row, row, r, row, r)
                ws['BQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BH{}="","",IF(70+30*BM{}/$BM{}<20,20,70+30*BM{}/$BM${})),2),"")'.format(row, row, r, row, r)
                ws['BR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BI{}="","",IF(70+30*BN{}/$BN${}<20,20,70+30*BN{}/$BN${})),2),"")'.format(row, row, r, row, r)
                ws['BS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BJ{}="","",IF(70+30*BO{}/$BO${}<20,20,70+30*BO{}/$BO${})),2),"")'.format(row, row, r, row, r)

                ws['BT{}'.format(row)] = '=IF(SUM(BP{}:BS{})=0,"",SUM(BP{}:BS{}))'.format(
                    row, row, row, row)
                ws['BU{}'.format(row)] = '=IF(BT{}="","",RANK(BT{},$BT$2:$BT${}))'.format(
                    row, row, q)
                ws['BV{}'.format(
                    row)] = '=IF(BU{}="","",COUNTIFS($BF$2:$BF${},BF{},$BU$2:$BU${},"<"&BU{})+1)'.format(row, q, row, q, row)
                # TAMBAHAN
                ws['BW{}'.format(row)] = '=IF($G${}=25,IF(AND(BG{}>4,BP{}=20),1,""),IF($G${}=30,IF(AND(BG{}>5,BP{}=20),1,""),IF($G${}=35,IF(AND(BG{}>6,BP{}=20),1,""),IF($G${}=40,IF(AND(BG{}>7,BP{}=20),1,""),IF($G${}=45,IF(AND(BG{}>8,BP{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BX{}'.format(row)] = '=IF($H${}=25,IF(AND(BH{}>4,BQ{}=20),1,""),IF($H${}=30,IF(AND(BH{}>5,BQ{}=20),1,""),IF($H${}=35,IF(AND(BH{}>6,BQ{}=20),1,""),IF($H${}=40,IF(AND(BH{}>7,BQ{}=20),1,""),IF($H${}=45,IF(AND(BH{}>8,BQ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BY{}'.format(row)] = '=IF($I${}=25,IF(AND(BI{}>4,BR{}=20),1,""),IF($I${}=30,IF(AND(BI{}>5,BR{}=20),1,""),IF($I${}=35,IF(AND(BI{}>6,BR{}=20),1,""),IF($I${}=40,IF(AND(BI{}>7,BR{}=20),1,""),IF($I${}=45,IF(AND(BI{}>8,BR{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BZ{}'.format(row)] = '=IF($J${}=25,IF(AND(BJ{}>4,BS{}=20),1,""),IF($J${}=30,IF(AND(BJ{}>5,BS{}=20),1,""),IF($J${}=35,IF(AND(BJ{}>6,BS{}=20),1,""),IF($J${}=40,IF(AND(BJ{}>7,BS{}=20),1,""),IF($J${}=45,IF(AND(BJ{}>8,BS{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score iterasi 3
            ws['CB1'] = 'NAMA SISWA_D'
            ws['CC1'] = 'NOMOR NF_D'
            ws['CD1'] = 'KELAS_D'
            ws['CE1'] = 'NAMA SEKOLAH_D'
            ws['CF1'] = 'LOKASI_D'
            ws['CG1'] = 'MAT_D'
            ws['CH1'] = 'FIS_D'
            ws['CI1'] = 'KIM_D'
            ws['CJ1'] = 'BIO_D'
            ws['CK1'] = 'JML_D'
            ws['CL1'] = 'Z_MAT_D'
            ws['CM1'] = 'Z_FIS_D'
            ws['CN1'] = 'Z_KIM_D'
            ws['CO1'] = 'Z_BIO_D'
            ws['CP1'] = 'S_MAT_D'
            ws['CQ1'] = 'S_FIS_D'
            ws['CR1'] = 'S_KIM_D'
            ws['CS1'] = 'S_BIO_D'
            ws['CT1'] = 'S_JML_D'
            ws['CU1'] = 'RANK NAS._D'
            ws['CV1'] = 'RANK LOK._D'

            ws['CL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CV1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['CB1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CC1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CD1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CE1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CF1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CG1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CH1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CI1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CJ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CK1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CL1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CM1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CN1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CO1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CP1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CQ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CR1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CS1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CT1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CU1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CV1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')

            # tambahan
            ws['CW1'] = 'MAT_20_D'
            ws['CX1'] = 'FIS_20_D'
            ws['CY1'] = 'KIM_20_D'
            ws['CZ1'] = 'BIO_20_D'
            ws['CW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CW1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CX1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CY1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CZ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')

            for row in range(2, q+1):
                ws['CB{}'.format(row)] = '=BB{}'.format(row)
                ws['CC{}'.format(row)] = '=BC{}'.format(row, row)
                ws['CD{}'.format(row)] = '=BD{}'.format(row, row)
                ws['CE{}'.format(row)] = '=BE{}'.format(row, row)
                ws['CF{}'.format(row)] = '=BF{}'.format(row, row)
                ws['CG{}'.format(row)] = '=IF(BG{}="","",BG{})'.format(
                    row, row)
                ws['CH{}'.format(row)] = '=IF(BH{}="","",BH{})'.format(
                    row, row)
                ws['CI{}'.format(row)] = '=IF(BI{}="","",BI{})'.format(
                    row, row)
                ws['CJ{}'.format(row)] = '=IF(BJ{}="","",BJ{})'.format(
                    row, row)
                ws['CK{}'.format(row)] = '=IF(BK{}="","",BK{})'.format(
                    row, row)

                ws['CL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CG{}="","",(CG{}-CG${})/CG${}),2),"")'.format(row, row, r, s)
                ws['CM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CH{}="","",(CH{}-CH${})/CH${}),2),"")'.format(row, row, r, s)
                ws['CN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CI{}="","",(CI{}-CI${})/CI${}),2),"")'.format(row, row, r, s)
                ws['CO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CJ{}="","",(CJ{}-CJ${})/CJ${}),2),"")'.format(row, row, r, s)

                ws['CP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CG{}="","",IF(70+30*CL{}/$CL${}<20,20,70+30*CL{}/$CL${})),2),"")'.format(row, row, r, row, r)
                ws['CQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CH{}="","",IF(70+30*CM{}/$CM{}<20,20,70+30*CM{}/$CM${})),2),"")'.format(row, row, r, row, r)
                ws['CR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CI{}="","",IF(70+30*CN{}/$CN${}<20,20,70+30*CN{}/$CN${})),2),"")'.format(row, row, r, row, r)
                ws['CS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CJ{}="","",IF(70+30*CO{}/$CO${}<20,20,70+30*CO{}/$CO${})),2),"")'.format(row, row, r, row, r)

                ws['CT{}'.format(row)] = '=IF(SUM(CP{}:CS{})=0,"",SUM(CP{}:CS{}))'.format(
                    row, row, row, row)
                ws['CU{}'.format(row)] = '=IF(CT{}="","",RANK(CT{},$CT$2:$CT${}))'.format(
                    row, row, q)
                ws['CV{}'.format(
                    row)] = '=IF(CU{}="","",COUNTIFS($CF$2:$CF${},CF{},$CU$2:$CU${},"<"&CU{})+1)'.format(row, q, row, q, row)
                # TAMBAHAN
                ws['CW{}'.format(row)] = '=IF($G${}=25,IF(AND(CG{}>4,CP{}=20),1,""),IF($G${}=30,IF(AND(CG{}>5,CP{}=20),1,""),IF($G${}=35,IF(AND(CG{}>6,CP{}=20),1,""),IF($G${}=40,IF(AND(CG{}>7,CP{}=20),1,""),IF($G${}=45,IF(AND(CG{}>8,CP{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CX{}'.format(row)] = '=IF($H${}=25,IF(AND(CH{}>4,CQ{}=20),1,""),IF($H${}=30,IF(AND(CH{}>5,CQ{}=20),1,""),IF($H${}=35,IF(AND(CH{}>6,CQ{}=20),1,""),IF($H${}=40,IF(AND(CH{}>7,CQ{}=20),1,""),IF($H${}=45,IF(AND(CH{}>8,CQ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CY{}'.format(row)] = '=IF($I${}=25,IF(AND(CI{}>4,CR{}=20),1,""),IF($I${}=30,IF(AND(CI{}>5,CR{}=20),1,""),IF($I${}=35,IF(AND(CI{}>6,CR{}=20),1,""),IF($I${}=40,IF(AND(CI{}>7,CR{}=20),1,""),IF($I${}=45,IF(AND(CI{}>8,CR{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CZ{}'.format(row)] = '=IF($J${}=25,IF(AND(CJ{}>4,CS{}=20),1,""),IF($J${}=30,IF(AND(CJ{}>5,CS{}=20),1,""),IF($J${}=35,IF(AND(CJ{}>6,CS{}=20),1,""),IF($J${}=40,IF(AND(CJ{}>7,CS{}=20),1,""),IF($J${}=45,IF(AND(CJ{}>8,CS{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score iterasi 4
            ws['DB1'] = 'NAMA SISWA_E'
            ws['DC1'] = 'NOMOR NF_E'
            ws['DD1'] = 'KELAS_E'
            ws['DE1'] = 'NAMA SEKOLAH_E'
            ws['DF1'] = 'LOKASI_E'
            ws['DG1'] = 'MAT_E'
            ws['DH1'] = 'FIS_E'
            ws['DI1'] = 'KIM_E'
            ws['DJ1'] = 'BIO_E'
            ws['DK1'] = 'JML_E'
            ws['DL1'] = 'Z_MAT_E'
            ws['DM1'] = 'Z_FIS_E'
            ws['DN1'] = 'Z_KIM_E'
            ws['DO1'] = 'Z_BIO_E'
            ws['DP1'] = 'S_MAT_E'
            ws['DQ1'] = 'S_FIS_E'
            ws['DR1'] = 'S_KIM_E'
            ws['DS1'] = 'S_BIO_E'
            ws['DT1'] = 'S_JML_E'
            ws['DU1'] = 'RANK NAS._E'
            ws['DV1'] = 'RANK LOK._E'

            ws['DL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DV1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['DB1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DC1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DD1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DE1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DF1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DG1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DH1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DI1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DJ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DK1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DL1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DM1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DN1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DO1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DP1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DQ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DR1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DS1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DT1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DU1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DV1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')

            # tambahan
            ws['DW1'] = 'MAT_20'
            ws['DX1'] = 'FIS_20'
            ws['DY1'] = 'KIM_20'
            ws['DZ1'] = 'BIO_20'
            ws['DW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DW1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DX1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DY1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DZ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')

            for row in range(2, q+1):
                # Tambahan
                ws['DB{}'.format(row)] = '=CB{}'.format(row)
                ws['DC{}'.format(row)] = '=CC{}'.format(row, row)
                ws['DD{}'.format(row)] = '=CD{}'.format(row, row)
                ws['DE{}'.format(row)] = '=CE{}'.format(row, row)
                ws['DF{}'.format(row)] = '=CF{}'.format(row, row)
                ws['DG{}'.format(row)] = '=IF(CG{}="","",CG{})'.format(
                    row, row)
                ws['DH{}'.format(row)] = '=IF(CH{}="","",CH{})'.format(
                    row, row)
                ws['DI{}'.format(row)] = '=IF(CI{}="","",CI{})'.format(
                    row, row)
                ws['DJ{}'.format(row)] = '=IF(CJ{}="","",CJ{})'.format(
                    row, row)
                ws['DK{}'.format(row)] = '=IF(CK{}="","",CK{})'.format(
                    row, row)

                ws['DL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DG{}="","",(DG{}-DG${})/DG${}),2),"")'.format(row, row, r, s)
                ws['DM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DH{}="","",(DH{}-DH${})/DH${}),2),"")'.format(row, row, r, s)
                ws['DN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DI{}="","",(DI{}-DI${})/DI${}),2),"")'.format(row, row, r, s)
                ws['DO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DJ{}="","",(DJ{}-DJ${})/DJ${}),2),"")'.format(row, row, r, s)

                ws['DP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DG{}="","",IF(70+30*DL{}/$DL${}<20,20,70+30*DL{}/$DL${})),2),"")'.format(row, row, r, row, r)
                ws['DQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DH{}="","",IF(70+30*DM{}/$DM{}<20,20,70+30*DM{}/$DM${})),2),"")'.format(row, row, r, row, r)
                ws['DR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DI{}="","",IF(70+30*DN{}/$DN${}<20,20,70+30*DN{}/$DN${})),2),"")'.format(row, row, r, row, r)
                ws['DS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DJ{}="","",IF(70+30*DO{}/$DO${}<20,20,70+30*DO{}/$DO${})),2),"")'.format(row, row, r, row, r)

                ws['DT{}'.format(row)] = '=IF(SUM(DP{}:DS{})=0,"",SUM(DP{}:DS{}))'.format(
                    row, row, row, row)
                ws['DU{}'.format(row)] = '=IF(DT{}="","",RANK(DT{},$DT$2:$DT${}))'.format(
                    row, row, q)
                ws['DV{}'.format(
                    row)] = '=IF(DU{}="","",COUNTIFS($DF$2:$DF${},DF{},$DU$2:$DU${},"<"&DU{})+1)'.format(row, q, row, q, row)
                # TAMBAHAN
                ws['DW{}'.format(row)] = '=IF($G${}=25,IF(AND(DG{}>4,DP{}=20),1,""),IF($G${}=30,IF(AND(DG{}>5,DP{}=20),1,""),IF($G${}=35,IF(AND(DG{}>6,DP{}=20),1,""),IF($G${}=40,IF(AND(DG{}>7,DP{}=20),1,""),IF($G${}=45,IF(AND(DG{}>8,DP{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DX{}'.format(row)] = '=IF($H${}=25,IF(AND(DH{}>4,DQ{}=20),1,""),IF($H${}=30,IF(AND(DH{}>5,DQ{}=20),1,""),IF($H${}=35,IF(AND(DH{}>6,DQ{}=20),1,""),IF($H${}=40,IF(AND(DH{}>7,DQ{}=20),1,""),IF($H${}=45,IF(AND(DH{}>8,DQ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DY{}'.format(row)] = '=IF($I${}=25,IF(AND(DI{}>4,DR{}=20),1,""),IF($I${}=30,IF(AND(DI{}>5,DR{}=20),1,""),IF($I${}=35,IF(AND(DI{}>6,DR{}=20),1,""),IF($I${}=40,IF(AND(DI{}>7,DR{}=20),1,""),IF($I${}=45,IF(AND(DI{}>8,DR{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DZ{}'.format(row)] = '=IF($J${}=25,IF(AND(DJ{}>4,DS{}=20),1,""),IF($J${}=30,IF(AND(DJ{}>5,DS{}=20),1,""),IF($J${}=35,IF(AND(DJ{}>6,DS{}=20),1,""),IF($J${}=40,IF(AND(DJ{}>7,DS{}=20),1,""),IF($J${}=45,IF(AND(DJ{}>8,DS{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score iterasi 5
            ws['EB1'] = 'NAMA SISWA'
            ws['EC1'] = 'NOMOR NF'
            ws['ED1'] = 'KELAS'
            ws['EE1'] = 'NAMA SEKOLAH'
            ws['EF1'] = 'LOKASI'
            ws['EG1'] = 'MAT'
            ws['EH1'] = 'FIS'
            ws['EI1'] = 'KIM'
            ws['EJ1'] = 'BIO'
            ws['EK1'] = 'JML'
            ws['EL1'] = 'Z_MAT'
            ws['EM1'] = 'Z_FIS'
            ws['EN1'] = 'Z_KIM'
            ws['EO1'] = 'Z_BIO'
            ws['EP1'] = 'S_MAT'
            ws['EQ1'] = 'S_FIS'
            ws['ER1'] = 'S_KIM'
            ws['ES1'] = 'S_BIO'
            ws['ET1'] = 'S_JML'
            ws['EU1'] = 'RANK NAS.'
            ws['EV1'] = 'RANK LOK.'

            ws['EL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ER1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ES1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ET1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EV1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['EB1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EC1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['ED1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EE1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EF1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EG1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EH1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EI1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EJ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EK1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EL1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EM1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EN1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EO1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EP1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EQ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['ER1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['ES1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['ET1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EU1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EV1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')

            # tambahan
            ws['EW1'] = 'MAT_20'
            ws['EX1'] = 'FIS_20'
            ws['EY1'] = 'KIM_20'
            ws['EZ1'] = 'BIO_20'
            ws['EW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EW1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EX1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EY1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EZ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')

            for row in range(2, q+1):
                # Tambahan
                ws['EB{}'.format(row)] = '=DB{}'.format(row)
                ws['EC{}'.format(row)] = '=DC{}'.format(row, row)
                ws['ED{}'.format(row)] = '=DD{}'.format(row, row)
                ws['EE{}'.format(row)] = '=DE{}'.format(row, row)
                ws['EF{}'.format(row)] = '=DF{}'.format(row, row)
                ws['EG{}'.format(row)] = '=IF(DG{}="","",DG{})'.format(
                    row, row)
                ws['EH{}'.format(row)] = '=IF(DH{}="","",DH{})'.format(
                    row, row)
                ws['EI{}'.format(row)] = '=IF(DI{}="","",DI{})'.format(
                    row, row)
                ws['EJ{}'.format(row)] = '=IF(DJ{}="","",DJ{})'.format(
                    row, row)
                ws['EK{}'.format(row)] = '=IF(DK{}="","",DK{})'.format(
                    row, row)

                ws['EL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EG{}="","",(EG{}-EG${})/EG${}),2),"")'.format(row, row, r, s)
                ws['EM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EH{}="","",(EH{}-EH${})/EH${}),2),"")'.format(row, row, r, s)
                ws['EN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EI{}="","",(EI{}-EI${})/EI${}),2),"")'.format(row, row, r, s)
                ws['EO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EJ{}="","",(EJ{}-EJ${})/EJ${}),2),"")'.format(row, row, r, s)

                ws['EP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EG{}="","",IF(70+30*EL{}/$EL${}<20,20,70+30*EL{}/$EL${})),2),"")'.format(row, row, r, row, r)
                ws['EQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EH{}="","",IF(70+30*EM{}/$EM{}<20,20,70+30*EM{}/$EM${})),2),"")'.format(row, row, r, row, r)
                ws['ER{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EI{}="","",IF(70+30*EN{}/$EN${}<20,20,70+30*EN{}/$EN${})),2),"")'.format(row, row, r, row, r)
                ws['ES{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EJ{}="","",IF(70+30*EO{}/$EO${}<20,20,70+30*EO{}/$EO${})),2),"")'.format(row, row, r, row, r)

                ws['ET{}'.format(row)] = '=IF(SUM(EP{}:ES{})=0,"",SUM(EP{}:ES{}))'.format(
                    row, row, row, row)
                ws['EU{}'.format(row)] = '=IF(ET{}="","",RANK(ET{},$ET$2:$ET${}))'.format(
                    row, row, q)
                ws['EV{}'.format(
                    row)] = '=IF(EU{}="","",COUNTIFS($EF$2:$EF${},EF{},$EU$2:$EU${},"<"&EU{})+1)'.format(row, q, row, q, row)
                # TAMBAHAN
                ws['EW{}'.format(row)] = '=IF($G${}=25,IF(AND(EG{}>4,EP{}=20),1,""),IF($G${}=30,IF(AND(EG{}>5,EP{}=20),1,""),IF($G${}=35,IF(AND(EG{}>6,EP{}=20),1,""),IF($G${}=40,IF(AND(EG{}>7,EP{}=20),1,""),IF($G${}=45,IF(AND(EG{}>8,EP{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['EX{}'.format(row)] = '=IF($H${}=25,IF(AND(EH{}>4,EQ{}=20),1,""),IF($H${}=30,IF(AND(EH{}>5,EQ{}=20),1,""),IF($H${}=35,IF(AND(EH{}>6,EQ{}=20),1,""),IF($H${}=40,IF(AND(EH{}>7,EQ{}=20),1,""),IF($H${}=45,IF(AND(EH{}>8,EQ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['EY{}'.format(row)] = '=IF($I${}=25,IF(AND(EI{}>4,ER{}=20),1,""),IF($I${}=30,IF(AND(EI{}>5,ER{}=20),1,""),IF($I${}=35,IF(AND(EI{}>6,ER{}=20),1,""),IF($I${}=40,IF(AND(EI{}>7,ER{}=20),1,""),IF($I${}=45,IF(AND(EI{}>8,ER{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['EZ{}'.format(row)] = '=IF($J${}=25,IF(AND(EJ{}>4,ES{}=20),1,""),IF($J${}=30,IF(AND(EJ{}>5,ES{}=20),1,""),IF($J${}=35,IF(AND(EJ{}>6,ES{}=20),1,""),IF($J${}=40,IF(AND(EJ{}>7,ES{}=20),1,""),IF($J${}=45,IF(AND(EJ{}>8,ES{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Mengubah 'KELAS' sesuai dengan nilai yang dipilih dari selectbox 'KELAS'
            kelas = KELAS.lower().replace(" ", "")
            semester = SEMESTER.lower()
            tahun = TAHUN.replace("-", "")
            penilaian = PENILAIAN.lower()
            kurikulum = KURIKULUM.lower()

            path_file = f"{kelas}_{penilaian}_{semester}_{kurikulum}_{tahun}_nilai_std.xlsx"

            # Simpan file ke direktori temporer
            temp_dir = tempfile.gettempdir()
            file_path = temp_dir + '/' + path_file
            wb.save(file_path)

            st.success(
                "File siap diunduh!")

            # Tombol unduh file
            with open(file_path, "rb") as f:
                bytes_data = f.read()
            st.download_button(label="Unduh File", data=bytes_data,
                               file_name=path_file)

            st.warning(
                "Buka file unduhan, klik 'Enable Editing' dan 'Save'")
    if selected_file == "Nilai Std. 10, 11 IPS":
        # menghilangkan hamburger
        st.markdown("""
        <style>
        .css-1rs6os.edgvbvh3
        {
            visibility:hidden;
        }
        .css-1lsmgbg.egzxvld0
        {
            visibility:hidden;
        }
        </style>
        """, unsafe_allow_html=True)

        image = Image.open('logo resmi nf resize.png')
        st.image(image)

        st.title("Olah Nilai Standar K13")
        st.header("10-11 SMA IPS")

        col6 = st.container()

        with col6:
            KELAS = st.selectbox(
                "KELAS",
                ("--Pilih Kelas--", "10 SMA IPS", "11 SMA IPS"))

        col7 = st.container()

        with col7:
            SEMESTER = st.selectbox(
                "SEMESTER",
                ("--Pilih Semester--", "SEMESTER 1", "SEMESTER 2"))

        col8 = st.container()

        with col8:
            PENILAIAN = st.selectbox(
                "PENILAIAN",
                ("--Pilih Penilaian--", "PENILAIAN TENGAH SEMESTER", "PENILAIAN AKHIR SEMESTER", "PENILAIAN AKHIR TAHUN"))

        col9 = st.container()

        with col9:
            KURIKULUM = st.selectbox(
                "KURIKULUM",
                ("--Pilih Kurikulum--", "K13"))

        TAHUN = st.text_input("Masukkan Tahun Ajaran",
                              placeholder="contoh: 2022-2023")

        col1, col2, col3, col4, col5, col6, col7 = st.columns(7)

        with col1:
            MTK = st.selectbox(
                "JML. SOAL MAT.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col2:
            IND = st.selectbox(
                "JML. SOAL IND.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col3:
            ENG = st.selectbox(
                "JML. SOAL ENG.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col4:
            SEJ = st.selectbox(
                "JML. SOAL SEJ.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col5:
            GEO = st.selectbox(
                "JML. SOAL GEO.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col6:
            EKO = st.selectbox(
                "JML. SOAL EKO.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col7:
            SOS = st.selectbox(
                "JML. SOAL SOS.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        JML_SOAL_MAT = MTK
        JML_SOAL_IND = IND
        JML_SOAL_ENG = ENG
        JML_SOAL_SEJ = SEJ
        JML_SOAL_GEO = GEO
        JML_SOAL_EKO = EKO
        JML_SOAL_SOS = SOS

        uploaded_file = st.file_uploader(
            'Letakkan file excel IPS', type='xlsx')

        if uploaded_file is not None:
            wb = openpyxl.load_workbook(uploaded_file)
            ws = wb['Sheet1']

            q = len(ws['K'])
            r = len(ws['K'])+2
            s = len(ws['K'])+3
            t = len(ws['K'])+4
            u = len(ws['K'])+5
            v = len(ws['K'])+6
            w = len(ws['K'])+7
            x = len(ws['K'])+8

            ws['G{}'.format(r)] = "=ROUND(AVERAGE(G2:G{}),2)".format(q)
            ws['H{}'.format(r)] = "=ROUND(AVERAGE(H2:H{}),2)".format(q)
            ws['I{}'.format(r)] = "=ROUND(AVERAGE(I2:I{}),2)".format(q)
            ws['J{}'.format(r)] = "=ROUND(AVERAGE(J2:J{}),2)".format(q)
            ws['K{}'.format(r)] = "=ROUND(AVERAGE(K2:K{}),2)".format(q)
            ws['L{}'.format(r)] = "=ROUND(AVERAGE(L2:L{}),2)".format(q)
            ws['M{}'.format(r)] = "=ROUND(AVERAGE(M2:M{}),2)".format(q)
            ws['N{}'.format(r)] = "=ROUND(AVERAGE(N2:N{}),2)".format(q)

            ws['G{}'.format(s)] = "=STDEV(G2:G{})".format(q)
            ws['H{}'.format(s)] = "=STDEV(H2:H{})".format(q)
            ws['I{}'.format(s)] = "=STDEV(I2:I{})".format(q)
            ws['J{}'.format(s)] = "=STDEV(J2:J{})".format(q)
            ws['K{}'.format(s)] = "=STDEV(K2:K{})".format(q)
            ws['L{}'.format(s)] = "=STDEV(L2:L{})".format(q)
            ws['M{}'.format(s)] = "=STDEV(M2:M{})".format(q)

            ws['G{}'.format(t)] = "=MAX(G2:G{})".format(q)
            ws['H{}'.format(t)] = "=MAX(H2:H{})".format(q)
            ws['I{}'.format(t)] = "=MAX(I2:I{})".format(q)
            ws['J{}'.format(t)] = "=MAX(J2:J{})".format(q)
            ws['K{}'.format(t)] = "=MAX(K2:K{})".format(q)
            ws['L{}'.format(t)] = "=MAX(L2:L{})".format(q)
            ws['M{}'.format(t)] = "=MAX(M2:M{})".format(q)

            ws['N{}'.format(r)] = "=MAX(N2:N{})".format(q)
            ws['O{}'.format(r)] = "=MAX(O2:O{})".format(q)
            ws['P{}'.format(r)] = "=MAX(P2:P{})".format(q)
            ws['Q{}'.format(r)] = "=MAX(Q2:Q{})".format(q)
            ws['R{}'.format(r)] = "=MAX(R2:R{})".format(q)
            ws['S{}'.format(r)] = "=MAX(S2:S{})".format(q)
            ws['T{}'.format(r)] = "=MAX(T2:T{})".format(q)
            ws['U{}'.format(r)] = "=MAX(U2:U{})".format(q)
            ws['V{}'.format(r)] = "=MAX(V2:V{})".format(q)
            ws['W{}'.format(r)] = "=MAX(W2:W{})".format(q)
            ws['X{}'.format(r)] = "=MAX(X2:X{})".format(q)
            ws['Y{}'.format(r)] = "=MAX(Y2:Y{})".format(q)
            ws['Z{}'.format(r)] = "=MAX(Z2:Z{})".format(q)
            ws['AA{}'.format(r)] = "=MAX(AA2:AA{})".format(q)

            ws['AB{}'.format(r)] = "=ROUND(MAX(AB2:AB{}),2)".format(q)
            ws['AC{}'.format(r)] = "=MAX(AC2:AC{})".format(q)
            ws['AD{}'.format(r)] = "=MAX(AD2:AD{})".format(q)

            ws['G{}'.format(u)] = "=MIN(G2:G{})".format(q)
            ws['H{}'.format(u)] = "=MIN(H2:H{})".format(q)
            ws['I{}'.format(u)] = "=MIN(I2:I{})".format(q)
            ws['J{}'.format(u)] = "=MIN(J2:J{})".format(q)
            ws['K{}'.format(u)] = "=MIN(K2:K{})".format(q)
            ws['L{}'.format(u)] = "=MIN(L2:L{})".format(q)
            ws['M{}'.format(u)] = "=MIN(M2:M{})".format(q)
            ws['N{}'.format(u)] = "=MIN(N2:N{})".format(q)

            ws['V{}'.format(s)] = "=MIN(V2:V{})".format(q)
            ws['W{}'.format(s)] = "=MIN(W2:W{})".format(q)
            ws['X{}'.format(s)] = "=MIN(X2:X{})".format(q)
            ws['Y{}'.format(s)] = "=MIN(Y2:Y{})".format(q)
            ws['Z{}'.format(s)] = "=MIN(Z2:Z{})".format(q)
            ws['AA{}'.format(s)] = "=MIN(AA2:AA{})".format(q)
            ws['AB{}'.format(s)] = "=MIN(AB2:AB{})".format(q)
            ws['AC{}'.format(s)] = "=MIN(AC2:AC{})".format(q)

            ws['V{}'.format(t)] = "=ROUND(AVERAGE(V2:V{}),2)".format(q)
            ws['W{}'.format(t)] = "=ROUND(AVERAGE(W2:W{}),2)".format(q)
            ws['X{}'.format(t)] = "=ROUND(AVERAGE(X2:X{}),2)".format(q)
            ws['Y{}'.format(t)] = "=ROUND(AVERAGE(Y2:Y{}),2)".format(q)
            ws['Z{}'.format(t)] = "=ROUND(AVERAGE(Z2:Z{}),2)".format(q)
            ws['AA{}'.format(t)] = "=ROUND(AVERAGE(AA2:AA{}),2)".format(q)
            ws['AB{}'.format(t)] = "=ROUND(AVERAGE(AB2:AB{}),2)".format(q)
            ws['AC{}'.format(t)] = "=ROUND(AVERAGE(AC2:AC{}),2)".format(q)

            ws['AF{}'.format(r)] = "=SUM(AF2:AF{})".format(q)
            ws['AG{}'.format(r)] = "=SUM(AG2:AG{})".format(q)
            ws['AH{}'.format(r)] = "=SUM(AH2:AH{})".format(q)
            ws['AI{}'.format(r)] = "=SUM(AI2:AI{})".format(q)
            ws['AJ{}'.format(r)] = "=SUM(AJ2:AJ{})".format(q)
            ws['AK{}'.format(r)] = "=SUM(AK2:AK{})".format(q)
            ws['AL{}'.format(r)] = "=SUM(AL2:AL{})".format(q)

            # Jumlah Soal
            ws['F{}'.format(v)] = 'JUMLAH SOAL'
            ws['G{}'.format(v)] = JML_SOAL_MAT
            ws['H{}'.format(v)] = JML_SOAL_IND
            ws['I{}'.format(v)] = JML_SOAL_ENG
            ws['J{}'.format(v)] = JML_SOAL_SEJ
            ws['K{}'.format(v)] = JML_SOAL_GEO
            ws['L{}'.format(v)] = JML_SOAL_EKO
            ws['M{}'.format(v)] = JML_SOAL_SOS

            # new
            # iterasi 1 rata-rata - 1
            # rata" MTK ke MTK tambahan dan mapel MTK awal
            ws['AS{}'.format(
                r)] = "=IF($AF${}=0,$G${},$G${}-1)".format(r, r, r)
            ws['AS{}'.format(s)] = "=STDEV(AS2:AS{})".format(q)
            ws['AS{}'.format(t)] = "=MAX(AS2:AS{})".format(q)
            ws['AS{}'.format(u)] = "=MIN(AS2:AS{})".format(q)
            # rata" IND ke IND tambahan dan mapel IND awal
            ws['AT{}'.format(
                r)] = "=IF($AG${}=0,$H${},$H${}-1)".format(r, r, r)
            ws['AT{}'.format(s)] = "=STDEV(AT2:AT{})".format(q)
            ws['AT{}'.format(t)] = "=MAX(AT2:AT{})".format(q)
            ws['AT{}'.format(u)] = "=MIN(AT2:AT{})".format(q)
            # rata" ENG ke ENG tambahan dan mapel ENG awal
            ws['AU{}'.format(
                r)] = "=IF($AH${}=0,$I${},$I${}-1)".format(r, r, r)
            ws['AU{}'.format(s)] = "=STDEV(AU2:AU{})".format(q)
            ws['AU{}'.format(t)] = "=MAX(AU2:AU{})".format(q)
            ws['AU{}'.format(u)] = "=MIN(AU2:AU{})".format(q)
            # rata" SEJ ke SEJ tambahan dan mapel SEJ awal
            ws['AV{}'.format(
                r)] = "=IF($AI${}=0,$J${},$J${}-1)".format(r, r, r)
            ws['AV{}'.format(s)] = "=STDEV(AV2:AV{})".format(q)
            ws['AV{}'.format(t)] = "=MAX(AV2:AV{})".format(q)
            ws['AV{}'.format(u)] = "=MIN(AV2:AV{})".format(q)
            # rata" GEO ke GEO tambahan dan mapel GEO awal
            ws['AW{}'.format(
                r)] = "=IF($AJ${}=0,$K${},$K${}-1)".format(r, r, r)
            ws['AW{}'.format(s)] = "=STDEV(AW2:AW{})".format(q)
            ws['AW{}'.format(t)] = "=MAX(AW2:AW{})".format(q)
            ws['AW{}'.format(u)] = "=MIN(AW2:AW{})".format(q)
            # rata" EKO ke EKO tambahan dan mapel EKO awal
            ws['AX{}'.format(
                r)] = "=IF($AK${}=0,$L${},$L${}-1)".format(r, r, r)
            ws['AX{}'.format(s)] = "=STDEV(AX2:AX{})".format(q)
            ws['AX{}'.format(t)] = "=MAX(AX2:AX{})".format(q)
            ws['AX{}'.format(u)] = "=MIN(AX2:AX{})".format(q)
            # rata" SOS ke SOS tambahan dan mapel SOS awal
            ws['AY{}'.format(
                r)] = "=IF($AL${}=0,$M${},$M${}-1)".format(r, r, r)
            ws['AY{}'.format(s)] = "=STDEV(AY2:AY{})".format(q)
            ws['AY{}'.format(t)] = "=MAX(AY2:AY{})".format(q)
            ws['AY{}'.format(u)] = "=MIN(AY2:AY{})".format(q)
            # jml MAPEL
            ws['AZ{}'.format(r)] = "=ROUND(AVERAGE(AZ2:AZ{}),2)".format(q)
            ws['AZ{}'.format(t)] = "=MAX(AZ2:AZ{})".format(q)
            ws['AZ{}'.format(u)] = "=MIN(AZ2:AZ{})".format(q)
            # MAX Z SCORE
            ws['BA{}'.format(r)] = "=MAX(BA2:BA{})".format(q)
            ws['BB{}'.format(r)] = "=MAX(BB2:BB{})".format(q)
            ws['BC{}'.format(r)] = "=MAX(BC2:BC{})".format(q)
            ws['BD{}'.format(r)] = "=MAX(BD2:BD{})".format(q)
            ws['BE{}'.format(r)] = "=MAX(BE2:BE{})".format(q)
            ws['BF{}'.format(r)] = "=MAX(BF2:BF{})".format(q)
            ws['BG{}'.format(r)] = "=MAX(BG2:BG{})".format(q)
            # NILAI STANDAR MTK
            ws['BH{}'.format(r)] = "=MAX(BH2:BH{})".format(q)
            ws['BH{}'.format(s)] = "=MIN(BH2:BH{})".format(q)
            ws['BH{}'.format(t)] = "=ROUND(AVERAGE(BH2:BH{}),2)".format(q)
            # NILAI STANDAR IND
            ws['BI{}'.format(r)] = "=MAX(BI2:BI{})".format(q)
            ws['BI{}'.format(s)] = "=MIN(BI2:BI{})".format(q)
            ws['BI{}'.format(t)] = "=ROUND(AVERAGE(BI2:BI{}),2)".format(q)
            # NILAI STANDAR ENG
            ws['BJ{}'.format(r)] = "=MAX(BJ2:BJ{})".format(q)
            ws['BJ{}'.format(s)] = "=MIN(BJ2:BJ{})".format(q)
            ws['BJ{}'.format(t)] = "=ROUND(AVERAGE(BJ2:BJ{}),2)".format(q)
            # NILAI STANDAR SEJ
            ws['BK{}'.format(r)] = "=MAX(BK2:BK{})".format(q)
            ws['BK{}'.format(s)] = "=MIN(BK2:BK{})".format(q)
            ws['BK{}'.format(t)] = "=ROUND(AVERAGE(BK2:BK{}),2)".format(q)
            # NILAI STANDAR GEO
            ws['BL{}'.format(r)] = "=MAX(BL2:BL{})".format(q)
            ws['BL{}'.format(s)] = "=MIN(BL2:BL{})".format(q)
            ws['BL{}'.format(t)] = "=ROUND(AVERAGE(BL2:BL{}),2)".format(q)
            # NILAI STANDAR EKO
            ws['BM{}'.format(r)] = "=MAX(BM2:BM{})".format(q)
            ws['BM{}'.format(s)] = "=MIN(BM2:BM{})".format(q)
            ws['BM{}'.format(t)] = "=ROUND(AVERAGE(BM2:BM{}),2)".format(q)
            # NILAI STANDAR SOS
            ws['BN{}'.format(r)] = "=MAX(BN2:BN{})".format(q)
            ws['BN{}'.format(s)] = "=MIN(BN2:BN{})".format(q)
            ws['BN{}'.format(t)] = "=ROUND(AVERAGE(BN2:BN{}),2)".format(q)
            # NILAI STANDAR S.JML
            ws['BO{}'.format(r)] = "=MAX(BO2:BO{})".format(q)
            ws['BO{}'.format(s)] = "=MIN(BO2:BO{})".format(q)
            ws['BO{}'.format(t)] = "=ROUND(AVERAGE(BO2:BO{}),2)".format(q)

            # TAMBAHAN
            ws['BR{}'.format(r)] = "=SUM(BR2:BR{})".format(q)
            ws['BS{}'.format(r)] = "=SUM(BS2:BS{})".format(q)
            ws['BT{}'.format(r)] = "=SUM(BT2:BT{})".format(q)
            ws['BU{}'.format(r)] = "=SUM(BU2:BU{})".format(q)
            ws['BV{}'.format(r)] = "=SUM(BV2:BV{})".format(q)
            ws['BW{}'.format(r)] = "=SUM(BW2:BW{})".format(q)
            ws['BX{}'.format(r)] = "=SUM(BX2:BX{})".format(q)

            # iterasi 2 rata-rata - 2
            # rata" MTK ke MTK tambahan dan mapel MTK awal
            ws['CE{}'.format(
                r)] = "=IF($BR${}=0,$AS${},$AS${}-1)".format(r, r, r)
            ws['CE{}'.format(s)] = "=STDEV(CE2:CE{})".format(q)
            ws['CE{}'.format(t)] = "=MAX(CE2:CE{})".format(q)
            ws['CE{}'.format(u)] = "=MIN(CE2:CE{})".format(q)
            # rata" IND ke IND tambahan dan mapel IND awal
            ws['CF{}'.format(
                r)] = "=IF($BS${}=0,$AT${},$AT${}-1)".format(r, r, r)
            ws['CF{}'.format(s)] = "=STDEV(CF2:CF{})".format(q)
            ws['CF{}'.format(t)] = "=MAX(CF2:CF{})".format(q)
            ws['CF{}'.format(u)] = "=MIN(CF2:CF{})".format(q)
            # rata" ENG ke ENG tambahan dan mapel ENG awal
            ws['CG{}'.format(
                r)] = "=IF($BT${}=0,$AU${},$AU${}-1)".format(r, r, r)
            ws['CG{}'.format(s)] = "=STDEV(CG2:CG{})".format(q)
            ws['CG{}'.format(t)] = "=MAX(CG2:CG{})".format(q)
            ws['CG{}'.format(u)] = "=MIN(CG2:CG{})".format(q)
            # rata" SEJ ke SEJ tambahan dan mapel SEJ awal
            ws['CH{}'.format(
                r)] = "=IF($BU${}=0,$AV${},$AV${}-1)".format(r, r, r)
            ws['CH{}'.format(s)] = "=STDEV(CH2:CH{})".format(q)
            ws['CH{}'.format(t)] = "=MAX(CH2:CH{})".format(q)
            ws['CH{}'.format(u)] = "=MIN(CH2:CH{})".format(q)
            # rata" GEO ke GEO tambahan dan mapel GEO awal
            ws['CI{}'.format(
                r)] = "=IF($BV${}=0,$AW${},$AW${}-1)".format(r, r, r)
            ws['CI{}'.format(s)] = "=STDEV(CI2:CI{})".format(q)
            ws['CI{}'.format(t)] = "=MAX(CI2:CI{})".format(q)
            ws['CI{}'.format(u)] = "=MIN(CI2:CI{})".format(q)
            # rata" EKO ke EKO tambahan dan mapel EKO awal
            ws['CJ{}'.format(
                r)] = "=IF($BW${}=0,$AX${},$AX${}-1)".format(r, r, r)
            ws['CJ{}'.format(s)] = "=STDEV(CJ2:CJ{})".format(q)
            ws['CJ{}'.format(t)] = "=MAX(CJ2:CJ{})".format(q)
            ws['CJ{}'.format(u)] = "=MIN(CJ2:CJ{})".format(q)
            # rata" SOS ke SOS tambahan dan mapel SOS awal
            ws['CK{}'.format(
                r)] = "=IF($BX${}=0,$AY${},$AY${}-1)".format(r, r, r)
            ws['CK{}'.format(s)] = "=STDEV(CK2:CK{})".format(q)
            ws['CK{}'.format(t)] = "=MAX(CK2:CK{})".format(q)
            ws['CK{}'.format(u)] = "=MIN(CK2:CK{})".format(q)
            # jml MAPEL
            ws['CL{}'.format(r)] = "=ROUND(AVERAGE(CL2:CL{}),2)".format(q)
            ws['CL{}'.format(t)] = "=MAX(CL2:CL{})".format(q)
            ws['CL{}'.format(u)] = "=MIN(CL2:CL{})".format(q)
            # MAX Z SCORE
            ws['CM{}'.format(r)] = "=MAX(CM2:CM{})".format(q)
            ws['CN{}'.format(r)] = "=MAX(CN2:CN{})".format(q)
            ws['CO{}'.format(r)] = "=MAX(CO2:CO{})".format(q)
            ws['CP{}'.format(r)] = "=MAX(CP2:CP{})".format(q)
            ws['CQ{}'.format(r)] = "=MAX(CQ2:CQ{})".format(q)
            ws['CR{}'.format(r)] = "=MAX(CR2:CR{})".format(q)
            ws['CS{}'.format(r)] = "=MAX(CS2:CS{})".format(q)
            # NILAI STANDAR MTK
            ws['CT{}'.format(r)] = "=MAX(CT2:CT{})".format(q)
            ws['CT{}'.format(s)] = "=MIN(CT2:CT{})".format(q)
            ws['CT{}'.format(t)] = "=ROUND(AVERAGE(CT2:CT{}),2)".format(q)
            # NILAI STANDAR IND
            ws['CU{}'.format(r)] = "=MAX(CU2:CU{})".format(q)
            ws['CU{}'.format(s)] = "=MIN(CU2:CU{})".format(q)
            ws['CU{}'.format(t)] = "=ROUND(AVERAGE(CU2:CU{}),2)".format(q)
            # NILAI STANDAR ENG
            ws['CV{}'.format(r)] = "=MAX(CV2:CV{})".format(q)
            ws['CV{}'.format(s)] = "=MIN(CV2:CV{})".format(q)
            ws['CV{}'.format(t)] = "=ROUND(AVERAGE(CV2:CV{}),2)".format(q)
            # NILAI STANDAR SEJ
            ws['CW{}'.format(r)] = "=MAX(CW2:CW{})".format(q)
            ws['CW{}'.format(s)] = "=MIN(CW2:CW{})".format(q)
            ws['CW{}'.format(t)] = "=ROUND(AVERAGE(CW2:CW{}),2)".format(q)
            # NILAI STANDAR GEO
            ws['CX{}'.format(r)] = "=MAX(CX2:CX{})".format(q)
            ws['CX{}'.format(s)] = "=MIN(CX2:CX{})".format(q)
            ws['CX{}'.format(t)] = "=ROUND(AVERAGE(CX2:CX{}),2)".format(q)
            # NILAI STANDAR EKO
            ws['CY{}'.format(r)] = "=MAX(CY2:CY{})".format(q)
            ws['CY{}'.format(s)] = "=MIN(CY2:CY{})".format(q)
            ws['CY{}'.format(t)] = "=ROUND(AVERAGE(CY2:CY{}),2)".format(q)
            # NILAI STANDAR SOS
            ws['CZ{}'.format(r)] = "=MAX(CZ2:CZ{})".format(q)
            ws['CZ{}'.format(s)] = "=MIN(CZ2:CZ{})".format(q)
            ws['CZ{}'.format(t)] = "=ROUND(AVERAGE(CZ2:CZ{}),2)".format(q)
            # NILAI STANDAR S.JML
            ws['DA{}'.format(r)] = "=MAX(DA2:DA{})".format(q)
            ws['DA{}'.format(s)] = "=MIN(DA2:DA{})".format(q)
            ws['DA{}'.format(t)] = "=ROUND(AVERAGE(DA2:DA{}),2)".format(q)

            # TAMBAHAN
            ws['DD{}'.format(r)] = "=SUM(DD2:DD{})".format(q)
            ws['DE{}'.format(r)] = "=SUM(DE2:DE{})".format(q)
            ws['DF{}'.format(r)] = "=SUM(DF2:DF{})".format(q)
            ws['DG{}'.format(r)] = "=SUM(DG2:DG{})".format(q)
            ws['DH{}'.format(r)] = "=SUM(DH2:DH{})".format(q)
            ws['DI{}'.format(r)] = "=SUM(DI2:DI{})".format(q)
            ws['DJ{}'.format(r)] = "=SUM(DJ2:DJ{})".format(q)

            # iterasi 3 rata-rata - 3
            # rata" MTK ke MTK tambahan dan mapel MTK awal
            ws['DQ{}'.format(
                r)] = "=IF($DD${}=0,$CE${},$CE${}-1)".format(r, r, r)
            ws['DQ{}'.format(s)] = "=STDEV(DQ2:DQ{})".format(q)
            ws['DQ{}'.format(t)] = "=MAX(DQ2:DQ{})".format(q)
            ws['DQ{}'.format(u)] = "=MIN(DQ2:DQ{})".format(q)
            # rata" IND ke IND tambahan dan mapel IND awal
            ws['DR{}'.format(
                r)] = "=IF($DE${}=0,$CF${},$CF${}-1)".format(r, r, r)
            ws['DR{}'.format(s)] = "=STDEV(DR2:DR{})".format(q)
            ws['DR{}'.format(t)] = "=MAX(DR2:DR{})".format(q)
            ws['DR{}'.format(u)] = "=MIN(DR2:DR{})".format(q)
            # rata" ENG ke ENG tambahan dan mapel ENG awal
            ws['DS{}'.format(
                r)] = "=IF($DF${}=0,$CG${},$CG{}-1)".format(r, r, r)
            ws['DS{}'.format(s)] = "=STDEV(DS2:DS{})".format(q)
            ws['DS{}'.format(t)] = "=MAX(DS2:DS{})".format(q)
            ws['DS{}'.format(u)] = "=MIN(DS2:DS{})".format(q)
            # rata" SEJ ke SEJ tambahan dan mapel SEJ awal
            ws['DT{}'.format(
                r)] = "=IF($DG${}=0,$CH${},$CH${}-1)".format(r, r, r)
            ws['DT{}'.format(s)] = "=STDEV(DT2:DT{})".format(q)
            ws['DT{}'.format(t)] = "=MAX(DT2:DT{})".format(q)
            ws['DT{}'.format(u)] = "=MIN(DT2:DT{})".format(q)
            # rata" GEO ke GEO tambahan dan mapel GEO awal
            ws['DU{}'.format(
                r)] = "=IF($DH${}=0,$CI${},$CI${}-1)".format(r, r, r)
            ws['DU{}'.format(s)] = "=STDEV(DU2:DU{})".format(q)
            ws['DU{}'.format(t)] = "=MAX(DU2:DU{})".format(q)
            ws['DU{}'.format(u)] = "=MIN(DU2:DU{})".format(q)
            # rata" EKO ke EKO tambahan dan mapel EKO awal
            ws['DV{}'.format(
                r)] = "=IF($DI${}=0,$CJ${},$CJ${}-1)".format(r, r, r)
            ws['DV{}'.format(s)] = "=STDEV(DV2:DV{})".format(q)
            ws['DV{}'.format(t)] = "=MAX(DV2:DV{})".format(q)
            ws['DV{}'.format(u)] = "=MIN(DV2:DV{})".format(q)
            # rata" SOS ke SOS tambahan dan mapel SOS awal
            ws['DW{}'.format(
                r)] = "=IF($DJ${}=0,$CK${},$CK${}-1)".format(r, r, r)
            ws['DW{}'.format(s)] = "=STDEV(DW2:DW{})".format(q)
            ws['DW{}'.format(t)] = "=MAX(DW2:DW{})".format(q)
            ws['DW{}'.format(u)] = "=MIN(DW2:DW{})".format(q)
            # jml MAPEL
            ws['DX{}'.format(r)] = "=ROUND(AVERAGE(DX2:DX{}),2)".format(q)
            ws['DX{}'.format(t)] = "=MAX(DX2:DX{})".format(q)
            ws['DX{}'.format(u)] = "=MIN(DX2:DX{})".format(q)
            # MAX Z SCORE
            ws['DY{}'.format(r)] = "=MAX(DY2:DY{})".format(q)
            ws['DZ{}'.format(r)] = "=MAX(DZ2:DZ{})".format(q)
            ws['EA{}'.format(r)] = "=MAX(EA2:EA{})".format(q)
            ws['EB{}'.format(r)] = "=MAX(EB2:EB{})".format(q)
            ws['EC{}'.format(r)] = "=MAX(EC2:EC{})".format(q)
            ws['ED{}'.format(r)] = "=MAX(ED2:ED{})".format(q)
            ws['EE{}'.format(r)] = "=MAX(EE2:EE{})".format(q)
            # NILAI STANDAR MTK
            ws['EF{}'.format(r)] = "=MAX(EF2:EF{})".format(q)
            ws['EF{}'.format(s)] = "=MIN(EF2:EF{})".format(q)
            ws['EF{}'.format(t)] = "=ROUND(AVERAGE(EF2:EF{}),2)".format(q)
            # NILAI STANDAR IND
            ws['EG{}'.format(r)] = "=MAX(EG2:EG{})".format(q)
            ws['EG{}'.format(s)] = "=MIN(EG2:EG{})".format(q)
            ws['EG{}'.format(t)] = "=ROUND(AVERAGE(EG2:EG{}),2)".format(q)
            # NILAI STANDAR ENG
            ws['EH{}'.format(r)] = "=MAX(EH2:EH{})".format(q)
            ws['EH{}'.format(s)] = "=MIN(EH2:EH{})".format(q)
            ws['EH{}'.format(t)] = "=ROUND(AVERAGE(EH2:EH{}),2)".format(q)
            # NILAI STANDAR SEJ
            ws['EI{}'.format(r)] = "=MAX(EI2:EI{})".format(q)
            ws['EI{}'.format(s)] = "=MIN(EI2:EI{})".format(q)
            ws['EI{}'.format(t)] = "=ROUND(AVERAGE(EI2:EI{}),2)".format(q)
            # NILAI STANDAR GEO
            ws['EJ{}'.format(r)] = "=MAX(EJ2:EJ{})".format(q)
            ws['EJ{}'.format(s)] = "=MIN(EJ2:EJ{})".format(q)
            ws['EJ{}'.format(t)] = "=ROUND(AVERAGE(EJ2:EJ{}),2)".format(q)
            # NILAI STANDAR EKO
            ws['EK{}'.format(r)] = "=MAX(EK2:EK{})".format(q)
            ws['EK{}'.format(s)] = "=MIN(EK2:EK{})".format(q)
            ws['EK{}'.format(t)] = "=ROUND(AVERAGE(EK2:EK{}),2)".format(q)
            # NILAI STANDAR SOS
            ws['EL{}'.format(r)] = "=MAX(EL2:EL{})".format(q)
            ws['EL{}'.format(s)] = "=MIN(EL2:EL{})".format(q)
            ws['EL{}'.format(t)] = "=ROUND(AVERAGE(EL2:EL{}),2)".format(q)
            # NILAI STANDAR S.JML
            ws['EM{}'.format(r)] = "=MAX(EM2:EM{})".format(q)
            ws['EM{}'.format(s)] = "=MIN(EM2:EM{})".format(q)
            ws['EM{}'.format(t)] = "=ROUND(AVERAGE(EM2:EM{}),2)".format(q)

            # TAMBAHAN
            ws['EP{}'.format(r)] = "=SUM(EP2:EP{})".format(q)
            ws['EQ{}'.format(r)] = "=SUM(EQ2:EQ{})".format(q)
            ws['ER{}'.format(r)] = "=SUM(ER2:ER{})".format(q)
            ws['ES{}'.format(r)] = "=SUM(ES2:ES{})".format(q)
            ws['ET{}'.format(r)] = "=SUM(ET2:ET{})".format(q)
            ws['EU{}'.format(r)] = "=SUM(EU2:EU{})".format(q)
            ws['EV{}'.format(r)] = "=SUM(EV2:EV{})".format(q)

            # iterasi 4 rata-rata - 4
            # rata" MTK ke MTK tambahan dan mapel MTK awal
            ws['FC{}'.format(
                r)] = "=IF($EP${}=0,$DQ${},$DQ${}-1)".format(r, r, r)
            ws['FC{}'.format(s)] = "=STDEV(FC2:FC{})".format(q)
            ws['FC{}'.format(t)] = "=MAX(FC2:FC{})".format(q)
            ws['FC{}'.format(u)] = "=MIN(FC2:FC{})".format(q)
            # rata" IND ke IND tambahan dan mapel IND awal
            ws['FD{}'.format(
                r)] = "=IF($EQ${}=0,$DR${},$DR${}-1)".format(r, r, r)
            ws['FD{}'.format(s)] = "=STDEV(FD2:FD{})".format(q)
            ws['FD{}'.format(t)] = "=MAX(FD2:FD{})".format(q)
            ws['FD{}'.format(u)] = "=MIN(FD2:FD{})".format(q)
            # rata" ENG ke ENG tambahan dan mapel ENG awal
            ws['FE{}'.format(
                r)] = "=IF($ER${}=0,$DS${},$DS{}-1)".format(r, r, r)
            ws['FE{}'.format(s)] = "=STDEV(FE2:FE{})".format(q)
            ws['FE{}'.format(t)] = "=MAX(FE2:FE{})".format(q)
            ws['FE{}'.format(u)] = "=MIN(FE2:FE{})".format(q)
            # rata" SEJ ke SEJ tambahan dan mapel SEJ awal
            ws['FF{}'.format(
                r)] = "=IF($ES${}=0,$DT${},$DT${}-1)".format(r, r, r)
            ws['FF{}'.format(s)] = "=STDEV(FF2:FF{})".format(q)
            ws['FF{}'.format(t)] = "=MAX(FF2:FF{})".format(q)
            ws['FF{}'.format(u)] = "=MIN(FF2:FF{})".format(q)
            # rata" GEO ke GEO tambahan dan mapel GEO awal
            ws['FG{}'.format(
                r)] = "=IF($ET${}=0,$DU${},$DU${}-1)".format(r, r, r)
            ws['FG{}'.format(s)] = "=STDEV(FG2:FG{})".format(q)
            ws['FG{}'.format(t)] = "=MAX(FG2:FG{})".format(q)
            ws['FG{}'.format(u)] = "=MIN(FG2:FG{})".format(q)
            # rata" EKO ke EKO tambahan dan mapel EKO awal
            ws['FH{}'.format(
                r)] = "=IF($EU${}=0,$DV${},$DV${}-1)".format(r, r, r)
            ws['FH{}'.format(s)] = "=STDEV(FH2:FH{})".format(q)
            ws['FH{}'.format(t)] = "=MAX(FH2:FH{})".format(q)
            ws['FH{}'.format(u)] = "=MIN(FH2:FH{})".format(q)
            # rata" SOS ke SOS tambahan dan mapel SOS awal
            ws['FI{}'.format(
                r)] = "=IF($EV${}=0,$DW${},$DW${}-1)".format(r, r, r)
            ws['FI{}'.format(s)] = "=STDEV(FI2:FI{})".format(q)
            ws['FI{}'.format(t)] = "=MAX(FI2:FI{})".format(q)
            ws['FI{}'.format(u)] = "=MIN(FI2:FI{})".format(q)
            # jml MAPEL
            ws['FJ{}'.format(r)] = "=ROUND(AVERAGE(FJ2:FJ{}),2)".format(q)
            ws['FJ{}'.format(t)] = "=MAX(FJ2:FJ{})".format(q)
            ws['FJ{}'.format(u)] = "=MIN(FJ2:FJ{})".format(q)
            # MAX Z SCORE
            ws['FK{}'.format(r)] = "=MAX(FK2:FK{})".format(q)
            ws['FL{}'.format(r)] = "=MAX(FL2:FL{})".format(q)
            ws['FM{}'.format(r)] = "=MAX(FM2:FM{})".format(q)
            ws['FN{}'.format(r)] = "=MAX(FN2:FN{})".format(q)
            ws['FO{}'.format(r)] = "=MAX(FO2:FO{})".format(q)
            ws['FP{}'.format(r)] = "=MAX(FP2:FP{})".format(q)
            ws['FQ{}'.format(r)] = "=MAX(FQ2:FQ{})".format(q)
            # NILAI STANDAR MTK
            ws['FR{}'.format(r)] = "=MAX(FR2:FR{})".format(q)
            ws['FR{}'.format(s)] = "=MIN(FR2:FR{})".format(q)
            ws['FR{}'.format(t)] = "=ROUND(AVERAGE(FR2:FR{}),2)".format(q)
            # NILAI STANDAR IND
            ws['FS{}'.format(r)] = "=MAX(FS2:FS{})".format(q)
            ws['FS{}'.format(s)] = "=MIN(FS2:FS{})".format(q)
            ws['FS{}'.format(t)] = "=ROUND(AVERAGE(FS2:FS{}),2)".format(q)
            # NILAI STANDAR ENG
            ws['FT{}'.format(r)] = "=MAX(FT2:FT{})".format(q)
            ws['FT{}'.format(s)] = "=MIN(FT2:FT{})".format(q)
            ws['FT{}'.format(t)] = "=ROUND(AVERAGE(FT2:FT{}),2)".format(q)
            # NILAI STANDAR SEJ
            ws['FU{}'.format(r)] = "=MAX(FU2:FU{})".format(q)
            ws['FU{}'.format(s)] = "=MIN(FU2:FU{})".format(q)
            ws['FU{}'.format(t)] = "=ROUND(AVERAGE(FU2:FU{}),2)".format(q)
            # NILAI STANDAR GEO
            ws['FV{}'.format(r)] = "=MAX(FV2:FV{})".format(q)
            ws['FV{}'.format(s)] = "=MIN(FV2:FV{})".format(q)
            ws['FV{}'.format(t)] = "=ROUND(AVERAGE(FV2:FV{}),2)".format(q)
            # NILAI STANDAR EKO
            ws['FW{}'.format(r)] = "=MAX(FW2:FW{})".format(q)
            ws['FW{}'.format(s)] = "=MIN(FW2:FW{})".format(q)
            ws['FW{}'.format(t)] = "=ROUND(AVERAGE(FW2:FW{}),2)".format(q)
            # NILAI STANDAR SOS
            ws['FX{}'.format(r)] = "=MAX(FX2:FX{})".format(q)
            ws['FX{}'.format(s)] = "=MIN(FX2:FX{})".format(q)
            ws['FX{}'.format(t)] = "=ROUND(AVERAGE(FX2:FX{}),2)".format(q)
            # NILAI STANDAR S.JML
            ws['FY{}'.format(r)] = "=MAX(FY2:FY{})".format(q)
            ws['FY{}'.format(s)] = "=MIN(FY2:FY{})".format(q)
            ws['FY{}'.format(t)] = "=ROUND(AVERAGE(FY2:FY{}),2)".format(q)

            # TAMBAHAN
            ws['GB{}'.format(r)] = "=SUM(GB2:GB{})".format(q)
            ws['GC{}'.format(r)] = "=SUM(GC2:GC{})".format(q)
            ws['GD{}'.format(r)] = "=SUM(GD2:GD{})".format(q)
            ws['GE{}'.format(r)] = "=SUM(GE2:GE{})".format(q)
            ws['GF{}'.format(r)] = "=SUM(GF2:GF{})".format(q)
            ws['GG{}'.format(r)] = "=SUM(GG2:GG{})".format(q)
            ws['GH{}'.format(r)] = "=SUM(GH2:GH{})".format(q)

            # iterasi 5 rata-rata - 5
            # rata" MTK ke MTK tambahan dan mapel MTK awal
            ws['GO{}'.format(
                r)] = "=IF($GB${}=0,$FC${},$FC${}-1)".format(r, r, r)
            ws['GO{}'.format(s)] = "=STDEV(GO2:GO{})".format(q)
            ws['GO{}'.format(t)] = "=MAX(GO2:GO{})".format(q)
            ws['GO{}'.format(u)] = "=MIN(GO2:GO{})".format(q)
            # rata" IND ke IND tambahan dan mapel IND awal
            ws['GP{}'.format(
                r)] = "=IF($GC${}=0,$FD${},$FD${}-1)".format(r, r, r)
            ws['GP{}'.format(s)] = "=STDEV(GP2:GP{})".format(q)
            ws['GP{}'.format(t)] = "=MAX(GP2:GP{})".format(q)
            ws['GP{}'.format(u)] = "=MIN(GP2:GP{})".format(q)
            # rata" ENG ke ENG tambahan dan mapel ENG awal
            ws['GQ{}'.format(
                r)] = "=IF($GD${}=0,$FE${},$FE{}-1)".format(r, r, r)
            ws['GQ{}'.format(s)] = "=STDEV(GQ2:GQ{})".format(q)
            ws['GQ{}'.format(t)] = "=MAX(GQ2:GQ{})".format(q)
            ws['GQ{}'.format(u)] = "=MIN(GQ2:GQ{})".format(q)
            # rata" SEJ ke SEJ tambahan dan mapel SEJ awal
            ws['GR{}'.format(
                r)] = "=IF($GE${}=0,$FF${},$FF${}-1)".format(r, r, r)
            ws['GR{}'.format(s)] = "=STDEV(GR2:GR{})".format(q)
            ws['GR{}'.format(t)] = "=MAX(GR2:GR{})".format(q)
            ws['GR{}'.format(u)] = "=MIN(GR2:GR{})".format(q)
            # rata" GEO ke GEO tambahan dan mapel GEO awal
            ws['GS{}'.format(
                r)] = "=IF($GF${}=0,$FG${},$FG${}-1)".format(r, r, r)
            ws['GS{}'.format(s)] = "=STDEV(GS2:GS{})".format(q)
            ws['GS{}'.format(t)] = "=MAX(GS2:GS{})".format(q)
            ws['GS{}'.format(u)] = "=MIN(GS2:GS{})".format(q)
            # rata" EKO ke EKO tambahan dan mapel EKO awal
            ws['GT{}'.format(
                r)] = "=IF($GG${}=0,$FH${},$FH${}-1)".format(r, r, r)
            ws['GT{}'.format(s)] = "=STDEV(GT2:GT{})".format(q)
            ws['GT{}'.format(t)] = "=MAX(GT2:GT{})".format(q)
            ws['GT{}'.format(u)] = "=MIN(GT2:GT{})".format(q)
            # rata" SOS ke SOS tambahan dan mapel SOS awal
            ws['GU{}'.format(
                r)] = "=IF($GH${}=0,$FI${},$FI${}-1)".format(r, r, r)
            ws['GU{}'.format(s)] = "=STDEV(GU2:GU{})".format(q)
            ws['GU{}'.format(t)] = "=MAX(GU2:GU{})".format(q)
            ws['GU{}'.format(u)] = "=MIN(GU2:GU{})".format(q)
            # jml MAPEL
            ws['GV{}'.format(r)] = "=ROUND(AVERAGE(GV2:GV{}),2)".format(q)
            ws['GV{}'.format(t)] = "=MAX(GV2:GV{})".format(q)
            ws['GV{}'.format(u)] = "=MIN(GV2:GV{})".format(q)
            # MAX Z SCORE
            ws['GW{}'.format(r)] = "=MAX(GW2:GW{})".format(q)
            ws['GX{}'.format(r)] = "=MAX(GX2:GX{})".format(q)
            ws['GY{}'.format(r)] = "=MAX(GY2:GY{})".format(q)
            ws['GZ{}'.format(r)] = "=MAX(GZ2:GZ{})".format(q)
            ws['HA{}'.format(r)] = "=MAX(HA2:HA{})".format(q)
            ws['HB{}'.format(r)] = "=MAX(HB2:HB{})".format(q)
            ws['HC{}'.format(r)] = "=MAX(HC2:HC{})".format(q)
            # NILAI STANDAR MTK
            ws['HD{}'.format(r)] = "=MAX(HD2:HD{})".format(q)
            ws['HD{}'.format(s)] = "=MIN(HD2:HD{})".format(q)
            ws['HD{}'.format(t)] = "=ROUND(AVERAGE(HD2:HD{}),2)".format(q)
            # NILAI STANDAR IND
            ws['HE{}'.format(r)] = "=MAX(HE2:HE{})".format(q)
            ws['HE{}'.format(s)] = "=MIN(HE2:HE{})".format(q)
            ws['HE{}'.format(t)] = "=ROUND(AVERAGE(HE2:HE{}),2)".format(q)
            # NILAI STANDAR ENG
            ws['HF{}'.format(r)] = "=MAX(HF2:HF{})".format(q)
            ws['HF{}'.format(s)] = "=MIN(HF2:HF{})".format(q)
            ws['HF{}'.format(t)] = "=ROUND(AVERAGE(HF2:HF{}),2)".format(q)
            # NILAI STANDAR SEJ
            ws['HG{}'.format(r)] = "=MAX(HG2:HG{})".format(q)
            ws['HG{}'.format(s)] = "=MIN(HG2:HG{})".format(q)
            ws['HG{}'.format(t)] = "=ROUND(AVERAGE(HG2:HG{}),2)".format(q)
            # NILAI STANDAR GEO
            ws['HH{}'.format(r)] = "=MAX(HH2:HH{})".format(q)
            ws['HH{}'.format(s)] = "=MIN(HH2:HH{})".format(q)
            ws['HH{}'.format(t)] = "=ROUND(AVERAGE(HH2:HH{}),2)".format(q)
            # NILAI STANDAR EKO
            ws['HI{}'.format(r)] = "=MAX(HI2:HI{})".format(q)
            ws['HI{}'.format(s)] = "=MIN(HI2:HI{})".format(q)
            ws['HI{}'.format(t)] = "=ROUND(AVERAGE(HI2:HI{}),2)".format(q)
            # NILAI STANDAR SOS
            ws['HJ{}'.format(r)] = "=MAX(HJ2:HJ{})".format(q)
            ws['HJ{}'.format(s)] = "=MIN(HJ2:HJ{})".format(q)
            ws['HJ{}'.format(t)] = "=ROUND(AVERAGE(HJ2:HJ{}),2)".format(q)
            # NILAI STANDAR S.JML
            ws['HK{}'.format(r)] = "=MAX(HK2:HK{})".format(q)
            ws['HK{}'.format(s)] = "=MIN(HK2:HK{})".format(q)
            ws['HK{}'.format(t)] = "=ROUND(AVERAGE(HK2:HK{}),2)".format(q)

            # TAMBAHAN
            ws['HN{}'.format(r)] = "=SUM(HN2:HN{})".format(q)
            ws['HO{}'.format(r)] = "=SUM(HO2:HO{})".format(q)
            ws['HP{}'.format(r)] = "=SUM(HP2:HP{})".format(q)
            ws['HQ{}'.format(r)] = "=SUM(HQ2:HQ{})".format(q)
            ws['HR{}'.format(r)] = "=SUM(HR2:HR{})".format(q)
            ws['HS{}'.format(r)] = "=SUM(HS2:HS{})".format(q)
            ws['HT{}'.format(r)] = "=SUM(HT2:HT{})".format(q)

            # Z Score [1]
            ws['B1'] = 'NAMA SISWA_A'
            ws['C1'] = 'NOMOR NF_A'
            ws['D1'] = 'KELAS_A'
            ws['E1'] = 'NAMA SEKOLAH_A'
            ws['F1'] = 'LOKASI_A'

            ws['G1'] = 'MAT_A'
            ws['H1'] = 'IND_A'
            ws['I1'] = 'ENG_A'
            ws['J1'] = 'SEJ_A'
            ws['K1'] = 'GEO_A'
            ws['L1'] = 'EKO_A'
            ws['M1'] = 'SOS_A'
            ws['N1'] = 'JML_A'

            ws['O1'] = 'Z_MAT_A'
            ws['P1'] = 'Z_IND_A'
            ws['Q1'] = 'Z_ENG_A'
            ws['R1'] = 'Z_SEJ_A'
            ws['S1'] = 'Z_GEO_A'
            ws['T1'] = 'Z_EKO_A'
            ws['U1'] = 'Z_SOS_A'

            ws['V1'] = 'S_MAT_A'
            ws['W1'] = 'S_IND_A'
            ws['X1'] = 'S_ENG_A'
            ws['Y1'] = 'S_SEJ_A'
            ws['Z1'] = 'S_GEO_A'
            ws['AA1'] = 'S_EKO_A'
            ws['AB1'] = 'S_SOS_A'
            ws['AC1'] = 'S_JML_A'

            ws['AD1'] = 'RANK NAS._A'
            ws['AE1'] = 'RANK LOK._A'

            ws['O1'].font = Font(bold=False, name='Calibri', size=11)
            ws['P1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Q1'].font = Font(bold=False, name='Calibri', size=11)
            ws['R1'].font = Font(bold=False, name='Calibri', size=11)
            ws['S1'].font = Font(bold=False, name='Calibri', size=11)
            ws['T1'].font = Font(bold=False, name='Calibri', size=11)
            ws['U1'].font = Font(bold=False, name='Calibri', size=11)
            ws['V1'].font = Font(bold=False, name='Calibri', size=11)
            ws['W1'].font = Font(bold=False, name='Calibri', size=11)
            ws['X1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Y1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Z1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AE1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['B1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['C1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['D1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['E1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['F1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['G1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['H1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['I1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['J1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['K1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['L1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['M1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['N1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['O1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['P1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Q1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['R1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['S1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['T1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['U1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['V1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['W1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['X1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Y1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Z1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AA1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AB1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AC1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AD1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AE1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')

            # tambahan
            ws['AF1'] = 'MAT_20_A'
            ws['AG1'] = 'IND_20_A'
            ws['AH1'] = 'ENG_20_A'
            ws['AI1'] = 'SEJ_20_A'
            ws['AJ1'] = 'GEO_20_A'
            ws['AK1'] = 'EKO_20_A'
            ws['AL1'] = 'SOS_20_A'

            ws['AF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AL1'].font = Font(bold=False, name='Calibri', size=11)

            ws['AF1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AG1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AH1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AI1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AJ1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AK1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AL1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')

            for row in range(2, q+1):
                ws['N{}'.format(
                    row)] = '=SUM(G{}:M{})'.format(row, row, row)
                ws['O{}'.format(
                    row)] = '=IFERROR(ROUND(IF(G{}="","",(G{}-G${})/G${}),2),"")'.format(row, row, r, s)
                ws['P{}'.format(
                    row)] = '=IFERROR(ROUND(IF(H{}="","",(H{}-H${})/H${}),2),"")'.format(row, row, r, s)
                ws['Q{}'.format(
                    row)] = '=IFERROR(ROUND(IF(I{}="","",(I{}-I${})/I${}),2),"")'.format(row, row, r, s)
                ws['R{}'.format(
                    row)] = '=IFERROR(ROUND(IF(J{}="","",(J{}-J${})/J${}),2),"")'.format(row, row, r, s)
                ws['S{}'.format(
                    row)] = '=IFERROR(ROUND(IF(K{}="","",(K{}-K${})/K${}),2),"")'.format(row, row, r, s)
                ws['T{}'.format(
                    row)] = '=IFERROR(ROUND(IF(L{}="","",(L{}-L${})/L${}),2),"")'.format(row, row, r, s)
                ws['U{}'.format(
                    row)] = '=IFERROR(ROUND(IF(M{}="","",(M{}-M${})/M${}),2),"")'.format(row, row, r, s)

                ws['V{}'.format(
                    row)] = '=IFERROR(ROUND(IF(G{}="","",IF(70+30*O{}/$O${}<20,20,70+30*O{}/$O${})),2),"")'.format(row, row, r, row, r)
                ws['W{}'.format(
                    row)] = '=IFERROR(ROUND(IF(H{}="","",IF(70+30*P{}/$P${}<20,20,70+30*P{}/$P${})),2),"")'.format(row, row, r, row, r)
                ws['X{}'.format(
                    row)] = '=IFERROR(ROUND(IF(I{}="","",IF(70+30*Q{}/$Q${}<20,20,70+30*Q{}/$Q${})),2),"")'.format(row, row, r, row, r)
                ws['Y{}'.format(
                    row)] = '=IFERROR(ROUND(IF(J{}="","",IF(70+30*R{}/$R${}<20,20,70+30*R{}/$R${})),2),"")'.format(row, row, r, row, r)
                ws['Z{}'.format(
                    row)] = '=IFERROR(ROUND(IF(K{}="","",IF(70+30*S{}/$S${}<20,20,70+30*S{}/$S${})),2),"")'.format(row, row, r, row, r)
                ws['AA{}'.format(
                    row)] = '=IFERROR(ROUND(IF(L{}="","",IF(70+30*T{}/$T${}<20,20,70+30*T{}/$T${})),2),"")'.format(row, row, r, row, r)
                ws['AB{}'.format(
                    row)] = '=IFERROR(ROUND(IF(M{}="","",IF(70+30*U{}/$U${}<20,20,70+30*U{}/$U${})),2),"")'.format(row, row, r, row, r)

                ws['AC{}'.format(row)] = '=IF(SUM(V{}:AB{})=0,"",SUM(V{}:AB{}))'.format(
                    row, row, row, row)
                ws['AD{}'.format(row)] = '=IF(AC{}="","",RANK(AC{},$AC$2:$AC${}))'.format(
                    row, row, q)
                ws['AE{}'.format(
                    row)] = '=IF(AD{}="","",COUNTIFS($F$2:$F${},F{},$AD$2:$AD${},"<"&AD{})+1)'.format(row, q, row, q, row)
            # TAMBAHAN
                ws['AF{}'.format(row)] = '=IF($G${}=20,IF(AND(G{}>3,V{}=20),1,""),IF($G${}=25,IF(AND(G{}>4,V{}=20),1,""),IF($G${}=30,IF(AND(G{}>5,V{}=20),1,""),IF($G${}=35,IF(AND(G{}>6,V{}=20),1,""),IF($G${}=40,IF(AND(G{}>7,V{}=20),1,""),IF($G${}=45,IF(AND(G{}>8,V{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AG{}'.format(row)] = '=IF($H${}=20,IF(AND(H{}>3,W{}=20),1,""),IF($H${}=25,IF(AND(H{}>4,W{}=20),1,""),IF($H${}=30,IF(AND(H{}>5,W{}=20),1,""),IF($H${}=35,IF(AND(H{}>6,W{}=20),1,""),IF($H${}=40,IF(AND(H{}>7,W{}=20),1,""),IF($H${}=45,IF(AND(H{}>8,W{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AH{}'.format(row)] = '=IF($I${}=20,IF(AND(I{}>3,X{}=20),1,""),IF($I${}=25,IF(AND(I{}>4,X{}=20),1,""),IF($I${}=30,IF(AND(I{}>5,X{}=20),1,""),IF($I${}=35,IF(AND(I{}>6,X{}=20),1,""),IF($I${}=40,IF(AND(I{}>7,X{}=20),1,""),IF($I${}=45,IF(AND(I{}>8,X{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AI{}'.format(row)] = '=IF($J${}=20,IF(AND(J{}>3,Y{}=20),1,""),IF($J${}=25,IF(AND(J{}>4,Y{}=20),1,""),IF($J${}=30,IF(AND(J{}>5,Y{}=20),1,""),IF($J${}=35,IF(AND(J{}>6,Y{}=20),1,""),IF($J${}=40,IF(AND(J{}>7,Y{}=20),1,""),IF($J${}=45,IF(AND(J{}>8,Y{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AJ{}'.format(row)] = '=IF($K${}=20,IF(AND(K{}>3,Z{}=20),1,""),IF($K${}=25,IF(AND(K{}>4,Z{}=20),1,""),IF($K${}=30,IF(AND(K{}>5,Z{}=20),1,""),IF($K${}=35,IF(AND(K{}>6,Z{}=20),1,""),IF($K${}=40,IF(AND(K{}>7,Z{}=20),1,""),IF($K${}=45,IF(AND(K{}>8,Z{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AK{}'.format(row)] = '=IF($L${}=20,IF(AND(L{}>3,AA{}=20),1,""),IF($L${}=25,IF(AND(L{}>4,AA{}=20),1,""),IF($L${}=30,IF(AND(L{}>5,AA{}=20),1,""),IF($L${}=35,IF(AND(L{}>6,AA{}=20),1,""),IF($L${}=40,IF(AND(L{}>7,AA{}=20),1,""),IF($L${}=45,IF(AND(L{}>8,AA{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AL{}'.format(row)] = '=IF($M${}=20,IF(AND(M{}>3,AB{}=20),1,""),IF($M${}=25,IF(AND(M{}>4,AB{}=20),1,""),IF($M${}=30,IF(AND(M{}>5,AB{}=20),1,""),IF($M${}=35,IF(AND(M{}>6,AB{}=20),1,""),IF($M${}=40,IF(AND(M{}>7,AB{}=20),1,""),IF($M${}=45,IF(AND(M{}>8,AB{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score [2]
            ws['AN1'] = 'NAMA SISWA_B'
            ws['AO1'] = 'NOMOR NF_B'
            ws['AP1'] = 'KELAS_B'
            ws['AQ1'] = 'NAMA SEKOLAH_B'
            ws['AR1'] = 'LOKASI_B'

            ws['AS1'] = 'MAT_B'
            ws['AT1'] = 'IND_B'
            ws['AU1'] = 'ENG_B'
            ws['AV1'] = 'SEJ_B'
            ws['AW1'] = 'GEO_B'
            ws['AX1'] = 'EKO_B'
            ws['AY1'] = 'SOS_B'
            ws['AZ1'] = 'JML_B'

            ws['BA1'] = 'Z_MAT_B'
            ws['BB1'] = 'Z_IND_B'
            ws['BC1'] = 'Z_ENG_B'
            ws['BD1'] = 'Z_SEJ_B'
            ws['BE1'] = 'Z_GEO_B'
            ws['BF1'] = 'Z_EKO_B'
            ws['BG1'] = 'Z_SOS_B'

            ws['BH1'] = 'S_MAT_B'
            ws['BI1'] = 'S_IND_B'
            ws['BJ1'] = 'S_ENG_B'
            ws['BK1'] = 'S_SEJ_B'
            ws['BL1'] = 'S_GEO_B'
            ws['BM1'] = 'S_EKO_B'
            ws['BN1'] = 'S_SOS_B'
            ws['BO1'] = 'S_JML_B'

            ws['BP1'] = 'RANK NAS._B'
            ws['BQ1'] = 'RANK LOK._B'

            ws['BA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BQ1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['AN1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AO1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AP1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AQ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AR1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AS1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AT1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AU1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AV1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AW1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AX1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AY1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AZ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BA1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BB1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BC1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BD1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BE1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BF1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BG1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BH1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BI1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BJ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BK1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BL1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BM1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BN1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BO1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BP1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BQ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')

            # tambahan
            ws['BR1'] = 'MAT_20_B'
            ws['BS1'] = 'IND_20_B'
            ws['BT1'] = 'ENG_20_B'
            ws['BU1'] = 'SEJ_20_B'
            ws['BV1'] = 'GEO_20_B'
            ws['BW1'] = 'EKO_20_B'
            ws['BX1'] = 'SOS_20_B'

            ws['BR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BX1'].font = Font(bold=False, name='Calibri', size=11)

            ws['BR1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BS1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BT1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BU1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BV1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BW1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BX1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')

            for row in range(2, q+1):
                ws['AN{}'.format(row)] = '=B{}'.format(row)
                ws['AO{}'.format(row)] = '=C{}'.format(row, row)
                ws['AP{}'.format(row)] = '=D{}'.format(row, row)
                ws['AQ{}'.format(row)] = '=E{}'.format(row, row)
                ws['AR{}'.format(row)] = '=F{}'.format(row, row)
                ws['AS{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['AT{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['AU{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['AV{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['AW{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['AX{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['AY{}'.format(row)] = '=IF(M{}="","",M{})'.format(row, row)
                ws['AZ{}'.format(row)] = '=IF(N{}="","",N{})'.format(row, row)
            # Z Ke mapel
                ws['BA{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AS{}="","",(AS{}-AS${})/AS${}),2),"")'.format(row, row, r, s)
                ws['BB{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AT{}="","",(AT{}-AT${})/AT${}),2),"")'.format(row, row, r, s)
                ws['BC{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AU{}="","",(AU{}-AU${})/AU${}),2),"")'.format(row, row, r, s)
                ws['BD{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AV{}="","",(AV{}-AV${})/AV${}),2),"")'.format(row, row, r, s)
                ws['BE{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AW{}="","",(AW{}-AW${})/AW${}),2),"")'.format(row, row, r, s)
                ws['BF{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AX{}="","",(AX{}-AX${})/AX${}),2),"")'.format(row, row, r, s)
                ws['BG{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AY{}="","",(AY{}-AY${})/AY${}),2),"")'.format(row, row, r, s)
            # NILAI STANDAR ke mapel dan Z score
                ws['BH{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AS{}="","",IF(70+30*BA{}/$BA${}<20,20,70+30*BA{}/$BA${})),2),"")'.format(row, row, r, row, r)
                ws['BI{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AT{}="","",IF(70+30*BB{}/$BB${}<20,20,70+30*BB{}/$BB${})),2),"")'.format(row, row, r, row, r)
                ws['BJ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AU{}="","",IF(70+30*BC{}/$BC${}<20,20,70+30*BC{}/$BC${})),2),"")'.format(row, row, r, row, r)
                ws['BK{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AV{}="","",IF(70+30*BD{}/$BD${}<20,20,70+30*BD{}/$BD${})),2),"")'.format(row, row, r, row, r)
                ws['BL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AW{}="","",IF(70+30*BE{}/$BE${}<20,20,70+30*BE{}/$BE${})),2),"")'.format(row, row, r, row, r)
                ws['BM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AX{}="","",IF(70+30*BF{}/$BF${}<20,20,70+30*BF{}/$BF${})),2),"")'.format(row, row, r, row, r)
                ws['BN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AY{}="","",IF(70+30*BG{}/$BG${}<20,20,70+30*BG{}/$BG${})),2),"")'.format(row, row, r, row, r)
            # JUMLAH SELURUH NILAI STANDAR
                ws['BO{}'.format(row)] = '=IF(SUM(BH{}:BN{})=0,"",SUM(BH{}:BN{}))'.format(
                    row, row, row, row)
                ws['BP{}'.format(row)] = '=IF(BO{}="","",RANK(BO{},$BO$2:$BO${}))'.format(
                    row, row, q)
                ws['BQ{}'.format(
                    row)] = '=IF(BP{}="","",COUNTIFS($AR$2:$AR${},AR{},$BP$2:$BP${},"<"&BP{})+1)'.format(row, q, row, q, row)
            # TAMBAHAN, MAPEL DAN NILAI STANDAR
                ws['BR{}'.format(row)] = '=IF($G${}=20,IF(AND(AS{}>3,BH{}=20),1,""),IF($G${}=25,IF(AND(AS{}>4,BH{}=20),1,""),IF($G${}=30,IF(AND(AS{}>5,BH{}=20),1,""),IF($G${}=35,IF(AND(AS{}>6,BH{}=20),1,""),IF($G${}=40,IF(AND(AS{}>7,BH{}=20),1,""),IF($G${}=45,IF(AND(AS{}>8,BH{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BS{}'.format(row)] = '=IF($H${}=20,IF(AND(AT{}>3,BI{}=20),1,""),IF($H${}=25,IF(AND(AT{}>4,BI{}=20),1,""),IF($H${}=30,IF(AND(AT{}>5,BI{}=20),1,""),IF($H${}=35,IF(AND(AT{}>6,BI{}=20),1,""),IF($H${}=40,IF(AND(AT{}>7,BI{}=20),1,""),IF($H${}=45,IF(AND(AT{}>8,BI{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BT{}'.format(row)] = '=IF($I${}=20,IF(AND(AU{}>3,BJ{}=20),1,""),IF($I${}=25,IF(AND(AU{}>4,BJ{}=20),1,""),IF($I${}=30,IF(AND(AU{}>5,BJ{}=20),1,""),IF($I${}=35,IF(AND(AU{}>6,BJ{}=20),1,""),IF($I${}=40,IF(AND(AU{}>7,BJ{}=20),1,""),IF($I${}=45,IF(AND(AU{}>8,BJ{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BU{}'.format(row)] = '=IF($J${}=20,IF(AND(AV{}>3,BK{}=20),1,""),IF($J${}=25,IF(AND(AV{}>4,BK{}=20),1,""),IF($J${}=30,IF(AND(AV{}>5,BK{}=20),1,""),IF($J${}=35,IF(AND(AV{}>6,BK{}=20),1,""),IF($J${}=40,IF(AND(AV{}>7,BK{}=20),1,""),IF($J${}=45,IF(AND(AV{}>8,BK{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BV{}'.format(row)] = '=IF($K${}=20,IF(AND(AW{}>3,BL{}=20),1,""),IF($K${}=25,IF(AND(AW{}>4,BL{}=20),1,""),IF($K${}=30,IF(AND(AW{}>5,BL{}=20),1,""),IF($K${}=35,IF(AND(AW{}>6,BL{}=20),1,""),IF($K${}=40,IF(AND(AW{}>7,BL{}=20),1,""),IF($K${}=45,IF(AND(AW{}>8,BL{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BW{}'.format(row)] = '=IF($L${}=20,IF(AND(AX{}>3,BM{}=20),1,""),IF($L${}=25,IF(AND(AX{}>4,BM{}=20),1,""),IF($L${}=30,IF(AND(AX{}>5,BM{}=20),1,""),IF($L${}=35,IF(AND(AX{}>6,BM{}=20),1,""),IF($L${}=40,IF(AND(AX{}>7,BM{}=20),1,""),IF($L${}=45,IF(AND(AX{}>8,BM{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BX{}'.format(row)] = '=IF($M${}=20,IF(AND(AY{}>3,BN{}=20),1,""),IF($M${}=25,IF(AND(AY{}>4,BN{}=20),1,""),IF($M${}=30,IF(AND(AY{}>5,BN{}=20),1,""),IF($M${}=35,IF(AND(AY{}>6,BN{}=20),1,""),IF($M${}=40,IF(AND(AY{}>7,BN{}=20),1,""),IF($M${}=45,IF(AND(AY{}>8,BN{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score [3]
            ws['BZ1'] = 'NAMA SISWA_C'
            ws['CA1'] = 'NOMOR NF_C'
            ws['CB1'] = 'KELAS_C'
            ws['CC1'] = 'NAMA SEKOLAH_C'
            ws['CD1'] = 'LOKASI_C'

            ws['CE1'] = 'MAT_C'
            ws['CF1'] = 'IND_C'
            ws['CG1'] = 'ENG_C'
            ws['CH1'] = 'SEJ_C'
            ws['CI1'] = 'GEO_C'
            ws['CJ1'] = 'EKO_C'
            ws['CK1'] = 'SOS_C'
            ws['CL1'] = 'JML_C'

            ws['CM1'] = 'Z_MAT_C'
            ws['CN1'] = 'Z_IND_C'
            ws['CO1'] = 'Z_ENG_C'
            ws['CP1'] = 'Z_SEJ_C'
            ws['CQ1'] = 'Z_GEO_C'
            ws['CR1'] = 'Z_EKO_C'
            ws['CS1'] = 'Z_SOS_C'

            ws['CT1'] = 'S_MAT_C'
            ws['CU1'] = 'S_IND_C'
            ws['CV1'] = 'S_ENG_C'
            ws['CW1'] = 'S_SEJ_C'
            ws['CX1'] = 'S_GEO_C'
            ws['CY1'] = 'S_EKO_C'
            ws['CZ1'] = 'S_SOS_C'
            ws['DA1'] = 'S_JML_C'

            ws['DB1'] = 'RANK NAS._C'
            ws['DC1'] = 'RANK LOK._C'

            ws['CM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DC1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['BZ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CA1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CB1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CC1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CD1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CE1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CF1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CG1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CH1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CI1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CJ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CK1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CL1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CM1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CN1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CO1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CP1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CQ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CR1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CS1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CT1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CU1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CV1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CW1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CX1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CY1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CZ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DA1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DB1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DC1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')

            # tambahan
            ws['DD1'] = 'MAT_20_C'
            ws['DE1'] = 'IND_20_C'
            ws['DF1'] = 'ENG_20_C'
            ws['DG1'] = 'SEJ_20_C'
            ws['DH1'] = 'GEO_20_C'
            ws['DI1'] = 'EKO_20_C'
            ws['DJ1'] = 'SOS_20_C'

            ws['DD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DJ1'].font = Font(bold=False, name='Calibri', size=11)

            ws['DD1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DE1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DF1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DG1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DH1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DI1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DJ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')

            for row in range(2, q+1):
                ws['BZ{}'.format(row)] = '=AN{}'.format(row)
                ws['CA{}'.format(row)] = '=AO{}'.format(row, row)
                ws['CB{}'.format(row)] = '=AP{}'.format(row, row)
                ws['CC{}'.format(row)] = '=AQ{}'.format(row, row)
                ws['CD{}'.format(row)] = '=AR{}'.format(row, row)
                ws['CE{}'.format(row)] = '=IF(AS{}="","",AS{})'.format(
                    row, row)
                ws['CF{}'.format(row)] = '=IF(AT{}="","",AT{})'.format(
                    row, row)
                ws['CG{}'.format(row)] = '=IF(AU{}="","",AU{})'.format(
                    row, row)
                ws['CH{}'.format(row)] = '=IF(AV{}="","",AV{})'.format(
                    row, row)
                ws['CI{}'.format(row)] = '=IF(AW{}="","",AW{})'.format(
                    row, row)
                ws['CJ{}'.format(row)] = '=IF(AX{}="","",AX{})'.format(
                    row, row)
                ws['CK{}'.format(row)] = '=IF(AY{}="","",AY{})'.format(
                    row, row)
                ws['CL{}'.format(row)] = '=IF(AZ{}="","",AZ{})'.format(
                    row, row)
            # Z Ke mapel
                ws['CM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CE{}="","",(CE{}-CE${})/CE${}),2),"")'.format(row, row, r, s)
                ws['CN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CF{}="","",(CF{}-CF${})/CF${}),2),"")'.format(row, row, r, s)
                ws['CO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CG{}="","",(CG{}-CG${})/CG${}),2),"")'.format(row, row, r, s)
                ws['CP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CH{}="","",(CH{}-CH${})/CH${}),2),"")'.format(row, row, r, s)
                ws['CQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CI{}="","",(CI{}-CI${})/CI${}),2),"")'.format(row, row, r, s)
                ws['CR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CJ{}="","",(CJ{}-CJ${})/CJ${}),2),"")'.format(row, row, r, s)
                ws['CS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CK{}="","",(CK{}-CK${})/CK${}),2),"")'.format(row, row, r, s)
            # NILAI STANDAR ke mapel dan Z score
                ws['CT{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CE{}="","",IF(70+30*CM{}/$CM${}<20,20,70+30*CM{}/$CM${})),2),"")'.format(row, row, r, row, r)
                ws['CU{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CF{}="","",IF(70+30*CN{}/$CN${}<20,20,70+30*CN{}/$CN${})),2),"")'.format(row, row, r, row, r)
                ws['CV{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CG{}="","",IF(70+30*CO{}/$CO${}<20,20,70+30*CO{}/$CO${})),2),"")'.format(row, row, r, row, r)
                ws['CW{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CH{}="","",IF(70+30*CP{}/$CP${}<20,20,70+30*CP{}/$CP${})),2),"")'.format(row, row, r, row, r)
                ws['CX{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CI{}="","",IF(70+30*CQ{}/$CQ${}<20,20,70+30*CQ{}/$CQ${})),2),"")'.format(row, row, r, row, r)
                ws['CY{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CJ{}="","",IF(70+30*CR{}/$CR${}<20,20,70+30*CR{}/$CR${})),2),"")'.format(row, row, r, row, r)
                ws['CZ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CK{}="","",IF(70+30*CS{}/$CS${}<20,20,70+30*CS{}/$CS${})),2),"")'.format(row, row, r, row, r)
            # JUMLAH SELURUH NILAI STANDAR
                ws['DA{}'.format(row)] = '=IF(SUM(CT{}:CZ{})=0,"",SUM(CT{}:CZ{}))'.format(
                    row, row, row, row)
                ws['DB{}'.format(row)] = '=IF(DA{}="","",RANK(DA{},$DA$2:$DA${}))'.format(
                    row, row, q)
                ws['DC{}'.format(
                    row)] = '=IF(DB{}="","",COUNTIFS($CD$2:$CD${},CD{},$DB$2:$DB${},"<"&DB{})+1)'.format(row, q, row, q, row)
            # TAMBAHAN, MAPEL DAN NILAI STANDAR
                ws['DD{}'.format(row)] = '=IF($G${}=20,IF(AND(CE{}>3,CT{}=20),1,""),IF($G${}=25,IF(AND(CE{}>4,CT{}=20),1,""),IF($G${}=30,IF(AND(CE{}>5,CT{}=20),1,""),IF($G${}=35,IF(AND(CE{}>6,CT{}=20),1,""),IF($G${}=40,IF(AND(CE{}>7,CT{}=20),1,""),IF($G${}=45,IF(AND(CE{}>8,CT{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DE{}'.format(row)] = '=IF($H${}=20,IF(AND(CF{}>3,CU{}=20),1,""),IF($H${}=25,IF(AND(CF{}>4,CU{}=20),1,""),IF($H${}=30,IF(AND(CF{}>5,CU{}=20),1,""),IF($H${}=35,IF(AND(CF{}>6,CU{}=20),1,""),IF($H${}=40,IF(AND(CF{}>7,CU{}=20),1,""),IF($H${}=45,IF(AND(CF{}>8,CU{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DF{}'.format(row)] = '=IF($I${}=20,IF(AND(CG{}>3,CV{}=20),1,""),IF($I${}=25,IF(AND(CG{}>4,CV{}=20),1,""),IF($I${}=30,IF(AND(CG{}>5,CV{}=20),1,""),IF($I${}=35,IF(AND(CG{}>6,CV{}=20),1,""),IF($I${}=40,IF(AND(CG{}>7,CV{}=20),1,""),IF($I${}=45,IF(AND(CG{}>8,CV{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DG{}'.format(row)] = '=IF($J${}=20,IF(AND(CH{}>3,CW{}=20),1,""),IF($J${}=25,IF(AND(CH{}>4,CW{}=20),1,""),IF($J${}=30,IF(AND(CH{}>5,CW{}=20),1,""),IF($J${}=35,IF(AND(CH{}>6,CW{}=20),1,""),IF($J${}=40,IF(AND(CH{}>7,CW{}=20),1,""),IF($J${}=45,IF(AND(CH{}>8,CW{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DH{}'.format(row)] = '=IF($K${}=20,IF(AND(CI{}>3,CX{}=20),1,""),IF($K${}=25,IF(AND(CI{}>4,CX{}=20),1,""),IF($K${}=30,IF(AND(CI{}>5,CX{}=20),1,""),IF($K${}=35,IF(AND(CI{}>6,CX{}=20),1,""),IF($K${}=40,IF(AND(CI{}>7,CX{}=20),1,""),IF($K${}=45,IF(AND(CI{}>8,CX{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DI{}'.format(row)] = '=IF($L${}=20,IF(AND(CJ{}>3,CY{}=20),1,""),IF($L${}=25,IF(AND(CJ{}>4,CY{}=20),1,""),IF($L${}=30,IF(AND(CJ{}>5,CY{}=20),1,""),IF($L${}=35,IF(AND(CJ{}>6,CY{}=20),1,""),IF($L${}=40,IF(AND(CJ{}>7,CY{}=20),1,""),IF($L${}=45,IF(AND(CJ{}>8,CY{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DJ{}'.format(row)] = '=IF($M${}=20,IF(AND(CK{}>3,CZ{}=20),1,""),IF($M${}=25,IF(AND(CK{}>4,CZ{}=20),1,""),IF($M${}=30,IF(AND(CK{}>5,CZ{}=20),1,""),IF($M${}=35,IF(AND(CK{}>6,CZ{}=20),1,""),IF($M${}=40,IF(AND(CK{}>7,CZ{}=20),1,""),IF($M${}=45,IF(AND(CK{}>8,CZ{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score [4]
            ws['DL1'] = 'NAMA SISWA_D'
            ws['DM1'] = 'NOMOR NF_D'
            ws['DN1'] = 'KELAS_D'
            ws['DO1'] = 'NAMA SEKOLAH_D'
            ws['DP1'] = 'LOKASI_D'

            ws['DQ1'] = 'MAT_D'
            ws['DR1'] = 'IND_D'
            ws['DS1'] = 'ENG_D'
            ws['DT1'] = 'SEJ_D'
            ws['DU1'] = 'GEO_D'
            ws['DV1'] = 'EKO_D'
            ws['DW1'] = 'SOS_D'
            ws['DX1'] = 'JML_D'

            ws['DY1'] = 'Z_MAT_D'
            ws['DZ1'] = 'Z_IND_D'
            ws['EA1'] = 'Z_ENG_D'
            ws['EB1'] = 'Z_SEJ_D'
            ws['EC1'] = 'Z_GEO_D'
            ws['ED1'] = 'Z_EKO_D'
            ws['EE1'] = 'Z_SOS_D'

            ws['EF1'] = 'S_MAT_D'
            ws['EG1'] = 'S_IND_D'
            ws['EH1'] = 'S_ENG_D'
            ws['EI1'] = 'S_SEJ_D'
            ws['EJ1'] = 'S_GEO_D'
            ws['EK1'] = 'S_EKO_D'
            ws['EL1'] = 'S_SOS_D'
            ws['EM1'] = 'S_JML_D'

            ws['EN1'] = 'RANK NAS._D'
            ws['EO1'] = 'RANK LOK._D'

            ws['DY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ED1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EO1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['DL1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DM1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DN1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DO1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DP1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DQ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DR1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DS1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DT1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DU1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DV1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DW1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DX1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DY1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DZ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EA1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EB1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EC1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['ED1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EE1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EF1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EG1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EH1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EI1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EJ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EK1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EL1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EM1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EN1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EO1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')

            # tambahan
            ws['EP1'] = 'MAT_20_D'
            ws['EQ1'] = 'IND_20_D'
            ws['ER1'] = 'ENG_20_D'
            ws['ES1'] = 'SEJ_20_D'
            ws['ET1'] = 'GEO_20_D'
            ws['EU1'] = 'EKO_20_D'
            ws['EV1'] = 'SOS_20_D'

            ws['EP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ER1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ES1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ET1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EV1'].font = Font(bold=False, name='Calibri', size=11)

            ws['EP1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EQ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['ER1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['ES1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['ET1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EU1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['EV1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')

            for row in range(2, q+1):
                ws['DL{}'.format(row)] = '=BZ{}'.format(row)
                ws['DM{}'.format(row)] = '=CA{}'.format(row, row)
                ws['DN{}'.format(row)] = '=CB{}'.format(row, row)
                ws['DO{}'.format(row)] = '=CC{}'.format(row, row)
                ws['DP{}'.format(row)] = '=CD{}'.format(row, row)
                ws['DQ{}'.format(row)] = '=IF(CE{}="","",CE{})'.format(
                    row, row)
                ws['DR{}'.format(row)] = '=IF(CF{}="","",CF{})'.format(
                    row, row)
                ws['DS{}'.format(row)] = '=IF(CG{}="","",CG{})'.format(
                    row, row)
                ws['DT{}'.format(row)] = '=IF(CH{}="","",CH{})'.format(
                    row, row)
                ws['DU{}'.format(row)] = '=IF(CI{}="","",CI{})'.format(
                    row, row)
                ws['DV{}'.format(row)] = '=IF(CJ{}="","",CJ{})'.format(
                    row, row)
                ws['DW{}'.format(row)] = '=IF(CK{}="","",CK{})'.format(
                    row, row)
                ws['DX{}'.format(row)] = '=IF(CL{}="","",CL{})'.format(
                    row, row)
            # Z Ke mapel
                ws['DY{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DQ{}="","",(DQ{}-DQ${})/DQ${}),2),"")'.format(row, row, r, s)
                ws['DZ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DR{}="","",(DR{}-DR${})/DR${}),2),"")'.format(row, row, r, s)
                ws['EA{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DS{}="","",(DS{}-DS${})/DS${}),2),"")'.format(row, row, r, s)
                ws['EB{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DT{}="","",(DT{}-DT${})/DT${}),2),"")'.format(row, row, r, s)
                ws['EC{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DU{}="","",(DU{}-DU${})/DU${}),2),"")'.format(row, row, r, s)
                ws['ED{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DV{}="","",(DV{}-DV${})/DV${}),2),"")'.format(row, row, r, s)
                ws['EE{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DW{}="","",(DW{}-DW${})/DW${}),2),"")'.format(row, row, r, s)
            # NILAI STANDAR ke mapel dan Z score
                ws['EF{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DQ{}="","",IF(70+30*DY{}/$DY${}<20,20,70+30*DY{}/$DY${})),2),"")'.format(row, row, r, row, r)
                ws['EG{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DR{}="","",IF(70+30*DZ{}/$DZ${}<20,20,70+30*DZ{}/$DZ${})),2),"")'.format(row, row, r, row, r)
                ws['EH{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DS{}="","",IF(70+30*EA{}/$EA${}<20,20,70+30*EA{}/$EA${})),2),"")'.format(row, row, r, row, r)
                ws['EI{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DT{}="","",IF(70+30*EB{}/$EB${}<20,20,70+30*EB{}/$EB${})),2),"")'.format(row, row, r, row, r)
                ws['EJ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DU{}="","",IF(70+30*EC{}/$EC${}<20,20,70+30*EC{}/$EC${})),2),"")'.format(row, row, r, row, r)
                ws['EK{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DV{}="","",IF(70+30*ED{}/$ED${}<20,20,70+30*ED{}/$ED${})),2),"")'.format(row, row, r, row, r)
                ws['EL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DW{}="","",IF(70+30*EE{}/$EE${}<20,20,70+30*EE{}/$EE${})),2),"")'.format(row, row, r, row, r)
            # JUMLAH SELURUH NILAI STANDAR
                ws['EM{}'.format(row)] = '=IF(SUM(EF{}:EL{})=0,"",SUM(EF{}:EL{}))'.format(
                    row, row, row, row)
                ws['EN{}'.format(row)] = '=IF(EM{}="","",RANK(EM{},$EM$2:$EM${}))'.format(
                    row, row, q)
                ws['EO{}'.format(
                    row)] = '=IF(EN{}="","",COUNTIFS($DP$2:$DP${},DP{},$EN$2:$EN${},"<"&EN{})+1)'.format(row, q, row, q, row)
            # TAMBAHAN, MAPEL DAN NILAI STANDAR
                ws['EP{}'.format(row)] = '=IF($G${}=20,IF(AND(DQ{}>3,EF{}=20),1,""),IF($G${}=25,IF(AND(DQ{}>4,EF{}=20),1,""),IF($G${}=30,IF(AND(DQ{}>5,EF{}=20),1,""),IF($G${}=35,IF(AND(DQ{}>6,EF{}=20),1,""),IF($G${}=40,IF(AND(DQ{}>7,EF{}=20),1,""),IF($G${}=45,IF(AND(DQ{}>8,EF{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['EQ{}'.format(row)] = '=IF($H${}=20,IF(AND(DR{}>3,EG{}=20),1,""),IF($H${}=25,IF(AND(DR{}>4,EG{}=20),1,""),IF($H${}=30,IF(AND(DR{}>5,EG{}=20),1,""),IF($H${}=35,IF(AND(DR{}>6,EG{}=20),1,""),IF($H${}=40,IF(AND(DR{}>7,EG{}=20),1,""),IF($H${}=45,IF(AND(DR{}>8,EG{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['ER{}'.format(row)] = '=IF($I${}=20,IF(AND(DS{}>3,EH{}=20),1,""),IF($I${}=25,IF(AND(DS{}>4,EH{}=20),1,""),IF($I${}=30,IF(AND(DS{}>5,EH{}=20),1,""),IF($I${}=35,IF(AND(DS{}>6,EH{}=20),1,""),IF($I${}=40,IF(AND(DS{}>7,EH{}=20),1,""),IF($I${}=45,IF(AND(DS{}>8,EH{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['ES{}'.format(row)] = '=IF($J${}=20,IF(AND(DT{}>3,EI{}=20),1,""),IF($J${}=25,IF(AND(DT{}>4,EI{}=20),1,""),IF($J${}=30,IF(AND(DT{}>5,EI{}=20),1,""),IF($J${}=35,IF(AND(DT{}>6,EI{}=20),1,""),IF($J${}=40,IF(AND(DT{}>7,EI{}=20),1,""),IF($J${}=45,IF(AND(DT{}>8,EI{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['ET{}'.format(row)] = '=IF($K${}=20,IF(AND(DU{}>3,EJ{}=20),1,""),IF($K${}=25,IF(AND(DU{}>4,EJ{}=20),1,""),IF($K${}=30,IF(AND(DU{}>5,EJ{}=20),1,""),IF($K${}=35,IF(AND(DU{}>6,EJ{}=20),1,""),IF($K${}=40,IF(AND(DU{}>7,EJ{}=20),1,""),IF($K${}=45,IF(AND(DU{}>8,EJ{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['EU{}'.format(row)] = '=IF($L${}=20,IF(AND(DV{}>3,EK{}=20),1,""),IF($L${}=25,IF(AND(DV{}>4,EK{}=20),1,""),IF($L${}=30,IF(AND(DV{}>5,EK{}=20),1,""),IF($L${}=35,IF(AND(DV{}>6,EK{}=20),1,""),IF($L${}=40,IF(AND(DV{}>7,EK{}=20),1,""),IF($L${}=45,IF(AND(DV{}>8,EK{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['EV{}'.format(row)] = '=IF($M${}=20,IF(AND(DW{}>3,EL{}=20),1,""),IF($M${}=25,IF(AND(DW{}>4,EL{}=20),1,""),IF($M${}=30,IF(AND(DW{}>5,EL{}=20),1,""),IF($M${}=35,IF(AND(DW{}>6,EL{}=20),1,""),IF($M${}=40,IF(AND(DW{}>7,EL{}=20),1,""),IF($M${}=45,IF(AND(DW{}>8,EL{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score [5]
            ws['EX1'] = 'NAMA SISWA_E'
            ws['EY1'] = 'NOMOR NF_E'
            ws['EZ1'] = 'KELAS_E'
            ws['FA1'] = 'NAMA SEKOLAH_E'
            ws['FB1'] = 'LOKASI_E'

            ws['FC1'] = 'MAT_E'
            ws['FD1'] = 'IND_E'
            ws['FE1'] = 'ENG_E'
            ws['FF1'] = 'SEJ_E'
            ws['FG1'] = 'GEO_E'
            ws['FH1'] = 'EKO_E'
            ws['FI1'] = 'SOS_E'
            ws['FJ1'] = 'JML_E'

            ws['FK1'] = 'Z_MAT_E'
            ws['FL1'] = 'Z_IND_E'
            ws['FM1'] = 'Z_ENG_E'
            ws['FN1'] = 'Z_SEJ_E'
            ws['FO1'] = 'Z_GEO_E'
            ws['FP1'] = 'Z_EKO_E'
            ws['FQ1'] = 'Z_SOS_E'

            ws['FR1'] = 'S_MAT_E'
            ws['FS1'] = 'S_IND_E'
            ws['FT1'] = 'S_ENG_E'
            ws['FU1'] = 'S_SEJ_E'
            ws['FV1'] = 'S_GEO_E'
            ws['FW1'] = 'S_EKO_E'
            ws['FX1'] = 'S_SOS_E'
            ws['FY1'] = 'S_JML_E'

            ws['FZ1'] = 'RANK NAS._E'
            ws['GA1'] = 'RANK LOK._E'

            ws['FK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GA1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['EX1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EY1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EZ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FA1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FB1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FC1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FD1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FE1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FF1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FG1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FH1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FI1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FJ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FK1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FL1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FM1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FN1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FO1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FP1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FQ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FR1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FS1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FT1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FU1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FV1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FW1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FX1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FY1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['FZ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['GA1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')

            # tambahan
            ws['GB1'] = 'MAT_20_E'
            ws['GC1'] = 'IND_20_E'
            ws['GD1'] = 'ENG_20_E'
            ws['GE1'] = 'SEJ_20_E'
            ws['GF1'] = 'GEO_20_E'
            ws['GG1'] = 'EKO_20_E'
            ws['GH1'] = 'SOS_20_E'

            ws['GB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GH1'].font = Font(bold=False, name='Calibri', size=11)

            ws['GB1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['GC1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['GD1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['GE1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['GF1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['GG1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['GH1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')

            for row in range(2, q+1):
                ws['EX{}'.format(row)] = '=DL{}'.format(row)
                ws['EY{}'.format(row)] = '=DM{}'.format(row, row)
                ws['EZ{}'.format(row)] = '=DN{}'.format(row, row)
                ws['FA{}'.format(row)] = '=DO{}'.format(row, row)
                ws['FB{}'.format(row)] = '=DP{}'.format(row, row)
                ws['FC{}'.format(row)] = '=IF(DQ{}="","",DQ{})'.format(
                    row, row)
                ws['FD{}'.format(row)] = '=IF(DR{}="","",DR{})'.format(
                    row, row)
                ws['FE{}'.format(row)] = '=IF(DS{}="","",DS{})'.format(
                    row, row)
                ws['FF{}'.format(row)] = '=IF(DT{}="","",DT{})'.format(
                    row, row)
                ws['FG{}'.format(row)] = '=IF(DU{}="","",DU{})'.format(
                    row, row)
                ws['FH{}'.format(row)] = '=IF(DV{}="","",DV{})'.format(
                    row, row)
                ws['FI{}'.format(row)] = '=IF(DW{}="","",DW{})'.format(
                    row, row)
                ws['FJ{}'.format(row)] = '=IF(DX{}="","",DX{})'.format(
                    row, row)
            # Z Ke mapel
                ws['FK{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FC{}="","",(FC{}-FC${})/FC${}),2),"")'.format(row, row, r, s)
                ws['FL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FD{}="","",(FD{}-FD${})/FD${}),2),"")'.format(row, row, r, s)
                ws['FM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FE{}="","",(FE{}-FE${})/FE${}),2),"")'.format(row, row, r, s)
                ws['FN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FF{}="","",(FF{}-FF${})/FF${}),2),"")'.format(row, row, r, s)
                ws['FO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FG{}="","",(FG{}-FG${})/FG${}),2),"")'.format(row, row, r, s)
                ws['FP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FH{}="","",(FH{}-FH${})/FH${}),2),"")'.format(row, row, r, s)
                ws['FQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FI{}="","",(FI{}-FI${})/FI${}),2),"")'.format(row, row, r, s)
            # NILAI STANDAR ke mapel dan Z score
                ws['FR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FC{}="","",IF(70+30*FK{}/$FK${}<20,20,70+30*FK{}/$FK${})),2),"")'.format(row, row, r, row, r)
                ws['FS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FD{}="","",IF(70+30*FL{}/$FL${}<20,20,70+30*FL{}/$FL${})),2),"")'.format(row, row, r, row, r)
                ws['FT{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FE{}="","",IF(70+30*FM{}/$FM${}<20,20,70+30*FM{}/$FM${})),2),"")'.format(row, row, r, row, r)
                ws['FU{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FF{}="","",IF(70+30*FN{}/$FN${}<20,20,70+30*FN{}/$FN${})),2),"")'.format(row, row, r, row, r)
                ws['FV{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FG{}="","",IF(70+30*FO{}/$FO${}<20,20,70+30*FO{}/$FO${})),2),"")'.format(row, row, r, row, r)
                ws['FW{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FH{}="","",IF(70+30*FP{}/$FP${}<20,20,70+30*FP{}/$FP${})),2),"")'.format(row, row, r, row, r)
                ws['FX{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FI{}="","",IF(70+30*FQ{}/$FQ${}<20,20,70+30*FQ{}/$FQ${})),2),"")'.format(row, row, r, row, r)
            # JUMLAH SELURUH NILAI STANDAR
                ws['FY{}'.format(row)] = '=IF(SUM(FR{}:FX{})=0,"",SUM(FR{}:FX{}))'.format(
                    row, row, row, row)
                ws['FZ{}'.format(row)] = '=IF(FY{}="","",RANK(FY{},$FY$2:$FY${}))'.format(
                    row, row, q)
                ws['GA{}'.format(
                    row)] = '=IF(FZ{}="","",COUNTIFS($FB$2:$FB${},FB{},$FZ$2:$FZ${},"<"&FZ{})+1)'.format(row, q, row, q, row)
            # TAMBAHAN, MAPEL DAN NILAI STANDAR
                ws['GB{}'.format(row)] = '=IF($G${}=20,IF(AND(FC{}>3,FR{}=20),1,""),IF($G${}=25,IF(AND(FC{}>4,FR{}=20),1,""),IF($G${}=30,IF(AND(FC{}>5,FR{}=20),1,""),IF($G${}=35,IF(AND(FC{}>6,FR{}=20),1,""),IF($G${}=40,IF(AND(FC{}>7,FR{}=20),1,""),IF($G${}=45,IF(AND(FC{}>8,FR{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['GC{}'.format(row)] = '=IF($H${}=20,IF(AND(FD{}>3,FS{}=20),1,""),IF($H${}=25,IF(AND(FD{}>4,FS{}=20),1,""),IF($H${}=30,IF(AND(FD{}>5,FS{}=20),1,""),IF($H${}=35,IF(AND(FD{}>6,FS{}=20),1,""),IF($H${}=40,IF(AND(FD{}>7,FS{}=20),1,""),IF($H${}=45,IF(AND(FD{}>8,FS{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['GD{}'.format(row)] = '=IF($I${}=20,IF(AND(FE{}>3,FT{}=20),1,""),IF($I${}=25,IF(AND(FE{}>4,FT{}=20),1,""),IF($I${}=30,IF(AND(FE{}>5,FT{}=20),1,""),IF($I${}=35,IF(AND(FE{}>6,FT{}=20),1,""),IF($I${}=40,IF(AND(FE{}>7,FT{}=20),1,""),IF($I${}=45,IF(AND(FE{}>8,FT{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['GE{}'.format(row)] = '=IF($J${}=20,IF(AND(FF{}>3,FU{}=20),1,""),IF($J${}=25,IF(AND(FF{}>4,FU{}=20),1,""),IF($J${}=30,IF(AND(FF{}>5,FU{}=20),1,""),IF($J${}=35,IF(AND(FF{}>6,FU{}=20),1,""),IF($J${}=40,IF(AND(FF{}>7,FU{}=20),1,""),IF($J${}=45,IF(AND(FF{}>8,FU{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['GF{}'.format(row)] = '=IF($K${}=20,IF(AND(FG{}>3,FV{}=20),1,""),IF($K${}=25,IF(AND(FG{}>4,FV{}=20),1,""),IF($K${}=30,IF(AND(FG{}>5,FV{}=20),1,""),IF($K${}=35,IF(AND(FG{}>6,FV{}=20),1,""),IF($K${}=40,IF(AND(FG{}>7,FV{}=20),1,""),IF($K${}=45,IF(AND(FG{}>8,FV{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['GG{}'.format(row)] = '=IF($L${}=20,IF(AND(FH{}>3,FW{}=20),1,""),IF($L${}=25,IF(AND(FH{}>4,FW{}=20),1,""),IF($L${}=30,IF(AND(FH{}>5,FW{}=20),1,""),IF($L${}=35,IF(AND(FH{}>6,FW{}=20),1,""),IF($L${}=40,IF(AND(FH{}>7,FW{}=20),1,""),IF($L${}=45,IF(AND(FH{}>8,FW{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['GH{}'.format(row)] = '=IF($M${}=20,IF(AND(FI{}>3,FX{}=20),1,""),IF($M${}=25,IF(AND(FI{}>4,FX{}=20),1,""),IF($M${}=30,IF(AND(FI{}>5,FX{}=20),1,""),IF($M${}=35,IF(AND(FI{}>6,FX{}=20),1,""),IF($M${}=40,IF(AND(FI{}>7,FX{}=20),1,""),IF($M${}=45,IF(AND(FI{}>8,FX{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score
            ws['GJ1'] = 'NAMA SISWA'
            ws['GK1'] = 'NOMOR NF'
            ws['GL1'] = 'KELAS'
            ws['GM1'] = 'NAMA SEKOLAH'
            ws['GN1'] = 'LOKASI'

            ws['GO1'] = 'MAT'
            ws['GP1'] = 'IND'
            ws['GQ1'] = 'ENG'
            ws['GR1'] = 'SEJ'
            ws['GS1'] = 'GEO'
            ws['GT1'] = 'EKO'
            ws['GU1'] = 'SOS'
            ws['GV1'] = 'JML'

            ws['GW1'] = 'Z_MAT'
            ws['GX1'] = 'Z_IND'
            ws['GY1'] = 'Z_ENG'
            ws['GZ1'] = 'Z_SEJ'
            ws['HA1'] = 'Z_GEO'
            ws['HB1'] = 'Z_EKO'
            ws['HC1'] = 'Z_SOS'

            ws['HD1'] = 'S_MAT'
            ws['HE1'] = 'S_IND'
            ws['HF1'] = 'S_ENG'
            ws['HG1'] = 'S_SEJ'
            ws['HH1'] = 'S_GEO'
            ws['HI1'] = 'S_EKO'
            ws['HJ1'] = 'S_SOS'
            ws['HK1'] = 'S_JML'

            ws['HL1'] = 'RANK NAS.'
            ws['HM1'] = 'RANK LOK.'

            ws['GW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HM1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['GJ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GK1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GL1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GM1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GN1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GO1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GP1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GQ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GR1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GS1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GT1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GU1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GV1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GW1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GX1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GY1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['GZ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HA1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HB1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HC1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HD1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HE1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HF1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HG1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HH1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HI1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HJ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HK1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HL1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HM1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')

            # tambahan
            ws['HN1'] = 'MAT_20'
            ws['HO1'] = 'IND_20'
            ws['HP1'] = 'ENG_20'
            ws['HQ1'] = 'SEJ_20'
            ws['HR1'] = 'GEO_20'
            ws['HS1'] = 'EKO_20'
            ws['HT1'] = 'SOS_20'

            ws['HN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HT1'].font = Font(bold=False, name='Calibri', size=11)

            ws['HN1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HO1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HP1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HQ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HR1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HS1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['HT1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')

            for row in range(2, q+1):
                ws['GJ{}'.format(row)] = '=EX{}'.format(row)
                ws['GK{}'.format(row)] = '=EY{}'.format(row, row)
                ws['GL{}'.format(row)] = '=EZ{}'.format(row, row)
                ws['GM{}'.format(row)] = '=FA{}'.format(row, row)
                ws['GN{}'.format(row)] = '=FB{}'.format(row, row)
                ws['GO{}'.format(row)] = '=IF(FC{}="","",FC{})'.format(
                    row, row)
                ws['GP{}'.format(row)] = '=IF(FD{}="","",FD{})'.format(
                    row, row)
                ws['GQ{}'.format(row)] = '=IF(FE{}="","",FE{})'.format(
                    row, row)
                ws['GR{}'.format(row)] = '=IF(FF{}="","",FF{})'.format(
                    row, row)
                ws['GS{}'.format(row)] = '=IF(FG{}="","",FG{})'.format(
                    row, row)
                ws['GT{}'.format(row)] = '=IF(FH{}="","",FH{})'.format(
                    row, row)
                ws['GU{}'.format(row)] = '=IF(FI{}="","",FI{})'.format(
                    row, row)
                ws['GV{}'.format(row)] = '=IF(FJ{}="","",FJ{})'.format(
                    row, row)
            # Z Ke mapel
                ws['GW{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GO{}="","",(GO{}-GO${})/GO${}),2),"")'.format(row, row, r, s)
                ws['GX{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GP{}="","",(GP{}-GP${})/GP${}),2),"")'.format(row, row, r, s)
                ws['GY{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GQ{}="","",(GQ{}-GQ${})/GQ${}),2),"")'.format(row, row, r, s)
                ws['GZ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GR{}="","",(GR{}-GR${})/GR${}),2),"")'.format(row, row, r, s)
                ws['HA{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GS{}="","",(GS{}-GS${})/GS${}),2),"")'.format(row, row, r, s)
                ws['HB{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GT{}="","",(GT{}-GT${})/GT${}),2),"")'.format(row, row, r, s)
                ws['HC{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GU{}="","",(GU{}-GU${})/GU${}),2),"")'.format(row, row, r, s)
            # NILAI STANDAR ke mapel dan Z score
                ws['HD{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GO{}="","",IF(70+30*GW{}/$GW${}<20,20,70+30*GW{}/$GW${})),2),"")'.format(row, row, r, row, r)
                ws['HE{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GP{}="","",IF(70+30*GX{}/$GX${}<20,20,70+30*GX{}/$GX${})),2),"")'.format(row, row, r, row, r)
                ws['HF{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GQ{}="","",IF(70+30*GY{}/$GY${}<20,20,70+30*GY{}/$GY${})),2),"")'.format(row, row, r, row, r)
                ws['HG{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GR{}="","",IF(70+30*GZ{}/$GZ${}<20,20,70+30*GZ{}/$GZ${})),2),"")'.format(row, row, r, row, r)
                ws['HH{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GS{}="","",IF(70+30*HA{}/$HA${}<20,20,70+30*HA{}/$HA${})),2),"")'.format(row, row, r, row, r)
                ws['HI{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GT{}="","",IF(70+30*HB{}/$HB${}<20,20,70+30*HB{}/$HB${})),2),"")'.format(row, row, r, row, r)
                ws['HJ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(GU{}="","",IF(70+30*HC{}/$HC${}<20,20,70+30*HC{}/$HC${})),2),"")'.format(row, row, r, row, r)
            # JUMLAH SELURUH NILAI STANDAR
                ws['HK{}'.format(row)] = '=IF(SUM(HD{}:HJ{})=0,"",SUM(HD{}:HJ{}))'.format(
                    row, row, row, row)
                ws['HL{}'.format(row)] = '=IF(HK{}="","",RANK(HK{},$HK$2:$HK${}))'.format(
                    row, row, q)
                ws['HM{}'.format(
                    row)] = '=IF(HL{}="","",COUNTIFS($GN$2:$GN${},GN{},$HL$2:$HL${},"<"&HL{})+1)'.format(row, q, row, q, row)
            # TAMBAHAN, MAPEL DAN NILAI STANDAR
                ws['HN{}'.format(row)] = '=IF($G${}=20,IF(AND(GO{}>3,HD{}=20),1,""),IF($G${}=25,IF(AND(GO{}>4,HD{}=20),1,""),IF($G${}=30,IF(AND(GO{}>5,HD{}=20),1,""),IF($G${}=35,IF(AND(GO{}>6,HD{}=20),1,""),IF($G${}=40,IF(AND(GO{}>7,HD{}=20),1,""),IF($G${}=45,IF(AND(GO{}>8,HD{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['HO{}'.format(row)] = '=IF($H${}=20,IF(AND(GP{}>3,HE{}=20),1,""),IF($H${}=25,IF(AND(GP{}>4,HE{}=20),1,""),IF($H${}=30,IF(AND(GP{}>5,HE{}=20),1,""),IF($H${}=35,IF(AND(GP{}>6,HE{}=20),1,""),IF($H${}=40,IF(AND(GP{}>7,HE{}=20),1,""),IF($H${}=45,IF(AND(GP{}>8,HE{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['HP{}'.format(row)] = '=IF($I${}=20,IF(AND(GQ{}>3,HF{}=20),1,""),IF($I${}=25,IF(AND(GQ{}>4,HF{}=20),1,""),IF($I${}=30,IF(AND(GQ{}>5,HF{}=20),1,""),IF($I${}=35,IF(AND(GQ{}>6,HF{}=20),1,""),IF($I${}=40,IF(AND(GQ{}>7,HF{}=20),1,""),IF($I${}=45,IF(AND(GQ{}>8,HF{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['HQ{}'.format(row)] = '=IF($J${}=20,IF(AND(GR{}>3,HG{}=20),1,""),IF($J${}=25,IF(AND(GR{}>4,HG{}=20),1,""),IF($J${}=30,IF(AND(GR{}>5,HG{}=20),1,""),IF($J${}=35,IF(AND(GR{}>6,HG{}=20),1,""),IF($J${}=40,IF(AND(GR{}>7,HG{}=20),1,""),IF($J${}=45,IF(AND(GR{}>8,HG{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['HR{}'.format(row)] = '=IF($K${}=20,IF(AND(GS{}>3,HH{}=20),1,""),IF($K${}=25,IF(AND(GS{}>4,HH{}=20),1,""),IF($K${}=30,IF(AND(GS{}>5,HH{}=20),1,""),IF($K${}=35,IF(AND(GS{}>6,HH{}=20),1,""),IF($K${}=40,IF(AND(GS{}>7,HH{}=20),1,""),IF($K${}=45,IF(AND(GS{}>8,HH{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['HS{}'.format(row)] = '=IF($L${}=20,IF(AND(GT{}>3,HI{}=20),1,""),IF($L${}=25,IF(AND(GT{}>4,HI{}=20),1,""),IF($L${}=30,IF(AND(GT{}>5,HI{}=20),1,""),IF($L${}=35,IF(AND(GT{}>6,HI{}=20),1,""),IF($L${}=40,IF(AND(GT{}>7,HI{}=20),1,""),IF($L${}=45,IF(AND(GT{}>8,HI{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['HT{}'.format(row)] = '=IF($M${}=20,IF(AND(GU{}>3,HJ{}=20),1,""),IF($M${}=25,IF(AND(GU{}>4,HJ{}=20),1,""),IF($M${}=30,IF(AND(GU{}>5,HJ{}=20),1,""),IF($M${}=35,IF(AND(GU{}>6,HJ{}=20),1,""),IF($M${}=40,IF(AND(GU{}>7,HJ{}=20),1,""),IF($M${}=45,IF(AND(GU{}>8,HJ{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Mengubah 'KELAS' sesuai dengan nilai yang dipilih dari selectbox 'KELAS'
            kelas = KELAS.lower().replace(" ", "")
            semester = SEMESTER.lower()
            tahun = TAHUN.replace("-", "")
            penilaian = PENILAIAN.lower()
            kurikulum = KURIKULUM.lower()

            path_file = f"{kelas}_{penilaian}_{semester}_{kurikulum}_{tahun}_nilai_std.xlsx"

            # Simpan file ke direktori temporer
            temp_dir = tempfile.gettempdir()
            file_path = temp_dir + '/' + path_file
            wb.save(file_path)

            st.success(
                "File siap diunduh!")

            # Tombol unduh file
            with open(file_path, "rb") as f:
                bytes_data = f.read()
            st.download_button(label="Unduh File", data=bytes_data,
                               file_name=path_file)

            st.warning(
                "Buka file unduhan, klik 'Enable Editing' dan 'Save'")
    if selected_file == "Nilai Std. PPLS, RONIN IPS":
        # menghilangkan hamburger
        st.markdown("""
        <style>
        .css-1rs6os.edgvbvh3
        {
            visibility:hidden;
        }
        .css-1lsmgbg.egzxvld0
        {
            visibility:hidden;
        }
        </style>
        """, unsafe_allow_html=True)

        image = Image.open('logo resmi nf resize.png')
        st.image(image)

        st.title("Olah Nilai Standar")

        # url = 'https://docs.google.com/document/d/1xjkgcq86pMfLieqwBGTmV0kB6mWO_R1L7042razShlk/edit?usp=sharing'

        # st.warning("Harap dibaca terlebih dahulu panduannya")
        # if st.button("Panduan"):
        #     webbrowser.open_new_tab(url)

        st.header("PPLS-RONIN IPS")

        col6 = st.container()

        with col6:
            KELAS = st.selectbox(
                "KELAS",
                ("--Pilih Kelas--", "PPLS IPS", "RONIN IPS"))

        col7 = st.container()

        with col7:
            SEMESTER = st.selectbox(
                "SEMESTER",
                ("--Pilih Semester--", "SEMESTER 1", "SEMESTER 2"))

        col8 = st.container()

        with col8:
            PENILAIAN = st.selectbox(
                "PENILAIAN",
                ("--Pilih Penilaian--", "PENILAIAN TENGAH SEMESTER", "PENILAIAN AKHIR SEMESTER", "PENILAIAN AKHIR TAHUN", "TES EVALUASI"))

        TAHUN = st.text_input("Masukkan Tahun Ajaran",
                              placeholder="contoh: 2022-2023")

        col1, col2, col3, col4, col5 = st.columns(5)

        with col1:
            MTK = st.selectbox(
                "JML. SOAL MAT.",
                ("--Pilih--", 25, 30, 35, 40, 45))

        with col2:
            SEJ = st.selectbox(
                "JML. SOAL SEJ.",
                ("--Pilih--", 25, 30, 35, 40, 45))

        with col3:
            GEO = st.selectbox(
                "JML. SOAL GEO.",
                ("--Pilih--", 25, 30, 35, 40, 45))

        with col4:
            EKO = st.selectbox(
                "JML. SOAL EKO.",
                ("--Pilih--", 25, 30, 35, 40, 45))

        with col5:
            SOS = st.selectbox(
                "JML. SOAL SOS.",
                ("--Pilih--", 25, 30, 35, 40, 45))

        JML_SOAL_MAT = MTK
        JML_SOAL_SEJ = SEJ
        JML_SOAL_GEO = GEO
        JML_SOAL_EKO = EKO
        JML_SOAL_SOS = SOS

        uploaded_file = st.file_uploader(
            'Letakkan file excel', type='xlsx')

        if uploaded_file is not None:

            wb = openpyxl.load_workbook(uploaded_file)
            ws = wb['Sheet1']

            q = len(ws['K'])
            r = len(ws['K'])+2
            s = len(ws['K'])+3
            t = len(ws['K'])+4
            u = len(ws['K'])+5
            v = len(ws['K'])+6
            w = len(ws['K'])+7
            x = len(ws['K'])+8

            ws['G{}'.format(r)] = "=ROUND(AVERAGE(G2:G{}),2)".format(q)
            ws['H{}'.format(r)] = "=ROUND(AVERAGE(H2:H{}),2)".format(q)
            ws['I{}'.format(r)] = "=ROUND(AVERAGE(I2:I{}),2)".format(q)
            ws['J{}'.format(r)] = "=ROUND(AVERAGE(J2:J{}),2)".format(q)
            ws['K{}'.format(r)] = "=ROUND(AVERAGE(K2:K{}),2)".format(q)
            ws['L{}'.format(r)] = "=ROUND(AVERAGE(L2:L{}),2)".format(q)
            ws['G{}'.format(s)] = "=STDEV(G2:G{})".format(q)
            ws['H{}'.format(s)] = "=STDEV(H2:H{})".format(q)
            ws['I{}'.format(s)] = "=STDEV(I2:I{})".format(q)
            ws['J{}'.format(s)] = "=STDEV(J2:J{})".format(q)
            ws['K{}'.format(s)] = "=STDEV(K2:K{})".format(q)
            ws['G{}'.format(t)] = "=MAX(G2:G{})".format(q)
            ws['H{}'.format(t)] = "=MAX(H2:H{})".format(q)
            ws['I{}'.format(t)] = "=MAX(I2:I{})".format(q)
            ws['J{}'.format(t)] = "=MAX(J2:J{})".format(q)
            ws['K{}'.format(t)] = "=MAX(K2:K{})".format(q)
            ws['L{}'.format(t)] = "=MAX(L2:L{})".format(q)
            ws['M{}'.format(r)] = "=MAX(M2:M{})".format(q)
            ws['N{}'.format(r)] = "=MAX(N2:N{})".format(q)
            ws['O{}'.format(r)] = "=MAX(O2:O{})".format(q)
            ws['P{}'.format(r)] = "=MAX(P2:P{})".format(q)
            ws['Q{}'.format(r)] = "=MAX(Q2:Q{})".format(q)
            ws['R{}'.format(r)] = "=MAX(R2:R{})".format(q)
            ws['S{}'.format(r)] = "=MAX(S2:S{})".format(q)
            ws['T{}'.format(r)] = "=MAX(T2:T{})".format(q)
            ws['U{}'.format(r)] = "=MAX(U2:U{})".format(q)
            ws['V{}'.format(r)] = "=MAX(V2:V{})".format(q)
            ws['W{}'.format(r)] = "=ROUND(MAX(W2:W{}),2)".format(q)
            ws['G{}'.format(u)] = "=MIN(G2:G{})".format(q)
            ws['H{}'.format(u)] = "=MIN(H2:H{})".format(q)
            ws['I{}'.format(u)] = "=MIN(I2:I{})".format(q)
            ws['J{}'.format(u)] = "=MIN(J2:J{})".format(q)
            ws['K{}'.format(u)] = "=MIN(K2:K{})".format(q)
            ws['L{}'.format(u)] = "=MIN(L2:L{})".format(q)
            ws['R{}'.format(s)] = "=MIN(R2:R{})".format(q)
            ws['S{}'.format(s)] = "=MIN(S2:S{})".format(q)
            ws['T{}'.format(s)] = "=MIN(T2:T{})".format(q)
            ws['U{}'.format(s)] = "=MIN(U2:U{})".format(q)
            ws['V{}'.format(s)] = "=MIN(V2:V{})".format(q)
            ws['W{}'.format(s)] = "=MIN(W2:W{})".format(q)
            ws['R{}'.format(t)] = "=ROUND(AVERAGE(R2:R{}),2)".format(q)
            ws['S{}'.format(t)] = "=ROUND(AVERAGE(S2:S{}),2)".format(q)
            ws['T{}'.format(t)] = "=ROUND(AVERAGE(T2:T{}),2)".format(q)
            ws['U{}'.format(t)] = "=ROUND(AVERAGE(U2:U{}),2)".format(q)
            ws['V{}'.format(t)] = "=ROUND(AVERAGE(V2:V{}),2)".format(q)
            ws['W{}'.format(t)] = "=ROUND(AVERAGE(W2:W{}),2)".format(q)
            ws['X{}'.format(r)] = "=MAX(X2:X{})".format(q)
            ws['Z{}'.format(r)] = "=SUM(Z2:Z{})".format(q)
            ws['AA{}'.format(r)] = "=SUM(AA2:AA{})".format(q)
            ws['AB{}'.format(r)] = "=SUM(AB2:AB{})".format(q)
            ws['AC{}'.format(r)] = "=SUM(AC2:AC{})".format(q)
            ws['AD{}'.format(r)] = "=SUM(AD2:AD{})".format(q)
            # new
            # iterasi 1 rata-rata - 1
            ws['F{}'.format(v)] = 'JUMLAH SOAL'
            ws['G{}'.format(v)] = JML_SOAL_MAT
            ws['H{}'.format(v)] = JML_SOAL_SEJ
            ws['I{}'.format(v)] = JML_SOAL_GEO
            ws['J{}'.format(v)] = JML_SOAL_EKO
            ws['K{}'.format(v)] = JML_SOAL_SOS
            ws['AK{}'.format(r)] = "=IF($Z${}=0,$G${},$G${}-1)".format(r, r, r)
            ws['AK{}'.format(s)] = "=STDEV(AK2:AK{})".format(q)
            ws['AK{}'.format(t)] = "=MAX(AK2:AK{})".format(q)
            ws['AK{}'.format(u)] = "=MIN(AK2:AK{})".format(q)
            ws['AL{}'.format(
                r)] = "=IF($AA${}=0,$H${},$H${}-1)".format(r, r, r)
            ws['AL{}'.format(s)] = "=STDEV(AL2:AL{})".format(q)
            ws['AL{}'.format(t)] = "=MAX(AL2:AL{})".format(q)
            ws['AL{}'.format(u)] = "=MIN(AL2:AL{})".format(q)
            ws['AM{}'.format(
                r)] = "=IF($AB${}=0,$I${},$I${}-1)".format(r, r, r)
            ws['AM{}'.format(s)] = "=STDEV(AM2:AM{})".format(q)
            ws['AM{}'.format(t)] = "=MAX(AM2:AM{})".format(q)
            ws['AM{}'.format(u)] = "=MIN(AM2:AM{})".format(q)
            ws['AN{}'.format(
                r)] = "=IF($AC${}=0,$J${},$J${}-1)".format(r, r, r)
            ws['AN{}'.format(s)] = "=STDEV(AN2:AN{})".format(q)
            ws['AN{}'.format(t)] = "=MAX(AN2:AN{})".format(q)
            ws['AN{}'.format(u)] = "=MIN(AN2:AN{})".format(q)
            ws['AO{}'.format(
                r)] = "=IF($AD${}=0,$K${},$K${}-1)".format(r, r, r)
            ws['AO{}'.format(s)] = "=STDEV(AO2:AO{})".format(q)
            ws['AO{}'.format(t)] = "=MAX(AO2:AO{})".format(q)
            ws['AO{}'.format(u)] = "=MIN(AO2:AO{})".format(q)
            ws['AP{}'.format(r)] = "=ROUND(AVERAGE(AP2:AP{}),2)".format(q)
            ws['AP{}'.format(t)] = "=MAX(AP2:AP{})".format(q)
            ws['AP{}'.format(u)] = "=MIN(AP2:AP{})".format(q)
            ws['AQ{}'.format(r)] = "=MAX(AQ2:AQ{})".format(q)
            ws['AR{}'.format(r)] = "=MAX(AR2:AR{})".format(q)
            ws['AS{}'.format(r)] = "=MAX(AS2:AS{})".format(q)
            ws['AT{}'.format(r)] = "=MAX(AT2:AT{})".format(q)
            ws['AU{}'.format(r)] = "=MAX(AU2:AU{})".format(q)
            ws['AV{}'.format(r)] = "=MAX(AV2:AV{})".format(q)
            ws['AV{}'.format(s)] = "=MIN(AV2:AV{})".format(q)
            ws['AV{}'.format(t)] = "=ROUND(AVERAGE(AV2:AV{}),2)".format(q)
            ws['AW{}'.format(r)] = "=MAX(AW2:AW{})".format(q)
            ws['AW{}'.format(s)] = "=MIN(AW2:AW{})".format(q)
            ws['AW{}'.format(t)] = "=ROUND(AVERAGE(AW2:AW{}),2)".format(q)
            ws['AX{}'.format(r)] = "=MAX(AX2:AX{})".format(q)
            ws['AX{}'.format(s)] = "=MIN(AX2:AX{})".format(q)
            ws['AX{}'.format(t)] = "=ROUND(AVERAGE(AX2:AX{}),2)".format(q)
            ws['AY{}'.format(r)] = "=MAX(AY2:AY{})".format(q)
            ws['AY{}'.format(s)] = "=MIN(AY2:AY{})".format(q)
            ws['AY{}'.format(t)] = "=ROUND(AVERAGE(AY2:AY{}),2)".format(q)
            ws['AZ{}'.format(r)] = "=MAX(AZ2:AZ{})".format(q)
            ws['AZ{}'.format(s)] = "=MIN(AZ2:AZ{})".format(q)
            ws['AZ{}'.format(t)] = "=ROUND(AVERAGE(AZ2:AZ{}),2)".format(q)
            ws['BA{}'.format(r)] = "=MAX(BA2:BA{})".format(q)
            ws['BA{}'.format(s)] = "=MIN(BA2:BA{})".format(q)
            ws['BA{}'.format(t)] = "=ROUND(AVERAGE(BA2:BA{}),2)".format(q)
            ws['BD{}'.format(r)] = "=SUM(BD2:BD{})".format(q)
            ws['BE{}'.format(r)] = "=SUM(BE2:BE{})".format(q)
            ws['BF{}'.format(r)] = "=SUM(BF2:BF{})".format(q)
            ws['BG{}'.format(r)] = "=SUM(BG2:BG{})".format(q)
            ws['BH{}'.format(r)] = "=SUM(BH2:BH{})".format(q)

            # iterasi 2 rata-rata - 1
            ws['BO{}'.format(
                r)] = "=IF($BD${}=0,$AK${},$AK${}-1)".format(r, r, r)
            ws['BO{}'.format(s)] = "=STDEV(BO2:BO{})".format(q)
            ws['BO{}'.format(t)] = "=MAX(BO2:BO{})".format(q)
            ws['BO{}'.format(u)] = "=MIN(BO2:BO{})".format(q)
            ws['BP{}'.format(
                r)] = "=IF($BE${}=0,$AL${},$AL${}-1)".format(r, r, r)
            ws['BP{}'.format(s)] = "=STDEV(BP2:BP{})".format(q)
            ws['BP{}'.format(t)] = "=MAX(BP2:BP{})".format(q)
            ws['BP{}'.format(u)] = "=MIN(BP2:BP{})".format(q)
            ws['BQ{}'.format(
                r)] = "=IF($BF${}=0,$AM${},$AM${}-1)".format(r, r, r)
            ws['BQ{}'.format(s)] = "=STDEV(BQ2:BQ{})".format(q)
            ws['BQ{}'.format(t)] = "=MAX(BQ2:BQ{})".format(q)
            ws['BQ{}'.format(u)] = "=MIN(BQ2:BQ{})".format(q)
            ws['BR{}'.format(
                r)] = "=IF($BG${}=0,$AN${},$AN${}-1)".format(r, r, r)
            ws['BR{}'.format(s)] = "=STDEV(BR2:BR{})".format(q)
            ws['BR{}'.format(t)] = "=MAX(BR2:BR{})".format(q)
            ws['BR{}'.format(u)] = "=MIN(BR2:BR{})".format(q)
            ws['BS{}'.format(
                r)] = "=IF($BH${}=0,$AO${},$AO${}-1)".format(r, r, r)
            ws['BS{}'.format(s)] = "=STDEV(BS2:BS{})".format(q)
            ws['BS{}'.format(t)] = "=MAX(BS2:BS{})".format(q)
            ws['BS{}'.format(u)] = "=MIN(BS2:BS{})".format(q)
            ws['BT{}'.format(r)] = "=ROUND(AVERAGE(BT2:BT{}),2)".format(q)
            ws['BT{}'.format(t)] = "=MAX(BT2:BT{})".format(q)
            ws['BT{}'.format(u)] = "=MIN(BT2:BT{})".format(q)
            ws['BU{}'.format(r)] = "=MAX(BU2:BU{})".format(q)
            ws['BV{}'.format(r)] = "=MAX(BV2:BV{})".format(q)
            ws['BW{}'.format(r)] = "=MAX(BW2:BW{})".format(q)
            ws['BX{}'.format(r)] = "=MAX(BX2:BX{})".format(q)
            ws['BY{}'.format(r)] = "=MAX(BY2:BY{})".format(q)
            ws['BZ{}'.format(r)] = "=MAX(BZ2:BZ{})".format(q)
            ws['BZ{}'.format(s)] = "=MIN(BZ2:BZ{})".format(q)
            ws['BZ{}'.format(t)] = "=ROUND(AVERAGE(BZ2:BZ{}),2)".format(q)
            ws['CA{}'.format(r)] = "=MAX(CA2:CA{})".format(q)
            ws['CA{}'.format(s)] = "=MIN(CA2:CA{})".format(q)
            ws['CA{}'.format(t)] = "=ROUND(AVERAGE(CA2:CA{}),2)".format(q)
            ws['CB{}'.format(r)] = "=MAX(CB2:CB{})".format(q)
            ws['CB{}'.format(s)] = "=MIN(CB2:CB{})".format(q)
            ws['CB{}'.format(t)] = "=ROUND(AVERAGE(CB2:CB{}),2)".format(q)
            ws['CC{}'.format(r)] = "=MAX(CC2:CC{})".format(q)
            ws['CC{}'.format(s)] = "=MIN(CC2:CC{})".format(q)
            ws['CC{}'.format(t)] = "=ROUND(AVERAGE(CC2:CC{}),2)".format(q)
            ws['CD{}'.format(r)] = "=MAX(CD2:CD{})".format(q)
            ws['CD{}'.format(s)] = "=MIN(CD2:CD{})".format(q)
            ws['CD{}'.format(t)] = "=ROUND(AVERAGE(CD2:CD{}),2)".format(q)
            ws['CE{}'.format(r)] = "=MAX(CE2:CE{})".format(q)
            ws['CE{}'.format(s)] = "=MIN(CE2:CE{})".format(q)
            ws['CE{}'.format(t)] = "=ROUND(AVERAGE(CE2:CE{}),2)".format(q)
            ws['CH{}'.format(r)] = "=SUM(CH2:CH{})".format(q)
            ws['CI{}'.format(r)] = "=SUM(CI2:CI{})".format(q)
            ws['CJ{}'.format(r)] = "=SUM(CJ2:CJ{})".format(q)
            ws['CK{}'.format(r)] = "=SUM(CK2:CK{})".format(q)
            ws['CL{}'.format(r)] = "=SUM(CL2:CL{})".format(q)

            # iterasi 3 rata-rata - 1
            ws['CS{}'.format(
                r)] = "=IF($CH${}=0,$BO${},$BO${}-1)".format(r, r, r)
            ws['CS{}'.format(s)] = "=STDEV(CS2:CS{})".format(q)
            ws['CS{}'.format(t)] = "=MAX(CS2:CS{})".format(q)
            ws['CS{}'.format(u)] = "=MIN(CS2:CS{})".format(q)
            ws['CT{}'.format(
                r)] = "=IF($CI${}=0,$BP${},$BP${}-1)".format(r, r, r)
            ws['CT{}'.format(s)] = "=STDEV(CT2:CT{})".format(q)
            ws['CT{}'.format(t)] = "=MAX(CT2:CT{})".format(q)
            ws['CT{}'.format(u)] = "=MIN(CT2:CT{})".format(q)
            ws['CU{}'.format(
                r)] = "=IF($CJ${}=0,$BQ${},$BQ${}-1)".format(r, r, r)
            ws['CU{}'.format(s)] = "=STDEV(CU2:CU{})".format(q)
            ws['CU{}'.format(t)] = "=MAX(CU2:CU{})".format(q)
            ws['CU{}'.format(u)] = "=MIN(CU2:CU{})".format(q)
            ws['CV{}'.format(
                r)] = "=IF($CK${}=0,$BR${},$BR${}-1)".format(r, r, r)
            ws['CV{}'.format(s)] = "=STDEV(CV2:CV{})".format(q)
            ws['CV{}'.format(t)] = "=MAX(CV2:CV{})".format(q)
            ws['CV{}'.format(u)] = "=MIN(CV2:CV{})".format(q)
            ws['CW{}'.format(
                r)] = "=IF($CL${}=0,$BS${},$BS${}-1)".format(r, r, r)
            ws['CW{}'.format(s)] = "=STDEV(CW2:CW{})".format(q)
            ws['CW{}'.format(t)] = "=MAX(CW2:CW{})".format(q)
            ws['CW{}'.format(u)] = "=MIN(CW2:CW{})".format(q)
            ws['CX{}'.format(r)] = "=ROUND(AVERAGE(CX2:CX{}),2)".format(q)
            ws['CX{}'.format(t)] = "=MAX(CX2:CX{})".format(q)
            ws['CX{}'.format(u)] = "=MIN(CX2:CX{})".format(q)
            ws['CY{}'.format(r)] = "=MAX(CY2:CY{})".format(q)
            ws['CZ{}'.format(r)] = "=MAX(CZ2:CZ{})".format(q)
            ws['DA{}'.format(r)] = "=MAX(DA2:DA{})".format(q)
            ws['DB{}'.format(r)] = "=MAX(DB2:DB{})".format(q)
            ws['DC{}'.format(r)] = "=MAX(DC2:DC{})".format(q)
            ws['DD{}'.format(r)] = "=MAX(DD2:DD{})".format(q)
            ws['DD{}'.format(s)] = "=MIN(DD2:DD{})".format(q)
            ws['DD{}'.format(t)] = "=ROUND(AVERAGE(DD2:DD{}),2)".format(q)
            ws['DE{}'.format(r)] = "=MAX(DE2:DE{})".format(q)
            ws['DE{}'.format(s)] = "=MIN(DE2:DE{})".format(q)
            ws['DE{}'.format(t)] = "=ROUND(AVERAGE(DE2:DE{}),2)".format(q)
            ws['DF{}'.format(r)] = "=MAX(DF2:DF{})".format(q)
            ws['DF{}'.format(s)] = "=MIN(DF2:DF{})".format(q)
            ws['DF{}'.format(t)] = "=ROUND(AVERAGE(DF2:DF{}),2)".format(q)
            ws['DG{}'.format(r)] = "=MAX(DG2:DG{})".format(q)
            ws['DG{}'.format(s)] = "=MIN(DG2:DG{})".format(q)
            ws['DG{}'.format(t)] = "=ROUND(AVERAGE(DG2:DG{}),2)".format(q)
            ws['DH{}'.format(r)] = "=MAX(DH2:DH{})".format(q)
            ws['DH{}'.format(s)] = "=MIN(DH2:DH{})".format(q)
            ws['DH{}'.format(t)] = "=ROUND(AVERAGE(DH2:DH{}),2)".format(q)
            ws['DI{}'.format(r)] = "=MAX(DI2:DI{})".format(q)
            ws['DI{}'.format(s)] = "=MIN(DI2:DI{})".format(q)
            ws['DI{}'.format(t)] = "=ROUND(AVERAGE(DI2:DI{}),2)".format(q)
            ws['DL{}'.format(r)] = "=SUM(DL2:DL{})".format(q)
            ws['DM{}'.format(r)] = "=SUM(DM2:DM{})".format(q)
            ws['DN{}'.format(r)] = "=SUM(DN2:DN{})".format(q)
            ws['DO{}'.format(r)] = "=SUM(DO2:DO{})".format(q)
            ws['DP{}'.format(r)] = "=SUM(DP2:DP{})".format(q)

            # iterasi 4 rata-rata - 1
            ws['DW{}'.format(
                r)] = "=IF($DL${}=0,$CS${},$CS${}-1)".format(r, r, r)
            ws['DW{}'.format(s)] = "=STDEV(DW2:DW{})".format(q)
            ws['DW{}'.format(t)] = "=MAX(DW2:DW{})".format(q)
            ws['DW{}'.format(u)] = "=MIN(DW2:DW{})".format(q)
            ws['DX{}'.format(
                r)] = "=IF($DM${}=0,$CT${},$CT${}-1)".format(r, r, r)
            ws['DX{}'.format(s)] = "=STDEV(DX2:DX{})".format(q)
            ws['DX{}'.format(t)] = "=MAX(DX2:DX{})".format(q)
            ws['DX{}'.format(u)] = "=MIN(DX2:DX{})".format(q)
            ws['DY{}'.format(
                r)] = "=IF($DN${}=0,$CU${},$CU${}-1)".format(r, r, r)
            ws['DY{}'.format(s)] = "=STDEV(DY2:DY{})".format(q)
            ws['DY{}'.format(t)] = "=MAX(DY2:DY{})".format(q)
            ws['DY{}'.format(u)] = "=MIN(DY2:DY{})".format(q)
            ws['DZ{}'.format(
                r)] = "=IF($DO${}=0,$CV${},$CV${}-1)".format(r, r, r)
            ws['DZ{}'.format(s)] = "=STDEV(DZ2:DZ{})".format(q)
            ws['DZ{}'.format(t)] = "=MAX(DZ2:DZ{})".format(q)
            ws['DZ{}'.format(u)] = "=MIN(DZ2:DZ{})".format(q)
            ws['EA{}'.format(
                r)] = "=IF($DP${}=0,$CW${},$CW${}-1)".format(r, r, r)
            ws['EA{}'.format(s)] = "=STDEV(EA2:EA{})".format(q)
            ws['EA{}'.format(t)] = "=MAX(EA2:EA{})".format(q)
            ws['EA{}'.format(u)] = "=MIN(EA2:EA{})".format(q)
            ws['EB{}'.format(r)] = "=ROUND(AVERAGE(EB2:EB{}),2)".format(q)
            ws['EB{}'.format(t)] = "=MAX(EB2:EB{})".format(q)
            ws['EB{}'.format(u)] = "=MIN(EB2:EB{})".format(q)
            ws['EC{}'.format(r)] = "=MAX(EC2:EC{})".format(q)
            ws['ED{}'.format(r)] = "=MAX(ED2:ED{})".format(q)
            ws['EE{}'.format(r)] = "=MAX(EE2:EE{})".format(q)
            ws['EF{}'.format(r)] = "=MAX(EF2:EF{})".format(q)
            ws['EG{}'.format(r)] = "=MAX(EG2:EG{})".format(q)
            ws['EH{}'.format(r)] = "=MAX(EH2:EH{})".format(q)
            ws['EH{}'.format(s)] = "=MIN(EH2:EH{})".format(q)
            ws['EH{}'.format(t)] = "=ROUND(AVERAGE(EH2:EH{}),2)".format(q)
            ws['EI{}'.format(r)] = "=MAX(EI2:EI{})".format(q)
            ws['EI{}'.format(s)] = "=MIN(EI2:EI{})".format(q)
            ws['EI{}'.format(t)] = "=ROUND(AVERAGE(EI2:EI{}),2)".format(q)
            ws['EJ{}'.format(r)] = "=MAX(EJ2:EJ{})".format(q)
            ws['EJ{}'.format(s)] = "=MIN(EJ2:EJ{})".format(q)
            ws['EJ{}'.format(t)] = "=ROUND(AVERAGE(EJ2:EJ{}),2)".format(q)
            ws['EK{}'.format(r)] = "=MAX(EK2:EK{})".format(q)
            ws['EK{}'.format(s)] = "=MIN(EK2:EK{})".format(q)
            ws['EK{}'.format(t)] = "=ROUND(AVERAGE(EK2:EK{}),2)".format(q)
            ws['EL{}'.format(r)] = "=MAX(EL2:EL{})".format(q)
            ws['EL{}'.format(s)] = "=MIN(EL2:EL{})".format(q)
            ws['EL{}'.format(t)] = "=ROUND(AVERAGE(EL2:EL{}),2)".format(q)
            ws['EM{}'.format(r)] = "=MAX(EM2:EM{})".format(q)
            ws['EM{}'.format(s)] = "=MIN(EM2:EM{})".format(q)
            ws['EM{}'.format(t)] = "=ROUND(AVERAGE(EM2:EM{}),2)".format(q)
            ws['EP{}'.format(r)] = "=SUM(EP2:EP{})".format(q)
            ws['EQ{}'.format(r)] = "=SUM(EQ2:EQ{})".format(q)
            ws['ER{}'.format(r)] = "=SUM(ER2:ER{})".format(q)
            ws['ES{}'.format(r)] = "=SUM(ES2:ES{})".format(q)
            ws['ET{}'.format(r)] = "=SUM(ET2:ET{})".format(q)

            # iterasi 5 rata-rata - 1
            ws['FA{}'.format(
                r)] = "=IF($EP${}=0,$DW${},$DW${}-1)".format(r, r, r)
            ws['FA{}'.format(s)] = "=STDEV(FA2:FA{})".format(q)
            ws['FA{}'.format(t)] = "=MAX(FA2:FA{})".format(q)
            ws['FA{}'.format(u)] = "=MIN(FA2:FA{})".format(q)
            ws['FB{}'.format(
                r)] = "=IF($EQ${}=0,$DX${},$DX${}-1)".format(r, r, r)
            ws['FB{}'.format(s)] = "=STDEV(FB2:FB{})".format(q)
            ws['FB{}'.format(t)] = "=MAX(FB2:FB{})".format(q)
            ws['FB{}'.format(u)] = "=MIN(FB2:FB{})".format(q)
            ws['FC{}'.format(
                r)] = "=IF($ER${}=0,$DY${},$DY${}-1)".format(r, r, r)
            ws['FC{}'.format(s)] = "=STDEV(FC2:FC{})".format(q)
            ws['FC{}'.format(t)] = "=MAX(FC2:FC{})".format(q)
            ws['FC{}'.format(u)] = "=MIN(FC2:FC{})".format(q)
            ws['FD{}'.format(
                r)] = "=IF($ES${}=0,$DZ${},$DZ${}-1)".format(r, r, r)
            ws['FD{}'.format(s)] = "=STDEV(FD2:FD{})".format(q)
            ws['FD{}'.format(t)] = "=MAX(FD2:FD{})".format(q)
            ws['FD{}'.format(u)] = "=MIN(FD2:FD{})".format(q)
            ws['FE{}'.format(
                r)] = "=IF($ET${}=0,$EA${},$EA${}-1)".format(r, r, r)
            ws['FE{}'.format(s)] = "=STDEV(FE2:FE{})".format(q)
            ws['FE{}'.format(t)] = "=MAX(FE2:FE{})".format(q)
            ws['FE{}'.format(u)] = "=MIN(FE2:FE{})".format(q)
            ws['FF{}'.format(r)] = "=ROUND(AVERAGE(FF2:FF{}),2)".format(q)
            ws['FF{}'.format(t)] = "=MAX(FF2:FF{})".format(q)
            ws['FF{}'.format(u)] = "=MIN(FF2:FF{})".format(q)
            ws['FG{}'.format(r)] = "=MAX(FG2:FG{})".format(q)
            ws['FH{}'.format(r)] = "=MAX(FH2:FH{})".format(q)
            ws['FI{}'.format(r)] = "=MAX(FI2:FI{})".format(q)
            ws['FJ{}'.format(r)] = "=MAX(FJ2:FJ{})".format(q)
            ws['FK{}'.format(r)] = "=MAX(FK2:FK{})".format(q)
            ws['FL{}'.format(r)] = "=MAX(FL2:FL{})".format(q)
            ws['FL{}'.format(s)] = "=MIN(FL2:FL{})".format(q)
            ws['FL{}'.format(t)] = "=ROUND(AVERAGE(FL2:FL{}),2)".format(q)
            ws['FM{}'.format(r)] = "=MAX(FM2:FM{})".format(q)
            ws['FM{}'.format(s)] = "=MIN(FM2:FM{})".format(q)
            ws['FM{}'.format(t)] = "=ROUND(AVERAGE(FM2:FM{}),2)".format(q)
            ws['FN{}'.format(r)] = "=MAX(FN2:FN{})".format(q)
            ws['FN{}'.format(s)] = "=MIN(FN2:FN{})".format(q)
            ws['FN{}'.format(t)] = "=ROUND(AVERAGE(FN2:FN{}),2)".format(q)
            ws['FO{}'.format(r)] = "=MAX(FO2:FO{})".format(q)
            ws['FO{}'.format(s)] = "=MIN(FO2:FO{})".format(q)
            ws['FO{}'.format(t)] = "=ROUND(AVERAGE(FO2:FO{}),2)".format(q)
            ws['FP{}'.format(r)] = "=MAX(FP2:FP{})".format(q)
            ws['FP{}'.format(s)] = "=MIN(FP2:FP{})".format(q)
            ws['FP{}'.format(t)] = "=ROUND(AVERAGE(FP2:FP{}),2)".format(q)
            ws['FQ{}'.format(r)] = "=MAX(FQ2:FQ{})".format(q)
            ws['FQ{}'.format(s)] = "=MIN(FQ2:FQ{})".format(q)
            ws['FQ{}'.format(t)] = "=ROUND(AVERAGE(FQ2:FQ{}),2)".format(q)
            ws['FT{}'.format(r)] = "=SUM(FT2:FT{})".format(q)
            ws['FU{}'.format(r)] = "=SUM(FU2:FU{})".format(q)
            ws['FV{}'.format(r)] = "=SUM(FV2:FV{})".format(q)
            ws['FW{}'.format(r)] = "=SUM(FW2:FW{})".format(q)
            ws['FX{}'.format(r)] = "=SUM(FX2:FX{})".format(q)

            # Z Score
            ws['B1'] = 'NAMA_SISWA_1'
            ws['C1'] = 'NOMOR_NF_1'
            ws['D1'] = 'KELAS_1'
            ws['E1'] = 'NAMA_SEKOLAH_1'
            ws['F1'] = 'LOKASI_1'
            ws['G1'] = 'MAT_1'
            ws['H1'] = 'SEJ_1'
            ws['I1'] = 'GEO_1'
            ws['J1'] = 'EKO_1'
            ws['K1'] = 'SOS_1'
            ws['L1'] = 'JML_1'
            ws['M1'] = 'Z_MAT_1'
            ws['N1'] = 'Z_SEJ_1'
            ws['O1'] = 'Z_GEO_1'
            ws['P1'] = 'Z_EKO_1'
            ws['Q1'] = 'Z_SOS_1'
            ws['R1'] = 'S_MAT_1'
            ws['S1'] = 'S_SEJ_1'
            ws['T1'] = 'S_GEO_1'
            ws['U1'] = 'S_EKO_1'
            ws['V1'] = 'S_SOS_1'
            ws['W1'] = 'S_JML_1'
            ws['X1'] = 'RANK_NAS._1'
            ws['Y1'] = 'RANK_LOK._1'
            ws['M1'].font = Font(bold=False, name='Calibri', size=11)
            ws['N1'].font = Font(bold=False, name='Calibri', size=11)
            ws['O1'].font = Font(bold=False, name='Calibri', size=11)
            ws['P1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Q1'].font = Font(bold=False, name='Calibri', size=11)
            ws['R1'].font = Font(bold=False, name='Calibri', size=11)
            ws['S1'].font = Font(bold=False, name='Calibri', size=11)
            ws['T1'].font = Font(bold=False, name='Calibri', size=11)
            ws['U1'].font = Font(bold=False, name='Calibri', size=11)
            ws['V1'].font = Font(bold=False, name='Calibri', size=11)
            ws['W1'].font = Font(bold=False, name='Calibri', size=11)
            ws['X1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Y1'].font = Font(bold=False, name='Calibri', size=11)
        # FILL
            ws['B1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['C1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['D1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['E1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['F1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['G1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['H1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['I1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['J1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['K1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['L1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['M1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['N1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['O1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['P1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Q1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['R1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['S1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['T1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['U1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['V1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['W1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['X1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Y1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            # tambahan
            ws['Z1'] = 'MAT_20_1'
            ws['AA1'] = 'SEJ_20_1'
            ws['AB1'] = 'GEO_20_1'
            ws['AC1'] = 'EKO_20_1'
            ws['AD1'] = 'SOS_20_1'
            ws['Z1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Z1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AA1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AB1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AC1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AD1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')

            for row in range(2, q+1):
                ws['M{}'.format(
                    row)] = '=IFERROR(ROUND(IF(G{}="","",(G{}-G${})/G${}),2),"")'.format(row, row, r, s)
                ws['N{}'.format(
                    row)] = '=IFERROR(ROUND(IF(H{}="","",(H{}-H${})/H${}),2),"")'.format(row, row, r, s)
                ws['O{}'.format(
                    row)] = '=IFERROR(ROUND(IF(I{}="","",(I{}-I${})/I${}),2),"")'.format(row, row, r, s)
                ws['P{}'.format(
                    row)] = '=IFERROR(ROUND(IF(J{}="","",(J{}-J${})/J${}),2),"")'.format(row, row, r, s)
                ws['Q{}'.format(
                    row)] = '=IFERROR(ROUND(IF(K{}="","",(K{}-K${})/K${}),2),"")'.format(row, row, r, s)
                ws['R{}'.format(
                    row)] = '=IFERROR(ROUND(IF(G{}="","",IF(70+30*M{}/$M${}<20,20,70+30*M{}/$M${})),2),"")'.format(row, row, r, row, r)
                ws['S{}'.format(
                    row)] = '=IFERROR(ROUND(IF(H{}="","",IF(70+30*N{}/$N${}<20,20,70+30*N{}/$N${})),2),"")'.format(row, row, r, row, r)
                ws['T{}'.format(
                    row)] = '=IFERROR(ROUND(IF(I{}="","",IF(70+30*O{}/$O${}<20,20,70+30*O{}/$O${})),2),"")'.format(row, row, r, row, r)
                ws['U{}'.format(
                    row)] = '=IFERROR(ROUND(IF(J{}="","",IF(70+30*P{}/$P${}<20,20,70+30*P{}/$P${})),2),"")'.format(row, row, r, row, r)
                ws['V{}'.format(
                    row)] = '=IFERROR(ROUND(IF(K{}="","",IF(70+30*Q{}/$Q${}<20,20,70+30*Q{}/$Q${})),2),"")'.format(row, row, r, row, r)

                ws['W{}'.format(row)] = '=IF(SUM(R{}:V{})=0,"",SUM(R{}:V{}))'.format(
                    row, row, row, row)
                ws['X{}'.format(row)] = '=IF(W{}="","",RANK(W{},$W$2:$W${}))'.format(
                    row, row, q)
                ws['Y{}'.format(
                    row)] = '=IF(X{}="","",COUNTIFS($F$2:$F${},F{},$X$2:$X${},"<"&X{})+1)'.format(row, q, row, q, row)
                # TAMBAHAN
                ws['Z{}'.format(row)] = '=IF($G${}=25,IF(AND(G{}>4,R{}=20),1,""),IF($G${}=30,IF(AND(G{}>5,R{}=20),1,""),IF($G${}=35,IF(AND(G{}>6,R{}=20),1,""),IF($G${}=40,IF(AND(G{}>7,R{}=20),1,""),IF($G${}=45,IF(AND(G{}>8,R{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AA{}'.format(row)] = '=IF($H${}=25,IF(AND(H{}>4,S{}=20),1,""),IF($H${}=30,IF(AND(H{}>5,S{}=20),1,""),IF($H${}=35,IF(AND(H{}>6,S{}=20),1,""),IF($H${}=40,IF(AND(H{}>7,S{}=20),1,""),IF($H${}=45,IF(AND(H{}>8,S{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AB{}'.format(row)] = '=IF($I${}=25,IF(AND(I{}>4,T{}=20),1,""),IF($I${}=30,IF(AND(I{}>5,T{}=20),1,""),IF($I${}=35,IF(AND(I{}>6,T{}=20),1,""),IF($I${}=40,IF(AND(I{}>7,T{}=20),1,""),IF($I${}=45,IF(AND(I{}>8,T{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AC{}'.format(row)] = '=IF($J${}=25,IF(AND(J{}>4,U{}=20),1,""),IF($J${}=30,IF(AND(J{}>5,U{}=20),1,""),IF($J${}=35,IF(AND(J{}>6,U{}=20),1,""),IF($J${}=40,IF(AND(J{}>7,U{}=20),1,""),IF($J${}=45,IF(AND(J{}>8,U{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AD{}'.format(row)] = '=IF($K${}=25,IF(AND(K{}>4,V{}=20),1,""),IF($K${}=30,IF(AND(K{}>5,V{}=20),1,""),IF($K${}=35,IF(AND(K{}>6,V{}=20),1,""),IF($K${}=40,IF(AND(K{}>7,V{}=20),1,""),IF($K${}=45,IF(AND(K{}>8,V{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

        # new Z Score
            ws['AF1'] = 'NAMA_SISWA_2'
            ws['AG1'] = 'NOMOR_NF_2'
            ws['AH1'] = 'KELAS_2'
            ws['AI1'] = 'NAMA_SEKOLAH_2'
            ws['AJ1'] = 'LOKASI_2'
            ws['AK1'] = 'MAT_2'
            ws['AL1'] = 'SEJ_2'
            ws['AM1'] = 'GEO_2'
            ws['AN1'] = 'EKO_2'
            ws['AO1'] = 'SOS_2'
            ws['AP1'] = 'JML_2'
            ws['AQ1'] = 'Z_MAT_2'
            ws['AR1'] = 'Z_SEJ_2'
            ws['AS1'] = 'Z_GEO_2'
            ws['AT1'] = 'Z_EKO_2'
            ws['AU1'] = 'Z_SOS_2'
            ws['AV1'] = 'S_MAT_2'
            ws['AW1'] = 'S_SEJ_2'
            ws['AX1'] = 'S_GEO_2'
            ws['AY1'] = 'S_EKO_2'
            ws['AZ1'] = 'S_SOS_2'
            ws['BA1'] = 'S_JML_2'
            ws['BB1'] = 'RANK_NAS._2'
            ws['BC1'] = 'RANK_LOK._2'
            ws['AQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BC1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['AF1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AG1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AH1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AI1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AJ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AK1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AL1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AM1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AN1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AO1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AP1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AQ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AR1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AS1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AT1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AU1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AV1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AW1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AX1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AY1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['AZ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BA1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BB1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BC1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')

            # tambahan
            ws['BD1'] = 'MAT_20_2'
            ws['BE1'] = 'SEJ_20_2'
            ws['BF1'] = 'GEO_20_2'
            ws['BG1'] = 'EKO_20_2'
            ws['BH1'] = 'SOS_20_2'
            ws['BD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BD1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BE1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BF1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BG1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BH1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')

            for row in range(2, q+1):
                ws['AF{}'.format(row)] = '=B{}'.format(row)
                ws['AG{}'.format(row)] = '=C{}'.format(row, row)
                ws['AH{}'.format(row)] = '=D{}'.format(row, row)
                ws['AI{}'.format(row)] = '=E{}'.format(row, row)
                ws['AJ{}'.format(row)] = '=F{}'.format(row, row)
                ws['AK{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['AL{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['AM{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['AN{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['AO{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['AP{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['AQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AK{}="","",(AK{}-AK${})/AK${}),2),"")'.format(row, row, r, s)
                ws['AR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AL{}="","",(AL{}-AL${})/AL${}),2),"")'.format(row, row, r, s)
                ws['AS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AM{}="","",(AM{}-AM${})/AM${}),2),"")'.format(row, row, r, s)
                ws['AT{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AN{}="","",(AN{}-AN${})/AN${}),2),"")'.format(row, row, r, s)
                ws['AU{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AO{}="","",(AO{}-AO${})/AO${}),2),"")'.format(row, row, r, s)
                ws['AV{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AK{}="","",IF(70+30*AQ{}/$AQ${}<20,20,70+30*AQ{}/$AQ${})),2),"")'.format(row, row, r, row, r)
                ws['AW{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AL{}="","",IF(70+30*AR{}/$AR${}<20,20,70+30*AR{}/$AR${})),2),"")'.format(row, row, r, row, r)
                ws['AX{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AM{}="","",IF(70+30*AS{}/$AS${}<20,20,70+30*AS{}/$AS${})),2),"")'.format(row, row, r, row, r)
                ws['AY{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AN{}="","",IF(70+30*AT{}/$AT${}<20,20,70+30*AT{}/$AT${})),2),"")'.format(row, row, r, row, r)
                ws['AZ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(AO{}="","",IF(70+30*AU{}/$AU${}<20,20,70+30*AU{}/$AU${})),2),"")'.format(row, row, r, row, r)

                ws['BA{}'.format(row)] = '=IF(SUM(AV{}:AZ{})=0,"",SUM(AV{}:AZ{}))'.format(
                    row, row, row, row)
                ws['BB{}'.format(row)] = '=IF(BA{}="","",RANK(BA{},$BA$2:$BA${}))'.format(
                    row, row, q)
                ws['BC{}'.format(
                    row)] = '=IF(BB{}="","",COUNTIFS($AJ$2:$AJ${},F{},$BB$2:$BB${},"<"&BB{})+1)'.format(row, q, row, q, row)
            #     TAMBAHAN
                ws['BD{}'.format(row)] = '=IF($G${}=25,IF(AND(AK{}>4,AV{}=20),1,""),IF($G${}=30,IF(AND(AK{}>5,AV{}=20),1,""),IF($G${}=35,IF(AND(AK{}>6,AV{}=20),1,""),IF($G${}=40,IF(AND(AK{}>7,AV{}=20),1,""),IF($G${}=45,IF(AND(AK{}>8,AV{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BE{}'.format(row)] = '=IF($H${}=25,IF(AND(AL{}>4,AW{}=20),1,""),IF($H${}=30,IF(AND(AL{}>5,AW{}=20),1,""),IF($H${}=35,IF(AND(AL{}>6,AW{}=20),1,""),IF($H${}=40,IF(AND(AL{}>7,AW{}=20),1,""),IF($H${}=45,IF(AND(AL{}>8,AW{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BF{}'.format(row)] = '=IF($I${}=25,IF(AND(AM{}>4,AX{}=20),1,""),IF($I${}=30,IF(AND(AM{}>5,AX{}=20),1,""),IF($I${}=35,IF(AND(AM{}>6,AX{}=20),1,""),IF($I${}=40,IF(AND(AM{}>7,AX{}=20),1,""),IF($I${}=45,IF(AND(AM{}>8,AX{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BG{}'.format(row)] = '=IF($J${}=25,IF(AND(AN{}>4,AY{}=20),1,""),IF($J${}=30,IF(AND(AN{}>5,AY{}=20),1,""),IF($J${}=35,IF(AND(AN{}>6,AY{}=20),1,""),IF($J${}=40,IF(AND(AN{}>7,AY{}=20),1,""),IF($J${}=45,IF(AND(AN{}>8,AY{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BH{}'.format(row)] = '=IF($K${}=25,IF(AND(AO{}>4,AZ{}=20),1,""),IF($K${}=30,IF(AND(AO{}>5,AZ{}=20),1,""),IF($K${}=35,IF(AND(AO{}>6,AZ{}=20),1,""),IF($K${}=40,IF(AND(AO{}>7,AZ{}=20),1,""),IF($K${}=45,IF(AND(AO{}>8,AZ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

                # new Z Score [2]
            ws['BJ1'] = 'NAMA_SISWA_3'
            ws['BK1'] = 'NOMOR_NF_3'
            ws['BL1'] = 'KELAS_3'
            ws['BM1'] = 'NAMA_SEKOLAH_3'
            ws['BN1'] = 'LOKASI_3'
            ws['BO1'] = 'MAT_3'
            ws['BP1'] = 'SEJ_3'
            ws['BQ1'] = 'GEO_3'
            ws['BR1'] = 'EKO_3'
            ws['BS1'] = 'SOS_3'
            ws['BT1'] = 'JML_3'
            ws['BU1'] = 'Z_MAT_3'
            ws['BV1'] = 'Z_SEJ_3'
            ws['BW1'] = 'Z_GEO_3'
            ws['BX1'] = 'Z_EKO_3'
            ws['BY1'] = 'Z_SOS_3'
            ws['BZ1'] = 'S_MAT_3'
            ws['CA1'] = 'S_SEJ_3'
            ws['CB1'] = 'S_GEO_3'
            ws['CC1'] = 'S_EKO_3'
            ws['CD1'] = 'S_SOS_3'
            ws['CE1'] = 'S_JML_3'
            ws['CF1'] = 'RANK_NAS._3'
            ws['CG1'] = 'RANK_LOK._3'
            ws['BU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CG1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['BJ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BK1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BL1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BM1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BN1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BO1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BP1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BQ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BR1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BS1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BT1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BU1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BV1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BW1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BX1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BY1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['BZ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CA1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CB1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CC1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CD1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CE1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CF1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CG1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')

            # tambahan
            ws['CH1'] = 'MAT_20_3'
            ws['CI1'] = 'SEJ_20_3'
            ws['CJ1'] = 'GEO_20_3'
            ws['CK1'] = 'EKO_20_3'
            ws['CL1'] = 'SOS_20_3'
            ws['CH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CH1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CI1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CJ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CK1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['CL1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')

            for row in range(2, q+1):
                ws['BJ{}'.format(row)] = '=B{}'.format(row)
                ws['BK{}'.format(row)] = '=C{}'.format(row, row)
                ws['BL{}'.format(row)] = '=D{}'.format(row, row)
                ws['BM{}'.format(row)] = '=E{}'.format(row, row)
                ws['BN{}'.format(row)] = '=F{}'.format(row, row)
                ws['BO{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['BP{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['BQ{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['BR{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['BS{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['BT{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['BU{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BO{}="","",(BO{}-BO${})/BO${}),2),"")'.format(row, row, r, s)
                ws['BV{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BP{}="","",(BP{}-BP${})/BP${}),2),"")'.format(row, row, r, s)
                ws['BW{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BQ{}="","",(BQ{}-BQ${})/BQ${}),2),"")'.format(row, row, r, s)
                ws['BX{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BR{}="","",(BR{}-BR${})/BR${}),2),"")'.format(row, row, r, s)
                ws['BY{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BS{}="","",(BS{}-BS${})/BS${}),2),"")'.format(row, row, r, s)
                ws['BZ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BO{}="","",IF(70+30*BU{}/$BU${}<20,20,70+30*BU{}/$BU${})),2),"")'.format(row, row, r, row, r)
                ws['CA{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BP{}="","",IF(70+30*BV{}/$BV${}<20,20,70+30*BV{}/$BV${})),2),"")'.format(row, row, r, row, r)
                ws['CB{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BQ{}="","",IF(70+30*BW{}/$BW${}<20,20,70+30*BW{}/$BW${})),2),"")'.format(row, row, r, row, r)
                ws['CC{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BR{}="","",IF(70+30*BX{}/$BX${}<20,20,70+30*BX{}/$BX${})),2),"")'.format(row, row, r, row, r)
                ws['CD{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BS{}="","",IF(70+30*BY{}/$BY${}<20,20,70+30*BY{}/$BY${})),2),"")'.format(row, row, r, row, r)

                ws['CE{}'.format(row)] = '=IF(SUM(BZ{}:CD{})=0,"",SUM(BZ{}:CD{}))'.format(
                    row, row, row, row)
                ws['CF{}'.format(row)] = '=IF(CE{}="","",RANK(CE{},$CE$2:$CE${}))'.format(
                    row, row, q)
                ws['CG{}'.format(
                    row)] = '=IF(CF{}="","",COUNTIFS($BN$2:$BN${},F{},$CF$2:$CF${},"<"&CF{})+1)'.format(row, q, row, q, row)
                #     TAMBAHAN
                ws['CH{}'.format(row)] = '=IF($G${}=25,IF(AND(BO{}>4,BZ{}=20),1,""),IF($G${}=30,IF(AND(BO{}>5,BZ{}=20),1,""),IF($G${}=35,IF(AND(BO{}>6,BZ{}=20),1,""),IF($G${}=40,IF(AND(BO{}>7,BZ{}=20),1,""),IF($G${}=45,IF(AND(BO{}>8,BZ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CI{}'.format(row)] = '=IF($H${}=25,IF(AND(BP{}>4,CA{}=20),1,""),IF($H${}=30,IF(AND(BP{}>5,CA{}=20),1,""),IF($H${}=35,IF(AND(BP{}>6,CA{}=20),1,""),IF($H${}=40,IF(AND(BP{}>7,CA{}=20),1,""),IF($H${}=45,IF(AND(BP{}>8,CA{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CJ{}'.format(row)] = '=IF($I${}=25,IF(AND(BQ{}>4,CB{}=20),1,""),IF($I${}=30,IF(AND(BQ{}>5,CB{}=20),1,""),IF($I${}=35,IF(AND(BQ{}>6,CB{}=20),1,""),IF($I${}=40,IF(AND(BQ{}>7,CB{}=20),1,""),IF($I${}=45,IF(AND(BQ{}>8,CB{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CK{}'.format(row)] = '=IF($J${}=25,IF(AND(BR{}>4,CC{}=20),1,""),IF($J${}=30,IF(AND(BR{}>5,CC{}=20),1,""),IF($J${}=35,IF(AND(BR{}>6,CC{}=20),1,""),IF($J${}=40,IF(AND(BR{}>7,CC{}=20),1,""),IF($J${}=45,IF(AND(BR{}>8,CC{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CL{}'.format(row)] = '=IF($K${}=25,IF(AND(BS{}>4,CD{}=20),1,""),IF($K${}=30,IF(AND(BS{}>5,CD{}=20),1,""),IF($K${}=35,IF(AND(BS{}>6,CD{}=20),1,""),IF($K${}=40,IF(AND(BS{}>7,CD{}=20),1,""),IF($K${}=45,IF(AND(BS{}>8,CD{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

                # new Z Score [3]
            ws['CN1'] = 'NAMA_SISWA_4'
            ws['CO1'] = 'NOMOR_NF_4'
            ws['CP1'] = 'KELAS_4'
            ws['CQ1'] = 'NAMA_SEKOLAH_4'
            ws['CR1'] = 'LOKASI_4'
            ws['CS1'] = 'MAT_4'
            ws['CT1'] = 'SEJ_4'
            ws['CU1'] = 'GEO_4'
            ws['CV1'] = 'EKO_4'
            ws['CW1'] = 'SOS_4'
            ws['CX1'] = 'JML_4'
            ws['CY1'] = 'Z_MAT_4'
            ws['CZ1'] = 'Z_SEJ_4'
            ws['DA1'] = 'Z_GEO_4'
            ws['DB1'] = 'Z_EKO_4'
            ws['DC1'] = 'Z_SOS_4'
            ws['DD1'] = 'S_MAT_4'
            ws['DE1'] = 'S_SEJ_4'
            ws['DF1'] = 'S_GEO_4'
            ws['DG1'] = 'S_EKO_4'
            ws['DH1'] = 'S_SOS_4'
            ws['DI1'] = 'S_JML_4'
            ws['DJ1'] = 'RANK_NAS._4'
            ws['DK1'] = 'RANK_LOK._4'
            ws['CY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DK1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['CN1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CO1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CP1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CQ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CR1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CS1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CT1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CU1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CV1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CW1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CX1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CY1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['CZ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DA1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DB1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DC1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DD1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DE1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DF1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DG1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DH1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DI1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DJ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DK1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')

            # tambahan
            ws['DL1'] = 'MAT_20_4'
            ws['DM1'] = 'SEJ_20_4'
            ws['DN1'] = 'GEO_20_4'
            ws['DO1'] = 'EKO_20_4'
            ws['DP1'] = 'SOS_20_4'
            ws['DL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DL1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DM1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DN1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DO1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['DP1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')

            for row in range(2, q+1):
                ws['CN{}'.format(row)] = '=B{}'.format(row)
                ws['CO{}'.format(row)] = '=C{}'.format(row, row)
                ws['CP{}'.format(row)] = '=D{}'.format(row, row)
                ws['CQ{}'.format(row)] = '=E{}'.format(row, row)
                ws['CR{}'.format(row)] = '=F{}'.format(row, row)
                ws['CS{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['CT{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['CU{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['CV{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['CW{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['CX{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['CY{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CS{}="","",(CS{}-CS${})/CS${}),2),"")'.format(row, row, r, s)
                ws['CZ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CT{}="","",(CT{}-CT${})/CT${}),2),"")'.format(row, row, r, s)
                ws['DA{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CU{}="","",(CU{}-CU${})/CU${}),2),"")'.format(row, row, r, s)
                ws['DB{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CV{}="","",(CV{}-CV${})/CV${}),2),"")'.format(row, row, r, s)
                ws['DC{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CW{}="","",(CW{}-CW${})/CW${}),2),"")'.format(row, row, r, s)
                ws['DD{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CY{}="","",IF(70+30*CY{}/$CY${}<20,20,70+30*CY{}/$CY${})),2),"")'.format(row, row, r, row, r)
                ws['DE{}'.format(
                    row)] = '=IFERROR(ROUND(IF(CZ{}="","",IF(70+30*CZ{}/$CZ${}<20,20,70+30*CZ{}/$CZ${})),2),"")'.format(row, row, r, row, r)
                ws['DF{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DA{}="","",IF(70+30*DA{}/$DA${}<20,20,70+30*DA{}/$DA${})),2),"")'.format(row, row, r, row, r)
                ws['DG{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DB{}="","",IF(70+30*DB{}/$DB${}<20,20,70+30*DB{}/$DB${})),2),"")'.format(row, row, r, row, r)
                ws['DH{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DC{}="","",IF(70+30*DC{}/$DC${}<20,20,70+30*DC{}/$DC${})),2),"")'.format(row, row, r, row, r)

                ws['DI{}'.format(row)] = '=IF(SUM(DD{}:DH{})=0,"",SUM(DD{}:DH{}))'.format(
                    row, row, row, row)
                ws['DJ{}'.format(row)] = '=IF(DI{}="","",RANK(DI{},$DI$2:$DI${}))'.format(
                    row, row, q)
                ws['DK{}'.format(
                    row)] = '=IF(DJ{}="","",COUNTIFS($CR$2:$CR${},F{},$DJ$2:$DJ${},"<"&DJ{})+1)'.format(row, q, row, q, row)
                #     TAMBAHAN
                ws['DL{}'.format(row)] = '=IF($G${}=25,IF(AND(CS{}>4,DD{}=20),1,""),IF($G${}=30,IF(AND(CS{}>5,DD{}=20),1,""),IF($G${}=35,IF(AND(CS{}>6,DD{}=20),1,""),IF($G${}=40,IF(AND(CS{}>7,DD{}=20),1,""),IF($G${}=45,IF(AND(CS{}>8,DD{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DM{}'.format(row)] = '=IF($H${}=25,IF(AND(CT{}>4,DE{}=20),1,""),IF($H${}=30,IF(AND(CT{}>5,DE{}=20),1,""),IF($H${}=35,IF(AND(CT{}>6,DE{}=20),1,""),IF($H${}=40,IF(AND(CT{}>7,DE{}=20),1,""),IF($H${}=45,IF(AND(CT{}>8,DE{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DN{}'.format(row)] = '=IF($I${}=25,IF(AND(CU{}>4,DF{}=20),1,""),IF($I${}=30,IF(AND(CU{}>5,DF{}=20),1,""),IF($I${}=35,IF(AND(CU{}>6,DF{}=20),1,""),IF($I${}=40,IF(AND(CU{}>7,DF{}=20),1,""),IF($I${}=45,IF(AND(CU{}>8,DF{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DO{}'.format(row)] = '=IF($J${}=25,IF(AND(CV{}>4,DG{}=20),1,""),IF($J${}=30,IF(AND(CV{}>5,DG{}=20),1,""),IF($J${}=35,IF(AND(CV{}>6,DG{}=20),1,""),IF($J${}=40,IF(AND(CV{}>7,DG{}=20),1,""),IF($J${}=45,IF(AND(CV{}>8,DG{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DP{}'.format(row)] = '=IF($K${}=25,IF(AND(CW{}>4,DH{}=20),1,""),IF($K${}=30,IF(AND(CW{}>5,DH{}=20),1,""),IF($K${}=35,IF(AND(CW{}>6,DH{}=20),1,""),IF($K${}=40,IF(AND(CW{}>7,DH{}=20),1,""),IF($K${}=45,IF(AND(CW{}>8,DH{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # new Z Score [4]
            ws['DR1'] = 'NAMA_SISWA_5'
            ws['DS1'] = 'NOMOR_NF_5'
            ws['DT1'] = 'KELAS_5'
            ws['DU1'] = 'NAMA_SEKOLAH_5'
            ws['DV1'] = 'LOKASI_5'
            ws['DW1'] = 'MAT_5'
            ws['DX1'] = 'SEJ_5'
            ws['DY1'] = 'GEO_5'
            ws['DZ1'] = 'EKO_5'
            ws['EA1'] = 'SOS_5'
            ws['EB1'] = 'JML_5'
            ws['EC1'] = 'Z_MAT_5'
            ws['ED1'] = 'Z_SEJ_5'
            ws['EE1'] = 'Z_GEO_5'
            ws['EF1'] = 'Z_EKO_5'
            ws['EG1'] = 'Z_SOS_5'
            ws['EH1'] = 'S_MAT_5'
            ws['EI1'] = 'S_SEJ_5'
            ws['EJ1'] = 'S_GEO_5'
            ws['EK1'] = 'S_EKO_5'
            ws['EL1'] = 'S_SOS_5'
            ws['EM1'] = 'S_JML_5'
            ws['EN1'] = 'RANK_NAS._5'
            ws['EO1'] = 'RANK_LOK._5'
            ws['EC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ED1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EO1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['DR1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DS1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DT1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DU1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DV1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DW1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DX1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DY1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['DZ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EA1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EB1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EC1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['ED1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EE1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EF1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EG1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EH1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EI1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EJ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EK1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EL1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EM1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EN1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EO1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')

            # tambahan
            ws['EP1'] = 'MAT_20_5'
            ws['EQ1'] = 'SEJ_20_5'
            ws['ER1'] = 'GEO_20_5'
            ws['ES1'] = 'EKO_20_5'
            ws['ET1'] = 'SOS_20_5'
            ws['EP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ER1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ES1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ET1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EP1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['EQ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['ER1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['ES1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['ET1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')

            for row in range(2, q+1):
                ws['DR{}'.format(row)] = '=B{}'.format(row)
                ws['DS{}'.format(row)] = '=C{}'.format(row, row)
                ws['DT{}'.format(row)] = '=D{}'.format(row, row)
                ws['DU{}'.format(row)] = '=E{}'.format(row, row)
                ws['DV{}'.format(row)] = '=F{}'.format(row, row)
                ws['DW{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['DX{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['DY{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['DZ{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['EA{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['EB{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['EC{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DW{}="","",(DW{}-DW${})/DW${}),2),"")'.format(row, row, r, s)
                ws['ED{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DX{}="","",(DX{}-DX${})/DX${}),2),"")'.format(row, row, r, s)
                ws['EE{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DY{}="","",(DY{}-DY${})/DY${}),2),"")'.format(row, row, r, s)
                ws['EF{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DZ{}="","",(DZ{}-DZ${})/DZ${}),2),"")'.format(row, row, r, s)
                ws['EG{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EA{}="","",(EA{}-EA${})/EA${}),2),"")'.format(row, row, r, s)
                ws['EH{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EC{}="","",IF(70+30*EC{}/$EC${}<20,20,70+30*EC{}/$EC${})),2),"")'.format(row, row, r, row, r)
                ws['EI{}'.format(
                    row)] = '=IFERROR(ROUND(IF(ED{}="","",IF(70+30*ED{}/$ED${}<20,20,70+30*ED{}/$ED${})),2),"")'.format(row, row, r, row, r)
                ws['EJ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EE{}="","",IF(70+30*EE{}/$EE${}<20,20,70+30*EE{}/$EE${})),2),"")'.format(row, row, r, row, r)
                ws['EK{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EF{}="","",IF(70+30*EF{}/$EF${}<20,20,70+30*EF{}/$EF${})),2),"")'.format(row, row, r, row, r)
                ws['EL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(EG{}="","",IF(70+30*EG{}/$EG${}<20,20,70+30*EG{}/$EG${})),2),"")'.format(row, row, r, row, r)

                ws['EM{}'.format(row)] = '=IF(SUM(EH{}:EL{})=0,"",SUM(EH{}:EL{}))'.format(
                    row, row, row, row)
                ws['EN{}'.format(row)] = '=IF(EM{}="","",RANK(EM{},$EM$2:$EM${}))'.format(
                    row, row, q)
                ws['EO{}'.format(
                    row)] = '=IF(EN{}="","",COUNTIFS($DV$2:$DV${},F{},$EN$2:$EN${},"<"&EN{})+1)'.format(row, q, row, q, row)
                #     TAMBAHAN
                ws['EP{}'.format(row)] = '=IF($G${}=25,IF(AND(DW{}>4,EH{}=20),1,""),IF($G${}=30,IF(AND(DW{}>5,EH{}=20),1,""),IF($G${}=35,IF(AND(DW{}>6,EH{}=20),1,""),IF($G${}=40,IF(AND(DW{}>7,EH{}=20),1,""),IF($G${}=45,IF(AND(DW{}>8,EH{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['EQ{}'.format(row)] = '=IF($H${}=25,IF(AND(DX{}>4,EI{}=20),1,""),IF($H${}=30,IF(AND(DX{}>5,EI{}=20),1,""),IF($H${}=35,IF(AND(DX{}>6,EI{}=20),1,""),IF($H${}=40,IF(AND(DX{}>7,EI{}=20),1,""),IF($H${}=45,IF(AND(DX{}>8,EI{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['ER{}'.format(row)] = '=IF($I${}=25,IF(AND(DY{}>4,EJ{}=20),1,""),IF($I${}=30,IF(AND(DY{}>5,EJ{}=20),1,""),IF($I${}=35,IF(AND(DY{}>6,EJ{}=20),1,""),IF($I${}=40,IF(AND(DY{}>7,EJ{}=20),1,""),IF($I${}=45,IF(AND(DY{}>8,EJ{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['ES{}'.format(row)] = '=IF($J${}=25,IF(AND(DZ{}>4,EK{}=20),1,""),IF($J${}=30,IF(AND(DZ{}>5,EK{}=20),1,""),IF($J${}=35,IF(AND(DZ{}>6,EK{}=20),1,""),IF($J${}=40,IF(AND(DZ{}>7,EK{}=20),1,""),IF($J${}=45,IF(AND(DZ{}>8,EK{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['ET{}'.format(row)] = '=IF($K${}=25,IF(AND(EA{}>4,EL{}=20),1,""),IF($K${}=30,IF(AND(EA{}>5,EL{}=20),1,""),IF($K${}=35,IF(AND(EA{}>6,EL{}=20),1,""),IF($K${}=40,IF(AND(EA{}>7,EL{}=20),1,""),IF($K${}=45,IF(AND(EA{}>8,EL{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

                # new Z Score [5]
            ws['EV1'] = 'NAMA SISWA'
            ws['EW1'] = 'NOMOR NF'
            ws['EX1'] = 'KELAS'
            ws['EY1'] = 'NAMA SEKOLAH'
            ws['EZ1'] = 'LOKASI'
            ws['FA1'] = 'MAT'
            ws['FB1'] = 'SEJ'
            ws['FC1'] = 'GEO'
            ws['FD1'] = 'EKO'
            ws['FE1'] = 'SOS'
            ws['FF1'] = 'JML'
            ws['FG1'] = 'Z_MAT'
            ws['FH1'] = 'Z_SEJ'
            ws['FI1'] = 'Z_GEO'
            ws['FJ1'] = 'Z_EKO'
            ws['FK1'] = 'Z_SOS'
            ws['FL1'] = 'S_MAT'
            ws['FM1'] = 'S_SEJ'
            ws['FN1'] = 'S_GEO'
            ws['FO1'] = 'S_EKO'
            ws['FP1'] = 'S_SOS'
            ws['FQ1'] = 'S_JML'
            ws['FR1'] = 'RANK NAS.'
            ws['FS1'] = 'RANK LOK.'
            ws['FG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FS1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['EV1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EW1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EX1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EY1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['EZ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FA1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FB1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FC1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FD1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FE1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FF1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FG1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FH1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FI1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FJ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FK1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FL1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FM1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FN1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FO1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FP1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FQ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FR1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FS1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')

            # tambahan
            ws['FT1'] = 'MAT_20'
            ws['FU1'] = 'SEJ_20'
            ws['FV1'] = 'GEO_20'
            ws['FW1'] = 'EKO_20'
            ws['FX1'] = 'SOS_20'
            ws['FT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FT1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FU1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FV1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FW1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['FX1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')

            for row in range(2, q+1):
                ws['EV{}'.format(row)] = '=B{}'.format(row)
                ws['EW{}'.format(row)] = '=C{}'.format(row, row)
                ws['EX{}'.format(row)] = '=D{}'.format(row, row)
                ws['EY{}'.format(row)] = '=E{}'.format(row, row)
                ws['EZ{}'.format(row)] = '=F{}'.format(row, row)
                ws['FA{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['FB{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['FC{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['FD{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['FE{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['FF{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['FG{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FA{}="","",(FA{}-FA${})/FA${}),2),"")'.format(row, row, r, s)
                ws['FH{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FB{}="","",(FB{}-FB${})/FB${}),2),"")'.format(row, row, r, s)
                ws['FI{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FC{}="","",(FC{}-FC${})/FC${}),2),"")'.format(row, row, r, s)
                ws['FJ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FD{}="","",(FD{}-FD${})/FD${}),2),"")'.format(row, row, r, s)
                ws['FK{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FE{}="","",(FE{}-FE${})/FE${}),2),"")'.format(row, row, r, s)
                ws['FL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FG{}="","",IF(70+30*FG{}/$FG${}<20,20,70+30*FG{}/$FG${})),2),"")'.format(row, row, r, row, r)
                ws['FM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FH{}="","",IF(70+30*FH{}/$FH${}<20,20,70+30*FH{}/$FH${})),2),"")'.format(row, row, r, row, r)
                ws['FN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FI{}="","",IF(70+30*FI{}/$FI${}<20,20,70+30*FI{}/$FI${})),2),"")'.format(row, row, r, row, r)
                ws['FO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FJ{}="","",IF(70+30*FJ{}/$FJ${}<20,20,70+30*FJ{}/$FJ${})),2),"")'.format(row, row, r, row, r)
                ws['FP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FK{}="","",IF(70+30*FK{}/$FK${}<20,20,70+30*FK{}/$FK${})),2),"")'.format(row, row, r, row, r)

                ws['FQ{}'.format(row)] = '=IF(SUM(FL{}:FP{})=0,"",SUM(FL{}:FP{}))'.format(
                    row, row, row, row)
                ws['FR{}'.format(row)] = '=IF(FQ{}="","",RANK(FQ{},$FQ$2:$FQ${}))'.format(
                    row, row, q)
                ws['FS{}'.format(
                    row)] = '=IF(FR{}="","",COUNTIFS($EZ$2:$EZ${},F{},$FR$2:$FR${},"<"&FR{})+1)'.format(row, q, row, q, row)
                #     TAMBAHAN
                ws['FT{}'.format(row)] = '=IF($G${}=25,IF(AND(FA{}>4,FL{}=20),1,""),IF($G${}=30,IF(AND(FA{}>5,FL{}=20),1,""),IF($G${}=35,IF(AND(FA{}>6,FL{}=20),1,""),IF($G${}=40,IF(AND(FA{}>7,FL{}=20),1,""),IF($G${}=45,IF(AND(FA{}>8,FL{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['FU{}'.format(row)] = '=IF($H${}=25,IF(AND(FB{}>4,FM{}=20),1,""),IF($H${}=30,IF(AND(FB{}>5,FM{}=20),1,""),IF($H${}=35,IF(AND(FB{}>6,FM{}=20),1,""),IF($H${}=40,IF(AND(FB{}>7,FM{}=20),1,""),IF($H${}=45,IF(AND(FB{}>8,FM{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['FV{}'.format(row)] = '=IF($I${}=25,IF(AND(FC{}>4,FN{}=20),1,""),IF($I${}=30,IF(AND(FC{}>5,FN{}=20),1,""),IF($I${}=35,IF(AND(FC{}>6,FN{}=20),1,""),IF($I${}=40,IF(AND(FC{}>7,FN{}=20),1,""),IF($I${}=45,IF(AND(FC{}>8,FN{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['FW{}'.format(row)] = '=IF($J${}=25,IF(AND(FD{}>4,FO{}=20),1,""),IF($J${}=30,IF(AND(FD{}>5,FO{}=20),1,""),IF($J${}=35,IF(AND(FD{}>6,FO{}=20),1,""),IF($J${}=40,IF(AND(FD{}>7,FO{}=20),1,""),IF($J${}=45,IF(AND(FD{}>8,FO{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['FX{}'.format(row)] = '=IF($K${}=25,IF(AND(FE{}>4,FP{}=20),1,""),IF($K${}=30,IF(AND(FE{}>5,FP{}=20),1,""),IF($K${}=35,IF(AND(FE{}>6,FP{}=20),1,""),IF($K${}=40,IF(AND(FE{}>7,FP{}=20),1,""),IF($K${}=45,IF(AND(FE{}>8,FP{}=20),1,""))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Mengubah 'KELAS' sesuai dengan nilai yang dipilih dari selectbox 'KELAS'
            kelas = KELAS.lower().replace(" ", "")
            semester = SEMESTER.lower()
            tahun = TAHUN.replace("-", "")
            penilaian = PENILAIAN.lower()

            path_file = f"{kelas}_{penilaian}_{semester}_{tahun}_nilai_std.xlsx"

            # Simpan file ke direktori temporer
            temp_dir = tempfile.gettempdir()
            file_path = temp_dir + '/' + path_file
            wb.save(file_path)

            st.success(
                "File siap diunduh!")

            # Tombol unduh file
            with open(file_path, "rb") as f:
                bytes_data = f.read()
            st.download_button(label="Unduh File", data=bytes_data,
                               file_name=path_file)

            st.warning(
                "Buka file unduhan, klik 'Enable Editing' dan 'Save'")
    if selected_file == "Nilai Std. 11KM":
        # menghilangkan hamburger
        st.markdown("""
        <style>
        .css-1rs6os.edgvbvh3
        {
            visibility:hidden;
        }
        .css-1lsmgbg.egzxvld0
        {
            visibility:hidden;
        }
        </style>
        """, unsafe_allow_html=True)

        image = Image.open('logo resmi nf resize.png')
        st.image(image)

        st.title("Olah Nilai Standar")
        st.header("11 SMA KM")

        col6 = st.container()

        with col6:
            KELAS = st.selectbox(
                "KELAS",
                ("--Pilih Kelas--", "11 SMA"))

        col7 = st.container()

        with col7:
            SEMESTER = st.selectbox(
                "SEMESTER",
                ("--Pilih Semester--", "SEMESTER 1", "SEMESTER 2"))

        col8 = st.container()

        with col8:
            PENILAIAN = st.selectbox(
                "PENILAIAN",
                ("--Pilih Penilaian--", "PENILAIAN TENGAH SEMESTER", "PENILAIAN AKHIR SEMESTER", "PENILAIAN AKHIR TAHUN"))

        col9 = st.container()

        with col9:
            KURIKULUM = st.selectbox(
                "KURIKULUM",
                ("--Pilih Kurikulum--", "KM"))

        TAHUN = st.text_input("Masukkan Tahun Ajaran",
                              placeholder="contoh: 2022-2023")

        col1, col2, col3, col4, col5, col6 = st.columns(
            6)

        with col1:
            MTW = st.selectbox(
                "JML. SOAL MAW.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col2:
            MTP = st.selectbox(
                "JML. SOAL MAP.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col3:
            IND = st.selectbox(
                "JML. SOAL IND.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col4:
            ENG = st.selectbox(
                "JML. SOAL ENG.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col5:
            SEJ = st.selectbox(
                "JML. SOAL SEJ.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col6:
            GEO = st.selectbox(
                "JML. SOAL GEO.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        col7, col8, col9, col10, col11 = st.columns(5)

        with col7:
            EKO = st.selectbox(
                "JML. SOAL EKO.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col8:
            SOS = st.selectbox(
                "JML. SOAL SOS.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col9:
            FIS = st.selectbox(
                "JML. SOAL FIS.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col10:
            KIM = st.selectbox(
                "JML. SOAL KIM.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        with col11:
            BIO = st.selectbox(
                "JML. SOAL BIO.",
                (15, 20, 25, 30, 35, 40, 45, 50))

        JML_SOAL_MAW = MTW
        JML_SOAL_MAP = MTP
        JML_SOAL_IND = IND
        JML_SOAL_ENG = ENG
        JML_SOAL_SEJ = SEJ
        JML_SOAL_GEO = GEO
        JML_SOAL_EKO = EKO
        JML_SOAL_SOS = SOS
        JML_SOAL_FIS = FIS
        JML_SOAL_KIM = KIM
        JML_SOAL_BIO = BIO

        uploaded_file = st.file_uploader(
            'Letakkan file excel 11 SMA', type='xlsx')

        if uploaded_file is not None:
            wb = openpyxl.load_workbook(uploaded_file)
            ws = wb['Sheet1']

            q = len(ws['K'])
            r = len(ws['K'])+2
            s = len(ws['K'])+3
            t = len(ws['K'])+4
            u = len(ws['K'])+5
            v = len(ws['K'])+6
            w = len(ws['K'])+7
            x = len(ws['K'])+8

            ws['G{}'.format(r)] = "=ROUND(AVERAGE(G2:G{}),2)".format(q)
            ws['H{}'.format(r)] = "=ROUND(AVERAGE(H2:H{}),2)".format(q)
            ws['I{}'.format(r)] = "=ROUND(AVERAGE(I2:I{}),2)".format(q)
            ws['J{}'.format(r)] = "=ROUND(AVERAGE(J2:J{}),2)".format(q)
            ws['K{}'.format(r)] = "=ROUND(AVERAGE(K2:K{}),2)".format(q)
            ws['L{}'.format(r)] = "=ROUND(AVERAGE(L2:L{}),2)".format(q)
            ws['M{}'.format(r)] = "=ROUND(AVERAGE(M2:M{}),2)".format(q)
            ws['N{}'.format(r)] = "=ROUND(AVERAGE(N2:N{}),2)".format(q)
            ws['O{}'.format(r)] = "=ROUND(AVERAGE(O2:O{}),2)".format(q)
            ws['P{}'.format(r)] = "=ROUND(AVERAGE(P2:P{}),2)".format(q)
            ws['Q{}'.format(r)] = "=ROUND(AVERAGE(Q2:Q{}),2)".format(q)
            ws['R{}'.format(r)] = "=ROUND(AVERAGE(R2:R{}),2)".format(q)

            ws['G{}'.format(s)] = "=STDEV(G2:G{})".format(q)
            ws['H{}'.format(s)] = "=STDEV(H2:H{})".format(q)
            ws['I{}'.format(s)] = "=STDEV(I2:I{})".format(q)
            ws['J{}'.format(s)] = "=STDEV(J2:J{})".format(q)
            ws['K{}'.format(s)] = "=STDEV(K2:K{})".format(q)
            ws['L{}'.format(s)] = "=STDEV(L2:L{})".format(q)
            ws['M{}'.format(s)] = "=STDEV(M2:M{})".format(q)
            ws['N{}'.format(s)] = "=STDEV(N2:N{})".format(q)
            ws['O{}'.format(s)] = "=STDEV(O2:O{})".format(q)
            ws['P{}'.format(s)] = "=STDEV(P2:P{})".format(q)
            ws['Q{}'.format(s)] = "=STDEV(Q2:Q{})".format(q)

            ws['G{}'.format(t)] = "=MAX(G2:G{})".format(q)
            ws['H{}'.format(t)] = "=MAX(H2:H{})".format(q)
            ws['I{}'.format(t)] = "=MAX(I2:I{})".format(q)
            ws['J{}'.format(t)] = "=MAX(J2:J{})".format(q)
            ws['K{}'.format(t)] = "=MAX(K2:K{})".format(q)
            ws['L{}'.format(t)] = "=MAX(L2:L{})".format(q)
            ws['M{}'.format(t)] = "=MAX(M2:M{})".format(q)
            ws['N{}'.format(t)] = "=MAX(N2:N{})".format(q)
            ws['O{}'.format(t)] = "=MAX(O2:O{})".format(q)
            ws['P{}'.format(t)] = "=MAX(P2:P{})".format(q)
            ws['Q{}'.format(t)] = "=MAX(Q2:Q{})".format(q)

            ws['R{}'.format(r)] = "=MAX(R2:R{})".format(q)
            ws['S{}'.format(r)] = "=MAX(S2:S{})".format(q)
            ws['T{}'.format(r)] = "=MAX(T2:T{})".format(q)
            ws['U{}'.format(r)] = "=MAX(U2:U{})".format(q)
            ws['V{}'.format(r)] = "=MAX(V2:V{})".format(q)
            ws['W{}'.format(r)] = "=MAX(W2:W{})".format(q)
            ws['X{}'.format(r)] = "=MAX(X2:X{})".format(q)
            ws['Y{}'.format(r)] = "=MAX(Y2:Y{})".format(q)
            ws['Z{}'.format(r)] = "=MAX(Z2:Z{})".format(q)
            ws['AA{}'.format(r)] = "=MAX(AA2:AA{})".format(q)
            ws['AB{}'.format(r)] = "=MAX(AB2:AB{})".format(q)
            ws['AC{}'.format(r)] = "=MAX(AC2:AC{})".format(q)
            ws['AD{}'.format(r)] = "=MAX(AD2:AD{})".format(q)
            ws['AE{}'.format(r)] = "=MAX(AE2:AE{})".format(q)
            ws['AF{}'.format(r)] = "=MAX(AF2:AF{})".format(q)
            ws['AG{}'.format(r)] = "=MAX(AG2:AG{})".format(q)
            ws['AH{}'.format(r)] = "=MAX(AH2:AH{})".format(q)
            ws['AI{}'.format(r)] = "=MAX(AI2:AI{})".format(q)
            ws['AJ{}'.format(r)] = "=MAX(AJ2:AJ{})".format(q)
            ws['AK{}'.format(r)] = "=MAX(AK2:AK{})".format(q)
            ws['AL{}'.format(r)] = "=MAX(AL2:AL{})".format(q)
            ws['AM{}'.format(r)] = "=MAX(AM2:AM{})".format(q)
            ws['AN{}'.format(r)] = "=MAX(AN2:AN{})".format(q)
            ws['AO{}'.format(r)] = "=MAX(AO2:AO{})".format(q)
            ws['AP{}'.format(r)] = "=MAX(AP2:AP{})".format(q)

            ws['G{}'.format(u)] = "=MIN(G2:G{})".format(q)
            ws['H{}'.format(u)] = "=MIN(H2:H{})".format(q)
            ws['I{}'.format(u)] = "=MIN(I2:I{})".format(q)
            ws['J{}'.format(u)] = "=MIN(J2:J{})".format(q)
            ws['K{}'.format(u)] = "=MIN(K2:K{})".format(q)
            ws['L{}'.format(u)] = "=MIN(L2:L{})".format(q)
            ws['M{}'.format(u)] = "=MIN(M2:M{})".format(q)
            ws['N{}'.format(u)] = "=MIN(N2:N{})".format(q)
            ws['O{}'.format(u)] = "=MIN(O2:O{})".format(q)
            ws['P{}'.format(u)] = "=MIN(P2:P{})".format(q)
            ws['Q{}'.format(u)] = "=MIN(Q2:Q{})".format(q)
            ws['R{}'.format(u)] = "=MIN(R2:R{})".format(q)

            ws['AD{}'.format(s)] = "=MIN(AD2:AD{})".format(q)
            ws['AE{}'.format(s)] = "=MIN(AE2:AE{})".format(q)
            ws['AF{}'.format(s)] = "=MIN(AF2:AF{})".format(q)
            ws['AG{}'.format(s)] = "=MIN(AG2:AG{})".format(q)
            ws['AH{}'.format(s)] = "=MIN(AH2:AH{})".format(q)
            ws['AI{}'.format(s)] = "=MIN(AI2:AI{})".format(q)
            ws['AJ{}'.format(s)] = "=MIN(AJ2:AJ{})".format(q)
            ws['AK{}'.format(s)] = "=MIN(AK2:AK{})".format(q)
            ws['AL{}'.format(s)] = "=MIN(AL2:AL{})".format(q)
            ws['AM{}'.format(s)] = "=MIN(AM2:AM{})".format(q)
            ws['AN{}'.format(s)] = "=MIN(AN2:AN{})".format(q)
            ws['AO{}'.format(s)] = "=MIN(AO2:AO{})".format(q)

            ws['AD{}'.format(t)] = "=ROUND(AVERAGE(AD2:AD{}),2)".format(q)
            ws['AE{}'.format(t)] = "=ROUND(AVERAGE(AE2:AE{}),2)".format(q)
            ws['AF{}'.format(t)] = "=ROUND(AVERAGE(AF2:AF{}),2)".format(q)
            ws['AG{}'.format(t)] = "=ROUND(AVERAGE(AG2:AG{}),2)".format(q)
            ws['AH{}'.format(t)] = "=ROUND(AVERAGE(AH2:AH{}),2)".format(q)
            ws['AI{}'.format(t)] = "=ROUND(AVERAGE(AI2:AI{}),2)".format(q)
            ws['AJ{}'.format(t)] = "=ROUND(AVERAGE(AJ2:AJ{}),2)".format(q)
            ws['AK{}'.format(t)] = "=ROUND(AVERAGE(AK2:AK{}),2)".format(q)
            ws['AL{}'.format(t)] = "=ROUND(AVERAGE(AL2:AL{}),2)".format(q)
            ws['AM{}'.format(t)] = "=ROUND(AVERAGE(AM2:AM{}),2)".format(q)
            ws['AN{}'.format(t)] = "=ROUND(AVERAGE(AN2:AN{}),2)".format(q)
            ws['AO{}'.format(t)] = "=ROUND(AVERAGE(AO2:AO{}),2)".format(q)

            ws['AR{}'.format(r)] = "=SUM(AR2:AR{})".format(q)
            ws['AS{}'.format(r)] = "=SUM(AS2:AS{})".format(q)
            ws['AT{}'.format(r)] = "=SUM(AT2:AT{})".format(q)
            ws['AU{}'.format(r)] = "=SUM(AU2:AU{})".format(q)
            ws['AV{}'.format(r)] = "=SUM(AV2:AV{})".format(q)
            ws['AW{}'.format(r)] = "=SUM(AW2:AW{})".format(q)
            ws['AX{}'.format(r)] = "=SUM(AX2:AX{})".format(q)
            ws['AY{}'.format(r)] = "=SUM(AY2:AY{})".format(q)
            ws['AZ{}'.format(r)] = "=SUM(AZ2:AZ{})".format(q)
            ws['BA{}'.format(r)] = "=SUM(BA2:BA{})".format(q)
            ws['BB{}'.format(r)] = "=SUM(BB2:BB{})".format(q)

            # Jumlah Soal
            ws['F{}'.format(v)] = 'JUMLAH SOAL'
            ws['G{}'.format(v)] = JML_SOAL_MAW
            ws['H{}'.format(v)] = JML_SOAL_MAP
            ws['I{}'.format(v)] = JML_SOAL_IND
            ws['J{}'.format(v)] = JML_SOAL_ENG
            ws['K{}'.format(v)] = JML_SOAL_SEJ
            ws['L{}'.format(v)] = JML_SOAL_GEO
            ws['M{}'.format(v)] = JML_SOAL_EKO
            ws['N{}'.format(v)] = JML_SOAL_SOS
            ws['O{}'.format(v)] = JML_SOAL_FIS
            ws['P{}'.format(v)] = JML_SOAL_KIM
            ws['Q{}'.format(v)] = JML_SOAL_BIO

            # new
            # iterasi 1 rata-rata - 1
            # rata" MTW ke MTW tambahan dan mapel MTW awal
            ws['BI{}'.format(
                r)] = "=IF($AR${}=0,$G${},$G${}-1)".format(r, r, r)
            ws['BI{}'.format(s)] = "=STDEV(BI2:BI{})".format(q)
            ws['BI{}'.format(t)] = "=MAX(BI2:BI{})".format(q)
            ws['BI{}'.format(u)] = "=MIN(BI2:BI{})".format(q)
            # rata" MTP ke MTP tambahan dan mapel MTP awal
            ws['BJ{}'.format(
                r)] = "=IF($AS${}=0,$H${},$H${}-1)".format(r, r, r)
            ws['BJ{}'.format(s)] = "=STDEV(BJ2:BJ{})".format(q)
            ws['BJ{}'.format(t)] = "=MAX(BJ2:BJ{})".format(q)
            ws['BJ{}'.format(u)] = "=MIN(BJ2:BJ{})".format(q)
            # rata" IND ke IND tambahan dan mapel IND awal
            ws['BK{}'.format(
                r)] = "=IF($AT${}=0,$I${},$I${}-1)".format(r, r, r)
            ws['BK{}'.format(s)] = "=STDEV(BK2:BK{})".format(q)
            ws['BK{}'.format(t)] = "=MAX(BK2:BK{})".format(q)
            ws['BK{}'.format(u)] = "=MIN(BK2:BK{})".format(q)
            # rata" ENG ke ENG tambahan dan mapel ENG awal
            ws['BL{}'.format(
                r)] = "=IF($AU${}=0,$J${},$J${}-1)".format(r, r, r)
            ws['BL{}'.format(s)] = "=STDEV(BL2:BL{})".format(q)
            ws['BL{}'.format(t)] = "=MAX(BL2:BL{})".format(q)
            ws['BL{}'.format(u)] = "=MIN(BL2:BL{})".format(q)
            # rata" SEJ ke SEJ tambahan dan mapel SEJ awal
            ws['BM{}'.format(
                r)] = "=IF($AV${}=0,$K${},$K${}-1)".format(r, r, r)
            ws['BM{}'.format(s)] = "=STDEV(BM2:BM{})".format(q)
            ws['BM{}'.format(t)] = "=MAX(BM2:BM{})".format(q)
            ws['BM{}'.format(u)] = "=MIN(BM2:BM{})".format(q)
            # rata" GEO ke GEO tambahan dan mapel GEO awal
            ws['BN{}'.format(
                r)] = "=IF($AW${}=0,$L${},$L${}-1)".format(r, r, r)
            ws['BN{}'.format(s)] = "=STDEV(BN2:BN{})".format(q)
            ws['BN{}'.format(t)] = "=MAX(BN2:BN{})".format(q)
            ws['BN{}'.format(u)] = "=MIN(BN2:BN{})".format(q)
            # rata" EKO ke EKO tambahan dan mapel EKO awal
            ws['BO{}'.format(
                r)] = "=IF($AX${}=0,$M${},$M${}-1)".format(r, r, r)
            ws['BO{}'.format(s)] = "=STDEV(BO2:BO{})".format(q)
            ws['BO{}'.format(t)] = "=MAX(BO2:BO{})".format(q)
            ws['BO{}'.format(u)] = "=MIN(BO2:BO{})".format(q)
            # rata" SOS ke SOS tambahan dan mapel SOS awal
            ws['BP{}'.format(
                r)] = "=IF($AY${}=0,$N${},$N${}-1)".format(r, r, r)
            ws['BP{}'.format(s)] = "=STDEV(BP2:BP{})".format(q)
            ws['BP{}'.format(t)] = "=MAX(BP2:BP{})".format(q)
            ws['BP{}'.format(u)] = "=MIN(BP2:BP{})".format(q)
            # rata" FIS ke FIS tambahan dan mapel FIS awal
            ws['BQ{}'.format(
                r)] = "=IF($AZ${}=0,$O${},$O${}-1)".format(r, r, r)
            ws['BQ{}'.format(s)] = "=STDEV(BQ2:BQ{})".format(q)
            ws['BQ{}'.format(t)] = "=MAX(BQ2:BQ{})".format(q)
            ws['BQ{}'.format(u)] = "=MIN(BQ2:BQ{})".format(q)
            # rata" KIM ke KIM tambahan dan mapel KIM awal
            ws['BR{}'.format(
                r)] = "=IF($BA${}=0,$P${},$P${}-1)".format(r, r, r)
            ws['BR{}'.format(s)] = "=STDEV(BR2:BR{})".format(q)
            ws['BR{}'.format(t)] = "=MAX(BR2:BR{})".format(q)
            ws['BR{}'.format(u)] = "=MIN(BR2:BR{})".format(q)
            # rata" BIO ke BIO tambahan dan mapel BIO awal
            ws['BS{}'.format(
                r)] = "=IF($BB${}=0,$Q${},$Q${}-1)".format(r, r, r)
            ws['BS{}'.format(s)] = "=STDEV(BS2:BS{})".format(q)
            ws['BS{}'.format(t)] = "=MAX(BS2:BS{})".format(q)
            ws['BS{}'.format(u)] = "=MIN(BS2:BS{})".format(q)
            # jml MAPEL
            ws['BT{}'.format(r)] = "=ROUND(AVERAGE(BT2:BT{}),2)".format(q)
            ws['BT{}'.format(t)] = "=MAX(BT2:BT{})".format(q)
            ws['BT{}'.format(u)] = "=MIN(BT2:BT{})".format(q)
            # MAX Z SCORE
            ws['BU{}'.format(r)] = "=MAX(BU2:BU{})".format(q)
            ws['BV{}'.format(r)] = "=MAX(BV2:BV{})".format(q)
            ws['BW{}'.format(r)] = "=MAX(BW2:BW{})".format(q)
            ws['BX{}'.format(r)] = "=MAX(BX2:BX{})".format(q)
            ws['BY{}'.format(r)] = "=MAX(BY2:BY{})".format(q)
            ws['BZ{}'.format(r)] = "=MAX(BZ2:BZ{})".format(q)
            ws['CA{}'.format(r)] = "=MAX(CA2:CA{})".format(q)
            ws['CB{}'.format(r)] = "=MAX(CB2:CB{})".format(q)
            ws['CC{}'.format(r)] = "=MAX(CC2:CC{})".format(q)
            ws['CD{}'.format(r)] = "=MAX(CD2:CD{})".format(q)
            ws['CE{}'.format(r)] = "=MAX(CE2:CE{})".format(q)
            # NILAI STANDAR MTW
            ws['CF{}'.format(r)] = "=MAX(CF2:CF{})".format(q)
            ws['CF{}'.format(s)] = "=MIN(CF2:CF{})".format(q)
            ws['CF{}'.format(t)] = "=ROUND(AVERAGE(CF2:CF{}),2)".format(q)
            # NILAI STANDAR MTP
            ws['CG{}'.format(r)] = "=MAX(CG2:CG{})".format(q)
            ws['CG{}'.format(s)] = "=MIN(CG2:CG{})".format(q)
            ws['CG{}'.format(t)] = "=ROUND(AVERAGE(CG2:CG{}),2)".format(q)
            # NILAI STANDAR IND
            ws['CH{}'.format(r)] = "=MAX(CH2:CH{})".format(q)
            ws['CH{}'.format(s)] = "=MIN(CH2:CH{})".format(q)
            ws['CH{}'.format(t)] = "=ROUND(AVERAGE(CH2:CH{}),2)".format(q)
            # NILAI STANDAR ENG
            ws['CI{}'.format(r)] = "=MAX(CI2:CI{})".format(q)
            ws['CI{}'.format(s)] = "=MIN(CI2:CI{})".format(q)
            ws['CI{}'.format(t)] = "=ROUND(AVERAGE(CI2:CI{}),2)".format(q)
            # NILAI STANDAR SEJ
            ws['CJ{}'.format(r)] = "=MAX(CJ2:CJ{})".format(q)
            ws['CJ{}'.format(s)] = "=MIN(CJ2:CJ{})".format(q)
            ws['CJ{}'.format(t)] = "=ROUND(AVERAGE(CJ2:CJ{}),2)".format(q)
            # NILAI STANDAR GEO
            ws['CK{}'.format(r)] = "=MAX(CK2:CK{})".format(q)
            ws['CK{}'.format(s)] = "=MIN(CK2:CK{})".format(q)
            ws['CK{}'.format(t)] = "=ROUND(AVERAGE(CK2:CK{}),2)".format(q)
            # NILAI STANDAR EKO
            ws['CL{}'.format(r)] = "=MAX(CL2:CL{})".format(q)
            ws['CL{}'.format(s)] = "=MIN(CL2:CL{})".format(q)
            ws['CL{}'.format(t)] = "=ROUND(AVERAGE(CL2:CL{}),2)".format(q)
            # NILAI STANDAR SOS
            ws['CM{}'.format(r)] = "=MAX(CM2:CM{})".format(q)
            ws['CM{}'.format(s)] = "=MIN(CM2:CM{})".format(q)
            ws['CM{}'.format(t)] = "=ROUND(AVERAGE(CM2:CM{}),2)".format(q)
            # NILAI STANDAR FIS
            ws['CN{}'.format(r)] = "=MAX(CN2:CN{})".format(q)
            ws['CN{}'.format(s)] = "=MIN(CN2:CN{})".format(q)
            ws['CN{}'.format(t)] = "=ROUND(AVERAGE(CN2:CN{}),2)".format(q)
            # NILAI STANDAR KIM
            ws['CO{}'.format(r)] = "=MAX(CO2:CO{})".format(q)
            ws['CO{}'.format(s)] = "=MIN(CO2:CO{})".format(q)
            ws['CO{}'.format(t)] = "=ROUND(AVERAGE(CO2:CO{}),2)".format(q)
            # NILAI STANDAR BIO
            ws['CP{}'.format(r)] = "=MAX(CP2:CP{})".format(q)
            ws['CP{}'.format(s)] = "=MIN(CP2:CP{})".format(q)
            ws['CP{}'.format(t)] = "=ROUND(AVERAGE(CP2:CP{}),2)".format(q)
            # NILAI STANDAR S.JML
            ws['CQ{}'.format(r)] = "=MAX(CQ2:CQ{})".format(q)
            ws['CQ{}'.format(s)] = "=MIN(CQ2:CQ{})".format(q)
            ws['CQ{}'.format(t)] = "=ROUND(AVERAGE(CQ2:CQ{}),2)".format(q)

            # TAMBAHAN
            ws['CT{}'.format(r)] = "=SUM(CT2:CT{})".format(q)
            ws['CU{}'.format(r)] = "=SUM(CU2:CU{})".format(q)
            ws['CV{}'.format(r)] = "=SUM(CV2:CV{})".format(q)
            ws['CW{}'.format(r)] = "=SUM(CW2:CW{})".format(q)
            ws['CX{}'.format(r)] = "=SUM(CX2:CX{})".format(q)
            ws['CY{}'.format(r)] = "=SUM(CY2:CY{})".format(q)
            ws['CZ{}'.format(r)] = "=SUM(CZ2:CZ{})".format(q)
            ws['DA{}'.format(r)] = "=SUM(DA2:DA{})".format(q)
            ws['DB{}'.format(r)] = "=SUM(DB2:DB{})".format(q)
            ws['DC{}'.format(r)] = "=SUM(DC2:DC{})".format(q)
            ws['DD{}'.format(r)] = "=SUM(DD2:DD{})".format(q)

            # iterasi 2 rata-rata - 2
            # rata" MTW ke MTW tambahan dan mapel MTW awal
            ws['DK{}'.format(
                r)] = "=IF($CT${}=0,$BI${},$BI${}-1)".format(r, r, r)
            ws['DK{}'.format(s)] = "=STDEV(DK2:DK{})".format(q)
            ws['DK{}'.format(t)] = "=MAX(DK2:DK{})".format(q)
            ws['DK{}'.format(u)] = "=MIN(DK2:DK{})".format(q)
            # rata" MTP ke MTP tambahan dan mapel MTP awal
            ws['DL{}'.format(
                r)] = "=IF($CU${}=0,$BJ${},$BJ${}-1)".format(r, r, r)
            ws['DL{}'.format(s)] = "=STDEV(DL2:DL{})".format(q)
            ws['DL{}'.format(t)] = "=MAX(DL2:DL{})".format(q)
            ws['DL{}'.format(u)] = "=MIN(DL2:DL{})".format(q)
            # rata" IND ke IND tambahan dan mapel IND awal
            ws['DM{}'.format(
                r)] = "=IF($CV${}=0,$BK${},$BK${}-1)".format(r, r, r)
            ws['DM{}'.format(s)] = "=STDEV(DM2:DM{})".format(q)
            ws['DM{}'.format(t)] = "=MAX(DM2:DM{})".format(q)
            ws['DM{}'.format(u)] = "=MIN(DM2:DM{})".format(q)
            # rata" ENG ke ENG tambahan dan mapel ENG awal
            ws['DN{}'.format(
                r)] = "=IF($CW${}=0,$BL${},$BL${}-1)".format(r, r, r)
            ws['DN{}'.format(s)] = "=STDEV(DN2:DN{})".format(q)
            ws['DN{}'.format(t)] = "=MAX(DN2:DN{})".format(q)
            ws['DN{}'.format(u)] = "=MIN(DN2:DN{})".format(q)
            # rata" SEJ ke SEJ tambahan dan mapel SEJ awal
            ws['DO{}'.format(
                r)] = "=IF($CX${}=0,$BM${},$BM${}-1)".format(r, r, r)
            ws['DO{}'.format(s)] = "=STDEV(DO2:DO{})".format(q)
            ws['DO{}'.format(t)] = "=MAX(DO2:DO{})".format(q)
            ws['DO{}'.format(u)] = "=MIN(DO2:DO{})".format(q)
            # rata" GEO ke GEO tambahan dan mapel GEO awal
            ws['DP{}'.format(
                r)] = "=IF($CY${}=0,$BN${},$BN${}-1)".format(r, r, r)
            ws['DP{}'.format(s)] = "=STDEV(DP2:DP{})".format(q)
            ws['DP{}'.format(t)] = "=MAX(DP2:DP{})".format(q)
            ws['DP{}'.format(u)] = "=MIN(DP2:DP{})".format(q)
            # rata" EKO ke EKO tambahan dan mapel EKO awal
            ws['DQ{}'.format(
                r)] = "=IF($CZ${}=0,$BO${},$BO${}-1)".format(r, r, r)
            ws['DQ{}'.format(s)] = "=STDEV(DQ2:DQ{})".format(q)
            ws['DQ{}'.format(t)] = "=MAX(DQ2:DQ{})".format(q)
            ws['DQ{}'.format(u)] = "=MIN(DQ2:DQ{})".format(q)
            # rata" SOS ke SOS tambahan dan mapel SOS awal
            ws['DR{}'.format(
                r)] = "=IF($DA${}=0,$BP${},$BP${}-1)".format(r, r, r)
            ws['DR{}'.format(s)] = "=STDEV(DR2:DR{})".format(q)
            ws['DR{}'.format(t)] = "=MAX(DR2:DR{})".format(q)
            ws['DR{}'.format(u)] = "=MIN(DR2:DR{})".format(q)
            # rata" FIS ke FIS tambahan dan mapel FIS awal
            ws['DS{}'.format(
                r)] = "=IF($DB${}=0,$BQ${},$BQ${}-1)".format(r, r, r)
            ws['DS{}'.format(s)] = "=STDEV(DS2:DS{})".format(q)
            ws['DS{}'.format(t)] = "=MAX(DS2:DS{})".format(q)
            ws['DS{}'.format(u)] = "=MIN(DS2:DS{})".format(q)
            # rata" KIM ke KIM tambahan dan mapel KIM awal
            ws['DT{}'.format(
                r)] = "=IF($DC${}=0,$BR${},$BR${}-1)".format(r, r, r)
            ws['DT{}'.format(s)] = "=STDEV(DT2:DT{})".format(q)
            ws['DT{}'.format(t)] = "=MAX(DT2:DT{})".format(q)
            ws['DT{}'.format(u)] = "=MIN(DT2:DT{})".format(q)
            # rata" BIO ke BIO tambahan dan mapel BIO awal
            ws['DU{}'.format(
                r)] = "=IF($DD${}=0,$BS${},$BS${}-1)".format(r, r, r)
            ws['DU{}'.format(s)] = "=STDEV(DU2:DU{})".format(q)
            ws['DU{}'.format(t)] = "=MAX(DU2:DU{})".format(q)
            ws['DU{}'.format(u)] = "=MIN(DU2:DU{})".format(q)
            # jml MAPEL
            ws['DV{}'.format(r)] = "=ROUND(AVERAGE(DV2:DV{}),2)".format(q)
            ws['DV{}'.format(t)] = "=MAX(DV2:DV{})".format(q)
            ws['DV{}'.format(u)] = "=MIN(DV2:DV{})".format(q)
            # MAX Z SCORE
            ws['DW{}'.format(r)] = "=MAX(DW2:DW{})".format(q)
            ws['DX{}'.format(r)] = "=MAX(DX2:DX{})".format(q)
            ws['DY{}'.format(r)] = "=MAX(DY2:DY{})".format(q)
            ws['DZ{}'.format(r)] = "=MAX(DZ2:DZ{})".format(q)
            ws['EA{}'.format(r)] = "=MAX(EA2:EA{})".format(q)
            ws['EB{}'.format(r)] = "=MAX(EB2:EB{})".format(q)
            ws['EC{}'.format(r)] = "=MAX(EC2:EC{})".format(q)
            ws['ED{}'.format(r)] = "=MAX(ED2:ED{})".format(q)
            ws['EE{}'.format(r)] = "=MAX(EE2:EE{})".format(q)
            ws['EF{}'.format(r)] = "=MAX(EF2:EF{})".format(q)
            ws['EG{}'.format(r)] = "=MAX(EG2:EG{})".format(q)
            # NILAI STANDAR MTW
            ws['EH{}'.format(r)] = "=MAX(EH2:EH{})".format(q)
            ws['EH{}'.format(s)] = "=MIN(EH2:EH{})".format(q)
            ws['EH{}'.format(t)] = "=ROUND(AVERAGE(EH2:EH{}),2)".format(q)
            # NILAI STANDAR MTP
            ws['EI{}'.format(r)] = "=MAX(EI2:EI{})".format(q)
            ws['EI{}'.format(s)] = "=MIN(EI2:EI{})".format(q)
            ws['EI{}'.format(t)] = "=ROUND(AVERAGE(EI2:EI{}),2)".format(q)
            # NILAI STANDAR IND
            ws['EJ{}'.format(r)] = "=MAX(EJ2:EJ{})".format(q)
            ws['EJ{}'.format(s)] = "=MIN(EJ2:EJ{})".format(q)
            ws['EJ{}'.format(t)] = "=ROUND(AVERAGE(EJ2:EJ{}),2)".format(q)
            # NILAI STANDAR ENG
            ws['EK{}'.format(r)] = "=MAX(EK2:EK{})".format(q)
            ws['EK{}'.format(s)] = "=MIN(EK2:EK{})".format(q)
            ws['EK{}'.format(t)] = "=ROUND(AVERAGE(EK2:EK{}),2)".format(q)
            # NILAI STANDAR SEJ
            ws['EL{}'.format(r)] = "=MAX(EL2:EL{})".format(q)
            ws['EL{}'.format(s)] = "=MIN(EL2:EL{})".format(q)
            ws['EL{}'.format(t)] = "=ROUND(AVERAGE(EL2:EL{}),2)".format(q)
            # NILAI STANDAR GEO
            ws['EM{}'.format(r)] = "=MAX(EM2:EM{})".format(q)
            ws['EM{}'.format(s)] = "=MIN(EM2:EM{})".format(q)
            ws['EM{}'.format(t)] = "=ROUND(AVERAGE(EM2:EM{}),2)".format(q)
            # NILAI STANDAR EKO
            ws['EN{}'.format(r)] = "=MAX(EN2:EN{})".format(q)
            ws['EN{}'.format(s)] = "=MIN(EN2:EN{})".format(q)
            ws['EN{}'.format(t)] = "=ROUND(AVERAGE(EN2:EN{}),2)".format(q)
            # NILAI STANDAR SOS
            ws['EO{}'.format(r)] = "=MAX(EO2:EO{})".format(q)
            ws['EO{}'.format(s)] = "=MIN(EO2:EO{})".format(q)
            ws['EO{}'.format(t)] = "=ROUND(AVERAGE(EO2:EO{}),2)".format(q)
            # NILAI STANDAR FIS
            ws['EP{}'.format(r)] = "=MAX(EP2:EP{})".format(q)
            ws['EP{}'.format(s)] = "=MIN(EP2:EP{})".format(q)
            ws['EP{}'.format(t)] = "=ROUND(AVERAGE(EP2:EP{}),2)".format(q)
            # NILAI STANDAR KIM
            ws['EQ{}'.format(r)] = "=MAX(EQ2:EQ{})".format(q)
            ws['EQ{}'.format(s)] = "=MIN(EQ2:EQ{})".format(q)
            ws['EQ{}'.format(t)] = "=ROUND(AVERAGE(EQ2:EQ{}),2)".format(q)
            # NILAI STANDAR BIO
            ws['ER{}'.format(r)] = "=MAX(ER2:ER{})".format(q)
            ws['ER{}'.format(s)] = "=MIN(ER2:ER{})".format(q)
            ws['ER{}'.format(t)] = "=ROUND(AVERAGE(ER2:ER{}),2)".format(q)
            # NILAI STANDAR S.JML
            ws['ES{}'.format(r)] = "=MAX(ES2:ES{})".format(q)
            ws['ES{}'.format(s)] = "=MIN(ES2:ES{})".format(q)
            ws['ES{}'.format(t)] = "=ROUND(AVERAGE(ES2:ES{}),2)".format(q)

            # TAMBAHAN
            ws['EV{}'.format(r)] = "=SUM(EV2:EV{})".format(q)
            ws['EW{}'.format(r)] = "=SUM(EW2:EW{})".format(q)
            ws['EX{}'.format(r)] = "=SUM(EX2:EX{})".format(q)
            ws['EY{}'.format(r)] = "=SUM(EY2:EY{})".format(q)
            ws['EZ{}'.format(r)] = "=SUM(EZ2:EZ{})".format(q)
            ws['FA{}'.format(r)] = "=SUM(FA2:FA{})".format(q)
            ws['FB{}'.format(r)] = "=SUM(FB2:FB{})".format(q)
            ws['FC{}'.format(r)] = "=SUM(FC2:FC{})".format(q)
            ws['FD{}'.format(r)] = "=SUM(FD2:FD{})".format(q)
            ws['FE{}'.format(r)] = "=SUM(FE2:FE{})".format(q)
            ws['FF{}'.format(r)] = "=SUM(FF2:FF{})".format(q)

            # iterasi 3 rata-rata - 3
            # rata" MTW ke MTW tambahan dan mapel MTW awal
            ws['FM{}'.format(
                r)] = "=IF($EV${}=0,$DK${},$DK${}-1)".format(r, r, r)
            ws['FM{}'.format(s)] = "=STDEV(FM2:FM{})".format(q)
            ws['FM{}'.format(t)] = "=MAX(FM2:FM{})".format(q)
            ws['FM{}'.format(u)] = "=MIN(FM2:FM{})".format(q)
            # rata" MTP ke MTP tambahan dan mapel MTP awal
            ws['FN{}'.format(
                r)] = "=IF($EW${}=0,$DL${},$DL${}-1)".format(r, r, r)
            ws['FN{}'.format(s)] = "=STDEV(FN2:FN{})".format(q)
            ws['FN{}'.format(t)] = "=MAX(FN2:FN{})".format(q)
            ws['FN{}'.format(u)] = "=MIN(FN2:FN{})".format(q)
            # rata" IND ke IND tambahan dan mapel IND awal
            ws['FO{}'.format(
                r)] = "=IF($EX${}=0,$DM${},$DM${}-1)".format(r, r, r)
            ws['FO{}'.format(s)] = "=STDEV(FO2:FO{})".format(q)
            ws['FO{}'.format(t)] = "=MAX(FO2:FO{})".format(q)
            ws['FO{}'.format(u)] = "=MIN(FO2:FO{})".format(q)
            # rata" ENG ke ENG tambahan dan mapel ENG awal
            ws['FP{}'.format(
                r)] = "=IF($EY${}=0,$DN${},$DN${}-1)".format(r, r, r)
            ws['FP{}'.format(s)] = "=STDEV(FP2:FP{})".format(q)
            ws['FP{}'.format(t)] = "=MAX(FP2:FP{})".format(q)
            ws['FP{}'.format(u)] = "=MIN(FP2:FP{})".format(q)
            # rata" SEJ ke SEJ tambahan dan mapel SEJ awal
            ws['FQ{}'.format(
                r)] = "=IF($EZ${}=0,$DO${},$DO${}-1)".format(r, r, r)
            ws['FQ{}'.format(s)] = "=STDEV(FQ2:FQ{})".format(q)
            ws['FQ{}'.format(t)] = "=MAX(FQ2:FQ{})".format(q)
            ws['FQ{}'.format(u)] = "=MIN(FQ2:FQ{})".format(q)
            # rata" GEO ke GEO tambahan dan mapel GEO awal
            ws['FR{}'.format(
                r)] = "=IF($FA${}=0,$DP${},$DP${}-1)".format(r, r, r)
            ws['FR{}'.format(s)] = "=STDEV(FR2:FR{})".format(q)
            ws['FR{}'.format(t)] = "=MAX(FR2:FR{})".format(q)
            ws['FR{}'.format(u)] = "=MIN(FR2:FR{})".format(q)
            # rata" EKO ke EKO tambahan dan mapel EKO awal
            ws['FS{}'.format(
                r)] = "=IF($FB${}=0,$DQ${},$DQ${}-1)".format(r, r, r)
            ws['FS{}'.format(s)] = "=STDEV(FS2:FS{})".format(q)
            ws['FS{}'.format(t)] = "=MAX(FS2:FS{})".format(q)
            ws['FS{}'.format(u)] = "=MIN(FS2:FS{})".format(q)
            # rata" SOS ke SOS tambahan dan mapel SOS awal
            ws['FT{}'.format(
                r)] = "=IF($FC${}=0,$DR${},$DR${}-1)".format(r, r, r)
            ws['FT{}'.format(s)] = "=STDEV(FT2:FT{})".format(q)
            ws['FT{}'.format(t)] = "=MAX(FT2:FT{})".format(q)
            ws['FT{}'.format(u)] = "=MIN(FT2:FT{})".format(q)
            # rata" FIS ke FIS tambahan dan mapel FIS awal
            ws['FU{}'.format(
                r)] = "=IF($FD${}=0,$DS${},$DS${}-1)".format(r, r, r)
            ws['FU{}'.format(s)] = "=STDEV(FU2:FU{})".format(q)
            ws['FU{}'.format(t)] = "=MAX(FU2:FU{})".format(q)
            ws['FU{}'.format(u)] = "=MIN(FU2:FU{})".format(q)
            # rata" KIM ke KIM tambahan dan mapel KIM awal
            ws['FV{}'.format(
                r)] = "=IF($FE${}=0,$DT${},$DT${}-1)".format(r, r, r)
            ws['FV{}'.format(s)] = "=STDEV(FV2:FV{})".format(q)
            ws['FV{}'.format(t)] = "=MAX(FV2:FV{})".format(q)
            ws['FV{}'.format(u)] = "=MIN(FV2:FV{})".format(q)
            # rata" BIO ke BIO tambahan dan mapel BIO awal
            ws['FW{}'.format(
                r)] = "=IF($FF${}=0,$DU${},$DU${}-1)".format(r, r, r)
            ws['FW{}'.format(s)] = "=STDEV(FW2:FW{})".format(q)
            ws['FW{}'.format(t)] = "=MAX(FW2:FW{})".format(q)
            ws['FW{}'.format(u)] = "=MIN(FW2:FW{})".format(q)
            # jml MAPEL
            ws['FX{}'.format(r)] = "=ROUND(AVERAGE(FX2:FX{}),2)".format(q)
            ws['FX{}'.format(t)] = "=MAX(FX2:FX{})".format(q)
            ws['FX{}'.format(u)] = "=MIN(FX2:FX{})".format(q)
            # MAX Z SCORE
            ws['FY{}'.format(r)] = "=MAX(FY2:FY{})".format(q)
            ws['FZ{}'.format(r)] = "=MAX(FZ2:FZ{})".format(q)
            ws['GA{}'.format(r)] = "=MAX(GA2:GA{})".format(q)
            ws['GB{}'.format(r)] = "=MAX(GB2:GB{})".format(q)
            ws['GC{}'.format(r)] = "=MAX(GC2:GC{})".format(q)
            ws['GD{}'.format(r)] = "=MAX(GD2:GD{})".format(q)
            ws['GE{}'.format(r)] = "=MAX(GE2:GE{})".format(q)
            ws['GF{}'.format(r)] = "=MAX(GF2:GF{})".format(q)
            ws['GG{}'.format(r)] = "=MAX(GG2:GG{})".format(q)
            ws['GH{}'.format(r)] = "=MAX(GH2:GH{})".format(q)
            ws['GI{}'.format(r)] = "=MAX(GI2:GI{})".format(q)
            # NILAI STANDAR MTW
            ws['GJ{}'.format(r)] = "=MAX(GJ2:GJ{})".format(q)
            ws['GJ{}'.format(s)] = "=MIN(GJ2:GJ{})".format(q)
            ws['GJ{}'.format(t)] = "=ROUND(AVERAGE(GJ2:GJ{}),2)".format(q)
            # NILAI STANDAR MTP
            ws['GK{}'.format(r)] = "=MAX(GK2:GK{})".format(q)
            ws['GK{}'.format(s)] = "=MIN(GK2:GK{})".format(q)
            ws['GK{}'.format(t)] = "=ROUND(AVERAGE(GK2:GK{}),2)".format(q)
            # NILAI STANDAR IND
            ws['GL{}'.format(r)] = "=MAX(GL2:GL{})".format(q)
            ws['GL{}'.format(s)] = "=MIN(GL2:GL{})".format(q)
            ws['GL{}'.format(t)] = "=ROUND(AVERAGE(GL2:GL{}),2)".format(q)
            # NILAI STANDAR ENG
            ws['GM{}'.format(r)] = "=MAX(GM2:GM{})".format(q)
            ws['GM{}'.format(s)] = "=MIN(GM2:GM{})".format(q)
            ws['GM{}'.format(t)] = "=ROUND(AVERAGE(GM2:GM{}),2)".format(q)
            # NILAI STANDAR SEJ
            ws['GN{}'.format(r)] = "=MAX(GN2:GN{})".format(q)
            ws['GN{}'.format(s)] = "=MIN(GN2:GN{})".format(q)
            ws['GN{}'.format(t)] = "=ROUND(AVERAGE(GN2:GN{}),2)".format(q)
            # NILAI STANDAR GEO
            ws['GO{}'.format(r)] = "=MAX(GO2:GO{})".format(q)
            ws['GO{}'.format(s)] = "=MIN(GO2:GO{})".format(q)
            ws['GO{}'.format(t)] = "=ROUND(AVERAGE(GO2:GO{}),2)".format(q)
            # NILAI STANDAR EKO
            ws['GP{}'.format(r)] = "=MAX(GP2:GP{})".format(q)
            ws['GP{}'.format(s)] = "=MIN(GP2:GP{})".format(q)
            ws['GP{}'.format(t)] = "=ROUND(AVERAGE(GP2:GP{}),2)".format(q)
            # NILAI STANDAR SOS
            ws['GQ{}'.format(r)] = "=MAX(GQ2:GQ{})".format(q)
            ws['GQ{}'.format(s)] = "=MIN(GQ2:GQ{})".format(q)
            ws['GQ{}'.format(t)] = "=ROUND(AVERAGE(GQ2:GQ{}),2)".format(q)
            # NILAI STANDAR FIS
            ws['GR{}'.format(r)] = "=MAX(GR2:GR{})".format(q)
            ws['GR{}'.format(s)] = "=MIN(GR2:GR{})".format(q)
            ws['GR{}'.format(t)] = "=ROUND(AVERAGE(GR2:GR{}),2)".format(q)
            # NILAI STANDAR KIM
            ws['GS{}'.format(r)] = "=MAX(GS2:GS{})".format(q)
            ws['GS{}'.format(s)] = "=MIN(GS2:GS{})".format(q)
            ws['GS{}'.format(t)] = "=ROUND(AVERAGE(GS2:GS{}),2)".format(q)
            # NILAI STANDAR BIO
            ws['GT{}'.format(r)] = "=MAX(GT2:GT{})".format(q)
            ws['GT{}'.format(s)] = "=MIN(GT2:GT{})".format(q)
            ws['GT{}'.format(t)] = "=ROUND(AVERAGE(GT2:GT{}),2)".format(q)
            # NILAI STANDAR S.JML
            ws['GU{}'.format(r)] = "=MAX(GU2:GU{})".format(q)
            ws['GU{}'.format(s)] = "=MIN(GU2:GU{})".format(q)
            ws['GU{}'.format(t)] = "=ROUND(AVERAGE(GU2:GU{}),2)".format(q)

            # TAMBAHAN
            ws['GX{}'.format(r)] = "=SUM(GX2:GX{})".format(q)
            ws['GY{}'.format(r)] = "=SUM(GY2:GY{})".format(q)
            ws['GZ{}'.format(r)] = "=SUM(GZ2:GZ{})".format(q)
            ws['HA{}'.format(r)] = "=SUM(HA2:HA{})".format(q)
            ws['HB{}'.format(r)] = "=SUM(HB2:HB{})".format(q)
            ws['HC{}'.format(r)] = "=SUM(HC2:HC{})".format(q)
            ws['HD{}'.format(r)] = "=SUM(HD2:HD{})".format(q)
            ws['HE{}'.format(r)] = "=SUM(HE2:HE{})".format(q)
            ws['HF{}'.format(r)] = "=SUM(HF2:HF{})".format(q)
            ws['HG{}'.format(r)] = "=SUM(HG2:HG{})".format(q)
            ws['HH{}'.format(r)] = "=SUM(HH2:HH{})".format(q)

            # iterasi 4 rata-rata - 4
            # rata" MTW ke MTW tambahan dan mapel MTW awal
            ws['HO{}'.format(
                r)] = "=IF($GX${}=0,$FM${},$FM${}-1)".format(r, r, r)
            ws['HO{}'.format(s)] = "=STDEV(HO2:HO{})".format(q)
            ws['HO{}'.format(t)] = "=MAX(HO2:HO{})".format(q)
            ws['HO{}'.format(u)] = "=MIN(HO2:HO{})".format(q)
            # rata" MTP ke MTP tambahan dan mapel MTP awal
            ws['HP{}'.format(
                r)] = "=IF($GY${}=0,$FN${},$FN${}-1)".format(r, r, r)
            ws['HP{}'.format(s)] = "=STDEV(HP2:HP{})".format(q)
            ws['HP{}'.format(t)] = "=MAX(HP2:HP{})".format(q)
            ws['HP{}'.format(u)] = "=MIN(HP2:HP{})".format(q)
            # rata" IND ke IND tambahan dan mapel IND awal
            ws['HQ{}'.format(
                r)] = "=IF($GZ${}=0,$FO${},$FO${}-1)".format(r, r, r)
            ws['HQ{}'.format(s)] = "=STDEV(HQ2:HQ{})".format(q)
            ws['HQ{}'.format(t)] = "=MAX(HQ2:HQ{})".format(q)
            ws['HQ{}'.format(u)] = "=MIN(HQ2:HQ{})".format(q)
            # rata" ENG ke ENG tambahan dan mapel ENG awal
            ws['HR{}'.format(
                r)] = "=IF($HA${}=0,$FP${},$FP${}-1)".format(r, r, r)
            ws['HR{}'.format(s)] = "=STDEV(HR2:HR{})".format(q)
            ws['HR{}'.format(t)] = "=MAX(HR2:HR{})".format(q)
            ws['HR{}'.format(u)] = "=MIN(HR2:HR{})".format(q)
            # rata" SEJ ke SEJ tambahan dan mapel SEJ awal
            ws['HS{}'.format(
                r)] = "=IF($HB${}=0,$FQ${},$FQ${}-1)".format(r, r, r)
            ws['HS{}'.format(s)] = "=STDEV(HS2:HS{})".format(q)
            ws['HS{}'.format(t)] = "=MAX(HS2:HS{})".format(q)
            ws['HS{}'.format(u)] = "=MIN(HS2:HS{})".format(q)
            # rata" GEO ke GEO tambahan dan mapel GEO awal
            ws['HT{}'.format(
                r)] = "=IF($HC${}=0,$FR${},$FR${}-1)".format(r, r, r)
            ws['HT{}'.format(s)] = "=STDEV(HT2:HT{})".format(q)
            ws['HT{}'.format(t)] = "=MAX(HT2:HT{})".format(q)
            ws['HT{}'.format(u)] = "=MIN(HT2:HT{})".format(q)
            # rata" EKO ke EKO tambahan dan mapel EKO awal
            ws['HU{}'.format(
                r)] = "=IF($HD${}=0,$FS${},$FS${}-1)".format(r, r, r)
            ws['HU{}'.format(s)] = "=STDEV(HU2:HU{})".format(q)
            ws['HU{}'.format(t)] = "=MAX(HU2:HU{})".format(q)
            ws['HU{}'.format(u)] = "=MIN(HU2:HU{})".format(q)
            # rata" SOS ke SOS tambahan dan mapel SOS awal
            ws['HV{}'.format(
                r)] = "=IF($HE${}=0,$FT${},$FT${}-1)".format(r, r, r)
            ws['HV{}'.format(s)] = "=STDEV(HV2:HV{})".format(q)
            ws['HV{}'.format(t)] = "=MAX(HV2:HV{})".format(q)
            ws['HV{}'.format(u)] = "=MIN(HV2:HV{})".format(q)
            # rata" FIS ke FIS tambahan dan mapel FIS awal
            ws['HW{}'.format(
                r)] = "=IF($HF${}=0,$FU${},$FU${}-1)".format(r, r, r)
            ws['HW{}'.format(s)] = "=STDEV(HW2:HW{})".format(q)
            ws['HW{}'.format(t)] = "=MAX(HW2:HW{})".format(q)
            ws['HW{}'.format(u)] = "=MIN(HW2:HW{})".format(q)
            # rata" KIM ke KIM tambahan dan mapel KIM awal
            ws['HX{}'.format(
                r)] = "=IF($HG${}=0,$FV${},$FV${}-1)".format(r, r, r)
            ws['HX{}'.format(s)] = "=STDEV(HX2:HX{})".format(q)
            ws['HX{}'.format(t)] = "=MAX(HX2:HX{})".format(q)
            ws['HX{}'.format(u)] = "=MIN(HX2:HX{})".format(q)
            # rata" BIO ke BIO tambahan dan mapel BIO awal
            ws['HY{}'.format(
                r)] = "=IF($HH${}=0,$FW${},$FW${}-1)".format(r, r, r)
            ws['HY{}'.format(s)] = "=STDEV(HY2:HY{})".format(q)
            ws['HY{}'.format(t)] = "=MAX(HY2:HY{})".format(q)
            ws['HY{}'.format(u)] = "=MIN(HY2:HY{})".format(q)
            # jml MAPEL
            ws['HZ{}'.format(r)] = "=ROUND(AVERAGE(HZ2:HZ{}),2)".format(q)
            ws['HZ{}'.format(t)] = "=MAX(HZ2:HZ{})".format(q)
            ws['HZ{}'.format(u)] = "=MIN(HZ2:HZ{})".format(q)
            # MAX Z SCORE
            ws['IA{}'.format(r)] = "=MAX(IA2:IA{})".format(q)
            ws['IB{}'.format(r)] = "=MAX(IB2:IB{})".format(q)
            ws['IC{}'.format(r)] = "=MAX(IC2:IC{})".format(q)
            ws['ID{}'.format(r)] = "=MAX(ID2:ID{})".format(q)
            ws['IE{}'.format(r)] = "=MAX(IE2:IE{})".format(q)
            ws['IF{}'.format(r)] = "=MAX(IF2:IF{})".format(q)
            ws['IG{}'.format(r)] = "=MAX(IG2:IG{})".format(q)
            ws['IH{}'.format(r)] = "=MAX(IH2:IH{})".format(q)
            ws['II{}'.format(r)] = "=MAX(II2:II{})".format(q)
            ws['IJ{}'.format(r)] = "=MAX(IJ2:IJ{})".format(q)
            ws['IK{}'.format(r)] = "=MAX(IK2:IK{})".format(q)
            # NILAI STANDAR MTA
            ws['IL{}'.format(r)] = "=MAX(IL2:IL{})".format(q)
            ws['IL{}'.format(s)] = "=MIN(IL2:IL{})".format(q)
            ws['IL{}'.format(t)] = "=ROUND(AVERAGE(IL2:IL{}),2)".format(q)
            # NILAI STANDAR MTP
            ws['IM{}'.format(r)] = "=MAX(IM2:IM{})".format(q)
            ws['IM{}'.format(s)] = "=MIN(IM2:IM{})".format(q)
            ws['IM{}'.format(t)] = "=ROUND(AVERAGE(IM2:IM{}),2)".format(q)
            # NILAI STANDAR IND
            ws['IN{}'.format(r)] = "=MAX(IN2:IN{})".format(q)
            ws['IN{}'.format(s)] = "=MIN(IN2:IN{})".format(q)
            ws['IN{}'.format(t)] = "=ROUND(AVERAGE(IN2:IN{}),2)".format(q)
            # NILAI STANDAR ENG
            ws['IO{}'.format(r)] = "=MAX(IO2:IO{})".format(q)
            ws['IO{}'.format(s)] = "=MIN(IO2:IO{})".format(q)
            ws['IO{}'.format(t)] = "=ROUND(AVERAGE(IO2:IO{}),2)".format(q)
            # NILAI STANDAR SEJ
            ws['IP{}'.format(r)] = "=MAX(IP2:IP{})".format(q)
            ws['IP{}'.format(s)] = "=MIN(IP2:IP{})".format(q)
            ws['IP{}'.format(t)] = "=ROUND(AVERAGE(IP2:IP{}),2)".format(q)
            # NILAI STANDAR GEO
            ws['IQ{}'.format(r)] = "=MAX(IQ2:IQ{})".format(q)
            ws['IQ{}'.format(s)] = "=MIN(IQ2:IQ{})".format(q)
            ws['IQ{}'.format(t)] = "=ROUND(AVERAGE(IQ2:IQ{}),2)".format(q)
            # NILAI STANDAR EKO
            ws['IR{}'.format(r)] = "=MAX(IR2:IR{})".format(q)
            ws['IR{}'.format(s)] = "=MIN(IR2:IR{})".format(q)
            ws['IR{}'.format(t)] = "=ROUND(AVERAGE(IR2:IR{}),2)".format(q)
            # NILAI STANDAR SOS
            ws['IS{}'.format(r)] = "=MAX(IS2:IS{})".format(q)
            ws['IS{}'.format(s)] = "=MIN(IS2:IS{})".format(q)
            ws['IS{}'.format(t)] = "=ROUND(AVERAGE(IS2:IS{}),2)".format(q)
            # NILAI STANDAR FIS
            ws['IT{}'.format(r)] = "=MAX(IT2:IT{})".format(q)
            ws['IT{}'.format(s)] = "=MIN(IT2:IT{})".format(q)
            ws['IT{}'.format(t)] = "=ROUND(AVERAGE(IT2:IT{}),2)".format(q)
            # NILAI STANDAR KIM
            ws['IU{}'.format(r)] = "=MAX(IU2:IU{})".format(q)
            ws['IU{}'.format(s)] = "=MIN(IU2:IU{})".format(q)
            ws['IU{}'.format(t)] = "=ROUND(AVERAGE(IU2:IU{}),2)".format(q)
            # NILAI STANDAR BIO
            ws['IV{}'.format(r)] = "=MAX(IV2:IV{})".format(q)
            ws['IV{}'.format(s)] = "=MIN(IV2:IV{})".format(q)
            ws['IV{}'.format(t)] = "=ROUND(AVERAGE(IV2:IV{}),2)".format(q)
            # NILAI STANDAR S.JML
            ws['IW{}'.format(r)] = "=MAX(IW2:IW{})".format(q)
            ws['IW{}'.format(s)] = "=MIN(IW2:IW{})".format(q)
            ws['IW{}'.format(t)] = "=ROUND(AVERAGE(IW2:IW{}),2)".format(q)

            # TAMBAHAN
            ws['IZ{}'.format(r)] = "=SUM(IZ2:IZ{})".format(q)
            ws['JA{}'.format(r)] = "=SUM(JA2:JA{})".format(q)
            ws['JB{}'.format(r)] = "=SUM(JB2:JB{})".format(q)
            ws['JC{}'.format(r)] = "=SUM(JC2:JC{})".format(q)
            ws['JD{}'.format(r)] = "=SUM(JD2:JD{})".format(q)
            ws['JE{}'.format(r)] = "=SUM(JE2:JE{})".format(q)
            ws['JF{}'.format(r)] = "=SUM(JF2:JF{})".format(q)
            ws['JG{}'.format(r)] = "=SUM(JG2:JG{})".format(q)
            ws['JH{}'.format(r)] = "=SUM(JH2:JH{})".format(q)
            ws['JI{}'.format(r)] = "=SUM(JI2:JI{})".format(q)
            ws['JJ{}'.format(r)] = "=SUM(JJ2:JJ{})".format(q)

            # iterasi 5 rata-rata - 5
            # rata" MTW ke MTW tambahan dan mapel MTW awal
            ws['JQ{}'.format(
                r)] = "=IF($IZ${}=0,$HO${},$HO${}-1)".format(r, r, r)
            ws['JQ{}'.format(s)] = "=STDEV(JQ2:JQ{})".format(q)
            ws['JQ{}'.format(t)] = "=MAX(JQ2:JQ{})".format(q)
            ws['JQ{}'.format(u)] = "=MIN(JQ2:JQ{})".format(q)
            # rata" MTP ke MTP tambahan dan mapel MTP awal
            ws['JR{}'.format(
                r)] = "=IF($JA${}=0,$HP${},$HP${}-1)".format(r, r, r)
            ws['JR{}'.format(s)] = "=STDEV(JR2:JR{})".format(q)
            ws['JR{}'.format(t)] = "=MAX(JR2:JR{})".format(q)
            ws['JR{}'.format(u)] = "=MIN(JR2:JR{})".format(q)
            # rata" IND ke IND tambahan dan mapel IND awal
            ws['JS{}'.format(
                r)] = "=IF($JB${}=0,$HQ${},$HQ${}-1)".format(r, r, r)
            ws['JS{}'.format(s)] = "=STDEV(JS2:JS{})".format(q)
            ws['JS{}'.format(t)] = "=MAX(JS2:JS{})".format(q)
            ws['JS{}'.format(u)] = "=MIN(JS2:JS{})".format(q)
            # rata" ENG ke ENG tambahan dan mapel ENG awal
            ws['JT{}'.format(
                r)] = "=IF($JC${}=0,$HR${},$HR${}-1)".format(r, r, r)
            ws['JT{}'.format(s)] = "=STDEV(JT2:JT{})".format(q)
            ws['JT{}'.format(t)] = "=MAX(JT2:JT{})".format(q)
            ws['JT{}'.format(u)] = "=MIN(JT2:JT{})".format(q)
            # rata" SEJ ke SEJ tambahan dan mapel SEJ awal
            ws['JU{}'.format(
                r)] = "=IF($JD${}=0,$HS${},$HS${}-1)".format(r, r, r)
            ws['JU{}'.format(s)] = "=STDEV(JU2:JU{})".format(q)
            ws['JU{}'.format(t)] = "=MAX(JU2:JU{})".format(q)
            ws['JU{}'.format(u)] = "=MIN(JU2:JU{})".format(q)
            # rata" GEO ke GEO tambahan dan mapel GEO awal
            ws['JV{}'.format(
                r)] = "=IF($JE${}=0,$HT${},$HT${}-1)".format(r, r, r)
            ws['JV{}'.format(s)] = "=STDEV(JV2:JV{})".format(q)
            ws['JV{}'.format(t)] = "=MAX(JV2:JV{})".format(q)
            ws['JV{}'.format(u)] = "=MIN(JV2:JV{})".format(q)
            # rata" EKO ke EKO tambahan dan mapel EKO awal
            ws['JW{}'.format(
                r)] = "=IF($JF${}=0,$HU${},$HU${}-1)".format(r, r, r)
            ws['JW{}'.format(s)] = "=STDEV(JW2:JW{})".format(q)
            ws['JW{}'.format(t)] = "=MAX(JW2:JW{})".format(q)
            ws['JW{}'.format(u)] = "=MIN(JW2:JW{})".format(q)
            # rata" SOS ke SOS tambahan dan mapel SOS awal
            ws['JX{}'.format(
                r)] = "=IF($JG${}=0,$HV${},$HV${}-1)".format(r, r, r)
            ws['JX{}'.format(s)] = "=STDEV(JX2:JX{})".format(q)
            ws['JX{}'.format(t)] = "=MAX(JX2:JX{})".format(q)
            ws['JX{}'.format(u)] = "=MIN(JX2:JX{})".format(q)
            # rata" FIS ke FIS tambahan dan mapel FIS awal
            ws['JY{}'.format(
                r)] = "=IF($JH${}=0,$HW${},$HW${}-1)".format(r, r, r)
            ws['JY{}'.format(s)] = "=STDEV(JY2:JY{})".format(q)
            ws['JY{}'.format(t)] = "=MAX(JY2:JY{})".format(q)
            ws['JY{}'.format(u)] = "=MIN(JY2:JY{})".format(q)
            # rata" KIM ke KIM tambahan dan mapel KIM awal
            ws['JZ{}'.format(
                r)] = "=IF($JI${}=0,$HX${},$HX${}-1)".format(r, r, r)
            ws['JZ{}'.format(s)] = "=STDEV(JZ2:JZ{})".format(q)
            ws['JZ{}'.format(t)] = "=MAX(JZ2:JZ{})".format(q)
            ws['JZ{}'.format(u)] = "=MIN(JZ2:JZ{})".format(q)
            # rata" BIO ke BIO tambahan dan mapel BIO awal
            ws['KA{}'.format(
                r)] = "=IF($JJ${}=0,$HY${},$HY${}-1)".format(r, r, r)
            ws['KA{}'.format(s)] = "=STDEV(KA2:KA{})".format(q)
            ws['KA{}'.format(t)] = "=MAX(KA2:KA{})".format(q)
            ws['KA{}'.format(u)] = "=MIN(KA2:KA{})".format(q)
            # jml MAPEL
            ws['KB{}'.format(r)] = "=ROUND(AVERAGE(KB2:KB{}),2)".format(q)
            ws['KB{}'.format(t)] = "=MAX(KB2:KB{})".format(q)
            ws['KB{}'.format(u)] = "=MIN(KB2:KB{})".format(q)
            # MAX Z SCORE
            ws['KC{}'.format(r)] = "=MAX(KC2:KC{})".format(q)
            ws['KD{}'.format(r)] = "=MAX(KD2:KD{})".format(q)
            ws['KE{}'.format(r)] = "=MAX(KE2:KE{})".format(q)
            ws['KF{}'.format(r)] = "=MAX(KF2:KF{})".format(q)
            ws['KG{}'.format(r)] = "=MAX(KG2:KG{})".format(q)
            ws['KH{}'.format(r)] = "=MAX(KH2:KH{})".format(q)
            ws['KI{}'.format(r)] = "=MAX(KI2:KI{})".format(q)
            ws['KJ{}'.format(r)] = "=MAX(KJ2:KJ{})".format(q)
            ws['KK{}'.format(r)] = "=MAX(KK2:KK{})".format(q)
            ws['KL{}'.format(r)] = "=MAX(KL2:KL{})".format(q)
            ws['KM{}'.format(r)] = "=MAX(KM2:KM{})".format(q)
            # NILAI STANDAR MTA
            ws['KN{}'.format(r)] = "=MAX(KN2:KN{})".format(q)
            ws['KN{}'.format(s)] = "=MIN(KN2:KN{})".format(q)
            ws['KN{}'.format(t)] = "=ROUND(AVERAGE(KN2:KN{}),2)".format(q)
            # NILAI STANDAR MTP
            ws['KO{}'.format(r)] = "=MAX(KO2:KO{})".format(q)
            ws['KO{}'.format(s)] = "=MIN(KO2:KO{})".format(q)
            ws['KO{}'.format(t)] = "=ROUND(AVERAGE(KO2:KO{}),2)".format(q)
            # NILAI STANDAR IND
            ws['KP{}'.format(r)] = "=MAX(KP2:KP{})".format(q)
            ws['KP{}'.format(s)] = "=MIN(KP2:KP{})".format(q)
            ws['KP{}'.format(t)] = "=ROUND(AVERAGE(KP2:KP{}),2)".format(q)
            # NILAI STANDAR ENG
            ws['KQ{}'.format(r)] = "=MAX(KQ2:KQ{})".format(q)
            ws['KQ{}'.format(s)] = "=MIN(KQ2:KQ{})".format(q)
            ws['KQ{}'.format(t)] = "=ROUND(AVERAGE(KQ2:KQ{}),2)".format(q)
            # NILAI STANDAR SEJ
            ws['KR{}'.format(r)] = "=MAX(KR2:KR{})".format(q)
            ws['KR{}'.format(s)] = "=MIN(KR2:KR{})".format(q)
            ws['KR{}'.format(t)] = "=ROUND(AVERAGE(KR2:KR{}),2)".format(q)
            # NILAI STANDAR GEO
            ws['KS{}'.format(r)] = "=MAX(KS2:KS{})".format(q)
            ws['KS{}'.format(s)] = "=MIN(KS2:KS{})".format(q)
            ws['KS{}'.format(t)] = "=ROUND(AVERAGE(KS2:KS{}),2)".format(q)
            # NILAI STANDAR EKO
            ws['KT{}'.format(r)] = "=MAX(KT2:KT{})".format(q)
            ws['KT{}'.format(s)] = "=MIN(KT2:KT{})".format(q)
            ws['KT{}'.format(t)] = "=ROUND(AVERAGE(KT2:KT{}),2)".format(q)
            # NILAI STANDAR SOS
            ws['KU{}'.format(r)] = "=MAX(KU2:KU{})".format(q)
            ws['KU{}'.format(s)] = "=MIN(KU2:KU{})".format(q)
            ws['KU{}'.format(t)] = "=ROUND(AVERAGE(KU2:KU{}),2)".format(q)
            # NILAI STANDAR FIS
            ws['KV{}'.format(r)] = "=MAX(KV2:KV{})".format(q)
            ws['KV{}'.format(s)] = "=MIN(KV2:KV{})".format(q)
            ws['KV{}'.format(t)] = "=ROUND(AVERAGE(KV2:KV{}),2)".format(q)
            # NILAI STANDAR KIM
            ws['KW{}'.format(r)] = "=MAX(KW2:KW{})".format(q)
            ws['KW{}'.format(s)] = "=MIN(KW2:KW{})".format(q)
            ws['KW{}'.format(t)] = "=ROUND(AVERAGE(KW2:KW{}),2)".format(q)
            # NILAI STANDAR BIO
            ws['KX{}'.format(r)] = "=MAX(KX2:KX{})".format(q)
            ws['KX{}'.format(s)] = "=MIN(KX2:KX{})".format(q)
            ws['KX{}'.format(t)] = "=ROUND(AVERAGE(KX2:KX{}),2)".format(q)
            # NILAI STANDAR S.JML
            ws['KY{}'.format(r)] = "=MAX(KY2:KY{})".format(q)
            ws['KY{}'.format(s)] = "=MIN(KY2:KY{})".format(q)
            ws['KY{}'.format(t)] = "=ROUND(AVERAGE(KY2:KY{}),2)".format(q)

            # TAMBAHAN
            ws['LB{}'.format(r)] = "=SUM(LB2:LB{})".format(q)
            ws['LC{}'.format(r)] = "=SUM(LC2:LC{})".format(q)
            ws['LD{}'.format(r)] = "=SUM(LD2:LD{})".format(q)
            ws['LE{}'.format(r)] = "=SUM(LE2:LE{})".format(q)
            ws['LF{}'.format(r)] = "=SUM(LF2:LF{})".format(q)
            ws['LG{}'.format(r)] = "=SUM(LG2:LG{})".format(q)
            ws['LH{}'.format(r)] = "=SUM(LH2:LH{})".format(q)
            ws['LI{}'.format(r)] = "=SUM(LI2:LI{})".format(q)
            ws['LJ{}'.format(r)] = "=SUM(LJ2:LJ{})".format(q)
            ws['LK{}'.format(r)] = "=SUM(LK2:LK{})".format(q)
            ws['LL{}'.format(r)] = "=SUM(LL2:LL{})".format(q)

            # Z Score [1]
            ws['B1'] = 'NAMA SISWA_A'
            ws['C1'] = 'NOMOR NF_A'
            ws['D1'] = 'KELAS_A'
            ws['E1'] = 'NAMA SEKOLAH_A'
            ws['F1'] = 'LOKASI_A'

            ws['G1'] = 'MAW_A'
            ws['H1'] = 'MAP_A'
            ws['I1'] = 'IND_A'
            ws['J1'] = 'ENG_A'
            ws['K1'] = 'SEJ_A'
            ws['L1'] = 'GEO_A'
            ws['M1'] = 'EKO_A'
            ws['N1'] = 'SOS_A'
            ws['O1'] = 'FIS_A'
            ws['P1'] = 'KIM_A'
            ws['Q1'] = 'BIO_A'
            ws['R1'] = 'JML_A'

            ws['S1'] = 'Z_MAW_A'
            ws['T1'] = 'Z_MAP_A'
            ws['U1'] = 'Z_IND_A'
            ws['V1'] = 'Z_ENG_A'
            ws['W1'] = 'Z_SEJ_A'
            ws['X1'] = 'Z_GEO_A'
            ws['Y1'] = 'Z_EKO_A'
            ws['Z1'] = 'Z_SOS_A'
            ws['AA1'] = 'Z_FIS_A'
            ws['AB1'] = 'Z_KIM_A'
            ws['AC1'] = 'Z_BIO_A'

            ws['AD1'] = 'S_MAW_A'
            ws['AE1'] = 'S_MAP_A'
            ws['AF1'] = 'S_IND_A'
            ws['AG1'] = 'S_ENG_A'
            ws['AH1'] = 'S_SEJ_A'
            ws['AI1'] = 'S_GEO_A'
            ws['AJ1'] = 'S_EKO_A'
            ws['AK1'] = 'S_SOS_A'
            ws['AL1'] = 'S_FIS_A'
            ws['AM1'] = 'S_KIM_A'
            ws['AN1'] = 'S_BIO_A'
            ws['AO1'] = 'S_JML_A'

            ws['AP1'] = 'RANK NAS._A'
            ws['AQ1'] = 'RANK LOK._A'

            ws['S1'].font = Font(bold=False, name='Calibri', size=11)
            ws['T1'].font = Font(bold=False, name='Calibri', size=11)
            ws['U1'].font = Font(bold=False, name='Calibri', size=11)
            ws['V1'].font = Font(bold=False, name='Calibri', size=11)
            ws['W1'].font = Font(bold=False, name='Calibri', size=11)
            ws['X1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Y1'].font = Font(bold=False, name='Calibri', size=11)
            ws['Z1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AQ1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['B1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['C1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['D1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['E1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['F1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['G1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['H1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['I1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['J1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['K1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['L1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['M1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['N1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['O1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['P1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Q1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['R1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['S1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['T1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['U1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['V1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['W1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['X1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Y1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['Z1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AA1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AB1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AC1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AD1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AE1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AF1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AG1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AH1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AI1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AJ1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AK1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AL1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AM1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AN1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AO1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AP1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AQ1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')

            # tambahan
            ws['AR1'] = 'MAW_20_A'
            ws['AS1'] = 'MAP_20_A'
            ws['AT1'] = 'IND_20_A'
            ws['AU1'] = 'ENG_20_A'
            ws['AV1'] = 'SEJ_20_A'
            ws['AW1'] = 'GEO_20_A'
            ws['AX1'] = 'EKO_20_A'
            ws['AY1'] = 'SOS_20_A'
            ws['AZ1'] = 'FIS_20_A'
            ws['BA1'] = 'KIM_20_A'
            ws['BB1'] = 'BIO_20_A'

            ws['AR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['AZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BB1'].font = Font(bold=False, name='Calibri', size=11)

            ws['AR1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AS1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AT1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AU1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AV1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AW1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AX1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AY1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['AZ1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['BA1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')
            ws['BB1'].fill = PatternFill(
                fill_type='solid', start_color='00FF6600', end_color='00FF6600')

            for row in range(2, q+1):
                ws['S{}'.format(
                    row)] = '=IFERROR(ROUND(IF(G{}="","",(G{}-G${})/G${}),2),"")'.format(row, row, r, s)
                ws['T{}'.format(
                    row)] = '=IFERROR(ROUND(IF(H{}="","",(H{}-H${})/H${}),2),"")'.format(row, row, r, s)
                ws['U{}'.format(
                    row)] = '=IFERROR(ROUND(IF(I{}="","",(I{}-I${})/I${}),2),"")'.format(row, row, r, s)
                ws['V{}'.format(
                    row)] = '=IFERROR(ROUND(IF(J{}="","",(J{}-J${})/J${}),2),"")'.format(row, row, r, s)
                ws['W{}'.format(
                    row)] = '=IFERROR(ROUND(IF(K{}="","",(K{}-K${})/K${}),2),"")'.format(row, row, r, s)
                ws['X{}'.format(
                    row)] = '=IFERROR(ROUND(IF(L{}="","",(L{}-L${})/L${}),2),"")'.format(row, row, r, s)
                ws['Y{}'.format(
                    row)] = '=IFERROR(ROUND(IF(M{}="","",(M{}-M${})/M${}),2),"")'.format(row, row, r, s)
                ws['Z{}'.format(
                    row)] = '=IFERROR(ROUND(IF(N{}="","",(N{}-N${})/N${}),2),"")'.format(row, row, r, s)
                ws['AA{}'.format(
                    row)] = '=IFERROR(ROUND(IF(O{}="","",(O{}-O${})/O${}),2),"")'.format(row, row, r, s)
                ws['AB{}'.format(
                    row)] = '=IFERROR(ROUND(IF(P{}="","",(P{}-P${})/P${}),2),"")'.format(row, row, r, s)
                ws['AC{}'.format(
                    row)] = '=IFERROR(ROUND(IF(Q{}="","",(Q{}-Q${})/Q${}),2),"")'.format(row, row, r, s)

                ws['AD{}'.format(
                    row)] = '=IFERROR(ROUND(IF(G{}="","",IF(70+30*S{}/$S${}<20,20,70+30*S{}/$S${})),2),"")'.format(row, row, r, row, r)
                ws['AE{}'.format(
                    row)] = '=IFERROR(ROUND(IF(H{}="","",IF(70+30*T{}/$T${}<20,20,70+30*T{}/$T${})),2),"")'.format(row, row, r, row, r)
                ws['AF{}'.format(
                    row)] = '=IFERROR(ROUND(IF(I{}="","",IF(70+30*U{}/$U${}<20,20,70+30*U{}/$U${})),2),"")'.format(row, row, r, row, r)
                ws['AG{}'.format(
                    row)] = '=IFERROR(ROUND(IF(J{}="","",IF(70+30*V{}/$V${}<20,20,70+30*V{}/$V${})),2),"")'.format(row, row, r, row, r)
                ws['AH{}'.format(
                    row)] = '=IFERROR(ROUND(IF(K{}="","",IF(70+30*W{}/$W${}<20,20,70+30*W{}/$W${})),2),"")'.format(row, row, r, row, r)
                ws['AI{}'.format(
                    row)] = '=IFERROR(ROUND(IF(L{}="","",IF(70+30*X{}/$X${}<20,20,70+30*X{}/$X${})),2),"")'.format(row, row, r, row, r)
                ws['AJ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(M{}="","",IF(70+30*Y{}/$Y${}<20,20,70+30*Y{}/$Y${})),2),"")'.format(row, row, r, row, r)
                ws['AK{}'.format(
                    row)] = '=IFERROR(ROUND(IF(N{}="","",IF(70+30*Z{}/$Z${}<20,20,70+30*Z{}/$Z${})),2),"")'.format(row, row, r, row, r)
                ws['AL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(O{}="","",IF(70+30*AA{}/$AA${}<20,20,70+30*AA{}/$AA${})),2),"")'.format(row, row, r, row, r)
                ws['AM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(P{}="","",IF(70+30*AB{}/$AB${}<20,20,70+30*AB{}/$AB${})),2),"")'.format(row, row, r, row, r)
                ws['AN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(Q{}="","",IF(70+30*AC{}/$AC${}<20,20,70+30*AC{}/$AC${})),2),"")'.format(row, row, r, row, r)

                ws['AO{}'.format(row)] = '=IF(SUM(AD{}:AN{})=0,"",SUM(AD{}:AN{}))'.format(
                    row, row, row, row)
                ws['AP{}'.format(row)] = '=IF(AO{}="","",RANK(AO{},$AO$2:$AO${}))'.format(
                    row, row, q)
                ws['AQ{}'.format(
                    row)] = '=IF(AP{}="","",COUNTIFS($F$2:$F${},F{},$AP$2:$AP${},"<"&AP{})+1)'.format(row, q, row, q, row)

            # TAMBAHAN
                ws['AR{}'.format(row)] = '=IF($G${}=20,IF(AND(G{}>3,AD{}=20),1,""),IF($G${}=25,IF(AND(G{}>4,AD{}=20),1,""),IF($G${}=30,IF(AND(G{}>5,AD{}=20),1,""),IF($G${}=35,IF(AND(G{}>6,AD{}=20),1,""),IF($G${}=40,IF(AND(G{}>7,AD{}=20),1,""),IF($G${}=45,IF(AND(G{}>8,AD{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AS{}'.format(row)] = '=IF($H${}=20,IF(AND(H{}>3,AE{}=20),1,""),IF($H${}=25,IF(AND(H{}>4,AE{}=20),1,""),IF($H${}=30,IF(AND(H{}>5,AE{}=20),1,""),IF($H${}=35,IF(AND(H{}>6,AE{}=20),1,""),IF($H${}=40,IF(AND(H{}>7,AE{}=20),1,""),IF($H${}=45,IF(AND(H{}>8,AE{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AT{}'.format(row)] = '=IF($I${}=20,IF(AND(I{}>3,AF{}=20),1,""),IF($I${}=25,IF(AND(I{}>4,AF{}=20),1,""),IF($I${}=30,IF(AND(I{}>5,AF{}=20),1,""),IF($I${}=35,IF(AND(I{}>6,AF{}=20),1,""),IF($I${}=40,IF(AND(I{}>7,AF{}=20),1,""),IF($I${}=45,IF(AND(I{}>8,AF{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AU{}'.format(row)] = '=IF($J${}=20,IF(AND(J{}>3,AG{}=20),1,""),IF($J${}=25,IF(AND(J{}>4,AG{}=20),1,""),IF($J${}=30,IF(AND(J{}>5,AG{}=20),1,""),IF($J${}=35,IF(AND(J{}>6,AG{}=20),1,""),IF($J${}=40,IF(AND(J{}>7,AG{}=20),1,""),IF($J${}=45,IF(AND(J{}>8,AG{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AV{}'.format(row)] = '=IF($K${}=20,IF(AND(K{}>3,AH{}=20),1,""),IF($K${}=25,IF(AND(K{}>4,AH{}=20),1,""),IF($K${}=30,IF(AND(K{}>5,AH{}=20),1,""),IF($K${}=35,IF(AND(K{}>6,AH{}=20),1,""),IF($K${}=40,IF(AND(K{}>7,AH{}=20),1,""),IF($K${}=45,IF(AND(K{}>8,AH{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AW{}'.format(row)] = '=IF($L${}=20,IF(AND(L{}>3,AI{}=20),1,""),IF($L${}=25,IF(AND(L{}>4,AI{}=20),1,""),IF($L${}=30,IF(AND(L{}>5,AI{}=20),1,""),IF($L${}=35,IF(AND(L{}>6,AI{}=20),1,""),IF($L${}=40,IF(AND(L{}>7,AI{}=20),1,""),IF($L${}=45,IF(AND(L{}>8,AI{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AX{}'.format(row)] = '=IF($M${}=20,IF(AND(M{}>3,AJ{}=20),1,""),IF($M${}=25,IF(AND(M{}>4,AJ{}=20),1,""),IF($M${}=30,IF(AND(M{}>5,AJ{}=20),1,""),IF($M${}=35,IF(AND(M{}>6,AJ{}=20),1,""),IF($M${}=40,IF(AND(M{}>7,AJ{}=20),1,""),IF($M${}=45,IF(AND(M{}>8,AJ{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AY{}'.format(row)] = '=IF($N${}=20,IF(AND(N{}>3,AK{}=20),1,""),IF($N${}=25,IF(AND(N{}>4,AK{}=20),1,""),IF($N${}=30,IF(AND(N{}>5,AK{}=20),1,""),IF($N${}=35,IF(AND(N{}>6,AK{}=20),1,""),IF($N${}=40,IF(AND(N{}>7,AK{}=20),1,""),IF($N${}=45,IF(AND(N{}>8,AK{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['AZ{}'.format(row)] = '=IF($O${}=20,IF(AND(O{}>3,AL{}=20),1,""),IF($O${}=25,IF(AND(O{}>4,AL{}=20),1,""),IF($O${}=30,IF(AND(O{}>5,AL{}=20),1,""),IF($O${}=35,IF(AND(O{}>6,AL{}=20),1,""),IF($O${}=40,IF(AND(O{}>7,AL{}=20),1,""),IF($O${}=45,IF(AND(O{}>8,AL{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BA{}'.format(row)] = '=IF($P${}=20,IF(AND(P{}>3,AM{}=20),1,""),IF($P${}=25,IF(AND(P{}>4,AM{}=20),1,""),IF($P${}=30,IF(AND(P{}>5,AM{}=20),1,""),IF($P${}=35,IF(AND(P{}>6,AM{}=20),1,""),IF($P${}=40,IF(AND(P{}>7,AM{}=20),1,""),IF($P${}=45,IF(AND(P{}>8,AM{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['BB{}'.format(row)] = '=IF($Q${}=20,IF(AND(Q{}>3,AN{}=20),1,""),IF($Q${}=25,IF(AND(Q{}>4,AN{}=20),1,""),IF($Q${}=30,IF(AND(Q{}>5,AN{}=20),1,""),IF($Q${}=35,IF(AND(Q{}>6,AN{}=20),1,""),IF($Q${}=40,IF(AND(Q{}>7,AN{}=20),1,""),IF($Q${}=45,IF(AND(Q{}>8,AN{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score [2]
            ws['BD1'] = 'NAMA SISWA_A'
            ws['BE1'] = 'NOMOR NF_A'
            ws['BF1'] = 'KELAS_A'
            ws['BG1'] = 'NAMA SEKOLAH_A'
            ws['BH1'] = 'LOKASI_A'

            ws['BI1'] = 'MAW_A'
            ws['BJ1'] = 'MAP_A'
            ws['BK1'] = 'IND_A'
            ws['BL1'] = 'ENG_A'
            ws['BM1'] = 'SEJ_A'
            ws['BN1'] = 'GEO_A'
            ws['BO1'] = 'EKO_A'
            ws['BP1'] = 'SOS_A'
            ws['BQ1'] = 'FIS_A'
            ws['BR1'] = 'KIM_A'
            ws['BS1'] = 'BIO_A'
            ws['BT1'] = 'JML_A'

            ws['BU1'] = 'Z_MAW_A'
            ws['BV1'] = 'Z_MAP_A'
            ws['BW1'] = 'Z_IND_A'
            ws['BX1'] = 'Z_ENG_A'
            ws['BY1'] = 'Z_SEJ_A'
            ws['BZ1'] = 'Z_GEO_A'
            ws['CA1'] = 'Z_EKO_A'
            ws['CB1'] = 'Z_SOS_A'
            ws['CC1'] = 'Z_FIS_A'
            ws['CD1'] = 'Z_KIM_A'
            ws['CE1'] = 'Z_BIO_A'

            ws['CF1'] = 'S_MAW_A'
            ws['CG1'] = 'S_MAP_A'
            ws['CH1'] = 'S_IND_A'
            ws['CI1'] = 'S_ENG_A'
            ws['CJ1'] = 'S_SEJ_A'
            ws['CK1'] = 'S_GEO_A'
            ws['CL1'] = 'S_EKO_A'
            ws['CM1'] = 'S_SOS_A'
            ws['CN1'] = 'S_FIS_A'
            ws['CO1'] = 'S_KIM_A'
            ws['CP1'] = 'S_BIO_A'
            ws['CQ1'] = 'S_JML_A'

            ws['CR1'] = 'RANK NAS._A'
            ws['CS1'] = 'RANK LOK._A'

            ws['BU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['BZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CS1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['BD1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BE1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BF1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BG1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BH1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BI1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BJ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BK1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BL1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BM1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BN1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BO1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BP1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BQ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BR1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BS1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BT1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BU1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BV1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BW1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BX1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BY1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['BZ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CA1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CB1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CC1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CD1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CE1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CF1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CG1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CH1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CI1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CJ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CK1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CL1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CM1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CN1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CO1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CP1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CQ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CR1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CS1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')

            # tambahan
            ws['CT1'] = 'MAW_20_A'
            ws['CU1'] = 'MAP_20_A'
            ws['CV1'] = 'IND_20_A'
            ws['CW1'] = 'ENG_20_A'
            ws['CX1'] = 'SEJ_20_A'
            ws['CY1'] = 'GEO_20_A'
            ws['CZ1'] = 'EKO_20_A'
            ws['DA1'] = 'SOS_20_A'
            ws['DB1'] = 'FIS_20_A'
            ws['DC1'] = 'KIM_20_A'
            ws['DD1'] = 'BIO_20_A'

            ws['CT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['CZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DD1'].font = Font(bold=False, name='Calibri', size=11)

            ws['CT1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CU1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CV1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CW1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CX1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CY1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['CZ1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['DA1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['DB1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['DC1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')
            ws['DD1'].fill = PatternFill(
                fill_type='solid', start_color='31E1F7', end_color='31E1F7')

            for row in range(2, q+1):
                ws['BD{}'.format(row)] = '=B{}'.format(row)
                ws['BE{}'.format(row)] = '=C{}'.format(row, row)
                ws['BF{}'.format(row)] = '=D{}'.format(row, row)
                ws['BG{}'.format(row)] = '=E{}'.format(row, row)
                ws['BH{}'.format(row)] = '=F{}'.format(row, row)

                ws['BI{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['BJ{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['BK{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['BL{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['BM{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['BN{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['BO{}'.format(row)] = '=IF(M{}="","",M{})'.format(row, row)
                ws['BP{}'.format(row)] = '=IF(N{}="","",N{})'.format(row, row)
                ws['BQ{}'.format(row)] = '=IF(O{}="","",O{})'.format(row, row)
                ws['BR{}'.format(row)] = '=IF(P{}="","",P{})'.format(row, row)
                ws['BS{}'.format(row)] = '=IF(Q{}="","",Q{})'.format(row, row)
                ws['BT{}'.format(row)] = '=IF(R{}="","",R{})'.format(row, row)
                ws['BU{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BI{}="","",(BI{}-BI${})/BI${}),2),"")'.format(row, row, r, s)
                ws['BV{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BJ{}="","",(BJ{}-BJ${})/BJ${}),2),"")'.format(row, row, r, s)
                ws['BW{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BK{}="","",(BK{}-BK${})/BK${}),2),"")'.format(row, row, r, s)
                ws['BX{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BL{}="","",(BL{}-BL${})/BL${}),2),"")'.format(row, row, r, s)
                ws['BY{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BM{}="","",(BM{}-BM${})/BM${}),2),"")'.format(row, row, r, s)
                ws['BZ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BN{}="","",(BN{}-BN${})/BN${}),2),"")'.format(row, row, r, s)
                ws['CA{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BO{}="","",(BO{}-BO${})/BO${}),2),"")'.format(row, row, r, s)
                ws['CB{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BP{}="","",(BP{}-BP${})/BP${}),2),"")'.format(row, row, r, s)
                ws['CC{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BQ{}="","",(BQ{}-BQ${})/BQ${}),2),"")'.format(row, row, r, s)
                ws['CD{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BR{}="","",(BR{}-BR${})/BR${}),2),"")'.format(row, row, r, s)
                ws['CE{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BS{}="","",(BS{}-BS${})/BS${}),2),"")'.format(row, row, r, s)

                ws['CF{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BI{}="","",IF(70+30*BU{}/$BU${}<20,20,70+30*BU{}/$BU${})),2),"")'.format(row, row, r, row, r)
                ws['CG{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BJ{}="","",IF(70+30*BV{}/$BV${}<20,20,70+30*BV{}/$BV${})),2),"")'.format(row, row, r, row, r)
                ws['CH{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BK{}="","",IF(70+30*BW{}/$BW${}<20,20,70+30*BW{}/$BW${})),2),"")'.format(row, row, r, row, r)
                ws['CI{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BL{}="","",IF(70+30*BX{}/$BX${}<20,20,70+30*BX{}/$BX${})),2),"")'.format(row, row, r, row, r)
                ws['CJ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BM{}="","",IF(70+30*BY{}/$BY${}<20,20,70+30*BY{}/$BY${})),2),"")'.format(row, row, r, row, r)
                ws['CK{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BN{}="","",IF(70+30*BZ{}/$BZ${}<20,20,70+30*BZ{}/$BZ${})),2),"")'.format(row, row, r, row, r)
                ws['CL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BO{}="","",IF(70+30*CA{}/$CA${}<20,20,70+30*CA{}/$CA${})),2),"")'.format(row, row, r, row, r)
                ws['CM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BP{}="","",IF(70+30*CB{}/$CB${}<20,20,70+30*CB{}/$CB${})),2),"")'.format(row, row, r, row, r)
                ws['CN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BQ{}="","",IF(70+30*CC{}/$CC${}<20,20,70+30*CC{}/$CC${})),2),"")'.format(row, row, r, row, r)
                ws['CO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BR{}="","",IF(70+30*CD{}/$CD${}<20,20,70+30*CD{}/$CD${})),2),"")'.format(row, row, r, row, r)
                ws['CP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(BS{}="","",IF(70+30*CE{}/$CE${}<20,20,70+30*CE{}/$CE${})),2),"")'.format(row, row, r, row, r)

                ws['CQ{}'.format(row)] = '=IF(SUM(CF{}:CP{})=0,"",SUM(CF{}:CP{}))'.format(
                    row, row, row, row)
                ws['CR{}'.format(row)] = '=IF(CQ{}="","",RANK(CQ{},$CQ$2:$CQ${}))'.format(
                    row, row, q)
                ws['CS{}'.format(
                    row)] = '=IF(CR{}="","",COUNTIFS($BH$2:$BH${},BH{},$CR$2:$CR${},"<"&CR{})+1)'.format(row, q, row, q, row)

            # TAMBAHAN
                ws['CT{}'.format(row)] = '=IF($G${}=20,IF(AND(G{}>3,CF{}=20),1,""),IF($G${}=25,IF(AND(G{}>4,CF{}=20),1,""),IF($G${}=30,IF(AND(G{}>5,CF{}=20),1,""),IF($G${}=35,IF(AND(G{}>6,CF{}=20),1,""),IF($G${}=40,IF(AND(G{}>7,CF{}=20),1,""),IF($G${}=45,IF(AND(G{}>8,CF{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CU{}'.format(row)] = '=IF($H${}=20,IF(AND(H{}>3,CG{}=20),1,""),IF($H${}=25,IF(AND(H{}>4,CG{}=20),1,""),IF($H${}=30,IF(AND(H{}>5,CG{}=20),1,""),IF($H${}=35,IF(AND(H{}>6,CG{}=20),1,""),IF($H${}=40,IF(AND(H{}>7,CG{}=20),1,""),IF($H${}=45,IF(AND(H{}>8,CG{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CV{}'.format(row)] = '=IF($I${}=20,IF(AND(I{}>3,CH{}=20),1,""),IF($I${}=25,IF(AND(I{}>4,CH{}=20),1,""),IF($I${}=30,IF(AND(I{}>5,CH{}=20),1,""),IF($I${}=35,IF(AND(I{}>6,CH{}=20),1,""),IF($I${}=40,IF(AND(I{}>7,CH{}=20),1,""),IF($I${}=45,IF(AND(I{}>8,CH{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CW{}'.format(row)] = '=IF($J${}=20,IF(AND(J{}>3,CI{}=20),1,""),IF($J${}=25,IF(AND(J{}>4,CI{}=20),1,""),IF($J${}=30,IF(AND(J{}>5,CI{}=20),1,""),IF($J${}=35,IF(AND(J{}>6,CI{}=20),1,""),IF($J${}=40,IF(AND(J{}>7,CI{}=20),1,""),IF($J${}=45,IF(AND(J{}>8,CI{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CX{}'.format(row)] = '=IF($K${}=20,IF(AND(K{}>3,CJ{}=20),1,""),IF($K${}=25,IF(AND(K{}>4,CJ{}=20),1,""),IF($K${}=30,IF(AND(K{}>5,CJ{}=20),1,""),IF($K${}=35,IF(AND(K{}>6,CJ{}=20),1,""),IF($K${}=40,IF(AND(K{}>7,CJ{}=20),1,""),IF($K${}=45,IF(AND(K{}>8,CJ{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CY{}'.format(row)] = '=IF($L${}=20,IF(AND(L{}>3,CK{}=20),1,""),IF($L${}=25,IF(AND(L{}>4,CK{}=20),1,""),IF($L${}=30,IF(AND(L{}>5,CK{}=20),1,""),IF($L${}=35,IF(AND(L{}>6,CK{}=20),1,""),IF($L${}=40,IF(AND(L{}>7,CK{}=20),1,""),IF($L${}=45,IF(AND(L{}>8,CK{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['CZ{}'.format(row)] = '=IF($M${}=20,IF(AND(M{}>3,CL{}=20),1,""),IF($M${}=25,IF(AND(M{}>4,CL{}=20),1,""),IF($M${}=30,IF(AND(M{}>5,CL{}=20),1,""),IF($M${}=35,IF(AND(M{}>6,CL{}=20),1,""),IF($M${}=40,IF(AND(M{}>7,CL{}=20),1,""),IF($M${}=45,IF(AND(M{}>8,CL{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DA{}'.format(row)] = '=IF($N${}=20,IF(AND(N{}>3,CM{}=20),1,""),IF($N${}=25,IF(AND(N{}>4,CM{}=20),1,""),IF($N${}=30,IF(AND(N{}>5,CM{}=20),1,""),IF($N${}=35,IF(AND(N{}>6,CM{}=20),1,""),IF($N${}=40,IF(AND(N{}>7,CM{}=20),1,""),IF($N${}=45,IF(AND(N{}>8,CM{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DB{}'.format(row)] = '=IF($O${}=20,IF(AND(O{}>3,CN{}=20),1,""),IF($O${}=25,IF(AND(O{}>4,CN{}=20),1,""),IF($O${}=30,IF(AND(O{}>5,CN{}=20),1,""),IF($O${}=35,IF(AND(O{}>6,CN{}=20),1,""),IF($O${}=40,IF(AND(O{}>7,CN{}=20),1,""),IF($O${}=45,IF(AND(O{}>8,CN{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DC{}'.format(row)] = '=IF($P${}=20,IF(AND(P{}>3,CO{}=20),1,""),IF($P${}=25,IF(AND(P{}>4,CO{}=20),1,""),IF($P${}=30,IF(AND(P{}>5,CO{}=20),1,""),IF($P${}=35,IF(AND(P{}>6,CO{}=20),1,""),IF($P${}=40,IF(AND(P{}>7,CO{}=20),1,""),IF($P${}=45,IF(AND(P{}>8,CO{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['DD{}'.format(row)] = '=IF($Q${}=20,IF(AND(Q{}>3,CP{}=20),1,""),IF($Q${}=25,IF(AND(Q{}>4,CP{}=20),1,""),IF($Q${}=30,IF(AND(Q{}>5,CP{}=20),1,""),IF($Q${}=35,IF(AND(Q{}>6,CP{}=20),1,""),IF($Q${}=40,IF(AND(Q{}>7,CP{}=20),1,""),IF($Q${}=45,IF(AND(Q{}>8,CP{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score [3]
            ws['DF1'] = 'NAMA SISWA_A'
            ws['DG1'] = 'NOMOR NF_A'
            ws['DH1'] = 'KELAS_A'
            ws['DI1'] = 'NAMA SEKOLAH_A'
            ws['DJ1'] = 'LOKASI_A'

            ws['DK1'] = 'MAW_A'
            ws['DL1'] = 'MAP_A'
            ws['DM1'] = 'IND_A'
            ws['DN1'] = 'ENG_A'
            ws['DO1'] = 'SEJ_A'
            ws['DP1'] = 'GEO_A'
            ws['DQ1'] = 'EKO_A'
            ws['DR1'] = 'SOS_A'
            ws['DS1'] = 'FIS_A'
            ws['DT1'] = 'KIM_A'
            ws['DU1'] = 'BIO_A'
            ws['DV1'] = 'JML_A'

            ws['DW1'] = 'Z_MAW_A'
            ws['DX1'] = 'Z_MAP_A'
            ws['DY1'] = 'Z_IND_A'
            ws['DZ1'] = 'Z_ENG_A'
            ws['EA1'] = 'Z_SEJ_A'
            ws['EB1'] = 'Z_GEO_A'
            ws['EC1'] = 'Z_EKO_A'
            ws['ED1'] = 'Z_SOS_A'
            ws['EE1'] = 'Z_FIS_A'
            ws['EF1'] = 'Z_KIM_A'
            ws['EG1'] = 'Z_BIO_A'

            ws['EH1'] = 'S_MAW_A'
            ws['EI1'] = 'S_MAP_A'
            ws['EJ1'] = 'S_IND_A'
            ws['EK1'] = 'S_ENG_A'
            ws['EL1'] = 'S_SEJ_A'
            ws['EM1'] = 'S_GEO_A'
            ws['EN1'] = 'S_EKO_A'
            ws['EO1'] = 'S_SOS_A'
            ws['EP1'] = 'S_FIS_A'
            ws['EQ1'] = 'S_KIM_A'
            ws['ER1'] = 'S_BIO_A'
            ws['ES1'] = 'S_JML_A'

            ws['ET1'] = 'RANK NAS._A'
            ws['EU1'] = 'RANK LOK._A'

            ws['DW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['DZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ED1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ER1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ES1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ET1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EU1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['DF1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DG1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DH1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DI1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DJ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DK1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DL1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DM1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DN1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DO1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DP1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DQ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DR1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DS1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DT1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DU1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DV1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DW1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DX1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DY1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['DZ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EA1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EB1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EC1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['ED1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EE1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EF1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EG1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EH1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EI1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EJ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EK1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EL1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EM1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EN1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EO1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EP1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EQ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['ER1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['ES1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['ET1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EU1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')

            # tambahan
            ws['EV1'] = 'MAW_20_A'
            ws['EW1'] = 'MAP_20_A'
            ws['EX1'] = 'IND_20_A'
            ws['EY1'] = 'ENG_20_A'
            ws['EZ1'] = 'SEJ_20_A'
            ws['FA1'] = 'GEO_20_A'
            ws['FB1'] = 'EKO_20_A'
            ws['FC1'] = 'SOS_20_A'
            ws['FD1'] = 'FIS_20_A'
            ws['FE1'] = 'KIM_20_A'
            ws['FF1'] = 'BIO_20_A'

            ws['EV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['EZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FF1'].font = Font(bold=False, name='Calibri', size=11)

            ws['EV1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EW1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EX1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EY1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['EZ1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FA1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FB1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FC1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FD1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FE1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')
            ws['FF1'].fill = PatternFill(
                fill_type='solid', start_color='A1C298', end_color='A1C298')

            for row in range(2, q+1):
                ws['DF{}'.format(row)] = '=B{}'.format(row)
                ws['DG{}'.format(row)] = '=C{}'.format(row, row)
                ws['DH{}'.format(row)] = '=D{}'.format(row, row)
                ws['DI{}'.format(row)] = '=E{}'.format(row, row)
                ws['DJ{}'.format(row)] = '=F{}'.format(row, row)

                ws['DK{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['DL{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['DM{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['DN{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['DO{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['DP{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['DQ{}'.format(row)] = '=IF(M{}="","",M{})'.format(row, row)
                ws['DR{}'.format(row)] = '=IF(N{}="","",N{})'.format(row, row)
                ws['DS{}'.format(row)] = '=IF(O{}="","",O{})'.format(row, row)
                ws['DT{}'.format(row)] = '=IF(P{}="","",P{})'.format(row, row)
                ws['DU{}'.format(row)] = '=IF(Q{}="","",Q{})'.format(row, row)
                ws['DV{}'.format(row)] = '=IF(R{}="","",R{})'.format(row, row)
                ws['DW{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DK{}="","",(DK{}-DK${})/DK${}),2),"")'.format(row, row, r, s)
                ws['DX{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DL{}="","",(DL{}-DL${})/DL${}),2),"")'.format(row, row, r, s)
                ws['DY{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DM{}="","",(DM{}-DM${})/DM${}),2),"")'.format(row, row, r, s)
                ws['DZ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DN{}="","",(DN{}-DN${})/DN${}),2),"")'.format(row, row, r, s)
                ws['EA{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DO{}="","",(DO{}-DO${})/DO${}),2),"")'.format(row, row, r, s)
                ws['EB{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DP{}="","",(DP{}-DP${})/DP${}),2),"")'.format(row, row, r, s)
                ws['EC{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DQ{}="","",(DQ{}-DQ${})/DQ${}),2),"")'.format(row, row, r, s)
                ws['ED{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DR{}="","",(DR{}-DR${})/DR${}),2),"")'.format(row, row, r, s)
                ws['EE{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DS{}="","",(DS{}-DS${})/DS${}),2),"")'.format(row, row, r, s)
                ws['EF{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DT{}="","",(DT{}-DT${})/DT${}),2),"")'.format(row, row, r, s)
                ws['EG{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DU{}="","",(DU{}-DU${})/DU${}),2),"")'.format(row, row, r, s)

                ws['EH{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DK{}="","",IF(70+30*DW{}/$DW${}<20,20,70+30*DW{}/$DW${})),2),"")'.format(row, row, r, row, r)
                ws['EI{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DL{}="","",IF(70+30*DX{}/$DX${}<20,20,70+30*DX{}/$DX${})),2),"")'.format(row, row, r, row, r)
                ws['EJ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DM{}="","",IF(70+30*DY{}/$DY${}<20,20,70+30*DY{}/$DY${})),2),"")'.format(row, row, r, row, r)
                ws['EK{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DN{}="","",IF(70+30*DZ{}/$DZ${}<20,20,70+30*DZ{}/$DZ${})),2),"")'.format(row, row, r, row, r)
                ws['EL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DO{}="","",IF(70+30*EA{}/$EA${}<20,20,70+30*EA{}/$EA${})),2),"")'.format(row, row, r, row, r)
                ws['EM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DP{}="","",IF(70+30*EB{}/$EB${}<20,20,70+30*EB{}/$EB${})),2),"")'.format(row, row, r, row, r)
                ws['EN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DQ{}="","",IF(70+30*EC{}/$EC${}<20,20,70+30*EC{}/$EC${})),2),"")'.format(row, row, r, row, r)
                ws['EO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DR{}="","",IF(70+30*ED{}/$ED${}<20,20,70+30*ED{}/$ED${})),2),"")'.format(row, row, r, row, r)
                ws['EP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DS{}="","",IF(70+30*EF{}/$EF${}<20,20,70+30*EF{}/$EF${})),2),"")'.format(row, row, r, row, r)
                ws['EQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DT{}="","",IF(70+30*EG{}/$EG${}<20,20,70+30*EG{}/$EG${})),2),"")'.format(row, row, r, row, r)
                ws['ER{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DU{}="","",IF(70+30*EH{}/$EH${}<20,20,70+30*EH{}/$EH${})),2),"")'.format(row, row, r, row, r)

                ws['ES{}'.format(row)] = '=IF(SUM(EH{}:ER{})=0,"",SUM(EH{}:ER{}))'.format(
                    row, row, row, row)
                ws['ET{}'.format(row)] = '=IF(ES{}="","",RANK(ES{},$ES$2:$ES${}))'.format(
                    row, row, q)
                ws['EU{}'.format(
                    row)] = '=IF(ET{}="","",COUNTIFS($DJ$2:$DJ${},DJ{},$ET$2:$ET${},"<"&ET{})+1)'.format(row, q, row, q, row)

            # TAMBAHAN
                ws['EV{}'.format(row)] = '=IF($G${}=20,IF(AND(G{}>3,EH{}=20),1,""),IF($G${}=25,IF(AND(G{}>4,EH{}=20),1,""),IF($G${}=30,IF(AND(G{}>5,EH{}=20),1,""),IF($G${}=35,IF(AND(G{}>6,EH{}=20),1,""),IF($G${}=40,IF(AND(G{}>7,EH{}=20),1,""),IF($G${}=45,IF(AND(G{}>8,EH{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['EW{}'.format(row)] = '=IF($H${}=20,IF(AND(H{}>3,EI{}=20),1,""),IF($H${}=25,IF(AND(H{}>4,EI{}=20),1,""),IF($H${}=30,IF(AND(H{}>5,EI{}=20),1,""),IF($H${}=35,IF(AND(H{}>6,EI{}=20),1,""),IF($H${}=40,IF(AND(H{}>7,EI{}=20),1,""),IF($H${}=45,IF(AND(H{}>8,EI{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['EX{}'.format(row)] = '=IF($I${}=20,IF(AND(I{}>3,EJ{}=20),1,""),IF($I${}=25,IF(AND(I{}>4,EJ{}=20),1,""),IF($I${}=30,IF(AND(I{}>5,EJ{}=20),1,""),IF($I${}=35,IF(AND(I{}>6,EJ{}=20),1,""),IF($I${}=40,IF(AND(I{}>7,EJ{}=20),1,""),IF($I${}=45,IF(AND(I{}>8,EJ{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['EY{}'.format(row)] = '=IF($J${}=20,IF(AND(J{}>3,EK{}=20),1,""),IF($J${}=25,IF(AND(J{}>4,EK{}=20),1,""),IF($J${}=30,IF(AND(J{}>5,EK{}=20),1,""),IF($J${}=35,IF(AND(J{}>6,EK{}=20),1,""),IF($J${}=40,IF(AND(J{}>7,EK{}=20),1,""),IF($J${}=45,IF(AND(J{}>8,EK{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['EZ{}'.format(row)] = '=IF($K${}=20,IF(AND(K{}>3,EL{}=20),1,""),IF($K${}=25,IF(AND(K{}>4,EL{}=20),1,""),IF($K${}=30,IF(AND(K{}>5,EL{}=20),1,""),IF($K${}=35,IF(AND(K{}>6,EL{}=20),1,""),IF($K${}=40,IF(AND(K{}>7,EL{}=20),1,""),IF($K${}=45,IF(AND(K{}>8,EL{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['FA{}'.format(row)] = '=IF($L${}=20,IF(AND(L{}>3,EM{}=20),1,""),IF($L${}=25,IF(AND(L{}>4,EM{}=20),1,""),IF($L${}=30,IF(AND(L{}>5,EM{}=20),1,""),IF($L${}=35,IF(AND(L{}>6,EM{}=20),1,""),IF($L${}=40,IF(AND(L{}>7,EM{}=20),1,""),IF($L${}=45,IF(AND(L{}>8,EM{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['FB{}'.format(row)] = '=IF($M${}=20,IF(AND(M{}>3,EN{}=20),1,""),IF($M${}=25,IF(AND(M{}>4,EN{}=20),1,""),IF($M${}=30,IF(AND(M{}>5,EN{}=20),1,""),IF($M${}=35,IF(AND(M{}>6,EN{}=20),1,""),IF($M${}=40,IF(AND(M{}>7,EN{}=20),1,""),IF($M${}=45,IF(AND(M{}>8,EN{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['FC{}'.format(row)] = '=IF($N${}=20,IF(AND(N{}>3,EO{}=20),1,""),IF($N${}=25,IF(AND(N{}>4,EO{}=20),1,""),IF($N${}=30,IF(AND(N{}>5,EO{}=20),1,""),IF($N${}=35,IF(AND(N{}>6,EO{}=20),1,""),IF($N${}=40,IF(AND(N{}>7,EO{}=20),1,""),IF($N${}=45,IF(AND(N{}>8,EO{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['FD{}'.format(row)] = '=IF($O${}=20,IF(AND(O{}>3,EP{}=20),1,""),IF($O${}=25,IF(AND(O{}>4,EP{}=20),1,""),IF($O${}=30,IF(AND(O{}>5,EP{}=20),1,""),IF($O${}=35,IF(AND(O{}>6,EP{}=20),1,""),IF($O${}=40,IF(AND(O{}>7,EP{}=20),1,""),IF($O${}=45,IF(AND(O{}>8,EP{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['FE{}'.format(row)] = '=IF($P${}=20,IF(AND(P{}>3,EQ{}=20),1,""),IF($P${}=25,IF(AND(P{}>4,EQ{}=20),1,""),IF($P${}=30,IF(AND(P{}>5,EQ{}=20),1,""),IF($P${}=35,IF(AND(P{}>6,EQ{}=20),1,""),IF($P${}=40,IF(AND(P{}>7,EQ{}=20),1,""),IF($P${}=45,IF(AND(P{}>8,EQ{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['FF{}'.format(row)] = '=IF($Q${}=20,IF(AND(Q{}>3,ER{}=20),1,""),IF($Q${}=25,IF(AND(Q{}>4,ER{}=20),1,""),IF($Q${}=30,IF(AND(Q{}>5,ER{}=20),1,""),IF($Q${}=35,IF(AND(Q{}>6,ER{}=20),1,""),IF($Q${}=40,IF(AND(Q{}>7,ER{}=20),1,""),IF($Q${}=45,IF(AND(Q{}>8,ER{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score [4]
            ws['FH1'] = 'NAMA SISWA_A'
            ws['FI1'] = 'NOMOR NF_A'
            ws['FJ1'] = 'KELAS_A'
            ws['FK1'] = 'NAMA SEKOLAH_A'
            ws['FL1'] = 'LOKASI_A'

            ws['FM1'] = 'MAW_A'
            ws['FN1'] = 'MAP_A'
            ws['FO1'] = 'IND_A'
            ws['FP1'] = 'ENG_A'
            ws['FQ1'] = 'SEJ_A'
            ws['FR1'] = 'GEO_A'
            ws['FS1'] = 'EKO_A'
            ws['FT1'] = 'SOS_A'
            ws['FU1'] = 'FIS_A'
            ws['FV1'] = 'KIM_A'
            ws['FW1'] = 'BIO_A'
            ws['FX1'] = 'JML_A'

            ws['FY1'] = 'Z_MAW_A'
            ws['FZ1'] = 'Z_MAP_A'
            ws['GA1'] = 'Z_IND_A'
            ws['GB1'] = 'Z_ENG_A'
            ws['GC1'] = 'Z_SEJ_A'
            ws['GD1'] = 'Z_GEO_A'
            ws['GE1'] = 'Z_EKO_A'
            ws['GF1'] = 'Z_SOS_A'
            ws['GG1'] = 'Z_FIS_A'
            ws['GH1'] = 'Z_KIM_A'
            ws['GI1'] = 'Z_BIO_A'

            ws['GJ1'] = 'S_MAW_A'
            ws['GK1'] = 'S_MAP_A'
            ws['GL1'] = 'S_IND_A'
            ws['GM1'] = 'S_ENG_A'
            ws['GN1'] = 'S_SEJ_A'
            ws['GO1'] = 'S_GEO_A'
            ws['GP1'] = 'S_EKO_A'
            ws['GQ1'] = 'S_SOS_A'
            ws['GR1'] = 'S_FIS_A'
            ws['GS1'] = 'S_KIM_A'
            ws['GT1'] = 'S_BIO_A'
            ws['GU1'] = 'S_JML_A'

            ws['GV1'] = 'RANK NAS._A'
            ws['GW1'] = 'RANK LOK._A'

            ws['FY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['FZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GW1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['FH1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['FI1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['FJ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['FK1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['FL1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['FM1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['FN1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['FO1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['FP1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['FQ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['FR1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['FS1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['FT1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['FU1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['FV1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['FW1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['FX1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['FY1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['FZ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GA1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GB1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GC1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GD1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GE1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GF1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GG1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GH1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GI1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GJ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GK1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GL1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GM1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GN1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GO1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GP1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GQ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GR1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GS1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GT1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GU1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GV1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GW1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')

            # tambahan
            ws['GX1'] = 'MAW_20_A'
            ws['GY1'] = 'MAP_20_A'
            ws['GZ1'] = 'IND_20_A'
            ws['HA1'] = 'ENG_20_A'
            ws['HB1'] = 'SEJ_20_A'
            ws['HC1'] = 'GEO_20_A'
            ws['HD1'] = 'EKO_20_A'
            ws['HE1'] = 'SOS_20_A'
            ws['HF1'] = 'FIS_20_A'
            ws['HG1'] = 'KIM_20_A'
            ws['HH1'] = 'BIO_20_A'

            ws['GX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['GZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['HH1'].font = Font(bold=False, name='Calibri', size=11)

            ws['GX1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GY1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['GZ1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HA1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HB1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HC1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HD1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HE1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HF1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HG1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')
            ws['HH1'].fill = PatternFill(
                fill_type='solid', start_color='FFE9A0', end_color='FFE9A0')

            for row in range(2, q+1):
                ws['FH{}'.format(row)] = '=B{}'.format(row)
                ws['FI{}'.format(row)] = '=C{}'.format(row, row)
                ws['FJ{}'.format(row)] = '=D{}'.format(row, row)
                ws['FK{}'.format(row)] = '=E{}'.format(row, row)
                ws['FL{}'.format(row)] = '=F{}'.format(row, row)

                ws['FM{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['FN{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['FO{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['FP{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['FQ{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['FR{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['FS{}'.format(row)] = '=IF(M{}="","",M{})'.format(row, row)
                ws['FT{}'.format(row)] = '=IF(N{}="","",N{})'.format(row, row)
                ws['FU{}'.format(row)] = '=IF(O{}="","",O{})'.format(row, row)
                ws['FV{}'.format(row)] = '=IF(P{}="","",P{})'.format(row, row)
                ws['FW{}'.format(row)] = '=IF(Q{}="","",Q{})'.format(row, row)
                ws['FX{}'.format(row)] = '=IF(R{}="","",R{})'.format(row, row)
                ws['FY{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FM{}="","",(FM{}-FM${})/FM${}),2),"")'.format(row, row, r, s)
                ws['FZ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FN{}="","",(FN{}-FN${})/FN${}),2),"")'.format(row, row, r, s)
                ws['GA{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FO{}="","",(FO{}-FO${})/FO${}),2),"")'.format(row, row, r, s)
                ws['GB{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FP{}="","",(FP{}-FP${})/FP${}),2),"")'.format(row, row, r, s)
                ws['GC{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FQ{}="","",(FQ{}-FQ${})/FQ${}),2),"")'.format(row, row, r, s)
                ws['GD{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FR{}="","",(FR{}-FR${})/FR${}),2),"")'.format(row, row, r, s)
                ws['GE{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FS{}="","",(FS{}-FS${})/FS${}),2),"")'.format(row, row, r, s)
                ws['GF{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FT{}="","",(FT{}-FT${})/FT${}),2),"")'.format(row, row, r, s)
                ws['GG{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FU{}="","",(FU{}-FU${})/FU${}),2),"")'.format(row, row, r, s)
                ws['GH{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FV{}="","",(FV{}-FV${})/FV${}),2),"")'.format(row, row, r, s)
                ws['GI{}'.format(
                    row)] = '=IFERROR(ROUND(IF(DU{}="","",(DU{}-DU${})/DU${}),2),"")'.format(row, row, r, s)

                ws['GJ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FM{}="","",IF(70+30*FY{}/$FY${}<20,20,70+30*FY{}/$FY${})),2),"")'.format(row, row, r, row, r)
                ws['GK{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FN{}="","",IF(70+30*FZ{}/$FZ${}<20,20,70+30*FZ{}/$FZ${})),2),"")'.format(row, row, r, row, r)
                ws['GL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FO{}="","",IF(70+30*GA{}/$GA${}<20,20,70+30*GA{}/$GA${})),2),"")'.format(row, row, r, row, r)
                ws['GM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FP{}="","",IF(70+30*GB{}/$GB${}<20,20,70+30*GB{}/$GB${})),2),"")'.format(row, row, r, row, r)
                ws['GN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FQ{}="","",IF(70+30*GC{}/$GC${}<20,20,70+30*GC{}/$GC${})),2),"")'.format(row, row, r, row, r)
                ws['GO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FR{}="","",IF(70+30*GD{}/$GD${}<20,20,70+30*GD{}/$GD${})),2),"")'.format(row, row, r, row, r)
                ws['GP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FS{}="","",IF(70+30*GE{}/$GE${}<20,20,70+30*GE{}/$GE${})),2),"")'.format(row, row, r, row, r)
                ws['GQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FT{}="","",IF(70+30*GF{}/$GF${}<20,20,70+30*GF{}/$GF${})),2),"")'.format(row, row, r, row, r)
                ws['GR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FU{}="","",IF(70+30*GG{}/$GG${}<20,20,70+30*GG{}/$GG${})),2),"")'.format(row, row, r, row, r)
                ws['GS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FV{}="","",IF(70+30*GH{}/$GH${}<20,20,70+30*GH{}/$GH${})),2),"")'.format(row, row, r, row, r)
                ws['GT{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FW{}="","",IF(70+30*GI{}/$GI${}<20,20,70+30*GI{}/$GI${})),2),"")'.format(row, row, r, row, r)

                ws['GU{}'.format(row)] = '=IF(SUM(GJ{}:GT{})=0,"",SUM(GJ{}:GT{}))'.format(
                    row, row, row, row)
                ws['GV{}'.format(row)] = '=IF(GU{}="","",RANK(GU{},$GU$2:$GU${}))'.format(
                    row, row, q)
                ws['GW{}'.format(
                    row)] = '=IF(GV{}="","",COUNTIFS($FL$2:$FL${},FL{},$GV$2:$GV${},"<"&GV{})+1)'.format(row, q, row, q, row)

            # TAMBAHAN
                ws['GX{}'.format(row)] = '=IF($G${}=20,IF(AND(G{}>3,GJ{}=20),1,""),IF($G${}=25,IF(AND(G{}>4,GJ{}=20),1,""),IF($G${}=30,IF(AND(G{}>5,GJ{}=20),1,""),IF($G${}=35,IF(AND(G{}>6,GJ{}=20),1,""),IF($G${}=40,IF(AND(G{}>7,GJ{}=20),1,""),IF($G${}=45,IF(AND(G{}>8,GJ{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['GY{}'.format(row)] = '=IF($H${}=20,IF(AND(H{}>3,GK{}=20),1,""),IF($H${}=25,IF(AND(H{}>4,GK{}=20),1,""),IF($H${}=30,IF(AND(H{}>5,GK{}=20),1,""),IF($H${}=35,IF(AND(H{}>6,GK{}=20),1,""),IF($H${}=40,IF(AND(H{}>7,GK{}=20),1,""),IF($H${}=45,IF(AND(H{}>8,GK{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['GZ{}'.format(row)] = '=IF($I${}=20,IF(AND(I{}>3,GL{}=20),1,""),IF($I${}=25,IF(AND(I{}>4,GL{}=20),1,""),IF($I${}=30,IF(AND(I{}>5,GL{}=20),1,""),IF($I${}=35,IF(AND(I{}>6,GL{}=20),1,""),IF($I${}=40,IF(AND(I{}>7,GL{}=20),1,""),IF($I${}=45,IF(AND(I{}>8,GL{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['HA{}'.format(row)] = '=IF($J${}=20,IF(AND(J{}>3,GM{}=20),1,""),IF($J${}=25,IF(AND(J{}>4,GM{}=20),1,""),IF($J${}=30,IF(AND(J{}>5,GM{}=20),1,""),IF($J${}=35,IF(AND(J{}>6,GM{}=20),1,""),IF($J${}=40,IF(AND(J{}>7,GM{}=20),1,""),IF($J${}=45,IF(AND(J{}>8,GM{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['HB{}'.format(row)] = '=IF($K${}=20,IF(AND(K{}>3,GN{}=20),1,""),IF($K${}=25,IF(AND(K{}>4,GN{}=20),1,""),IF($K${}=30,IF(AND(K{}>5,GN{}=20),1,""),IF($K${}=35,IF(AND(K{}>6,GN{}=20),1,""),IF($K${}=40,IF(AND(K{}>7,GN{}=20),1,""),IF($K${}=45,IF(AND(K{}>8,GN{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['HC{}'.format(row)] = '=IF($L${}=20,IF(AND(L{}>3,GO{}=20),1,""),IF($L${}=25,IF(AND(L{}>4,GO{}=20),1,""),IF($L${}=30,IF(AND(L{}>5,GO{}=20),1,""),IF($L${}=35,IF(AND(L{}>6,GO{}=20),1,""),IF($L${}=40,IF(AND(L{}>7,GO{}=20),1,""),IF($L${}=45,IF(AND(L{}>8,GO{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['HD{}'.format(row)] = '=IF($M${}=20,IF(AND(M{}>3,GP{}=20),1,""),IF($M${}=25,IF(AND(M{}>4,GP{}=20),1,""),IF($M${}=30,IF(AND(M{}>5,GP{}=20),1,""),IF($M${}=35,IF(AND(M{}>6,GP{}=20),1,""),IF($M${}=40,IF(AND(M{}>7,GP{}=20),1,""),IF($M${}=45,IF(AND(M{}>8,GP{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['HE{}'.format(row)] = '=IF($N${}=20,IF(AND(N{}>3,GQ{}=20),1,""),IF($N${}=25,IF(AND(N{}>4,GQ{}=20),1,""),IF($N${}=30,IF(AND(N{}>5,GQ{}=20),1,""),IF($N${}=35,IF(AND(N{}>6,GQ{}=20),1,""),IF($N${}=40,IF(AND(N{}>7,GQ{}=20),1,""),IF($N${}=45,IF(AND(N{}>8,GQ{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['HF{}'.format(row)] = '=IF($O${}=20,IF(AND(O{}>3,GR{}=20),1,""),IF($O${}=25,IF(AND(O{}>4,GR{}=20),1,""),IF($O${}=30,IF(AND(O{}>5,GR{}=20),1,""),IF($O${}=35,IF(AND(O{}>6,GR{}=20),1,""),IF($O${}=40,IF(AND(O{}>7,GR{}=20),1,""),IF($O${}=45,IF(AND(O{}>8,GR{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['HG{}'.format(row)] = '=IF($P${}=20,IF(AND(P{}>3,GS{}=20),1,""),IF($P${}=25,IF(AND(P{}>4,GS{}=20),1,""),IF($P${}=30,IF(AND(P{}>5,GS{}=20),1,""),IF($P${}=35,IF(AND(P{}>6,GS{}=20),1,""),IF($P${}=40,IF(AND(P{}>7,GS{}=20),1,""),IF($P${}=45,IF(AND(P{}>8,GS{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['HH{}'.format(row)] = '=IF($Q${}=20,IF(AND(Q{}>3,GT{}=20),1,""),IF($Q${}=25,IF(AND(Q{}>4,GT{}=20),1,""),IF($Q${}=30,IF(AND(Q{}>5,GT{}=20),1,""),IF($Q${}=35,IF(AND(Q{}>6,GT{}=20),1,""),IF($Q${}=40,IF(AND(Q{}>7,GT{}=20),1,""),IF($Q${}=45,IF(AND(Q{}>8,GT{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score [5]
            ws['HJ1'] = 'NAMA SISWA_A'
            ws['HK1'] = 'NOMOR NF_A'
            ws['HL1'] = 'KELAS_A'
            ws['HM1'] = 'NAMA SEKOLAH_A'
            ws['HN1'] = 'LOKASI_A'

            ws['HO1'] = 'MAW_A'
            ws['HP1'] = 'MAP_A'
            ws['HQ1'] = 'IND_A'
            ws['HR1'] = 'ENG_A'
            ws['HS1'] = 'SEJ_A'
            ws['HT1'] = 'GEO_A'
            ws['HU1'] = 'EKO_A'
            ws['HV1'] = 'SOS_A'
            ws['HW1'] = 'FIS_A'
            ws['HX1'] = 'KIM_A'
            ws['HY1'] = 'BIO_A'
            ws['HZ1'] = 'JML_A'

            ws['IA1'] = 'Z_MAW_A'
            ws['IB1'] = 'Z_MAP_A'
            ws['IC1'] = 'Z_IND_A'
            ws['ID1'] = 'Z_ENG_A'
            ws['IE1'] = 'Z_SEJ_A'
            ws['IF1'] = 'Z_GEO_A'
            ws['IG1'] = 'Z_EKO_A'
            ws['IH1'] = 'Z_SOS_A'
            ws['II1'] = 'Z_FIS_A'
            ws['IJ1'] = 'Z_KIM_A'
            ws['IK1'] = 'Z_BIO_A'

            ws['IL1'] = 'S_MAW_A'
            ws['IM1'] = 'S_MAP_A'
            ws['IN1'] = 'S_IND_A'
            ws['IO1'] = 'S_ENG_A'
            ws['IP1'] = 'S_SEJ_A'
            ws['IQ1'] = 'S_GEO_A'
            ws['IR1'] = 'S_EKO_A'
            ws['IS1'] = 'S_SOS_A'
            ws['IT1'] = 'S_FIS_A'
            ws['IU1'] = 'S_KIM_A'
            ws['IV1'] = 'S_BIO_A'
            ws['IW1'] = 'S_JML_A'

            ws['IX1'] = 'RANK NAS._A'
            ws['IY1'] = 'RANK LOK._A'

            ws['IA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['ID1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['II1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['IY1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['HJ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['HK1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['HL1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['HM1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['HN1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['HO1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['HP1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['HQ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['HR1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['HS1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['HT1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['HU1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['HV1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['HW1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['HX1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['HY1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['HZ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IA1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IB1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IC1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['ID1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IE1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IF1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IG1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IH1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['II1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IJ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IK1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IL1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IM1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IN1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IO1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IP1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IQ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IR1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IS1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IT1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IU1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IV1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IW1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IX1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['IY1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')

            # tambahan
            ws['IZ1'] = 'MAW_20_A'
            ws['JA1'] = 'MAP_20_A'
            ws['JB1'] = 'IND_20_A'
            ws['JC1'] = 'ENG_20_A'
            ws['JD1'] = 'SEJ_20_A'
            ws['JE1'] = 'GEO_20_A'
            ws['JF1'] = 'EKO_20_A'
            ws['JG1'] = 'SOS_20_A'
            ws['JH1'] = 'FIS_20_A'
            ws['JI1'] = 'KIM_20_A'
            ws['JJ1'] = 'BIO_20_A'

            ws['IZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['JA1'].font = Font(bold=False, name='Calibri', size=11)
            ws['JB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['JC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['JD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['JE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['JF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['JG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['JH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['JI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['JJ1'].font = Font(bold=False, name='Calibri', size=11)

            ws['IZ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JA1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JB1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JC1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JD1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JE1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JF1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JG1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JH1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JI1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')
            ws['JJ1'].fill = PatternFill(
                fill_type='solid', start_color='ECC5FB', end_color='ECC5FB')

            for row in range(2, q+1):
                ws['HJ{}'.format(row)] = '=B{}'.format(row)
                ws['HK{}'.format(row)] = '=C{}'.format(row, row)
                ws['HL{}'.format(row)] = '=D{}'.format(row, row)
                ws['HM{}'.format(row)] = '=E{}'.format(row, row)
                ws['HN{}'.format(row)] = '=F{}'.format(row, row)

                ws['HO{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['HP{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['HQ{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['HR{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['HS{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['HT{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['HU{}'.format(row)] = '=IF(M{}="","",M{})'.format(row, row)
                ws['HV{}'.format(row)] = '=IF(N{}="","",N{})'.format(row, row)
                ws['HW{}'.format(row)] = '=IF(O{}="","",O{})'.format(row, row)
                ws['HX{}'.format(row)] = '=IF(P{}="","",P{})'.format(row, row)
                ws['HY{}'.format(row)] = '=IF(Q{}="","",Q{})'.format(row, row)
                ws['HZ{}'.format(row)] = '=IF(R{}="","",R{})'.format(row, row)
                ws['IA{}'.format(
                    row)] = '=IFERROR(ROUND(IF(HO{}="","",(HO{}-HO${})/HO${}),2),"")'.format(row, row, r, s)
                ws['IB{}'.format(
                    row)] = '=IFERROR(ROUND(IF(HP{}="","",(HP{}-HP${})/HP${}),2),"")'.format(row, row, r, s)
                ws['IC{}'.format(
                    row)] = '=IFERROR(ROUND(IF(HQ{}="","",(HQ{}-HQ${})/HQ${}),2),"")'.format(row, row, r, s)
                ws['ID{}'.format(
                    row)] = '=IFERROR(ROUND(IF(HR{}="","",(HR{}-HR${})/HR${}),2),"")'.format(row, row, r, s)
                ws['IE{}'.format(
                    row)] = '=IFERROR(ROUND(IF(HS{}="","",(HS{}-HS${})/HS${}),2),"")'.format(row, row, r, s)
                ws['IF{}'.format(
                    row)] = '=IFERROR(ROUND(IF(HT{}="","",(HT{}-HT${})/HT${}),2),"")'.format(row, row, r, s)
                ws['IG{}'.format(
                    row)] = '=IFERROR(ROUND(IF(HU{}="","",(HU{}-HU${})/HU${}),2),"")'.format(row, row, r, s)
                ws['IH{}'.format(
                    row)] = '=IFERROR(ROUND(IF(HV{}="","",(HV{}-HV${})/HV${}),2),"")'.format(row, row, r, s)
                ws['II{}'.format(
                    row)] = '=IFERROR(ROUND(IF(HW{}="","",(HW{}-HW${})/HW${}),2),"")'.format(row, row, r, s)
                ws['IJ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(HX{}="","",(HX{}-HX${})/HX${}),2),"")'.format(row, row, r, s)
                ws['IK{}'.format(
                    row)] = '=IFERROR(ROUND(IF(HY{}="","",(HY{}-HY${})/HY${}),2),"")'.format(row, row, r, s)

                ws['IL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(HO{}="","",IF(70+30*IA{}/$IA${}<20,20,70+30*IA{}/$IA${})),2),"")'.format(row, row, r, row, r)
                ws['IM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(HP{}="","",IF(70+30*IB{}/$IB${}<20,20,70+30*IB{}/$IB${})),2),"")'.format(row, row, r, row, r)
                ws['IN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(HQ{}="","",IF(70+30*IC{}/$IC${}<20,20,70+30*IC{}/$IC${})),2),"")'.format(row, row, r, row, r)
                ws['IO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(HR{}="","",IF(70+30*ID{}/$ID${}<20,20,70+30*ID{}/$ID${})),2),"")'.format(row, row, r, row, r)
                ws['IP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(HS{}="","",IF(70+30*IE{}/$IE${}<20,20,70+30*IE{}/$IE${})),2),"")'.format(row, row, r, row, r)
                ws['IQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(HT{}="","",IF(70+30*IF{}/$IF${}<20,20,70+30*IF{}/$IF${})),2),"")'.format(row, row, r, row, r)
                ws['IR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(HU{}="","",IF(70+30*IG{}/$IG${}<20,20,70+30*IG{}/$IG${})),2),"")'.format(row, row, r, row, r)
                ws['IS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(HV{}="","",IF(70+30*IH{}/$IH${}<20,20,70+30*IH{}/$IH${})),2),"")'.format(row, row, r, row, r)
                ws['IT{}'.format(
                    row)] = '=IFERROR(ROUND(IF(HW{}="","",IF(70+30*II{}/$II${}<20,20,70+30*II{}/$II${})),2),"")'.format(row, row, r, row, r)
                ws['IU{}'.format(
                    row)] = '=IFERROR(ROUND(IF(HX{}="","",IF(70+30*IJ{}/$IJ${}<20,20,70+30*IJ{}/$IJ${})),2),"")'.format(row, row, r, row, r)
                ws['IV{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FW{}="","",IF(70+30*IK{}/$IK${}<20,20,70+30*IK{}/$IK${})),2),"")'.format(row, row, r, row, r)

                ws['IW{}'.format(row)] = '=IF(SUM(IL{}:IV{})=0,"",SUM(IL{}:IV{}))'.format(
                    row, row, row, row)
                ws['IX{}'.format(row)] = '=IF(IW{}="","",RANK(IW{},$IW$2:$IW${}))'.format(
                    row, row, q)
                ws['IY{}'.format(
                    row)] = '=IF(IX{}="","",COUNTIFS($HN$2:$HN${},HN{},$IX$2:$IX${},"<"&IX{})+1)'.format(row, q, row, q, row)

            # TAMBAHAN
                ws['IZ{}'.format(row)] = '=IF($G${}=20,IF(AND(G{}>3,IL{}=20),1,""),IF($G${}=25,IF(AND(G{}>4,IL{}=20),1,""),IF($G${}=30,IF(AND(G{}>5,IL{}=20),1,""),IF($G${}=35,IF(AND(G{}>6,IL{}=20),1,""),IF($G${}=40,IF(AND(G{}>7,IL{}=20),1,""),IF($G${}=45,IF(AND(G{}>8,IL{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['JA{}'.format(row)] = '=IF($H${}=20,IF(AND(H{}>3,IM{}=20),1,""),IF($H${}=25,IF(AND(H{}>4,IM{}=20),1,""),IF($H${}=30,IF(AND(H{}>5,IM{}=20),1,""),IF($H${}=35,IF(AND(H{}>6,IM{}=20),1,""),IF($H${}=40,IF(AND(H{}>7,IM{}=20),1,""),IF($H${}=45,IF(AND(H{}>8,IM{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['JB{}'.format(row)] = '=IF($I${}=20,IF(AND(I{}>3,IN{}=20),1,""),IF($I${}=25,IF(AND(I{}>4,IN{}=20),1,""),IF($I${}=30,IF(AND(I{}>5,IN{}=20),1,""),IF($I${}=35,IF(AND(I{}>6,IN{}=20),1,""),IF($I${}=40,IF(AND(I{}>7,IN{}=20),1,""),IF($I${}=45,IF(AND(I{}>8,IN{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['JC{}'.format(row)] = '=IF($J${}=20,IF(AND(J{}>3,IO{}=20),1,""),IF($J${}=25,IF(AND(J{}>4,IO{}=20),1,""),IF($J${}=30,IF(AND(J{}>5,IO{}=20),1,""),IF($J${}=35,IF(AND(J{}>6,IO{}=20),1,""),IF($J${}=40,IF(AND(J{}>7,IO{}=20),1,""),IF($J${}=45,IF(AND(J{}>8,IO{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['JD{}'.format(row)] = '=IF($K${}=20,IF(AND(K{}>3,IP{}=20),1,""),IF($K${}=25,IF(AND(K{}>4,IP{}=20),1,""),IF($K${}=30,IF(AND(K{}>5,IP{}=20),1,""),IF($K${}=35,IF(AND(K{}>6,IP{}=20),1,""),IF($K${}=40,IF(AND(K{}>7,IP{}=20),1,""),IF($K${}=45,IF(AND(K{}>8,IP{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['JE{}'.format(row)] = '=IF($L${}=20,IF(AND(L{}>3,IQ{}=20),1,""),IF($L${}=25,IF(AND(L{}>4,IQ{}=20),1,""),IF($L${}=30,IF(AND(L{}>5,IQ{}=20),1,""),IF($L${}=35,IF(AND(L{}>6,IQ{}=20),1,""),IF($L${}=40,IF(AND(L{}>7,IQ{}=20),1,""),IF($L${}=45,IF(AND(L{}>8,IQ{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['JF{}'.format(row)] = '=IF($M${}=20,IF(AND(M{}>3,IR{}=20),1,""),IF($M${}=25,IF(AND(M{}>4,IR{}=20),1,""),IF($M${}=30,IF(AND(M{}>5,IR{}=20),1,""),IF($M${}=35,IF(AND(M{}>6,IR{}=20),1,""),IF($M${}=40,IF(AND(M{}>7,IR{}=20),1,""),IF($M${}=45,IF(AND(M{}>8,IR{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['JG{}'.format(row)] = '=IF($N${}=20,IF(AND(N{}>3,IS{}=20),1,""),IF($N${}=25,IF(AND(N{}>4,IS{}=20),1,""),IF($N${}=30,IF(AND(N{}>5,IS{}=20),1,""),IF($N${}=35,IF(AND(N{}>6,IS{}=20),1,""),IF($N${}=40,IF(AND(N{}>7,IS{}=20),1,""),IF($N${}=45,IF(AND(N{}>8,IS{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['JH{}'.format(row)] = '=IF($O${}=20,IF(AND(O{}>3,IT{}=20),1,""),IF($O${}=25,IF(AND(O{}>4,IT{}=20),1,""),IF($O${}=30,IF(AND(O{}>5,IT{}=20),1,""),IF($O${}=35,IF(AND(O{}>6,IT{}=20),1,""),IF($O${}=40,IF(AND(O{}>7,IT{}=20),1,""),IF($O${}=45,IF(AND(O{}>8,IT{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['JI{}'.format(row)] = '=IF($P${}=20,IF(AND(P{}>3,IU{}=20),1,""),IF($P${}=25,IF(AND(P{}>4,IU{}=20),1,""),IF($P${}=30,IF(AND(P{}>5,IU{}=20),1,""),IF($P${}=35,IF(AND(P{}>6,IU{}=20),1,""),IF($P${}=40,IF(AND(P{}>7,IU{}=20),1,""),IF($P${}=45,IF(AND(P{}>8,IU{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['JJ{}'.format(row)] = '=IF($Q${}=20,IF(AND(Q{}>3,IV{}=20),1,""),IF($Q${}=25,IF(AND(Q{}>4,IV{}=20),1,""),IF($Q${}=30,IF(AND(Q{}>5,IV{}=20),1,""),IF($Q${}=35,IF(AND(Q{}>6,IV{}=20),1,""),IF($Q${}=40,IF(AND(Q{}>7,IV{}=20),1,""),IF($Q${}=45,IF(AND(Q{}>8,IV{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Z Score
            ws['JL1'] = 'NAMA SISWA'
            ws['JM1'] = 'NOMOR NF'
            ws['JN1'] = 'KELAS'
            ws['JO1'] = 'NAMA SEKOLAH'
            ws['JP1'] = 'LOKASI'

            ws['JQ1'] = 'MAW'
            ws['JR1'] = 'MAP'
            ws['JS1'] = 'IND'
            ws['JT1'] = 'ENG'
            ws['JU1'] = 'SEJ'
            ws['JV1'] = 'GEO'
            ws['JW1'] = 'EKO'
            ws['JX1'] = 'SOS'
            ws['JY1'] = 'FIS'
            ws['JZ1'] = 'KIM'
            ws['KA1'] = 'BIO'
            ws['KB1'] = 'JML'

            ws['KC1'] = 'Z_MAW'
            ws['KD1'] = 'Z_MAP'
            ws['KE1'] = 'Z_IND'
            ws['KF1'] = 'Z_ENG'
            ws['KG1'] = 'Z_SEJ'
            ws['KH1'] = 'Z_GEO'
            ws['KI1'] = 'Z_EKO'
            ws['KJ1'] = 'Z_SOS'
            ws['KK1'] = 'Z_FIS'
            ws['KL1'] = 'Z_KIM'
            ws['KM1'] = 'Z_BIO'

            ws['KN1'] = 'S_MAW'
            ws['KO1'] = 'S_MAP'
            ws['KP1'] = 'S_IND'
            ws['KQ1'] = 'S_ENG'
            ws['KR1'] = 'S_SEJ'
            ws['KS1'] = 'S_GEO'
            ws['KT1'] = 'S_EKO'
            ws['KU1'] = 'S_SOS'
            ws['KV1'] = 'S_FIS'
            ws['KW1'] = 'S_KIM'
            ws['KX1'] = 'S_BIO'
            ws['KY1'] = 'S_JML'

            ws['KZ1'] = 'RANK NAS.'
            ws['LA1'] = 'RANK LOK.'

            ws['KC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KL1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KM1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KN1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KO1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KP1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KQ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KR1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KS1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KT1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KU1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KV1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KW1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KX1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KY1'].font = Font(bold=False, name='Calibri', size=11)
            ws['KZ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['LA1'].font = Font(bold=False, name='Calibri', size=11)

            # FILL
            ws['JL1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['JM1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['JN1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['JO1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['JP1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['JQ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['JR1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['JS1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['JT1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['JU1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['JV1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['JW1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['JX1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['JY1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['JZ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['KA1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['KB1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['KC1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['KD1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['KE1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['KF1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['KG1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['KH1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['KI1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['KJ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['KK1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['KL1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['KM1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['KN1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['KO1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['KP1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['KQ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['KR1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['KS1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['KT1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['KU1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['KV1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['KW1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['KX1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['KY1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['KZ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LA1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')

            # tambahan
            ws['LB1'] = 'MAW_20'
            ws['LC1'] = 'MAP_20'
            ws['LD1'] = 'IND_20'
            ws['LE1'] = 'ENG_20'
            ws['LF1'] = 'SEJ_20'
            ws['LG1'] = 'GEO_20'
            ws['LH1'] = 'EKO_20'
            ws['LI1'] = 'SOS_20'
            ws['LJ1'] = 'FIS_20'
            ws['LK1'] = 'KIM_20'
            ws['LL1'] = 'BIO_20'

            ws['LB1'].font = Font(bold=False, name='Calibri', size=11)
            ws['LC1'].font = Font(bold=False, name='Calibri', size=11)
            ws['LD1'].font = Font(bold=False, name='Calibri', size=11)
            ws['LE1'].font = Font(bold=False, name='Calibri', size=11)
            ws['LF1'].font = Font(bold=False, name='Calibri', size=11)
            ws['LG1'].font = Font(bold=False, name='Calibri', size=11)
            ws['LH1'].font = Font(bold=False, name='Calibri', size=11)
            ws['LI1'].font = Font(bold=False, name='Calibri', size=11)
            ws['LJ1'].font = Font(bold=False, name='Calibri', size=11)
            ws['LK1'].font = Font(bold=False, name='Calibri', size=11)
            ws['LL1'].font = Font(bold=False, name='Calibri', size=11)

            ws['LB1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LC1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LD1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LE1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LF1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LG1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LH1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LI1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LJ1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LK1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')
            ws['LL1'].fill = PatternFill(
                fill_type='solid', start_color='E1FFEE', end_color='E1FFEE')

            for row in range(2, q+1):
                ws['JL{}'.format(row)] = '=B{}'.format(row)
                ws['JM{}'.format(row)] = '=C{}'.format(row, row)
                ws['JN{}'.format(row)] = '=D{}'.format(row, row)
                ws['JO{}'.format(row)] = '=E{}'.format(row, row)
                ws['JP{}'.format(row)] = '=F{}'.format(row, row)

                ws['JQ{}'.format(row)] = '=IF(G{}="","",G{})'.format(row, row)
                ws['JR{}'.format(row)] = '=IF(H{}="","",H{})'.format(row, row)
                ws['JS{}'.format(row)] = '=IF(I{}="","",I{})'.format(row, row)
                ws['JT{}'.format(row)] = '=IF(J{}="","",J{})'.format(row, row)
                ws['JU{}'.format(row)] = '=IF(K{}="","",K{})'.format(row, row)
                ws['JV{}'.format(row)] = '=IF(L{}="","",L{})'.format(row, row)
                ws['JW{}'.format(row)] = '=IF(M{}="","",M{})'.format(row, row)
                ws['JX{}'.format(row)] = '=IF(N{}="","",N{})'.format(row, row)
                ws['JY{}'.format(row)] = '=IF(O{}="","",O{})'.format(row, row)
                ws['JZ{}'.format(row)] = '=IF(P{}="","",P{})'.format(row, row)
                ws['KA{}'.format(row)] = '=IF(Q{}="","",Q{})'.format(row, row)
                ws['KB{}'.format(row)] = '=IF(R{}="","",R{})'.format(row, row)
                ws['KC{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JQ{}="","",(JQ{}-JQ${})/JQ${}),2),"")'.format(row, row, r, s)
                ws['KD{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JR{}="","",(JR{}-JR${})/JR${}),2),"")'.format(row, row, r, s)
                ws['KE{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JS{}="","",(JS{}-JS${})/JS${}),2),"")'.format(row, row, r, s)
                ws['KF{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JT{}="","",(JT{}-JT${})/JT${}),2),"")'.format(row, row, r, s)
                ws['KG{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JU{}="","",(JU{}-JU${})/JU${}),2),"")'.format(row, row, r, s)
                ws['KH{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JV{}="","",(JV{}-JV${})/JV${}),2),"")'.format(row, row, r, s)
                ws['KI{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JW{}="","",(JW{}-JW${})/JW${}),2),"")'.format(row, row, r, s)
                ws['KJ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JX{}="","",(JX{}-JX${})/JX${}),2),"")'.format(row, row, r, s)
                ws['KK{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JY{}="","",(JY{}-JY${})/JY${}),2),"")'.format(row, row, r, s)
                ws['KL{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JZ{}="","",(JZ{}-JZ${})/JZ${}),2),"")'.format(row, row, r, s)
                ws['KM{}'.format(
                    row)] = '=IFERROR(ROUND(IF(KA{}="","",(KA{}-KA${})/KA${}),2),"")'.format(row, row, r, s)

                ws['KN{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JQ{}="","",IF(70+30*KC{}/$KC${}<20,20,70+30*KC{}/$KC${})),2),"")'.format(row, row, r, row, r)
                ws['KO{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JR{}="","",IF(70+30*KD{}/$KD${}<20,20,70+30*KD{}/$KD${})),2),"")'.format(row, row, r, row, r)
                ws['KP{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JS{}="","",IF(70+30*KE{}/$KE${}<20,20,70+30*KE{}/$KE${})),2),"")'.format(row, row, r, row, r)
                ws['KQ{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JT{}="","",IF(70+30*KF{}/$KF${}<20,20,70+30*KF{}/$KF${})),2),"")'.format(row, row, r, row, r)
                ws['KR{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JU{}="","",IF(70+30*KG{}/$KG${}<20,20,70+30*KG{}/$KG${})),2),"")'.format(row, row, r, row, r)
                ws['KS{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JV{}="","",IF(70+30*KH{}/$KH${}<20,20,70+30*KH{}/$KH${})),2),"")'.format(row, row, r, row, r)
                ws['KT{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JW{}="","",IF(70+30*KI{}/$KI${}<20,20,70+30*KI{}/$KI${})),2),"")'.format(row, row, r, row, r)
                ws['KU{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JX{}="","",IF(70+30*KJ{}/$KJ${}<20,20,70+30*KJ{}/$KJ${})),2),"")'.format(row, row, r, row, r)
                ws['KV{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JY{}="","",IF(70+30*KK{}/$KK${}<20,20,70+30*KK{}/$KK${})),2),"")'.format(row, row, r, row, r)
                ws['KW{}'.format(
                    row)] = '=IFERROR(ROUND(IF(JZ{}="","",IF(70+30*KL{}/$KL${}<20,20,70+30*KL{}/$KL${})),2),"")'.format(row, row, r, row, r)
                ws['KX{}'.format(
                    row)] = '=IFERROR(ROUND(IF(FW{}="","",IF(70+30*KM{}/$KM${}<20,20,70+30*KM{}/$KM${})),2),"")'.format(row, row, r, row, r)

                ws['KY{}'.format(row)] = '=IF(SUM(KN{}:KX{})=0,"",SUM(KN{}:KX{}))'.format(
                    row, row, row, row)
                ws['KZ{}'.format(row)] = '=IF(KY{}="","",RANK(KY{},$KY$2:$KY${}))'.format(
                    row, row, q)
                ws['LA{}'.format(
                    row)] = '=IF(KZ{}="","",COUNTIFS($JP$2:$JP${},JP{},$KZ$2:$KZ${},"<"&KZ{})+1)'.format(row, q, row, q, row)

            # TAMBAHAN
                ws['LB{}'.format(row)] = '=IF($G${}=20,IF(AND(G{}>3,KN{}=20),1,""),IF($G${}=25,IF(AND(G{}>4,KN{}=20),1,""),IF($G${}=30,IF(AND(G{}>5,KN{}=20),1,""),IF($G${}=35,IF(AND(G{}>6,KN{}=20),1,""),IF($G${}=40,IF(AND(G{}>7,KN{}=20),1,""),IF($G${}=45,IF(AND(G{}>8,KN{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['LC{}'.format(row)] = '=IF($H${}=20,IF(AND(H{}>3,KO{}=20),1,""),IF($H${}=25,IF(AND(H{}>4,KO{}=20),1,""),IF($H${}=30,IF(AND(H{}>5,KO{}=20),1,""),IF($H${}=35,IF(AND(H{}>6,KO{}=20),1,""),IF($H${}=40,IF(AND(H{}>7,KO{}=20),1,""),IF($H${}=45,IF(AND(H{}>8,KO{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['LD{}'.format(row)] = '=IF($I${}=20,IF(AND(I{}>3,KP{}=20),1,""),IF($I${}=25,IF(AND(I{}>4,KP{}=20),1,""),IF($I${}=30,IF(AND(I{}>5,KP{}=20),1,""),IF($I${}=35,IF(AND(I{}>6,KP{}=20),1,""),IF($I${}=40,IF(AND(I{}>7,KP{}=20),1,""),IF($I${}=45,IF(AND(I{}>8,KP{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['LE{}'.format(row)] = '=IF($J${}=20,IF(AND(J{}>3,KQ{}=20),1,""),IF($J${}=25,IF(AND(J{}>4,KQ{}=20),1,""),IF($J${}=30,IF(AND(J{}>5,KQ{}=20),1,""),IF($J${}=35,IF(AND(J{}>6,KQ{}=20),1,""),IF($J${}=40,IF(AND(J{}>7,KQ{}=20),1,""),IF($J${}=45,IF(AND(J{}>8,KQ{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['LF{}'.format(row)] = '=IF($K${}=20,IF(AND(K{}>3,KR{}=20),1,""),IF($K${}=25,IF(AND(K{}>4,KR{}=20),1,""),IF($K${}=30,IF(AND(K{}>5,KR{}=20),1,""),IF($K${}=35,IF(AND(K{}>6,KR{}=20),1,""),IF($K${}=40,IF(AND(K{}>7,KR{}=20),1,""),IF($K${}=45,IF(AND(K{}>8,KR{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['LG{}'.format(row)] = '=IF($L${}=20,IF(AND(L{}>3,KS{}=20),1,""),IF($L${}=25,IF(AND(L{}>4,KS{}=20),1,""),IF($L${}=30,IF(AND(L{}>5,KS{}=20),1,""),IF($L${}=35,IF(AND(L{}>6,KS{}=20),1,""),IF($L${}=40,IF(AND(L{}>7,KS{}=20),1,""),IF($L${}=45,IF(AND(L{}>8,KS{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['LH{}'.format(row)] = '=IF($M${}=20,IF(AND(M{}>3,KT{}=20),1,""),IF($M${}=25,IF(AND(M{}>4,KT{}=20),1,""),IF($M${}=30,IF(AND(M{}>5,KT{}=20),1,""),IF($M${}=35,IF(AND(M{}>6,KT{}=20),1,""),IF($M${}=40,IF(AND(M{}>7,KT{}=20),1,""),IF($M${}=45,IF(AND(M{}>8,KT{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['LI{}'.format(row)] = '=IF($N${}=20,IF(AND(N{}>3,KU{}=20),1,""),IF($N${}=25,IF(AND(N{}>4,KU{}=20),1,""),IF($N${}=30,IF(AND(N{}>5,KU{}=20),1,""),IF($N${}=35,IF(AND(N{}>6,KU{}=20),1,""),IF($N${}=40,IF(AND(N{}>7,KU{}=20),1,""),IF($N${}=45,IF(AND(N{}>8,KU{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['LJ{}'.format(row)] = '=IF($O${}=20,IF(AND(O{}>3,KV{}=20),1,""),IF($O${}=25,IF(AND(O{}>4,KV{}=20),1,""),IF($O${}=30,IF(AND(O{}>5,KV{}=20),1,""),IF($O${}=35,IF(AND(O{}>6,KV{}=20),1,""),IF($O${}=40,IF(AND(O{}>7,KV{}=20),1,""),IF($O${}=45,IF(AND(O{}>8,KV{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['LK{}'.format(row)] = '=IF($P${}=20,IF(AND(P{}>3,KW{}=20),1,""),IF($P${}=25,IF(AND(P{}>4,KW{}=20),1,""),IF($P${}=30,IF(AND(P{}>5,KW{}=20),1,""),IF($P${}=35,IF(AND(P{}>6,KW{}=20),1,""),IF($P${}=40,IF(AND(P{}>7,KW{}=20),1,""),IF($P${}=45,IF(AND(P{}>8,KW{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)
                ws['LL{}'.format(row)] = '=IF($Q${}=20,IF(AND(Q{}>3,KX{}=20),1,""),IF($Q${}=25,IF(AND(Q{}>4,KX{}=20),1,""),IF($Q${}=30,IF(AND(Q{}>5,KX{}=20),1,""),IF($Q${}=35,IF(AND(Q{}>6,KX{}=20),1,""),IF($Q${}=40,IF(AND(Q{}>7,KX{}=20),1,""),IF($Q${}=45,IF(AND(Q{}>8,KX{}=20),1,"")))))))'.format(
                    v, row, row, v, row, row, v, row, row, v, row, row, v, row, row, v, row, row)

            # Mengubah 'KELAS' sesuai dengan nilai yang dipilih dari selectbox 'KELAS'
            kelas = KELAS.lower().replace(" ", "")
            semester = SEMESTER.lower()
            tahun = TAHUN.replace("-", "")
            penilaian = PENILAIAN.lower()
            kurikulum = KURIKULUM.lower()

            path_file = f"{kelas}_{penilaian}_{semester}_{kurikulum}_{tahun}_nilai_std.xlsx"

            # Simpan file ke direktori temporer
            temp_dir = tempfile.gettempdir()
            file_path = temp_dir + '/' + path_file
            wb.save(file_path)

            st.success(
                "File siap diunduh!")

            # Tombol unduh file
            with open(file_path, "rb") as f:
                bytes_data = f.read()
            st.download_button(label="Unduh File", data=bytes_data,
                               file_name=path_file)

            st.warning(
                "Buka file unduhan, klik 'Enable Editing' dan 'Save'")
